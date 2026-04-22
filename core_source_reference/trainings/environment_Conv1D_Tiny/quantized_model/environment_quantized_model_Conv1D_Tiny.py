"""
Script: environment_quantized_model_Conv1D_Tiny.py
Module role:
    Train, convert, evaluate, and export the quantized Conv1D Tiny artifacts
    used in the LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, performs quantization-aware training, converts the
    model to TensorFlow Lite, evaluates the resulting artifact in normalized
    and reconstructed physical domains, and exports reference tables, figures,
    and firmware-support artifacts.

Inputs:
    - environment_dataset_Conv1D_Tiny.csv
    - Baseline and pruned model artifacts
    - Project path and versioning utilities

Outputs:
    - Versioned quantized TensorFlow Lite artifact
    - Rolling24 evaluation tables and reference spreadsheets
    - Debug workbooks, replay headers, and figure files

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys, random
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.model_selection import train_test_split
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr
import tensorflow_model_optimization as tfmot

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
rcParams["font.family"] = "Segoe UI Emoji"

# --- Bootstrap: allows importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base)
            break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_Conv1D_Tiny import (
    PROJECT_ROOT,
    DATASET_ENVIRONMENT,
    BASE_MODEL,
    PRUNED_MODEL,
    QUANTIZED_MODEL,
    QUANTIZED_MODEL_METRICS,
)
from utils.global_utils.versioning import (
    create_versioned_dir,
    ensure_dir,
    update_latest,
    write_manifest,
    list_runs,
    resolve_run,
    resolve_latest,
)
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Call at the top of the script, BEFORE model creation

# === Versioned directories for the current execution ===
run_dir = create_versioned_dir(QUANTIZED_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(QUANTIZED_MODEL_METRICS / run_dir.name)
# --- Standardized subdirectories ---
results_dir  = ensure_dir(metrics_run_dir / "quantization_metrics_results")
graphics_dir = ensure_dir(metrics_run_dir / "quantization_graphics")

base_version_path = resolve_latest(BASE_MODEL)
pruned_version_path = resolve_latest(PRUNED_MODEL)

# ---------------------
# Paths (Conv1D Tiny)
# ---------------------
model_path = pruned_version_path / "environment_pruned_model_Conv1D_Tiny.keras"
scaler_X_path = base_version_path / "environment_base_model_Conv1D_Tiny_scaler_X.pkl"
scaler_y_path = base_version_path / "environment_base_model_Conv1D_Tiny_scaler_y.pkl"
dataset_path = DATASET_ENVIRONMENT / "environment_dataset_Conv1D_Tiny.csv"
quantized_path = run_dir / "environment_quantized_model_Conv1D_Tiny.tflite"

# ---------------------
# Load dataset
# ---------------------
df_raw = pd.read_csv(dataset_path)

# Sort temporally to build windows and preserve the raw-real timeline
# required for the 2+47 replay/header exports.
df_raw = df_raw.sort_values("datetime").reset_index(drop=True)
df_raw["__row_id_raw"] = np.arange(len(df_raw), dtype=np.int64)

df_raw["datetime"] = pd.to_datetime(df_raw["datetime"])
for _c in ["T_out", "H_out", "T_in", "H_in"]:
    if _c in df_raw.columns:
        df_raw[_c] = pd.to_numeric(df_raw[_c], errors="coerce")

# Raw-real timeline preserved for replay/header exports.
df_raw_real = df_raw.copy()

# Processed dataframe used by the offline Conv1D Tiny pipeline.
df = df_raw.copy()

# Temporal feature engineering
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (as in Conv1D Tiny) ===
HIN_EMA_ALPHA = 0.08
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=HIN_EMA_ALPHA, adjust=False).mean()

# Cyclical features
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lags (first!)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Add lag2 only for indoor variables to keep 12 + weekday/month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)
df_processed_row_ids = df["__row_id_raw"].astype(np.int64).to_numpy()

# Separate features (12 features) 
features = [
    "T_out", "H_out",          # current outdoor
    "T_in_lag1", "H_in_lag1",  # indoor lag1
    "T_out_lag1", "H_out_lag1",# outdoor lag1
    "T_in_lag2", "H_in_lag2",  # indoor lag2
    "sin_hour", "cos_hour",    # cyclical hour
    "weekday", "month",        # calendar
]

targets = ["T_in", "H_in"]

X_all_phys = df[features].values.astype(np.float32)
y_abs_all = df[targets].values.astype(np.float32)

n_features = len(features)  # = 12

# ---------------------
# Load scalers
# ---------------------
scaler_X = joblib.load(scaler_X_path)
scaler_y = joblib.load(scaler_y_path)

# Scale X exactly as in training (feature-wise) with physical-domain clamping before the scaler
X_all_2d = X_all_phys.reshape(-1, n_features)
X_all_2d = np.clip(X_all_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_all_scaled_2d = scaler_X.transform(X_all_2d)
X_all_scaled_flat = X_all_scaled_2d

# Build the residual target (official contract: pure residual, without robust clipping)
baseline_all = df[["T_in_lag1", "H_in_lag1"]].values.astype(np.float32)
y_resid_raw_all = (y_abs_all - baseline_all).astype(np.float32)

# Scale y with the baseline scaler (the same one used in pruning)
y_scaled_all = scaler_y.transform(y_resid_raw_all)

# ---------------------
# Temporal sliding-window construction (24 h -> 1 prediction)
# ---------------------
WINDOW_STEPS = 24
N_total = len(df)
if N_total < WINDOW_STEPS:
    raise ValueError("Dataset too short for a 24 h sliding window.")

X_seq = []
y_seq = []
y_abs_seq = []
Tprev_seq = []
Hprev_seq = []

for t in range(WINDOW_STEPS - 1, N_total):
    X_window = X_all_scaled_flat[t - WINDOW_STEPS + 1: t + 1, :]
    X_seq.append(X_window)
    y_seq.append(y_scaled_all[t])
    y_abs_seq.append(y_abs_all[t])
    Tprev_seq.append(df["T_in_lag1"].values.astype(np.float32)[t])
    Hprev_seq.append(df["H_in_lag1"].values.astype(np.float32)[t])

X_all = np.stack(X_seq, axis=0).astype(np.float32)      # (N_seq, 24, 12)
y_all = np.stack(y_seq, axis=0).astype(np.float32)      # (N_seq, 2) residual target
y_abs_seq = np.stack(y_abs_seq, axis=0).astype(np.float32)
Tprev_seq = np.asarray(Tprev_seq, dtype=np.float32)
Hprev_seq = np.asarray(Hprev_seq, dtype=np.float32)

N_seq = X_all.shape[0]

# ---------------------
# Split temporal 60/20/20
# ---------------------
i1 = int(0.6 * N_seq)
i2 = int(0.8 * N_seq)
# --- idx_seq: maps each sequence (window) to the original dataframe index ---
# Sequence k corresponds to index t = (WINDOW_STEPS-1) + k in the dataframe after dropna.
idx_seq = np.arange(WINDOW_STEPS - 1, N_total, dtype=np.int64)  # len == N_seq
idx_train = idx_seq[:i1]
idx_val   = idx_seq[i1:i2]
idx_test  = idx_seq[i2:]


X_train, y_train = X_all[:i1], y_all[:i1]
X_val,   y_val   = X_all[i1:i2], y_all[i1:i2]
X_test,  y_test  = X_all[i2:],   y_all[i2:]

yabs_train = y_abs_seq[:i1]
yabs_val   = y_abs_seq[i1:i2]
yabs_test  = y_abs_seq[i2:]
Tprev_train = Tprev_seq[:i1]
Tprev_val   = Tprev_seq[i1:i2]
Tprev_test  = Tprev_seq[i2:]
Hprev_train = Hprev_seq[:i1]
Hprev_val   = Hprev_seq[i1:i2]
Hprev_test  = Hprev_seq[i2:]

# Targets multi-head
y_train_T = y_train[:, 0:1]
y_train_H = y_train[:, 1:2]
y_val_T   = y_val[:,   0:1]
y_val_H   = y_val[:,   1:2]
y_test_T  = y_test[:,  0:1]
y_test_H  = y_test[:,  1:2]

# Conv1D format
X_train_conv = X_train.reshape(-1, WINDOW_STEPS, n_features)
X_val_conv   = X_val.reshape(-1, WINDOW_STEPS, n_features)
X_test_conv  = X_test.reshape(-1, WINDOW_STEPS, n_features)

# ---------------------
# ---------------------
# Load the pruned Conv1D model
# ---------------------
with tfmot.sparsity.keras.prune_scope():
    base_model = tf.keras.models.load_model(model_path, compile=False)

output_names = base_model.output_names
print("Conv1D Tiny model outputs (quantization):", output_names)
if len(output_names) == 2:
    head_T_name, head_H_name = output_names
else:
    head_T_name, head_H_name = output_names[0], output_names[-1]

# ============================================================
# Light PRUNE fine-tuning stage (BEFORE quantization)
# ============================================================
# Objective: align the gap calculation with the PRUNE stage (history-based),
# and allow a light recovery of the pruned model on the current dataset.
PRUNE_FINETUNE_ENABLE = True
PRUNE_FINETUNE_EPOCHS = 40
PRUNE_FINETUNE_BATCH  = 128
PRUNE_FINETUNE_LR     = 1e-5

losses = {
    head_T_name: tf.keras.losses.Huber(delta=0.25),
    head_H_name: tf.keras.losses.Huber(delta=0.5),
}

#  MUST match PRUNE final weighting (strip+recovery stage)
loss_weights = {
    head_T_name: 1.0,
    head_H_name: 2.0,
}

metrics_dict = {
    head_T_name: [tf.keras.metrics.MeanAbsoluteError(name="mae_T")],
    head_H_name: [tf.keras.metrics.MeanAbsoluteError(name="mae_H")],
}

base_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=PRUNE_FINETUNE_LR),
    loss=losses,
    loss_weights=loss_weights,
    metrics=metrics_dict,
)

history_prune_ft = None
if PRUNE_FINETUNE_ENABLE:
    print(" PRUNED model fine-tuning (recovery) before quantization...")
    callbacks = [
        tf.keras.callbacks.ReduceLROnPlateau(
            monitor="val_loss", factor=0.5, patience=6, min_lr=1e-5, verbose=1
        ),
        tf.keras.callbacks.EarlyStopping(
            monitor="val_loss", patience=12, restore_best_weights=True, verbose=1
        ),
    ]
    history_prune_ft = base_model.fit(
        X_train_conv,
        {head_T_name: y_train_T, head_H_name: y_train_H},
        validation_data=(X_val_conv, {head_T_name: y_val_T, head_H_name: y_val_H}),
        epochs=PRUNE_FINETUNE_EPOCHS,
        batch_size=PRUNE_FINETUNE_BATCH,
        callbacks=callbacks,
        verbose=1,
    )

    # Save a copy of the fine-tuned PRUNE model inside run_dir (does not alter the original PRUNED_MODEL)
    try:
        pruned_ft_path = run_dir / "environment_pruned_model_Conv1D_Tiny.keras"
        base_model.save(pruned_ft_path)
        print(f" Fine-tuned PRUNE model saved: {pruned_ft_path}")
    except Exception as _e:
        print("[WARN] Could not save the fine-tuned PRUNE model:", _e)
else:
    print(" Using the already trained pruned Conv1D Tiny model (without fine-tuning).")

# Global evaluation
train_eval = base_model.evaluate(
    X_train_conv,
    {head_T_name: y_train_T, head_H_name: y_train_H},
    verbose=0,
    return_dict=True,
)
val_eval = base_model.evaluate(
    X_val_conv,
    {head_T_name: y_val_T, head_H_name: y_val_H},
    verbose=0,
    return_dict=True,
)

train_loss_global = float(train_eval["loss"])
val_loss_global   = float(val_eval["loss"])

train_mae_global = float(
    0.5 * (
        train_eval.get(f"{head_T_name}_mae_T", 0.0)
        + train_eval.get(f"{head_H_name}_mae_H", 0.0)
    )
)
val_mae_global = float(
    0.5 * (
        val_eval.get(f"{head_T_name}_mae_T", 0.0)
        + val_eval.get(f"{head_H_name}_mae_H", 0.0)
    )
)

print(f" Global evaluation - Training: loss={train_loss_global:.4f}, MAE≈{train_mae_global:.4f}")
print(f" Global evaluation - Validation:    loss={val_loss_global:.4f}, MAE≈{val_mae_global:.4f}")

# ============================================
# INT8 quantization + TFLite inference
# ============================================
X_calib = X_train_conv.astype(np.float32, copy=False)

def representative_data_gen():
    for i in range(min(500, len(X_calib))):
        yield [X_calib[i: i + 1]]

converter = tf.lite.TFLiteConverter.from_keras_model(base_model)
converter.optimizations = [tf.lite.Optimize.DEFAULT]
converter.representative_dataset = representative_data_gen
converter.target_spec.supported_ops = [tf.lite.OpsSet.TFLITE_BUILTINS_INT8]
converter.inference_input_type = tf.int8
converter.inference_output_type = tf.float32

quantized_model = converter.convert()
with open(quantized_path, "wb") as f:
    f.write(quantized_model)
print(" INT8-quantized model (weights/activations, float32 output) saved successfully:", quantized_path)

interpreter = tf.lite.Interpreter(model_path=str(quantized_path))
interpreter.allocate_tensors()
input_details = interpreter.get_input_details()
output_details = interpreter.get_output_details()

in_info = input_details[0]
in_scale, in_zp = in_info["quantization"]

print(f"[quant] Input dtype={in_info['dtype']}  scale={in_scale}  zp={in_zp}")
for i, od in enumerate(output_details):
    s, zp = od["quantization"]
    print(f"[quant] Output[{i}] name={od['name']} dtype={od['dtype']} scale={s} zp={zp}")

if in_scale == 0:
    raise RuntimeError(
        "Missing calibration at the input (scale=0). Check whether 'representative_dataset' was configured "
        "and whether the calibration data use dtype=float32 in the same domain as training."
    )

DBG_MODEL_RAW_MAX_DIMS = 8
DBG_MODEL_RAW_MAX_BYTES = 32

def _tflite_dtype_code(dtype) -> int:
    _dt = np.dtype(dtype)
    _map = {
        np.dtype(np.float32): 1,
        np.dtype(np.int32): 2,
        np.dtype(np.uint8): 3,
        np.dtype(np.int64): 4,
        np.dtype(np.str_): 5,
        np.dtype(np.bool_): 6,
        np.dtype(np.int16): 7,
        np.dtype(np.complex64): 8,
        np.dtype(np.int8): 9,
        np.dtype(np.float16): 10,
        np.dtype(np.float64): 11,
        np.dtype(np.complex128): 12,
        np.dtype(np.uint64): 13,
        np.dtype(np.uint32): 16,
        np.dtype(np.uint16): 17,
    }
    return int(_map.get(_dt, -1))


def _tflite_dtype_name(dtype) -> str:
    try:
        return np.dtype(dtype).name
    except Exception:
        return str(dtype)


def _capture_tflite_output_raw_records(max_dims: int = DBG_MODEL_RAW_MAX_DIMS,
                                       max_dump_bytes: int = DBG_MODEL_RAW_MAX_BYTES):
    records = []
    for out_idx, od in enumerate(output_details):
        y_q = interpreter.get_tensor(od["index"])
        y_arr = np.asarray(y_q)
        y_arr_c = np.ascontiguousarray(y_arr)
        raw_bytes = y_arr_c.tobytes(order="C")
        shape = list(y_arr_c.shape)
        bytes_total = int(len(raw_bytes))
        bytes_dumped = int(min(bytes_total, int(max_dump_bytes)))

        rec = {
            "out_idx": int(out_idx),
            "tensor_index": int(od["index"]),
            "type_code": _tflite_dtype_code(y_arr_c.dtype),
            "type_name": _tflite_dtype_name(y_arr_c.dtype),
            "bytes_total": bytes_total,
            "bytes_dumped": bytes_dumped,
            "dims_size": int(len(shape)),
        }

        for di in range(int(max_dims)):
            rec[f"dim{di:02d}"] = int(shape[di]) if di < len(shape) else -1

        for bi in range(int(max_dump_bytes)):
            rec[f"b{bi:02d}_hex"] = f"0x{raw_bytes[bi]:02X}" if bi < bytes_dumped else ""

        records.append(rec)

    return records


def quantize(x_f32: np.ndarray, scale: float, zp: int, dtype) -> np.ndarray:
    x_q = np.round(x_f32 / scale + zp)
    if dtype == np.int8:
        x_q = np.clip(x_q, -128, 127)
    elif dtype == np.uint8:
        x_q = np.clip(x_q, 0, 255)
    else:
        x_q = np.clip(x_q, -128, 127)
    return x_q.astype(dtype)

def dequantize_output(y_q: np.ndarray, scale: float, zp: int, dtype) -> np.ndarray:
    if dtype in (np.int8, np.uint8) and scale != 0.0:
        return (y_q.astype(np.float32) - zp) * scale
    return y_q.astype(np.float32)


def _read_tflite_output_payload_and_float():
    payload_parts = []
    float_parts = []
    for od in output_details:
        y_q = interpreter.get_tensor(od["index"])
        y_q_arr = np.asarray(y_q)
        payload_parts.append(y_q_arr)

        s, zp = od["quantization"]
        y_f_arr = dequantize_output(y_q_arr, s, zp, od["dtype"])
        float_parts.append(np.asarray(y_f_arr, dtype=np.float32))

    if len(payload_parts) == 1:
        payload_flat = np.asarray(payload_parts[0][0], dtype=np.float32).reshape(-1)
        float_flat = np.asarray(float_parts[0][0], dtype=np.float32).reshape(-1)
    else:
        payload_flat = np.asarray([
            np.asarray(part[0]).reshape(-1)[0] for part in payload_parts
        ], dtype=np.float32)
        float_flat = np.asarray([
            np.asarray(part[0], dtype=np.float32).reshape(-1)[0] for part in float_parts
        ], dtype=np.float32)

    return payload_flat.astype(np.float32, copy=False), float_flat.astype(np.float32, copy=False)


def _predict_tflite(X3D, capture_io: bool = False):
    preds = []
    captured_input_payload = []
    captured_input_float = []
    captured_output_payload = []
    captured_output_float = []
    captured_output_raw = []

    input_dtype = in_info["dtype"]
    for x in X3D:
        x_f32 = x.reshape(1, WINDOW_STEPS, X3D.shape[-1]).astype(np.float32)
        x_in = x_f32 if input_dtype == np.float32 else quantize(x_f32, in_scale, in_zp, input_dtype)
        interpreter.set_tensor(in_info["index"], x_in)
        interpreter.invoke()

        output_raw_records = _capture_tflite_output_raw_records()
        y_payload, y_float = _read_tflite_output_payload_and_float()
        preds.append(y_float)

        if capture_io:
            x_payload_flat = np.asarray(x_in[0], dtype=np.float32).reshape(-1)
            x_float_flat = np.asarray(x_f32[0], dtype=np.float32).reshape(-1)
            captured_input_payload.append(x_payload_flat.astype(np.float32, copy=False))
            captured_input_float.append(np.asarray(x_float_flat, dtype=np.float32))
            captured_output_payload.append(np.asarray(y_payload, dtype=np.float32))
            captured_output_float.append(np.asarray(y_float, dtype=np.float32))
            captured_output_raw.append(output_raw_records)

    preds = np.asarray(preds, dtype=np.float32)
    if not capture_io:
        return preds

    return {
        "preds": preds,
        "input_payload": np.asarray(captured_input_payload, dtype=np.float32),
        "input_float": np.asarray(captured_input_float, dtype=np.float32),
        "output_payload": np.asarray(captured_output_payload, dtype=np.float32),
        "output_float": np.asarray(captured_output_float, dtype=np.float32),
        "output_raw": captured_output_raw,
    }



def _dbg_f32(x) -> float:
    return float(np.asarray(x, dtype=np.float32))


def _float32_to_hex_bits(x) -> str:
    _u = np.asarray([x], dtype=np.float32).view(np.uint32)[0]
    return f"0x{int(_u):08X}"


def _build_replay_canonical_processed(raw_replay_df, ema_prev, alpha):
    """Rebuild the replay-processed path with float32 arithmetic, mirroring the
    firmware contract from the exported raw 2+47 block plus the EMA seed.
    """
    _df = raw_replay_df.copy().reset_index(drop=True)

    for _c in ["T_out", "H_out", "T_in", "H_in"]:
        _df[_c] = pd.to_numeric(_df[_c], errors="coerce").astype(np.float32)

    _dt = pd.to_datetime(_df["datetime"])
    _hour = _dt.dt.hour.astype(np.int64).to_numpy()
    _weekday = _dt.dt.weekday.astype(np.int64).to_numpy()
    _month = _dt.dt.month.astype(np.int64).to_numpy()

    _df["hour"] = _hour
    _df["weekday"] = _weekday.astype(np.float32)
    _df["month"] = _month.astype(np.float32)
    _df["sin_hour"] = np.sin(2.0 * np.pi * _hour / 24.0).astype(np.float32)
    _df["cos_hour"] = np.cos(2.0 * np.pi * _hour / 24.0).astype(np.float32)

    _alpha = np.float32(alpha)
    _one_minus = np.float32(1.0) - _alpha
    _prev = np.float32(ema_prev)
    _h_raw = _df["H_in"].to_numpy(dtype=np.float32, copy=True)
    _h_filt = np.empty_like(_h_raw, dtype=np.float32)
    for _i, _raw in enumerate(_h_raw):
        _curr = np.float32(_raw)
        _filt = np.float32(_alpha * _curr + _one_minus * _prev)
        _h_filt[_i] = _filt
        _prev = _filt
    _df["H_in"] = _h_filt

    _df["T_in_lag1"] = _df["T_in"].shift(1).astype(np.float32)
    _df["H_in_lag1"] = _df["H_in"].shift(1).astype(np.float32)
    _df["T_out_lag1"] = _df["T_out"].shift(1).astype(np.float32)
    _df["H_out_lag1"] = _df["H_out"].shift(1).astype(np.float32)
    _df["T_in_lag2"] = _df["T_in"].shift(2).astype(np.float32)
    _df["H_in_lag2"] = _df["H_in"].shift(2).astype(np.float32)

    return _df


def _extract_canonical_replay_debug_inputs(sample_df_indices, replay_processed_df, replay_processed_df_indices, raw_replay_df=None):
    """Build canonical preprocessing windows from the replay-processed source
    mirrored by firmware during DBG_MODEL_IN_CSV, while also exposing explicit
    preprocessing checkpoints for auditability.
    """
    _index_map = {int(_df_idx): int(_pos) for _pos, _df_idx in enumerate(replay_processed_df_indices)}
    _pre_raw_seqs = []
    _pre_smooth_seqs = []
    _pre_lags_seqs = []
    _pre_time_seqs = []
    _pre_phys_seqs = []
    _step_epoch_seqs = []
    _y_abs_rows = []
    _tprev_rows = []
    _hprev_rows = []

    _raw_df = raw_replay_df if raw_replay_df is not None else replay_processed_df

    for _df_idx in sample_df_indices:
        _end = _index_map.get(int(_df_idx))
        if _end is None:
            raise KeyError(f"Replay canonical index not found for df index {_df_idx}")
        if _end < WINDOW_STEPS - 1:
            raise ValueError(
                f"Replay canonical window too short for df index {_df_idx}: end={_end}, window={WINDOW_STEPS}"
            )

        _start = _end - WINDOW_STEPS + 1
        _win_proc = replay_processed_df.iloc[_start : _end + 1]
        _win_raw = _raw_df.iloc[_start : _end + 1]

        if "epoch" in _win_raw.columns:
            _step_epochs = _win_raw["epoch"].to_numpy(dtype=np.int64, copy=True)
        else:
            _step_epochs = (pd.to_datetime(_win_raw["datetime"]).astype("int64") // 10**9).to_numpy(dtype=np.int64, copy=True)

        _pre_raw_seqs.append(
            _win_raw[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_smooth_seqs.append(
            _win_proc[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_lags_seqs.append(
            _win_proc[["T_in_lag1", "H_in_lag1", "T_out_lag1", "H_out_lag1", "T_in_lag2", "H_in_lag2"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_time_seqs.append(
            np.column_stack([
                _step_epochs.astype(np.float32, copy=False),
                _win_proc["sin_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["cos_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["weekday"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["month"].to_numpy(dtype=np.float32, copy=True),
            ]).astype(np.float32, copy=False)
        )
        _pre_phys = _win_proc[features].to_numpy(dtype=np.float32, copy=True)
        _pre_phys_seqs.append(_pre_phys)
        _step_epoch_seqs.append(_step_epochs)

        _row = replay_processed_df.iloc[_end]
        _y_abs_rows.append(np.asarray([_row["T_in"], _row["H_in"]], dtype=np.float32))
        _tprev_rows.append(np.float32(_row["T_in_lag1"]))
        _hprev_rows.append(np.float32(_row["H_in_lag1"]))

    return {
        "pre_raw": np.asarray(_pre_raw_seqs, dtype=np.float32),
        "pre_smooth": np.asarray(_pre_smooth_seqs, dtype=np.float32),
        "pre_lags": np.asarray(_pre_lags_seqs, dtype=np.float32),
        "pre_time": np.asarray(_pre_time_seqs, dtype=np.float32),
        "pre_phys": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "step_epoch": np.asarray(_step_epoch_seqs, dtype=np.int64),
        "x_raw": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "y_abs": np.asarray(_y_abs_rows, dtype=np.float32),
        "tprev": np.asarray(_tprev_rows, dtype=np.float32),
        "hprev": np.asarray(_hprev_rows, dtype=np.float32),
    }


def _build_dbg_model_input_reference(sample_test_positions, sample_df_indices, sample_epochs, x_raw_seqs, x_clip_seqs, x_scaled_seqs, x_tensor_payload_seqs, x_tensor_float_seqs, y_true_abs_rows):
    rows = []
    _n_features = len(features)
    for _idx, (_test_pos, _df_idx, _epoch_dbg) in enumerate(zip(sample_test_positions, sample_df_indices, sample_epochs)):
        _x_raw = np.asarray(x_raw_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _x_clip = np.asarray(x_clip_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _x_scaled = np.asarray(x_scaled_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _x_tensor_payload = np.asarray(x_tensor_payload_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _x_tensor_float = np.asarray(x_tensor_float_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _gt = np.asarray(y_true_abs_rows[_idx], dtype=np.float32).reshape(2)

        for _step in range(WINDOW_STEPS):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "gt_Tin_true": float(_gt[0]),
                "gt_Hin_true": _dbg_f32(_gt[1]),
                "state_Tout_phys_raw": float(_x_raw[_step, 0]),
                "state_Hout_phys_raw": float(_x_raw[_step, 1]),
                "state_Tin_lag1_phys_raw": float(_x_raw[_step, 2]),
                "state_Hin_lag1_phys_raw": _dbg_f32(_x_raw[_step, 3]),
                "state_Tout_lag1_phys_raw": float(_x_raw[_step, 4]),
                "state_Hout_lag1_phys_raw": float(_x_raw[_step, 5]),
                "state_Tin_lag2_phys_raw": float(_x_raw[_step, 6]),
                "state_Hin_lag2_phys_raw": _dbg_f32(_x_raw[_step, 7]),
                "state_sin_hour": float(_x_raw[_step, 8]),
                "state_cos_hour": float(_x_raw[_step, 9]),
                "state_weekday": float(_x_raw[_step, 10]),
                "state_month": float(_x_raw[_step, 11]),
            }
            for _feat in range(_n_features):
                _value = _x_raw[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_raw"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _value = _x_clip[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_clip"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _row[f"in_f{_feat:02d}_scaled"] = float(_x_scaled[_step, _feat])
            for _feat in range(_n_features):
                _row[f"in_x{_feat:02d}_float"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_payload"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_float_prequant"] = float(_x_tensor_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_stage_reference(sample_epochs, step_epoch_seqs, data_seqs, field_names):
    rows = []
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _data = np.asarray(data_seqs[_idx], dtype=np.float32)
        if _data.ndim == 1:
            _data = _data.reshape(WINDOW_STEPS, 1)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(WINDOW_STEPS)
        for _step in range(_data.shape[0]):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _j, _name in enumerate(field_names):
                _val = _data[_step, _j]
                _row[_name] = _dbg_f32(_val)
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_sample_reference(raw_df, smooth_df, seed_rows: int = 2):
    """Build PRE_RAW/PRE_SMOOTH at firmware-equivalent logical-sample granularity."""
    if len(raw_df) != len(smooth_df):
        raise ValueError(
            f"PRE_RAW/PRE_SMOOTH export length mismatch: raw={len(raw_df)} smooth={len(smooth_df)}"
        )
    if len(raw_df) <= seed_rows:
        raise ValueError(
            f"Need more than {seed_rows} rows to build logical PRE_RAW/PRE_SMOOTH exports"
        )

    rows_raw = []
    rows_smooth = []
    for _raw_idx in range(seed_rows, len(raw_df)):
        _idx = _raw_idx - seed_rows
        _epoch = int(raw_df.loc[_raw_idx, "epoch"]) if "epoch" in raw_df.columns else 0
        _row_raw = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_out"]),
            "Hout_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_out"]),
            "Tin_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_in"]),
            "Hin_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_in"]),
        }
        _row_smooth = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_out"]),
            "Hout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_out"]),
            "Tin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_in"]),
            "Hin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_in"]),
        }
        rows_raw.append(_row_raw)
        rows_smooth.append(_row_smooth)

    return pd.DataFrame(rows_raw), pd.DataFrame(rows_smooth)


def _build_preprocess_model_in_reference(sample_epochs, step_epoch_seqs, x_payload_seqs, x_float_seqs):
    rows = []
    _n_features = len(features)
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _x_payload = np.asarray(x_payload_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _x_float = np.asarray(x_float_seqs[_idx], dtype=np.float32).reshape(WINDOW_STEPS, _n_features)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(WINDOW_STEPS)
        for _step in range(WINDOW_STEPS):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _feat in range(_n_features):
                _row[f"x_payload_{_feat:02d}"] = float(_x_payload[_step, _feat])
                _row[f"x_float_prequant_{_feat:02d}"] = float(_x_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_dbg_model_output_reference(sample_test_positions, sample_epochs, output_payload, output_float, tprev_override=None, hprev_override=None):
    _y_scaled = np.asarray(output_float, dtype=np.float32).reshape(-1, 2)
    _d_pred = scaler_y.inverse_transform(_y_scaled).astype(np.float32, copy=False)
    if tprev_override is None:
        _tprev = np.asarray(Tprev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _tprev = np.asarray(tprev_override, dtype=np.float32).reshape(-1)
    if hprev_override is None:
        _hprev = np.asarray(Hprev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _hprev = np.asarray(hprev_override, dtype=np.float32).reshape(-1)

    _tprev = np.asarray(_tprev, dtype=np.float32).reshape(-1)
    _hprev = np.asarray(_hprev, dtype=np.float32).reshape(-1)
    _dT_pred = np.asarray(_d_pred[:, 0], dtype=np.float32)
    _dH_pred = np.asarray(_d_pred[:, 1], dtype=np.float32)
    _t_pred = (_tprev + _dT_pred).astype(np.float32, copy=False)
    _h_pred = (_hprev + _dH_pred).astype(np.float32, copy=False)

    _out_o0_float = np.asarray(output_float[:, 0], dtype=np.float32)
    _out_o1_float = np.asarray(output_float[:, 1], dtype=np.float32)
    _out_o0_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o0_float]
    _out_o1_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o1_float]

    return pd.DataFrame({
        "idx": np.arange(len(sample_epochs), dtype=np.int64),
        "epoch": np.asarray(sample_epochs, dtype=np.int64),
        "out_o0_tensor": np.asarray(output_payload[:, 0], dtype=np.float32),
        "out_o1_tensor": np.asarray(output_payload[:, 1], dtype=np.float32),
        "out_o0_float": _out_o0_float,
        "out_o1_float": _out_o1_float,
        "out_o0_bits_hex": _out_o0_bits_hex,
        "out_o1_bits_hex": _out_o1_bits_hex,
        "y_T_scaled": _out_o0_float,
        "y_H_scaled": _out_o1_float,
        "d_T_pred": _dT_pred,
        "d_H_pred": _dH_pred,
        "p_Tprev_phys": _tprev,
        "p_Hprev_phys": np.asarray(_hprev, dtype=np.float32),
        "p_T_pred": _t_pred,
        "p_H_pred": _h_pred,
    })


def _build_dbg_model_output_raw_reference(sample_epochs, output_raw_records):
    rows = []
    max_dims = int(DBG_MODEL_RAW_MAX_DIMS)
    max_bytes = int(DBG_MODEL_RAW_MAX_BYTES)

    for _idx, (_epoch_dbg, _sample_records) in enumerate(zip(sample_epochs, output_raw_records)):
        for _rec in _sample_records:
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "out_idx": int(_rec.get("out_idx", -1)),
                "tensor_index": int(_rec.get("tensor_index", -1)),
                "type_code": int(_rec.get("type_code", -1)),
                "type_name": str(_rec.get("type_name", "")),
                "bytes_total": int(_rec.get("bytes_total", 0)),
                "bytes_dumped": int(_rec.get("bytes_dumped", 0)),
                "dims_size": int(_rec.get("dims_size", 0)),
            }
            for _di in range(max_dims):
                _row[f"dim{_di:02d}"] = int(_rec.get(f"dim{_di:02d}", -1))
            for _bi in range(max_bytes):
                _row[f"b{_bi:02d}_hex"] = str(_rec.get(f"b{_bi:02d}_hex", ""))
            rows.append(_row)

    if rows:
        return pd.DataFrame(rows)

    _columns = [
        "idx", "epoch", "out_idx", "tensor_index", "type_code", "type_name",
        "bytes_total", "bytes_dumped", "dims_size",
    ]
    _columns += [f"dim{_di:02d}" for _di in range(max_dims)]
    _columns += [f"b{_bi:02d}_hex" for _bi in range(max_bytes)]
    return pd.DataFrame(columns=_columns)


start_time = time.time()
y_pred_tflite = _predict_tflite(X_test_conv)
end_time = time.time()
y_pred_tflite = np.asarray(y_pred_tflite, dtype=np.float32)

inference_time_total = (end_time - start_time) * 1000.0
inference_time_per_sample = inference_time_total / len(X_test_conv)
print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.4f} ms")

# ---------------------
# Model sizes
# ---------------------
original_model_size_kb = os.path.getsize(model_path) / 1024
model_size_kb = os.path.getsize(quantized_path) / 1024
print(f" Original model size (.keras): {original_model_size_kb:.2f} KB")
print(f" Quantized model size: {model_size_kb:.2f} KB")

# ---------------------
# Normalized evaluation (scaled residual)
# ---------------------
y_test_scaled_resid = y_test
y_pred_scaled_resid = y_pred_tflite

mse_scaled = mean_squared_error(y_test_scaled_resid, y_pred_scaled_resid)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled = mean_absolute_error(y_test_scaled_resid, y_pred_scaled_resid)
r2_scaled = r2_score(y_test_scaled_resid, y_pred_scaled_resid)

print("\n Results (normalized scale - residual [T_in, H_in]):")
print(f"MSE  = {mse_scaled:.4f}")
print(f"RMSE = {rmse_scaled:.4f}")
print(f"MAE  = {mae_scaled:.4f}")
print(f"R²   = {r2_scaled:.4f}")

# ---------------------
# Reconstruction on the original scale (same contract as the MLP)
# ---------------------
y_test_orig_res = scaler_y.inverse_transform(y_test_scaled_resid).astype(np.float32, copy=False)
y_pred_orig_raw_res = scaler_y.inverse_transform(y_pred_scaled_resid).astype(np.float32, copy=False)

# LiteML-Edge contract: Δ_pred = inverse_transform(scaler_y)
y_pred_res = y_pred_orig_raw_res.astype(np.float32, copy=False)

# Ensure prev arrays are float32
Tprev_test = Tprev_test.astype(np.float32, copy=False)
Hprev_test = Hprev_test.astype(np.float32, copy=False)

# Absolute reconstruction
T_pred = (Tprev_test + y_pred_res[:, 0]).astype(np.float32, copy=False)
H_pred = (Hprev_test + y_pred_res[:, 1]).astype(np.float32, copy=False)

y_pred_orig = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)
y_test_orig = yabs_test.astype(np.float32, copy=False)  # true absolute [T_in, H_in]

# Final absolute-domain variables for metrics and plots
T_true = y_test_orig[:, 0]
H_true = y_test_orig[:, 1]
T_pred = y_pred_orig[:, 0]
H_pred = y_pred_orig[:, 1]

# ---------------------
# Metrics on the original scale (AFTER linear calibration)
# ---------------------
mse = mean_squared_error(y_test_orig, y_pred_orig)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_orig, y_pred_orig)
r2 = r2_score(y_test_orig, y_pred_orig)

print("\n Results (original scale - joint [T_in, H_in]):")
mse_status = " MSE within threshold" if mse <= 1.0 else " MSE moderate/high"
rmse_status = " RMSE within threshold" if rmse <= 1.0 else " RMSE moderate/high"
mae_status = " MAE within threshold" if mae <= 0.8 else " MAE moderate/high"
r2_status = " R² within threshold" if r2 >= 0.85 else " R² moderate/low"

print(f"MSE  = {mse:.4f}   {mse_status}")
print(f"RMSE = {rmse:.4f}   {rmse_status}")
print(f"MAE  = {mae:.4f}   {mae_status}")
print(f"R²   = {r2:.4f}   {r2_status}")

# =====================
# Metrics by variable
# =====================
# (Original scale)
if 'T_true' not in globals():
    T_true = y_test_orig[:, 0]
if 'H_true' not in globals():
    H_true = y_test_orig[:, 1]
if 'T_pred' not in globals():
    T_pred = y_pred_orig[:, 0]
if 'H_pred' not in globals():
    H_pred = y_pred_orig[:, 1]

mse_T = mean_squared_error(T_true, T_pred)
rmse_T = np.sqrt(mse_T)
mae_T = mean_absolute_error(T_true, T_pred)
r2_T = r2_score(T_true, T_pred)

mse_H = mean_squared_error(H_true, H_pred)
rmse_H = np.sqrt(mse_H)
mae_H = mean_absolute_error(H_true, H_pred)
r2_H = r2_score(H_true, H_pred)

print("\n Per-variable metrics - Temperature (T_in):")
print(f"MSE_T  = {mse_T:.4f}")
print(f"RMSE_T = {rmse_T:.4f}")
print(f"MAE_T  = {mae_T:.4f}")
print(f"R²_T   = {r2_T:.4f}")

print("\n Per-variable metrics - Humidity (H_in):")
print(f"MSE_H  = {mse_H:.4f}")
print(f"RMSE_H = {rmse_H:.4f}")
print(f"MAE_H  = {mae_H:.4f}")
print(f"R²_H   = {r2_H:.4f}")

model_ok = all([mse <= 1.0, rmse <= 1.0, mae <= 0.8, r2 >= 0.85])
model_status = "Performance thresholds satisfied " if model_ok else "Performance thresholds not satisfied "
print("\n Resultado (joint, linearly calibrated):", model_status)

# ---------------------
# ---------------------
# Overfitting/underfitting diagnosis (PRUNE)  -  history-based (without chunks)
# ---------------------
# In PRUNE, the percentage gap comes from the history (loss/val_loss) over the last epochs.
# Here, the same criterion is reproduced using the fine-tuning history (if enabled).
train_loss_hist = []
val_loss_hist   = []

if history_prune_ft is not None and hasattr(history_prune_ft, "history"):
    train_loss_hist = [float(v) for v in history_prune_ft.history.get("loss", [])]
    val_loss_hist   = [float(v) for v in history_prune_ft.history.get("val_loss", [])]

# Fallback: if the history is unavailable for any reason, use global evaluation
if len(train_loss_hist) == 0 or len(val_loss_hist) == 0:
    train_eval = base_model.evaluate(
        X_train_conv,
        {head_T_name: y_train_T, head_H_name: y_train_H},
        verbose=0,
        return_dict=True,
    )
    val_eval = base_model.evaluate(
        X_val_conv,
        {head_T_name: y_val_T, head_H_name: y_val_H},
        verbose=0,
        return_dict=True,
    )
    train_loss_hist = [float(train_eval.get("loss", 0.0))]
    val_loss_hist   = [float(val_eval.get("loss", 0.0))]

# Mean over the last epochs (same PRUNE criterion)
K_LAST = 5
mean_train_loss = float(np.mean(train_loss_hist[-K_LAST:]))
mean_val_loss   = float(np.mean(val_loss_hist[-K_LAST:]))
gap = abs(mean_val_loss - mean_train_loss)
gap_pct = (gap / mean_train_loss) * 100 if mean_train_loss > 0 else 0.0

# ---------------------
# Generalization diagnosis (history/global)
# ---------------------
status_diag = "Undetermined"

if (mean_train_loss > 0.3) and (mean_val_loss > 0.3):
    status_diag = "Underfitting  (high training and validation loss)"
elif mean_val_loss < (0.8 * mean_train_loss):
    status_diag = "Potential underfitting  (validation loss substantially below training)"
elif (gap_pct > 50) or ((mean_val_loss > 1.2 * mean_train_loss) and (gap > 0.05)):
    status_diag = "Overfitting  (large generalization gap)"
elif gap_pct < 10:
    status_diag = "Well-adjusted model  (gap < 10%)"
elif gap_pct < 30:
    status_diag = "Possibly well-adjusted model  (gap < 30%)"
else:
    status_diag = "Model with mild overfitting  (moderate gap)"

print("\n Model diagnosis (PRUNE  -  history-based, without chunks):")
print(f"- mean_train_loss = {mean_train_loss:.6f}")
print(f"- mean_val_loss   = {mean_val_loss:.6f}")
print(f"- gap_abs         = {gap:.6f}")
print(f"- gap_pct         = {gap_pct:.2f} %")
print(f"- status_diag     = {status_diag}")

print("\n Model diagnostics (PRUNE - history-based, without chunks):")
print(f"- Training loss (mean last {K_LAST}):     {mean_train_loss:.6f}")
print(f"- Validation loss (mean last {K_LAST}):  {mean_val_loss:.6f}")
print(f"- Absolute gap:                          {gap:.6f}")
print(f"- Generalization gap:                        {gap_pct:.2f}%")

# =========================
# SAFETY GUARD: ensure `status` is always defined (diagnostic of fit)
# =========================
if 'status' not in globals():
    try:
        if mean_train_loss > 0.3 and mean_val_loss > 0.3:
            status = "Model with UNDERFITTING  (high losses)"
        elif mean_val_loss < mean_train_loss * 0.8:
            status = "Possible UNDERFITTING  (validation much lower)"
        elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
            status = "Model with OVERFITTING  (large gap)"
        elif gap_pct < 10:
            status = "Well-adjusted model  (gap < 10%)"
        elif gap_pct < 30:
            status = "Possibly well-adjusted model  (gap < 30%)"
        else:
            status = "Model with mild overfitting "
    except Exception:
        status = "Status unavailable"

print(f"- Status:                          {status}")

# ========================= Publication-grade figure saving (PNG+PDF+SVG) =========================
def _savefig_pub(stem: str):
    """Save current matplotlib figure into quantization_graphics as PNG (600 dpi) + PDF + SVG."""
    png_path = graphics_dir / f"{stem}.png"
    pdf_path = graphics_dir / f"{stem}.pdf"
    svg_path = graphics_dir / f"{stem}.svg"
    plt.savefig(png_path, dpi=600, bbox_inches="tight")
    plt.savefig(pdf_path, bbox_inches="tight")
    plt.savefig(svg_path, bbox_inches="tight")
    print(f" Figure saved: {png_path}")

# ========================= Training/Validation loss (chunks) =========================
plt.figure(figsize=(8, 5))
plt.plot(train_loss_hist, label="Training loss (chunk-wise)", marker="o")
plt.plot(val_loss_hist, label="Validation loss (chunk-wise)", marker="o")
plt.title("Training vs Validation Loss (chunk-wise)\n" + str(status_diag))
plt.xlabel("Chunk index")
plt.ylabel("MSE loss")
plt.legend()
plt.grid(True)
plt.tight_layout()
_savefig_pub("environment_Conv1D_Tiny_training_validation_loss_diagnosis")
plt.close()

# ========================= Bar chart (aggregate metrics) =========================
plt.figure(figsize=(8, 5))
labels = ["MSE", "RMSE", "MAE", "R²"]
values = [mse, rmse, mae, r2]
bars = plt.bar(labels, values)
plt.title("Evaluation Metrics  -  Quantized INT8 Conv1D Tiny")
plt.ylabel("Value")
plt.grid(axis="y", linestyle="--", alpha=0.7)

for bar in bars:
    height = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2,
        height,
        f"{height:.4f}",
        ha="center",
        va="bottom",
        fontsize=9,
    )

plt.tight_layout()
_savefig_pub("environment_bar_metrics_quantized_Conv1D_Tiny")
plt.close()

# ========================= Scatter: joint (T_in + H_in) =========================
plt.figure(figsize=(6, 6))
plt.scatter(y_test_orig[:, 0], y_pred_orig[:, 0], label="T_in", alpha=0.6)
plt.scatter(y_test_orig[:, 1], y_pred_orig[:, 1], label="H_in", alpha=0.6)
min_all = float(min(y_test_orig.min(), y_pred_orig.min()))
max_all = float(max(y_test_orig.max(), y_pred_orig.max()))
plt.plot([min_all, max_all], [min_all, max_all], "k--", lw=2)
plt.xlabel("Ground truth")
plt.ylabel("Prediction")
plt.title("Ground Truth vs Prediction  -  Quantized INT8 Conv1D Tiny (T_in + H_in)")
plt.legend()
plt.grid(True)
plt.tight_layout()
_savefig_pub("scatter_tflite_predictions_Conv1D_Tiny")
plt.close()

# ========================= Scatter: Temperature (T_in) =========================
plt.figure(figsize=(6, 6))
min_t = float(min(T_true.min(), T_pred.min()))
max_t = float(max(T_true.max(), T_pred.max()))
plt.scatter(T_true, T_pred, alpha=0.6)
plt.plot([min_t, max_t], [min_t, max_t], "k--", lw=2)
plt.xlabel("Ground truth T_in (°C)")
plt.ylabel("Prediction T_in (°C)")
plt.title("Ground Truth vs Prediction  -  Temperature (T_in)\nQuantized INT8 Conv1D Tiny")
plt.grid(True)
plt.tight_layout()
_savefig_pub("scatter_tflite_T_in_Conv1D_Tiny")
plt.close()

# ========================= Scatter: Humidity (H_in) =========================
plt.figure(figsize=(6, 6))
min_h = float(min(H_true.min(), H_pred.min()))
max_h = float(max(H_true.max(), H_pred.max()))
plt.scatter(H_true, H_pred, alpha=0.6)
plt.plot([min_h, max_h], [min_h, max_h], "k--", lw=2)
plt.xlabel("Ground truth H_in (%RH)")
plt.ylabel("Prediction H_in (%RH)")
plt.title("Ground Truth vs Prediction  -  Humidity (H_in)\nQuantized INT8 Conv1D Tiny")
plt.grid(True)
plt.tight_layout()
_savefig_pub("scatter_tflite_H_in_Conv1D_Tiny")
plt.close()

# ========================= Rolling(24) metrics  -  firmware-equivalent (1:1) =========================
ROLLING_N = 24
ROLLING_EXPORT_ROWS = 24  # export only the last 24 windows 

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """Replicate firmware logic (metrics.cpp): R² is NaN when variance is too small."""
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return np.nan

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))
    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return np.nan
    r2v = 1.0 - (ss_res / ss_tot)
    return float(r2v) if np.isfinite(r2v) else np.nan

def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # T_in
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T ** 2))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = _r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0])

    # H_in
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H ** 2))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = _r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1])

    # Aggregate metrics (same semantics as firmware printout)
    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H) if (np.isfinite(r2_T) and np.isfinite(r2_H)) else np.nan

    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

y_test_abs_final = y_test_orig
y_pred_abs = y_pred_orig

n_test = int(y_test_abs_final.shape[0])
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)

accepted_idxs = [i for i in range(n_test) if (invoked_mask[i] and is_rollover_mask[i])]
rolling_rows = []
df_roll = None

if len(accepted_idxs) >= ROLLING_N:
    for k in range(ROLLING_N - 1, len(accepted_idxs)):
        end = accepted_idxs[k]
        start = accepted_idxs[k - ROLLING_N + 1]
        win = accepted_idxs[k - ROLLING_N + 1 : k + 1]

        yt_w = y_test_abs_final[win]
        yp_w = y_pred_abs[win]

        mae_w, rmse_w, r2_w, mae_T_w, rmse_T_w, r2_T_w, mae_H_w, rmse_H_w, r2_H_w = _metrics_like_firmware(yt_w, yp_w)

        df_idx_end = int(idx_test[end]) if ('idx_test' in globals() and end < len(idx_test)) else int(end)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": (str(df.iloc[df_idx_end]["datetime"]) if "datetime" in df.columns else ""),
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": mae_T_w,
            "RMSE_T": rmse_T_w,
            "R2_T": r2_T_w,

            "MAE_H": mae_H_w,
            "RMSE_H": rmse_H_w,
            "R2_H": r2_H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # --- Rolling(24) distribution summary over the full test set ---
    try:
        _mae_mean = float(df_roll["MAE"].mean())
        _rmse_mean = float(df_roll["RMSE"].mean())
        _mae_p95 = float(df_roll["MAE"].quantile(0.95))
        _rmse_p95 = float(df_roll["RMSE"].quantile(0.95))
        print("\n Rolling(24) summary over the full test set:")
        print(f"  - MAE_mean={_mae_mean:.4f} | MAE_p95={_mae_p95:.4f}")
        print(f"  - RMSE_mean={_rmse_mean:.4f} | RMSE_p95={_rmse_p95:.4f}")
    except Exception as _e:
        print("[WARN] Could not summarize the Rolling(24) distribution:", _e)

    df_roll_24 = df_roll.tail(ROLLING_EXPORT_ROWS).copy().reset_index(drop=True)

    # -----------------------------
    # Build the minimal sample stream required to reproduce the exported rolling24 windows
    # -----------------------------
    pos_of = {v: i for i, v in enumerate(accepted_idxs)}
    sample_set = set()
    for _, row in df_roll_24.iterrows():
        end_v = int(row["window_end"])
        pos = pos_of.get(end_v, None)
        if pos is None:
            continue
        win = accepted_idxs[max(0, pos - ROLLING_N + 1):pos + 1]
        for v in win:
            sample_set.add(int(v))

    sample_vals = sorted(sample_set, key=lambda v: pos_of.get(v, 1_000_000_000))
    ROLL24_24_PACK_IDXS = list(sample_vals)
    sample_df_indices = [int(idx_test[v]) for v in sample_vals if 0 <= v < len(idx_test)]

    # -----------------------------
    # Replay 2+47 contract for the raw firmware header
    #   - 47 real rows: exact rows already used in Python for the exported replay
    #   - 2 seed rows : immediately previous processed rows (t-2, t-1)
    # This removes the need for a boot-time mirror in firmware.
    # -----------------------------
    REPLAY_SEED_ROWS = 2
    REPLAY_REAL_ROWS = 47
    REPLAY_TOTAL_RAW_ROWS = REPLAY_SEED_ROWS + REPLAY_REAL_ROWS

    if len(sample_vals) != REPLAY_REAL_ROWS:
        raise ValueError(
            f"Expected exactly {REPLAY_REAL_ROWS} replay rows from Rolling(24), got {len(sample_vals)}"
        )
    if not np.all(np.diff(np.asarray(sample_vals, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must be contiguous in TEST space for the 2+47 export")
    if not np.all(np.diff(np.asarray(sample_df_indices, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must map to contiguous processed dataframe rows for the 2+47 export")

    first_real_df_idx = int(sample_df_indices[0])
    last_real_df_idx  = int(sample_df_indices[-1])

    if first_real_df_idx < REPLAY_SEED_ROWS:
        raise ValueError(
            f"Need {REPLAY_SEED_ROWS} previous processed rows to export replay seeds, got first_real_df_idx={first_real_df_idx}"
        )

    raw_replay_df_indices = list(range(first_real_df_idx - REPLAY_SEED_ROWS, last_real_df_idx + 1))
    if len(raw_replay_df_indices) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw replay rows, got {len(raw_replay_df_indices)}"
        )

    # Map the processed 2+47 block back to the raw-real timeline.
    raw_replay_row_ids = df_processed_row_ids[np.asarray(raw_replay_df_indices, dtype=np.int64)]
    if not np.all(np.diff(raw_replay_row_ids) == 1):
        raise ValueError("Replay 2+47 rows must map to contiguous raw-real dataframe rows")

    raw_replay_start_row_id = int(raw_replay_row_ids[0])
    raw_replay_end_row_id   = int(raw_replay_row_ids[-1])
    raw_replay_real = df_raw_real.iloc[raw_replay_start_row_id:raw_replay_end_row_id + 1].copy().reset_index(drop=True)
    if len(raw_replay_real) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw-real replay rows, got {len(raw_replay_real)}"
        )

    if raw_replay_start_row_id <= 0:
        raise ValueError(
            "Need one previous raw-real row to export the initial H_in EMA state for firmware replay"
        )

    ema_prev_row_id = raw_replay_start_row_id - 1
    ema_prev_raw_row = df_raw_real.iloc[ema_prev_row_id].copy()
    hin_ema_prev = float(df.iloc[first_real_df_idx - REPLAY_SEED_ROWS - 1]["H_in"])

    # 47 real rows preserved for the processed analysis/debug export
    df_samples_roll = df.iloc[sample_df_indices].copy().reset_index(drop=True)

    # 49 raw-real rows (2 seeds + 47 real) used by the firmware replay header and its mirrors
    df_samples_replay_raw = raw_replay_real.copy().reset_index(drop=True)

    # 49 processed rows aligned with the replay block for Python-side filtered/lagged reference
    df_samples_replay_proc = df.iloc[raw_replay_df_indices].copy().reset_index(drop=True)

    # --- Ensure epoch (seconds) for firmware replay/debug exports ---
    for _df_export in (df_samples_roll, df_samples_replay_raw, df_samples_replay_proc):
        if "datetime" in _df_export.columns:
            try:
                _df_export["epoch"] = pd.to_datetime(_df_export["datetime"]).astype("int64") // 10**9
            except Exception:
                _df_export["epoch"] = 0
        else:
            _df_export["epoch"] = 0

    # -----------------------------
    # Export rolling24 metrics (CSV + XLSX) with four-decimal formatting
    # -----------------------------
    df_roll_24_export = df_roll_24.copy()
    for c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]:
        if c in df_roll_24_export.columns:
            df_roll_24_export[c] = pd.to_numeric(df_roll_24_export[c], errors="coerce")

    df_roll_24_export = df_roll_24_export.round(4)

    csv_roll = results_dir / "environment_quantized_metrics_rolling24_Conv1D_Tiny.csv"
    xlsx_roll = results_dir / "environment_quantized_metrics_rolling24_Conv1D_Tiny.xlsx"
    df_roll_24_export.to_csv(csv_roll, index=False, encoding="utf-8-sig", float_format="%.4f")
    df_roll_24_export.to_excel(xlsx_roll, index=False)

    try:
        wb = load_workbook(xlsx_roll)
        ws = wb.active
        num_fmt = "0.0000"
        num_cols = {c: (list(df_roll_24_export.columns).index(c) + 1) for c in df_roll_24_export.columns if c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
        for _, col_idx in num_cols.items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = num_fmt
        wb.save(xlsx_roll)
    except Exception as _e:
        print("[WARN] Could not apply XLSX number formats for rolling24:", _e)

    print(f" rolling24 exported: {csv_roll}")
    print(f" rolling24 exported: {xlsx_roll}")

    # -----------------------------
    # Export real replay rows (47, processed analysis only), raw replay rows (2+47),
    # replay reference (2+47 mixed), predictions table, and firmware header (.h)
    # -----------------------------
    def _autosize_xlsx(_xlsx_path):
        try:
            _wb = load_workbook(_xlsx_path)
            _ws = _wb.active
            for _col_idx, _col_cells in enumerate(_ws.columns, 1):
                _max_length = max(len(str(_cell.value)) if _cell.value is not None else 0 for _cell in _col_cells)
                _ws.column_dimensions[get_column_letter(_col_idx)].width = _max_length + 2
            _wb.save(_xlsx_path)
        except Exception as _e:
            print(f"[WARN] Could not autosize XLSX {_xlsx_path}:", _e)

    samples_csv_path  = results_dir / "environment_quantized_samples_rolling24_Conv1D_Tiny.csv"
    samples_xlsx_path = results_dir / "environment_quantized_samples_rolling24_Conv1D_Tiny.xlsx"
    if "datetime" in df_samples_roll.columns:
        df_samples_roll["datetime"] = df_samples_roll["datetime"].astype(str)
    df_samples_roll.to_csv(samples_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    df_samples_roll.to_excel(samples_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(samples_xlsx_path)
    print(f" rolling24 samples exported: {samples_csv_path} | rows={len(df_samples_roll)}")
    print(f" rolling24 samples exported: {samples_xlsx_path} | rows={len(df_samples_roll)}")

    replay_raw_export = df_samples_replay_raw.copy()
    if "datetime" in replay_raw_export.columns:
        replay_raw_export["datetime"] = replay_raw_export["datetime"].astype(str)
    replay_raw_csv_path  = results_dir / "environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.csv"
    replay_raw_xlsx_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.xlsx"
    replay_raw_export.to_csv(replay_raw_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    replay_raw_export.to_excel(replay_raw_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(replay_raw_xlsx_path)
    print(f" replay raw 2+47 samples exported: {replay_raw_csv_path} | rows={len(replay_raw_export)}")
    print(f" replay raw 2+47 samples exported: {replay_raw_xlsx_path} | rows={len(replay_raw_export)}")

    # Replay reference export (mixed: raw + processed reference + Python GT/prediction)
    replay_reference = pd.DataFrame({
        "replay_idx": np.arange(REPLAY_TOTAL_RAW_ROWS, dtype=np.int64),
        "is_seed": [1 if i < REPLAY_SEED_ROWS else 0 for i in range(REPLAY_TOTAL_RAW_ROWS)],
        "epoch": df_samples_replay_raw["epoch"].astype(np.int64),
        "datetime": df_samples_replay_raw["datetime"].astype(str),
        "T_out_raw": df_samples_replay_raw["T_out"].astype(np.float32),
        "H_out_raw": df_samples_replay_raw["H_out"].astype(np.float32),
        "T_in_raw": df_samples_replay_raw["T_in"].astype(np.float32),
        "H_in_raw": df_samples_replay_raw["H_in"].astype(np.float32),
        "T_in_ref": df_samples_replay_proc["T_in"].astype(np.float32),
        "H_in_filt_ref": df_samples_replay_proc["H_in"].astype(np.float32),
        "T_in_lag1_ref": df_samples_replay_proc["T_in_lag1"].astype(np.float32),
        "H_in_lag1_ref": df_samples_replay_proc["H_in_lag1"].astype(np.float32),
        "T_in_lag2_ref": df_samples_replay_proc["T_in_lag2"].astype(np.float32),
        "H_in_lag2_ref": df_samples_replay_proc["H_in_lag2"].astype(np.float32),
    })

    # Python reference outputs only for the 47 real rows.
    replay_reference["T_in_ground truth"] = np.nan
    replay_reference["T_in_pred_python"] = np.nan
    replay_reference["H_in_ground truth"] = np.nan
    replay_reference["H_in_pred_python"] = np.nan

    real_rows_slice = slice(REPLAY_SEED_ROWS, REPLAY_TOTAL_RAW_ROWS)
    y_true_roll_ref = y_test_orig[np.asarray(sample_vals, dtype=np.int64)]
    y_pred_roll_ref = y_pred_orig[np.asarray(sample_vals, dtype=np.int64)]
    replay_reference.loc[real_rows_slice, "T_in_ground truth"] = y_true_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "T_in_pred_python"] = y_pred_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "H_in_ground truth"] = y_true_roll_ref[:, 1]
    replay_reference.loc[real_rows_slice, "H_in_pred_python"] = y_pred_roll_ref[:, 1]

    replay_ref_csv_path  = results_dir / "environment_quantized_replay_reference_2plus47_Conv1D_Tiny.csv"
    replay_ref_xlsx_path = results_dir / "environment_quantized_replay_reference_2plus47_Conv1D_Tiny.xlsx"
    replay_reference.to_csv(replay_ref_csv_path, index=False, encoding="utf-8-sig", float_format="%.4f")
    replay_reference.to_excel(replay_ref_xlsx_path, index=False, float_format="%.4f")
    _autosize_xlsx(replay_ref_xlsx_path)
    print(f" replay reference 2+47 exported: {replay_ref_csv_path} | rows={len(replay_reference)}")
    print(f" replay reference 2+47 exported: {replay_ref_xlsx_path} | rows={len(replay_reference)}")

    # -----------------------------
    # model I/O reference (immediate tensors that enter/exit the model)
    # -----------------------------
    dbg_model_valid_idxs = (
        np.asarray(ROLL24_24_PACK_IDXS[-ROLLING_N:], dtype=np.int64)
        if "ROLL24_24_PACK_IDXS" in globals() and len(ROLL24_24_PACK_IDXS) > 0
        else np.arange(max(0, len(X_test_conv) - ROLLING_N), len(X_test_conv), dtype=np.int64)
    )
    if dbg_model_valid_idxs.size > ROLLING_N:
        dbg_model_valid_idxs = dbg_model_valid_idxs[-ROLLING_N:]

    dbg_model_df_indices = [int(idx_test[v]) for v in dbg_model_valid_idxs]
    dbg_model_epochs = []
    for _df_idx in dbg_model_df_indices:
        _dt_dbg = pd.to_datetime(df.iloc[_df_idx]["datetime"])
        dbg_model_epochs.append(int(_dt_dbg.value // 10**9))

    dbg_model_replay_processed = _build_replay_canonical_processed(
        df_samples_replay_raw,
        ema_prev=hin_ema_prev,
        alpha=HIN_EMA_ALPHA,
    )
    dbg_model_replay_canonical = _extract_canonical_replay_debug_inputs(
        sample_df_indices=dbg_model_df_indices,
        replay_processed_df=dbg_model_replay_processed,
        replay_processed_df_indices=np.asarray(raw_replay_df_indices, dtype=np.int64),
        raw_replay_df=df_samples_replay_raw,
    )
    dbg_model_capture = _predict_tflite(X_test_conv[dbg_model_valid_idxs], capture_io=True)
    dbg_model_input_payload = dbg_model_capture["input_payload"]
    dbg_model_input_float = dbg_model_capture["input_float"]
    dbg_model_output_payload = dbg_model_capture["output_payload"]
    dbg_model_output_float = dbg_model_capture["output_float"]
    dbg_model_output_raw = dbg_model_capture["output_raw"]

    dbg_model_pre_clip = np.clip(
        dbg_model_replay_canonical["pre_phys"],
        np.asarray(scaler_X.data_min_, dtype=np.float32),
        np.asarray(scaler_X.data_max_, dtype=np.float32),
    ).astype(np.float32, copy=False)
    dbg_model_pre_scaled = np.asarray(X_test_conv[dbg_model_valid_idxs], dtype=np.float32).reshape(-1, WINDOW_STEPS, len(features))

    dbg_model_input_reference = _build_dbg_model_input_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_df_indices=dbg_model_df_indices,
        sample_epochs=dbg_model_epochs,
        x_raw_seqs=dbg_model_replay_canonical["pre_phys"],
        x_clip_seqs=dbg_model_pre_clip,
        x_scaled_seqs=dbg_model_pre_scaled,
        x_tensor_payload_seqs=dbg_model_input_payload,
        x_tensor_float_seqs=dbg_model_input_float,
        y_true_abs_rows=dbg_model_replay_canonical["y_abs"],
    )
    dbg_model_output_reference = _build_dbg_model_output_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_epochs=dbg_model_epochs,
        output_payload=dbg_model_output_payload,
        output_float=dbg_model_output_float,
        tprev_override=dbg_model_replay_canonical["tprev"],
        hprev_override=dbg_model_replay_canonical["hprev"],
    )
    dbg_model_output_raw_reference = _build_dbg_model_output_raw_reference(
        sample_epochs=dbg_model_epochs,
        output_raw_records=dbg_model_output_raw,
    )

    preprocess_raw_reference, preprocess_smooth_reference = _build_preprocess_sample_reference(
        raw_df=df_samples_replay_raw,
        smooth_df=dbg_model_replay_processed,
        seed_rows=REPLAY_SEED_ROWS,
    )
    preprocess_raw_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_raw"],
        field_names=["Tout_raw", "Hout_raw", "Tin_raw", "Hin_raw"],
    )
    preprocess_smooth_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_smooth"],
        field_names=["Tout_smooth", "Hout_smooth", "Tin_smooth", "Hin_smooth"],
    )
    preprocess_lags_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_lags"],
        field_names=["Tin_lag1", "Hin_lag1", "Tout_lag1", "Hout_lag1", "Tin_lag2", "Hin_lag2"],
    )
    preprocess_time_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_time"],
        field_names=["step_epoch_ref", "sin_hour", "cos_hour", "weekday", "month"],
    )
    preprocess_phys_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_phys"],
        field_names=[f"f{_i:02d}_phys" for _i in range(len(features))],
    )
    preprocess_clip_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_clip,
        field_names=[f"f{_i:02d}_clip" for _i in range(len(features))],
    )
    preprocess_scaled_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_scaled,
        field_names=[f"f{_i:02d}_scaled" for _i in range(len(features))],
    )
    preprocess_model_in_reference = _build_preprocess_model_in_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        x_payload_seqs=dbg_model_input_payload,
        x_float_seqs=dbg_model_input_float,
    )

    dbg_model_input_csv_path = results_dir / "environment_quantized_dbg_model_input_reference_Conv1D_Tiny.csv"
    dbg_model_input_xlsx_path = results_dir / "environment_quantized_dbg_model_input_reference_Conv1D_Tiny.xlsx"
    dbg_model_output_csv_path = results_dir / "environment_quantized_dbg_model_output_reference_Conv1D_Tiny.csv"
    dbg_model_output_xlsx_path = results_dir / "environment_quantized_dbg_model_output_reference_Conv1D_Tiny.xlsx"
    dbg_model_output_raw_csv_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_Conv1D_Tiny.csv"
    dbg_model_output_raw_xlsx_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_Conv1D_Tiny.xlsx"
    dbg_model_workbook_path = results_dir / "environment_quantized_model_io_debug_reference_Conv1D_Tiny.xlsx"
    preprocess_workbook_path = results_dir / "environment_quantized_preprocess_debug_reference_Conv1D_Tiny.xlsx"

    dbg_model_input_reference.to_csv(dbg_model_input_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_input_reference.to_excel(dbg_model_input_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_reference.to_csv(dbg_model_output_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_output_reference.to_excel(dbg_model_output_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_raw_reference.to_csv(dbg_model_output_raw_csv_path, index=False, encoding="utf-8-sig")
    dbg_model_output_raw_reference.to_excel(dbg_model_output_raw_xlsx_path, index=False)
    with pd.ExcelWriter(preprocess_workbook_path, engine="openpyxl") as _writer_pre_dbg:
        preprocess_raw_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
    with pd.ExcelWriter(dbg_model_workbook_path, engine="openpyxl") as _writer_model_dbg:
        preprocess_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
        dbg_model_input_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_IN_CSV", float_format="%.8f")
        dbg_model_output_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_CSV", float_format="%.8f")
        dbg_model_output_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_RAW_CSV")

    _autosize_xlsx(dbg_model_input_xlsx_path)
    _autosize_xlsx(dbg_model_output_xlsx_path)
    _autosize_xlsx(dbg_model_output_raw_xlsx_path)
    _autosize_xlsx(preprocess_workbook_path)
    _autosize_xlsx(dbg_model_workbook_path)

    print(f" model input reference exported: {dbg_model_input_csv_path} | rows={len(dbg_model_input_reference)}")
    print(f" model input reference exported: {dbg_model_input_xlsx_path} | rows={len(dbg_model_input_reference)}")
    print(f" model output reference exported: {dbg_model_output_csv_path} | rows={len(dbg_model_output_reference)}")
    print(f" model output reference exported: {dbg_model_output_xlsx_path} | rows={len(dbg_model_output_reference)}")
    print(f" model output RAW reference exported: {dbg_model_output_raw_csv_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f" model output RAW reference exported: {dbg_model_output_raw_xlsx_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f" preprocess workbook exported: {preprocess_workbook_path}")
    print(f" model I/O debug workbook exported: {dbg_model_workbook_path}")

    hdr_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.h"

    def _fmt_f(x):
        try:
            xf = float(x)
        except Exception:
            xf = float("nan")
        if not np.isfinite(xf):
            return "NAN"
        if abs(xf) < 0.0005:
            xf = 0.0
        return f"{xf:.6f}f"

    def _get_col(df_, names):
        for n in names:
            if n in df_.columns:
                return n
        return None

    c_tin  = _get_col(df_samples_replay_raw, ["Tin","T_in","tin","T_IN"])
    c_hin  = _get_col(df_samples_replay_raw, ["Hin","H_in","hin","H_IN"])
    c_tout = _get_col(df_samples_replay_raw, ["Tout","T_out","tout","T_OUT"])
    c_hout = _get_col(df_samples_replay_raw, ["Hout","H_out","hout","H_OUT"])

    with open(hdr_path, "w", encoding="utf-8") as hf:
        hf.write("#pragma once\n")
        hf.write("#include <stdint.h>\n\n")
        hf.write("typedef struct {\n")
        hf.write("  uint32_t epoch;\n")
        hf.write("  float T_out; float H_out; float T_in; float H_in_raw;\n")
        hf.write("} liteml_sample_raw_t;\n\n")

        hf.write(f"static const uint16_t LITEML_REPLAY_SEED_ROWS = {REPLAY_SEED_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_REAL_ROWS = {REPLAY_REAL_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_TOTAL_RAW_ROWS = {REPLAY_TOTAL_RAW_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_ROLL24_24_N_SAMPLES = {len(df_samples_replay_raw)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_ALPHA = {_fmt_f(HIN_EMA_ALPHA)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_PREV = {_fmt_f(hin_ema_prev)};\n")
        hf.write("static const liteml_sample_raw_t LITEML_ROLL24_24_SAMPLES[] = {\n")

        if not all([c_tin, c_hin, c_tout, c_hout]):
            hf.write("  // [WARN] Expected columns were not found in the raw-real dataset.\n")
        else:
            for i in range(len(df_samples_replay_raw)):
                epoch = int(df_samples_replay_raw.loc[i, "epoch"]) if "epoch" in df_samples_replay_raw.columns else 0
                Tout = df_samples_replay_raw.loc[i, c_tout]
                Hout = df_samples_replay_raw.loc[i, c_hout]
                Tin  = df_samples_replay_raw.loc[i, c_tin]
                Hin_raw  = df_samples_replay_raw.loc[i, c_hin]

                if i == 0:
                    comment = " // seed prev2 (t-2)"
                elif i == 1:
                    comment = " // seed prev  (t-1)"
                else:
                    comment = f" // real row {i - REPLAY_SEED_ROWS:02d}"

                hf.write(
                    f"  {{ {epoch}u, {_fmt_f(Tout)}, {_fmt_f(Hout)}, {_fmt_f(Tin)}, {_fmt_f(Hin_raw)} }},{comment}\n"
                )

        hf.write("};\n")
    print(f" rolling24 firmware header exported (2+47 raw replay + EMA seed): {hdr_path}")

    # -----------------------------
    # Export predictions table aligned with the MLP export contract
    #   - window_start/window_end in TEST space
    #   - datetime_end via idx_test -> df
    #   - exported features correspond to the LAST step of the window (12 columns)
    # -----------------------------
    try:
        _PRED_N = 47

        if "ROLL24_24_PACK_IDXS" in globals() and isinstance(ROLL24_24_PACK_IDXS, (list, tuple, np.ndarray)) and len(ROLL24_24_PACK_IDXS) > 0:
            _pred_sel = np.asarray(ROLL24_24_PACK_IDXS, dtype=np.int64)
            if _pred_sel.size >= _PRED_N:
                _pred_sel = _pred_sel[-_PRED_N:]
            else:
                _pad = np.arange(max(0, len(X_test_conv) - (_PRED_N - _pred_sel.size)), len(X_test_conv), dtype=np.int64)
                _pred_sel = np.unique(np.concatenate([_pred_sel, _pad]))[-_PRED_N:]
        else:
            _pred_sel = np.arange(max(0, len(X_test_conv) - _PRED_N), len(X_test_conv), dtype=np.int64)

        # Select samples in TEST space
        X_pred_sel = X_test_conv[_pred_sel]
        y_true_sel = y_test_orig[_pred_sel]
        y_pred_sel = y_pred_orig[_pred_sel]

        # Window metadata (in TEST space)
        _window_end   = _pred_sel.astype(np.int64)
        _window_start = (_window_end - (ROLLING_N - 1)).astype(np.int64)
        _pred_df_indices = [int(idx_test[int(_i)]) for _i in _window_end]

        _datetime_end = []
        for _df_idx in _pred_df_indices:
            _datetime_end.append(str(df.iloc[_df_idx]["datetime"]))

        # Base dataframe with features taken DIRECTLY from the engineered raw timeline.
        # This guarantees that the spreadsheet shows the same real rows that feed the processed analysis exports.
        df_preds_source = df.iloc[_pred_df_indices].copy().reset_index(drop=True)
        _missing_feature_cols = [c for c in features if c not in df_preds_source.columns]
        if _missing_feature_cols:
            raise ValueError(f"[export] Missing raw feature columns in df_preds_source: {_missing_feature_cols}")
        df_preds = df_preds_source[features].copy()

        # Provenance / alignment check against the model-side last-step tensor.
        # This check is diagnostic only; the spreadsheet export uses the engineered dataframe.
        X_last_orig = X_pred_sel[:, -1, :].astype(np.float32)
        _raw_feature_values = df_preds[features].values.astype(np.float32)
        _feature_abs_diff = np.abs(_raw_feature_values - X_last_orig.astype(np.float32))
        _feature_max_abs_diff = float(np.nanmax(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
        _feature_mean_abs_diff = float(np.nanmean(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
        print(
            f" Feature provenance check (engineered df vs inverse-transformed model input): "
            f"max_abs_diff={_feature_max_abs_diff:.6f} | mean_abs_diff={_feature_mean_abs_diff:.6f}"
        )

        df_preds["T_in_ground truth"] = y_true_sel[:, 0]
        df_preds["T_in_pred"]         = y_pred_sel[:, 0]
        df_preds["H_in_ground truth"] = y_true_sel[:, 1]
        df_preds["H_in_pred"]         = y_pred_sel[:, 1]

        # Compatibility aliases kept for the downstream correlation block
        df_preds["T_in_real"] = df_preds["T_in_ground truth"]
        df_preds["H_in_real"] = df_preds["H_in_ground truth"]

        df_preds.insert(0, "datetime_end", _datetime_end)
        df_preds.insert(0, "window_end", _window_end)
        df_preds.insert(0, "window_start", _window_start)

        _export_cols = (
            ["window_start", "window_end", "datetime_end"]
            + list(features)
            + ["T_in_ground truth", "T_in_pred", "H_in_ground truth", "H_in_pred"]
        )

        _missing_cols = [c for c in _export_cols if c not in df_preds.columns]
        if _missing_cols:
            raise ValueError(f"[export] Missing columns in df_preds: {_missing_cols}")

        df_preds_export = df_preds[_export_cols].copy()

        # Export rounding (presentation only)
        num_cols = df_preds_export.select_dtypes(include=[np.number]).columns
        df_preds_export[num_cols] = (np.round(df_preds_export[num_cols].astype(np.float32) * 100.0) / 100.0)

        # Export using the original filenames
        excel_path_pred = results_dir / "environment_quantized_predictions_rolling24_Conv1D_Tiny.xlsx"
        csv_path_pred   = results_dir / "environment_quantized_predictions_rolling24_Conv1D_Tiny.csv"

        df_preds_export.to_excel(excel_path_pred, index=False, float_format="%.2f")
        df_preds_export.to_csv(csv_path_pred, index=False, encoding="utf-8-sig", float_format="%.2f")

        print(f" Files saved: {excel_path_pred.name} | {csv_path_pred.name} | rows={len(df_preds_export)}")

        # Adjust column widths in XLSX
        wb = load_workbook(excel_path_pred)
        ws = wb.active
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        wb.save(excel_path_pred)

    except Exception as _e:
        print("[WARN] Could not export the Rolling(24) predictions table:", _e)

    print(" Files generated:")
    print(" - environment_quantized_metrics_rolling24_Conv1D_Tiny.csv")
    print(" - environment_quantized_metrics_rolling24_Conv1D_Tiny.xlsx")
    print(" - environment_quantized_samples_rolling24_Conv1D_Tiny.csv")
    print(" - environment_quantized_samples_rolling24_Conv1D_Tiny.xlsx")
    print(" - environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.csv")
    print(" - environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.xlsx")
    print(" - environment_quantized_replay_reference_2plus47_Conv1D_Tiny.csv")
    print(" - environment_quantized_replay_reference_2plus47_Conv1D_Tiny.xlsx")
    print(" - environment_quantized_predictions_rolling24_Conv1D_Tiny.csv")
    print(" - environment_quantized_predictions_rolling24_Conv1D_Tiny.xlsx")
    print(" - environment_quantized_samples_replay_raw_2plus47_Conv1D_Tiny.h")

    # -----------------------------
    # Rolling24 plots (paper-ready, English)
    # -----------------------------
    try:
        roll_end_idxs = df_roll_24["window_end"].astype(int).to_list()

        yT_true_r = y_test_abs_final[roll_end_idxs, 0]
        yT_pred_r = y_pred_abs[roll_end_idxs, 0]
        yH_true_r = y_test_abs_final[roll_end_idxs, 1]
        yH_pred_r = y_pred_abs[roll_end_idxs, 1]

        # Scatter: T_in (rolling24)
        plt.figure(figsize=(6, 6))
        plt.scatter(yT_true_r, yT_pred_r, s=14, alpha=0.65)
        mn = float(min(yT_true_r.min(), yT_pred_r.min()))
        mx = float(max(yT_true_r.max(), yT_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], "k--", linewidth=1.0)
        plt.gca().set_aspect("equal", adjustable="box")
        mae_t_r = float(np.mean(np.abs(yT_pred_r - yT_true_r)))
        rmse_t_r = float(np.sqrt(np.mean((yT_pred_r - yT_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yT_true_r)}\nMAE={mae_t_r:.3f} °C\nRMSE={rmse_t_r:.3f} °C",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (T_in, °C)")
        plt.ylabel("Conv1D prediction (T_in, °C)")
        plt.title("Rolling(24): ground truth vs prediction (T_in)")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_T_in_rolling24_Conv1D_Tiny")
        plt.close()

        # Scatter: H_in (rolling24)
        plt.figure(figsize=(6, 6))
        plt.scatter(yH_true_r, yH_pred_r, s=14, alpha=0.65)
        mn = float(min(yH_true_r.min(), yH_pred_r.min()))
        mx = float(max(yH_true_r.max(), yH_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], "k--", linewidth=1.0)
        plt.gca().set_aspect("equal", adjustable="box")
        mae_h_r = float(np.mean(np.abs(yH_pred_r - yH_true_r)))
        rmse_h_r = float(np.sqrt(np.mean((yH_pred_r - yH_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yH_true_r)}\nMAE={mae_h_r:.3f} %\nRMSE={rmse_h_r:.3f} %",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (H_in, %)")
        plt.ylabel("Conv1D prediction (H_in, %)")
        plt.title("Rolling(24): ground truth vs prediction (H_in)")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_H_in_rolling24_Conv1D_Tiny")
        plt.close()

        # Rolling metrics over exported windows (MAE/RMSE)
        plt.figure(figsize=(8, 4))
        plt.plot(df_roll_24_export["MAE"].values, label="MAE (aggregate)")
        plt.plot(df_roll_24_export["RMSE"].values, label="RMSE (aggregate)")
        plt.xlabel("Rolling window index (last 24)")
        plt.ylabel("Error")
        plt.title("Rolling(24) metrics over the last 24 windows")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_rolling24_metrics_last24_Conv1D_Tiny")
        plt.close()

        # Time series: Rolling(24) endpoints (T_in)
        plt.figure(figsize=(10, 4))
        try:
            x_dt = pd.to_datetime(df_roll_24["datetime_end"], errors="coerce")
            if x_dt.isna().all():
                x_dt = np.arange(len(df_roll_24))
                plt.xlabel("Rolling window index (last 24)")
            else:
                plt.xlabel("Datetime (window end)")
        except Exception:
            x_dt = np.arange(len(df_roll_24))
            plt.xlabel("Rolling window index (last 24)")

        plt.plot(x_dt, yT_true_r, label="Ground truth (T_in)")
        plt.plot(x_dt, yT_pred_r, label="Prediction (T_in)")
        plt.ylabel("Temperature (°C)")
        plt.title("Rolling(24) time series at window endpoints: T_in")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_timeseries_T_in_rolling24_Conv1D_Tiny")
        plt.close()

        # Time series: Rolling(24) endpoints (H_in)
        plt.figure(figsize=(10, 4))
        try:
            x_dt = pd.to_datetime(df_roll_24["datetime_end"], errors="coerce")
            if x_dt.isna().all():
                x_dt = np.arange(len(df_roll_24))
                plt.xlabel("Rolling window index (last 24)")
            else:
                plt.xlabel("Datetime (window end)")
        except Exception:
            x_dt = np.arange(len(df_roll_24))
            plt.xlabel("Rolling window index (last 24)")

        plt.plot(x_dt, yH_true_r, label="Ground truth (H_in)")
        plt.plot(x_dt, yH_pred_r, label="Prediction (H_in)")
        plt.ylabel("Relative humidity (%)")
        plt.title("Rolling(24) time series at window endpoints: H_in")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_timeseries_H_in_rolling24_Conv1D_Tiny")
        plt.close()

    except Exception as _e:
        print("[WARN] Could not generate rolling24 plots:", _e)


# ---------------------
# Rolling(24) summary guards (avoid NameError when rolling is available)
# ---------------------
fw_last = {}
fw_mean = {}
fw_N_ALL = float('nan')

if len(accepted_idxs) >= ROLLING_N and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
    try:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols if k in _last.index}
        fw_mean = {
            k: float(df_roll[k].mean())
            for k in _cols
            if k in df_roll.columns
        }
        fw_N_ALL = float(_last.get("N", float("nan")))
    except Exception as _e:
        print("[WARN] Failed to extract rolling(24) summary:", _e)
else:
    print(" Not enough samples for ROLLING N=24.")

# --- Compat: sparsity/strip ---
def _sparsity_fraction(model) -> float:
    try:
        total, zeros = 0, 0
        for layer in getattr(model, "layers", []):
            for w in layer.get_weights():
                arr = np.asarray(w)
                total += arr.size
                zeros += int(np.sum(arr == 0))
        return (zeros / total) if total > 0 else float("nan")
    except Exception:
        return float("nan")

sparsity_before = float("nan")
sparsity_after  = float("nan")
try:
    sparsity_before = _sparsity_fraction(base_model)
    try:
        from tensorflow_model_optimization.sparsity.keras import strip_pruning
        _stripped = strip_pruning(base_model)
        sparsity_after = _sparsity_fraction(_stripped)
    except Exception:
        sparsity_after = sparsity_before
except Exception:
    pass

size_kb = float(model_size_kb) if 'model_size_kb' in globals() else float("nan")
size_original_kb = float(original_model_size_kb) if 'original_model_size_kb' in globals() else float("nan")

status = str(status_diag) if 'status_diag' in globals() else ""
model_status = "Performance thresholds satisfied " if bool(model_ok) else "Performance thresholds not satisfied "

mse_status  = " MSE within threshold"  if mse  <= 2.125 else " MSE above threshold"
rmse_status = " RMSE within threshold" if rmse <= 1.458 else " RMSE above threshold"
mae_status  = " MAE within threshold"  if mae  <= 0.925 else " MAE above threshold"
r2_status   = " R² within threshold"   if r2   >= 0.8   else " R² below threshold"

_metrics_names = [
    'MSE (normalized residual joint)',
    'RMSE (normalized residual joint)',
    'MAE (normalized residual joint)',
    'R² (normalized residual joint)',
    'MSE (original joint)',
    'RMSE (original joint)',
    'MAE (original joint)',
    'R² (original joint)',
    'MSE_T (T_in)',
    'RMSE_T (T_in)',
    'MAE_T (T_in)',
    'R²_T (T_in)',
    'MSE_H (H_in)',
    'RMSE_H (H_in)',
    'MAE_H (H_in)',
    'R²_H (H_in)',
    'N_ALL (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, last window)',
    'RMSE (rolling24 firmware, last window)',
    'R² (rolling24 firmware, last window)',
    'MAE_T (rolling24 firmware, last window)',
    'RMSE_T (rolling24 firmware, last window)',
    'R²_T (rolling24 firmware, last window)',
    'MAE_H (rolling24 firmware, last window)',
    'RMSE_H (rolling24 firmware, last window)',
    'R²_H (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, test mean)',
    'RMSE (rolling24 firmware, test mean)',
    'R² (rolling24 firmware, test mean)',
    'MAE_T (rolling24 firmware, test mean)',
    'RMSE_T (rolling24 firmware, test mean)',
    'R²_T (rolling24 firmware, test mean)',
    'MAE_H (rolling24 firmware, test mean)',
    'RMSE_H (rolling24 firmware, test mean)',
    'R²_H (rolling24 firmware, test mean)',
    'Sparsity before strip (%)',
    'Sparsity after strip (%)',
    'Model size (KB)',
    'Original model size (KB)',
    'Mean Training Loss',
    'Mean Validation Loss',
    'Absolute Gap',
    'Gap Percentage (%)',
    'Total Inference Time (ms)',
    'Inference Time per Sample (ms)',
    'Fit Status',
    'Model Status',
]

_significados  = [
    'Mean squared error on the normalized scale, aggregated over the residuals ΔT_in and ΔH_in.',
    'Root mean squared error on the normalized scale.',
    'Mean absolute error on the normalized scale.',
    'Coefficient of determination on the normalized scale.',
    'Mean squared error on the original scale (°C / %RH), joint over [T_in, H_in].',
    'Root mean squared error on the original scale  -  joint.',
    'Mean absolute error on the original scale  -  joint.',
    'Coefficient of determination on the original scale  -  joint.',
    'Mean squared error on the original scale for T_in.',
    'Root mean squared error on the original scale for T_in.',
    'Mean absolute error on the original scale for T_in.',
    'Coefficient of determination on the original scale for T_in.',
    'Mean squared error on the original scale for H_in.',
    'Root mean squared error on the original scale for H_in.',
    'Mean absolute error on the original scale for H_in.',
    'Coefficient of determination on the original scale for H_in.',
    'Rolling(24): number of samples (HOUR events) effectively present in the final window.',
    'Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window  -  comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).',
    'Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).',
    'Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).',
    'Rolling(24) T_in MEAN over the test set (mean across windows).',
    'Rolling(24) T_in MEAN over the test set (mean across windows).',
    'Rolling(24) T_in MEAN over the test set (mean across windows).',
    'Rolling(24) H_in MEAN over the test set (mean across windows).',
    'Rolling(24) H_in MEAN over the test set (mean across windows).',
    'Rolling(24) H_in MEAN over the test set (mean across windows).',
    'Percentage of zeros in the weights before the `strip_pruning` function.',
    'Percentage of zeros in the weights after pruning operations are removed.',
    'Final model file size in kilobytes (KB).',
    'Original model size (.keras) in kilobytes (KB).',
    'Mean training loss over the last epochs.',
    'Mean validation loss over the last epochs.',
    'Absolute difference between mean losses.',
    'Percentage gap between losses (validation vs training).',
    'Total time required to infer all test samples.',
    'Mean time required to infer one sample.',
    'Diagnosis based on the loss gap: well-fitted, overfitting, or underfitting.',
    'Overall diagnosis considering the MSE/RMSE/MAE/R² limits.',
]

_thresholds      = [
    '→ Lower is better.',
    '→ Lower is better.',
    '→ Lower is better.',
    '→ Ideally > 0.95.',
    '→ < 0.1 excelente (depends on the problem).',
    '→ < 0.32 as a reference.',
    '→ < 0.3 as a reference.',
    '→ > 0.8 desirable.',
    '→ Lower is better (T_in).',
    '→ Lower is better (T_in).',
    '→ Lower is better (T_in).',
    '→ Ideally > 0.8 (T_in).',
    '→ Lower is better (H_in).',
    '→ Lower is better (H_in).',
    '→ Lower is better (H_in).',
    '→ Ideally > 0.8 (H_in).',
    '→ Should reach 24 when the window is full (after warm-up).',
    '→ Lower is better (rolling24).',
    '→ Lower is better (rolling24).',
    '→ May be NaN if the variance is low (same as the firmware).',
    '→ Lower is better (rolling24 T_in).',
    '→ Lower is better (rolling24 T_in).',
    '→ May be NaN if the variance is low.',
    '→ Lower is better (rolling24 H_in).',
    '→ Lower is better (rolling24 H_in).',
    '→ May be NaN if the variance is low.',
    '→ Lower is better (mean across windows).',
    '→ Lower is better (mean across windows).',
    '→ May be NaN if the variance is low.',
    '→ Lower is better (mean across T_in windows).',
    '→ Lower is better (mean across T_in windows).',
    '→ May be NaN if the variance is low.',
    '→ Lower is better (mean across H_in windows).',
    '→ Lower is better (mean across H_in windows).',
    '→ May be NaN if the variance is low.',
    '→ In general, > 50% is desirable for meaningful compression gains.',
    '→ Ideally close to the final target sparsity.',
    '→ Prefer < 256 KB on constrained MCUs.',
    '→ Reference for comparison with the original model.',
    '→ Low (for example, < 0.01).',
    '→ Close to the training loss.',
    '→ < 0.05 is good.',
    '→ < 10% is excellent.',
    '→ Lower is better.',
    '→ < 1 ms is ideal in TinyML.',
    "→ 'Well-fitted model' when the gap is low and the losses remain stable.",
    "→ 'Performance thresholds satisfied' when the model remains within the predefined limits.",
]

_valores = [
    f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
    f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
    f"{mse_T:.4f}",      f"{rmse_T:.4f}",      f"{mae_T:.4f}",      f"{r2_T:.4f}",
    f"{mse_H:.4f}",      f"{rmse_H:.4f}",      f"{mae_H:.4f}",      f"{r2_H:.4f}",
    (f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else ""),
    (f"{fw_last.get('MAE', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE', float('nan'))) else ""),
    (f"{fw_last.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE', float('nan'))) else ""),
    (f"{fw_last.get('R2', float('nan')):.4f}" if np.isfinite(fw_last.get('R2', float('nan'))) else ""),
    (f"{fw_last.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_T', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_last.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_T', float('nan'))) else ""),
    (f"{fw_last.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_H', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_last.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_H', float('nan'))) else ""),
    (f"{fw_mean.get('MAE', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE', float('nan'))) else ""),
    (f"{fw_mean.get('R2', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_T', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_mean.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_T', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_H', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_mean.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_H', float('nan'))) else ""),
    (f"{sparsity_before * 100:.2f} %" if np.isfinite(sparsity_before) else ""),
    (f"{sparsity_after * 100:.2f} %" if np.isfinite(sparsity_after) else ""),
    (f"{size_kb:.2f} KB" if np.isfinite(size_kb) else ""),
    (f"{size_original_kb:.2f} KB" if np.isfinite(size_original_kb) else ""),
    (f"{mean_train_loss:.4f}" if 'mean_train_loss' in globals() and np.isfinite(mean_train_loss) else ""),
    (f"{mean_val_loss:.4f}"   if 'mean_val_loss' in globals() and np.isfinite(mean_val_loss) else ""),
    (f"{gap:.4f}"             if 'gap' in globals() and np.isfinite(gap) else ""),
    (f"{gap_pct:.2f} %"       if 'gap_pct' in globals() and np.isfinite(gap_pct) else ""),
    f"{inference_time_total:.2f} ms",
    f"{inference_time_per_sample:.2f} ms",
    "",
    ""
]

# Build Status column with the same layout pattern (no duplication)
_statuses = [
    "", "", "", "",
    mse_status, rmse_status, mae_status, r2_status,
] + ([""] * (len(_metrics_names) - 8 - 2)) + [status, model_status]

_max_len = max(len(_metrics_names), len(_valores), len(_statuses), len(_significados), len(_thresholds))
def _pad(lst):
    lst = list(lst)
    if len(lst) < _max_len:
        lst += [""] * (_max_len - len(lst))
    return lst

metrics_dist = {
    "Quantized Model Metrics": _pad(_metrics_names),
    "Value": _pad(_valores),
    "Status": _pad(_statuses),
    "Meaning": _pad(_significados),
    "Expected Values / Thresholds": _pad(_thresholds),
}

df_metrics = pd.DataFrame(metrics_dist)

excel_path_metrics = results_dir / "environment_quantized_model_metrics_summary_Conv1D_Tiny.xlsx"
df_metrics.to_csv(results_dir / "environment_quantized_model_metrics_summary_Conv1D_Tiny.csv", index=False, encoding="utf-8-sig")
df_metrics.to_excel(excel_path_metrics, index=False)

wb = load_workbook(excel_path_metrics)
ws = wb.active
for col_cells in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    col_letter = get_column_letter(col_cells[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2
wb.save(excel_path_metrics)
print(f" Saved: {excel_path_metrics}")

# ========================= Correlations (no heatmap) =========================
cols_temp = ["T_out", "T_in_lag1", "T_in_lag2", "T_in_real", "T_in_pred"]
corr_temp = df_preds[cols_temp].corr()

cols_hum = ["H_out", "H_in_lag1", "H_in_lag2", "H_in_real", "H_in_pred"]
corr_hum = df_preds[cols_hum].corr()

corr_val_t = float(corr_temp.loc["T_in_real", "T_in_pred"])
if corr_val_t >= 0.9:
    print(f"Temperature diagnostic:  Strong positive correlation (r = {corr_val_t:.2f})")
elif corr_val_t >= 0.75:
    print(f"Temperature diagnostic: Moderate positive correlation (r = {corr_val_t:.2f})")
elif corr_val_t <= -0.9:
    print(f"Temperature diagnostic: Strong negative correlation (r = {corr_val_t:.2f})")
elif corr_val_t <= -0.75:
    print(f"Temperature diagnostic: Moderate negative correlation (r = {corr_val_t:.2f})")
else:
    print(f"Temperature diagnostic: Weak or no correlation (r = {corr_val_t:.2f})")

corr_val_h = float(corr_hum.loc["H_in_real", "H_in_pred"])
if corr_val_h >= 0.9:
    print(f"Humidity diagnostic:  Strong positive correlation (r = {corr_val_h:.2f})")
elif corr_val_h >= 0.75:
    print(f"Humidity diagnostic: Moderate positive correlation (r = {corr_val_h:.2f})")
elif corr_val_h <= -0.9:
    print(f"Humidity diagnostic: Strong negative correlation (r = {corr_val_h:.2f})")
elif corr_val_h <= -0.75:
    print(f"Humidity diagnostic: Moderate negative correlation (r = {corr_val_h:.2f})")
else:
    print(f"Humidity diagnostic: Weak or no correlation (r = {corr_val_h:.2f})")

# ========================= Post-execution =========================
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Could not update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Could not write manifest.json:", _e)
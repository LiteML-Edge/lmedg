"""
Script: environment_pruned_model_Conv1D_Tiny.py
Module role:
    Train, prune, fine-tune, and evaluate the pruned Conv1D Tiny model adopted
    in the LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, performs pruning-aware training and fine-tuning,
    evaluates the resulting model in normalized and reconstructed physical
    domains, and exports metrics, figures, and versioned artifacts.

Inputs:
    - environment_dataset_Conv1D_Tiny.csv
    - Baseline model path utilities and versioning helpers

Outputs:
    - Versioned pruned-model artifacts
    - Evaluation tables, figures, and summary spreadsheets

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os,sys 
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.model_selection import train_test_split
import joblib
import tensorflow_model_optimization as tfmot
from tensorflow_model_optimization.sparsity import keras as sparsity
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
rcParams['font.family'] = 'Segoe UI Emoji'

# --- Bootstrap: allows importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_Conv1D_Tiny import PROJECT_ROOT, DATASET_ENVIRONMENT, BASE_MODEL, PRUNED_MODEL, PRUNED_MODEL_METRICS
from utils.global_utils.versioning import create_versioned_dir, ensure_dir, update_latest, write_manifest, list_runs, resolve_run, resolve_latest
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # FIX: already correct

# === Versioned directories for the current execution ===
run_dir = create_versioned_dir(PRUNED_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(PRUNED_MODEL_METRICS / run_dir.name)
version_path = resolve_latest(BASE_MODEL)

# ---------------------
# Paths
# ---------------------
model_path = version_path/"environment_base_model_Conv1D_Tiny.keras"
scaler_X_path = version_path/"environment_base_model_Conv1D_Tiny_scaler_X.pkl"
scaler_y_path = version_path/"environment_base_model_Conv1D_Tiny_scaler_y.pkl"
dataset_path = DATASET_ENVIRONMENT/"environment_dataset_Conv1D_Tiny.csv"

# ---------------------
# Load dataset
# ---------------------
df = pd.read_csv(dataset_path)

# Sort temporally to build windows
df = df.sort_values("datetime").reset_index(drop=True)

# Time-Based Attribute Engineering
df["datetime"] = pd.to_datetime(df["datetime"])
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (as in Conv1D Tiny) ===
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=0.08, adjust=False).mean()

# Cyclical features
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lags (first!)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Add lag2 only for indoor variables to keep 12 + weekday/month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)

# Separate features (12 features)
features = [
    "T_out", "H_out",          # current outdoor
    "T_in_lag1", "H_in_lag1",  # indoor lag1
    "T_out_lag1", "H_out_lag1",# outdoor lag1
    "T_in_lag2", "H_in_lag2",  # indoor lag2
    "sin_hour", "cos_hour",    # cyclical hour
    "weekday", "month",        # calendar
]

# Separate residual target [ΔT_in, ΔH_in]
y_resid = np.stack([
    (df['T_in'] - df['T_in_lag1']).values,
    (df['H_in'] - df['H_in_lag1']).values
], axis=1).astype(np.float32)

# Preserve the pure residual target [ΔT_in, ΔH_in] for adherence to the official LiteML contract
# (without robust clipping of the residuals).
y_resid = y_resid.astype(np.float32)

# ---------------------
# Temporal sliding-window construction (24 h of history -> 1 prediction)
# ---------------------
WINDOW_STEPS = 24  # 24 hours
X_source = df[features].values.astype(np.float32)
N_total = len(df)
if N_total < WINDOW_STEPS:
    raise ValueError("Dataset too short for a 24 h sliding window.")

X_seq = []
y_seq = []
idx_seq = []

for t in range(WINDOW_STEPS - 1, N_total):
    X_seq.append(X_source[t-WINDOW_STEPS+1 : t+1, :])
    y_seq.append(y_resid[t])
    idx_seq.append(t)

X_all = np.stack(X_seq, axis=0).astype(np.float32)        # (N_seq, WINDOW_STEPS, n_features)
y_all = np.stack(y_seq, axis=0).astype(np.float32)        # (N_seq, 2)
idx_seq = np.array(idx_seq, dtype=np.int64)
N_seq = X_all.shape[0]
n_features = X_all.shape[-1]

# ---------------------
# Load scalers trained on the base_model (same domain of features/residuals)
# ---------------------
scaler_X = joblib.load(scaler_X_path)
scaler_y = joblib.load(scaler_y_path)

# Clamp in the physical domain before normalization, aligned with the official LiteML contract
X_all_2d = X_all.reshape(-1, n_features)
X_all_2d_clamped = np.clip(X_all_2d, scaler_X.data_min_, scaler_X.data_max_)
X_all_scaled_2d = scaler_X.transform(X_all_2d_clamped)
X_all_scaled = X_all_scaled_2d.reshape(N_seq, WINDOW_STEPS, n_features)

# Normalize the residual target
y_all_scaled = scaler_y.transform(y_all)

# ---------------------
# Train/Val/Test split (temporal, without shuffling) over the generated windows
# ---------------------
T_prev_all = df['T_in_lag1'].values.astype(np.float32)
H_prev_all = df['H_in_lag1'].values.astype(np.float32)
y_abs_all  = df[['T_in', 'H_in']].values.astype(np.float32)

# Map window indices to previous / absolute values
T_prev_seq = T_prev_all[idx_seq]
H_prev_seq = H_prev_all[idx_seq]
y_abs_seq  = y_abs_all[idx_seq]

i1, i2 = int(0.6 * N_seq), int(0.8 * N_seq)

X_train, X_val, X_test = X_all_scaled[:i1], X_all_scaled[i1:i2], X_all_scaled[i2:]
y_train, y_val, y_test = y_all_scaled[:i1], y_all_scaled[i1:i2], y_all_scaled[i2:]

Tprev_train, Tprev_val, Tprev_test = T_prev_seq[:i1], T_prev_seq[i1:i2], T_prev_seq[i2:]
Hprev_train, Hprev_val, Hprev_test = H_prev_seq[:i1], H_prev_seq[i1:i2], H_prev_seq[i2:]
y_test_abs = y_abs_seq[i2:]  # absolute ground truth for the test set

# >>> NEW: test timestamps (to match the firmware HOUR event)
dt_seq = df['datetime'].values[idx_seq]
dt_test = dt_seq[i2:]

# === Targets separated by head (already on the normalized scale) ===
y_train_T = y_train[:, 0:1]
y_train_H = y_train[:, 1:2]
y_val_T   = y_val[:,   0:1]
y_val_H   = y_val[:,   1:2]
y_test_T  = y_test[:,  0:1]
y_test_H  = y_test[:,  1:2]

# Reshape for Conv1D (already in the format [N, WINDOW_STEPS, n_features], but enforced explicitly)
X_train = X_train.reshape((-1, WINDOW_STEPS, n_features))
X_val   = X_val.reshape((-1, WINDOW_STEPS, n_features))
X_test  = X_test.reshape((-1, WINDOW_STEPS, n_features))

# ---------------------
# Load the baseline model (Conv1D Tiny multi-head)  -  compile=False
# ---------------------
try:
    base_model = tf.keras.models.load_model(model_path, compile=False)
except TypeError:
    base_model = tf.keras.models.load_model(model_path, custom_objects={}, compile=False)

# ============================================================
# Loss/Metrics multi-task learning (aligned with the base model)
# ============================================================
def make_dual_weighted_huber(delta_T=0.25, delta_H=0.5, w_T=1.2, w_H=1.8, aux_mae=0.1):
    """
    Preserved for future compatibility (single-output mode).
    It is not used directly because the model is now multi-head.
    """
    huber_T = tf.keras.losses.Huber(delta=delta_T, reduction='none')
    huber_H = tf.keras.losses.Huber(delta=delta_H, reduction='none')
    mae     = tf.keras.losses.MeanAbsoluteError(reduction='none')

    @tf.function(experimental_relax_shapes=True)
    def loss(y_true, y_pred):
        y_true = tf.cast(y_true, tf.float32)
        y_pred = tf.cast(y_pred, tf.float32)
        yT_true, yH_true = y_true[..., 0], y_true[..., 1]
        yT_pred, yH_pred = y_pred[..., 0], y_pred[..., 1]

        lT = tf.reduce_mean(huber_T(yT_true, yT_pred))
        lH = tf.reduce_mean(huber_H(yH_true, yH_pred))
        mT = tf.reduce_mean(mae(yT_true, yT_pred))
        mH = tf.reduce_mean(mae(yH_true, yH_pred))

        total = w_T * (lT + aux_mae * mT) + w_H * (lH + aux_mae * mH)
        return tf.cast(total, tf.float32)

    return loss

# ============================================================
# TRAINING WITH PRUNING (without warm-up) -> Recovery
# ============================================================
batch_size = 512
epochs_prune = 100
steps_per_epoch = int(np.ceil(len(X_train) / batch_size))
end_step = steps_per_epoch * epochs_prune

# ---------- Selective pruning by layer type ----------
def make_schedule(init, final, begin_mul=5, end_step=end_step):
    return sparsity.PolynomialDecay(
        initial_sparsity=init,
        final_sparsity=final,
        begin_step=steps_per_epoch * begin_mul,
        end_step=end_step
    )

def prune_clone_fn(layer):
    """
    Conv1D: light reduction of final sparsity.
    Dense:
      - Humidity output layers (head_H) are not pruned.
      - Temperature output layers (head_T) receive light pruning.
      - Remaining Dense layers receive moderate pruning (less aggressive than before).
    """
    # Conv1D less aggressive
    if isinstance(layer, tf.keras.layers.Conv1D):
        sched = make_schedule(0.10, 0.40)  # before: 0.10 -> 0.45
        return sparsity.prune_low_magnitude(layer, pruning_schedule=sched)

    if isinstance(layer, tf.keras.layers.Dense):
        name = layer.name or ""
        # Do NOT prune the humidity head to preserve H_in as much as possible
        if "head_H" in name:
            return layer
        # Very light pruning in the temperature head
        if "head_T" in name:
            sched = make_schedule(0.08, 0.40)
            return sparsity.prune_low_magnitude(layer, pruning_schedule=sched)
        # Remaining Dense layers: moderate pruning (reduced relative to the original)
        sched = make_schedule(0.12, 0.55)  # before: 0.15 -> 0.65
        return sparsity.prune_low_magnitude(layer, pruning_schedule=sched)

    return layer

# Direct cloning from the multi-head base_model
pruned_model = tf.keras.models.clone_model(base_model, clone_function=prune_clone_fn)

# --- Actual output names after pruning ---
output_names_pruned = pruned_model.output_names
print("Pruned model outputs:", output_names_pruned)
if len(output_names_pruned) == 2:
    head_T_name_pruned, head_H_name_pruned = output_names_pruned
else:
    # fallback (not expected, but kept for robustness)
    head_T_name_pruned, head_H_name_pruned = output_names_pruned[0], output_names_pruned[-1]

# Per-head losses/metrics (using the actual pruned output names)
losses_pruned = {
    head_T_name_pruned: tf.keras.losses.Huber(delta=0.25),
    head_H_name_pruned: tf.keras.losses.Huber(delta=0.5),
}
# >>> Prioritize humidity: assign a larger weight to head_H
loss_weights_pruned = {
    head_T_name_pruned: 1.5,
    head_H_name_pruned: 1.5,
}
metrics_dict_pruned = {
    head_T_name_pruned: [tf.keras.metrics.MeanAbsoluteError(name="mae_T")],
    head_H_name_pruned: [tf.keras.metrics.MeanAbsoluteError(name="mae_H")],
}

pruned_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=3e-4, clipnorm=1.0),
    loss=losses_pruned,
    loss_weights=loss_weights_pruned,
    metrics=metrics_dict_pruned,
)

callbacks = [
    sparsity.UpdatePruningStep(),
    tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=15, restore_best_weights=True, verbose=1),
    tf.keras.callbacks.ReduceLROnPlateau(monitor='val_loss', factor=0.5, patience=5, min_lr=1e-5, verbose=1),
]

history = pruned_model.fit(
    X_train, {head_T_name_pruned: y_train_T, head_H_name_pruned: y_train_H},
    validation_data=(X_val, {head_T_name_pruned: y_val_T, head_H_name_pruned: y_val_H}),
    batch_size=batch_size,
    epochs=epochs_prune,
    shuffle=True,  # ok, the split is already temporal
    callbacks=callbacks,
    verbose=1
)

# ---------- strip + recovery fine-tuning ----------
final_model = sparsity.strip_pruning(pruned_model)

# After strip_pruning, output names may return to their original names
output_names_final = final_model.output_names
print("Model outputs after strip_pruning:", output_names_final)
if len(output_names_final) == 2:
    head_T_name_final, head_H_name_final = output_names_final
else:
    # fallback: reuses the names from the pruning phase.
    head_T_name_final, head_H_name_final = head_T_name_pruned, head_H_name_pruned

losses_final = {
    head_T_name_final: tf.keras.losses.Huber(delta=0.25),
    head_H_name_final: tf.keras.losses.Huber(delta=0.5),
}
# >>> Same prioritization for H in the final stage
loss_weights_final = {
    head_T_name_final: 1.0,
    head_H_name_final: 2.0,
}
metrics_dict_final = {
    head_T_name_final: [tf.keras.metrics.MeanAbsoluteError(name="mae_T")],
    head_H_name_final: [tf.keras.metrics.MeanAbsoluteError(name="mae_H")],
}

final_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=2e-4, clipnorm=1.0),
    loss=losses_final,
    loss_weights=loss_weights_final,
    metrics=metrics_dict_final,
)

recovery_cb = [
    tf.keras.callbacks.EarlyStopping(monitor="val_loss", patience=5, restore_best_weights=True, verbose=1),
    tf.keras.callbacks.ReduceLROnPlateau(monitor="val_loss", factor=0.5, patience=2, min_lr=5e-6, verbose=1),
]
history_recovery = final_model.fit(
    X_train, {head_T_name_final: y_train_T, head_H_name_final: y_train_H},
    validation_data=(X_val, {head_T_name_final: y_val_T, head_H_name_final: y_val_H}),
    batch_size=batch_size,
    epochs=20,
    shuffle=True,
    callbacks=recovery_cb,
    verbose=1
)

# === Light fine-tuning (5 epochs) while maintaining the structure ===
fine_cb = [
    tf.keras.callbacks.EarlyStopping(monitor="val_loss", patience=2, restore_best_weights=True, verbose=1),
]
history_finetune = final_model.fit(
    X_train, {head_T_name_final: y_train_T, head_H_name_final: y_train_H},
    validation_data=(X_val, {head_T_name_final: y_val_T, head_H_name_final: y_val_H}),
    batch_size=batch_size,
    epochs=5,
    shuffle=True,
    callbacks=fine_cb,
    verbose=1
)

# Save the final model (native Keras format) without include_optimizer
final_model.save(run_dir/"environment_pruned_model_Conv1D_Tiny.keras")

# ---------------------
# Actual sparsity
# ---------------------
def calc_sparsity(m):
    zeros = 0
    total = 0
    for w in m.get_weights():
        arr = np.asarray(w)
        zeros += np.sum(arr == 0)
        total += arr.size
    return zeros / total if total else 0.0

sparsity_before = calc_sparsity(pruned_model)
sparsity_after  = calc_sparsity(final_model)
print(f"\n Sparsity (before strip): {sparsity_before:.2%}")
print(f" Sparsity (after strip):  {sparsity_after:.2%}")

# ---------------------
# Latency + normalized predictions (multi-head -> concat)
# ---------------------
start_time = time.time()
preds = final_model.predict(X_test, verbose=0)
end_time = time.time()

inference_time_total = (end_time - start_time) * 1000
inference_time_per_sample = (inference_time_total / len(X_test))

print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.2f} ms")

# preds can be a list [head_T, head_H] or a single tensor (for safety)
if isinstance(preds, (list, tuple)):
    y_pred_T_scaled, y_pred_H_scaled = preds
    y_pred_test_scaled = np.concatenate([y_pred_T_scaled, y_pred_H_scaled], axis=1)
else:
    y_pred_test_scaled = preds

# ---------------------
# Model size
# ---------------------
saved_model_path = run_dir/"environment_pruned_model_Conv1D_Tiny.keras"
size_kb = np.nan
try:
    if Path(saved_model_path).is_file():
        size_kb = os.path.getsize(saved_model_path) / 1024
        print(f" Pruned model size: {size_kb:.2f} KB ")
    else:
        print(" Model file not found.")
except Exception as _e:
    print(" Failed to measure model size:", _e)

# ---------------------
# Evaluation (normalized scale - joint residual)
# ---------------------
mse_scaled = mean_squared_error(y_test, y_pred_test_scaled)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled = mean_absolute_error(y_test, y_pred_test_scaled)
r2_scaled  = r2_score(y_test, y_pred_test_scaled)

print("\n Results (residual, normalized scale):")
print(f"MSE  (normalized residual) = {mse_scaled:.8f}")
print(f"RMSE (normalized residual) = {rmse_scaled:.8f}")
print(f"MAE  (normalized residual) = {mae_scaled:.8f}")
print(f"R²   (normalized residual) = {r2_scaled:.8f}")

# ---------------------
# Evaluation (original scale absolute)
# ---------------------
y_pred_delta = scaler_y.inverse_transform(y_pred_test_scaled)
T_pred = Tprev_test + y_pred_delta[:, 0]
H_pred = Hprev_test + y_pred_delta[:, 1]
y_pred_orig = np.stack([T_pred, H_pred], axis=1)

# Reconstruction of the original y_test absolute values from the scaled residual
y_test_delta = scaler_y.inverse_transform(y_test)
T_true = Tprev_test + y_test_delta[:, 0]
H_true = Hprev_test + y_test_delta[:, 1]
y_test_orig = np.stack([T_true, H_true], axis=1)

# Global metrics (joint)
mse = mean_squared_error(y_test_orig, y_pred_orig)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_orig, y_pred_orig)
r2  = r2_score(y_test_orig, y_pred_orig)

# >>> Metrics by output
mse_T  = mean_squared_error(y_test_orig[:,0], y_pred_orig[:,0])
rmse_T = np.sqrt(mse_T)
mae_T  = mean_absolute_error(y_test_orig[:,0], y_pred_orig[:,0])
r2_T   = r2_score(y_test_orig[:,0], y_pred_orig[:,0])

mse_H  = mean_squared_error(y_test_orig[:,1], y_pred_orig[:,1])
rmse_H = np.sqrt(mse_H)
mae_H  = mean_absolute_error(y_test_orig[:,1], y_pred_orig[:,1])
r2_H   = r2_score(y_test_orig[:,1], y_pred_orig[:,1])

print("\n Per variable (original scale):")
print(f"[T_in] MSE={mse_T:.6f} RMSE={rmse_T:.6f} MAE={mae_T:.6f} R²={r2_T:.6f}")
print(f"[H_in] MSE={mse_H:.6f} RMSE={rmse_H:.6f} MAE={mae_H:.6f} R²={r2_H:.6f}")

print("\n Results (original scale absolute - joint):")
mse_status  = " MSE within threshold"  if mse  <= 0.1 else " MSE above threshold"
rmse_status = " RMSE within threshold" if rmse <= 0.32 else " RMSE above threshold"
mae_status  = " MAE within threshold"  if mae  <= 0.3 else " MAE above threshold"
r2_status   = " R² within threshold"   if r2   >= 0.8 else " R² below threshold"
print(f"MSE  = {mse:.8f}   {mse_status}")
print(f"RMSE = {rmse:.8f}   {rmse_status}")
print(f"MAE  = {mae:.8f}   {mae_status}")
print(f"R²   = {r2:.8f}   {r2_status}")

# Overall model status (based on the joint)
model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied " if model_ok else "Performance thresholds not satisfied "

# ============================================================
#  NEW: Rolling metrics over a 24-sample moving window (firmware-equivalent)
# ============================================================
ROLLING_N = 24  # N=24 samples (HOUR) as in metrics.cpp

# ============================================================
# GATING SAME AS FIRMWARE:
# - invoked_mask: there was real Invoke
# - is_rollover_mask: EVENT type=HOUR (TRIG does not enter the metrics)
# In offline mode (test split), everything is True by default.
# ============================================================
invoked_mask = np.ones(len(y_test_orig), dtype=bool)
is_rollover_mask = np.ones(len(y_test_orig), dtype=bool)

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """
    Replicate the firmware logic (metrics.cpp):
      - R² = 1 - SS_res/SS_tot
      - SS_tot = sum(y^2) - (sum(y)^2)/n
      - If n < 2: R² is undefined (NaN)
      - If SS_tot <= 1e-6: variance is too low → R² is undefined (NaN)
    """
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = y_true_1d.size
    if n < 2:
        return float("nan")

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))

    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return float("nan")

    return 1.0 - (ss_res / ss_tot)

def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    """
    Returns:
      (mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H)
    following the firmware exactly:
      mae = 0.5*(mae_T + mae_H)
      rmse = sqrt( 0.5*(mse_T + mse_H) )
      r2 = 0.5*(r2_T + r2_H)
    """
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # T
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_Tw = float(np.mean(np.abs(err_T)))
    mse_Tw = float(np.mean(err_T * err_T))
    rmse_Tw = float(np.sqrt(mse_Tw))
    r2_Tw = float(_r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0]))

    # H
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_Hw = float(np.mean(np.abs(err_H)))
    mse_Hw = float(np.mean(err_H * err_H))
    rmse_Hw = float(np.sqrt(mse_Hw))
    r2_Hw = float(_r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1]))

    mae_w = 0.5 * (mae_Tw + mae_Hw)
    rmse_w = float(np.sqrt(0.5 * (mse_Tw + mse_Hw)))
    r2_w = 0.5 * (r2_Tw + r2_Hw)
    return mae_w, rmse_w, r2_w, mae_Tw, rmse_Tw, r2_Tw, mae_Hw, rmse_Hw, r2_Hw

rolling_rows = []
n_test = y_test_orig.shape[0]

if n_test < ROLLING_N:
    print(f"\n Rolling N=24 cannot be computed: the test split has only {n_test} samples.")
    df_roll = pd.DataFrame()
else:
    for end in range(ROLLING_N - 1, n_test):
        start = end - ROLLING_N + 1

        # --- GATING (firmware-equivalent) ---
        if (not invoked_mask[end]) or (not is_rollover_mask[end]):
            rolling_rows.append({
                "window_start": int(start),
                "window_end": int(end),
                "datetime_end": pd.to_datetime(dt_test[end]) if 'dt_test' in globals() else pd.NaT,
                "N": int(ROLLING_N),
                "MAE": float("nan"),
                "RMSE": float("nan"),
                "R2": float("nan"),
                "MAE_T": float("nan"),
                "RMSE_T": float("nan"),
                "R2_T": float("nan"),
                "MAE_H": float("nan"),
                "RMSE_H": float("nan"),
                "R2_H": float("nan"),
            })
            continue

        yt_w = y_test_orig[start:end+1, :]
        yp_w = y_pred_orig[start:end+1, :]

        mae_w, rmse_w, r2_w, maeT_w, rmseT_w, r2T_w, maeH_w, rmseH_w, r2H_w = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": pd.to_datetime(dt_test[end]) if 'dt_test' in globals() else pd.NaT,
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": maeT_w,
            "RMSE_T": rmseT_w,
            "R2_T": r2T_w,

            "MAE_H": maeH_w,
            "RMSE_H": rmseH_w,
            "R2_H": r2H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # Useful summaries for comparison with the firmware
    last = df_roll.iloc[-1]
    mean_roll = df_roll[["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]].mean()

    print("\n Rolling N=24 (firmware-equivalent)  -  LAST window (current state):")
    print(f"MAE={last['MAE']:.4f} RMSE={last['RMSE']:.4f} R²={last['R2']:.4f} | "
          f"T: MAE={last['MAE_T']:.4f} RMSE={last['RMSE_T']:.4f} R²={last['R2_T']:.4f} | "
          f"H: MAE={last['MAE_H']:.4f} RMSE={last['RMSE_H']:.4f} R²={last['R2_H']:.4f}")

    print("\n Rolling N=24  -  MEAN across all TEST windows:")
    print(f"MAE={mean_roll['MAE']:.4f} RMSE={mean_roll['RMSE']:.4f} R²={mean_roll['R2']:.4f} | "
          f"T: MAE={mean_roll['MAE_T']:.4f} RMSE={mean_roll['RMSE_T']:.4f} R²={mean_roll['R2_T']:.4f} | "
          f"H: MAE={mean_roll['MAE_H']:.4f} RMSE={mean_roll['RMSE_H']:.4f} R²={mean_roll['R2_H']:.4f}")

    # Save rolling metrics to CSV/Excel (for 1:1 comparison with firmware logs)
    df_roll = df_roll.round(4)
    df_roll.to_csv(metrics_run_dir / "environment_pruned_model_Conv1D_Tiny_metrics_rolling24.csv",
                   index=False, encoding="utf-8-sig")
    excel_roll_path = metrics_run_dir / "environment_pruned_model_Conv1D_Tiny_metrics_rolling24.xlsx"
    df_roll.to_excel(excel_roll_path, index=False)

    # Auto-adjust Excel columns
    wb2 = load_workbook(excel_roll_path)
    ws2 = wb2.active
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb2.save(excel_roll_path)

    print(" Rolling files saved:")
    print(" - environment_pruned_model_Conv1D_Tiny_metrics_rolling24.csv")
    print(" - environment_pruned_model_Conv1D_Tiny_metrics_rolling24.xlsx")

    # Plot: rolling (MAE/RMSE)
    plt.figure(figsize=(10, 5))
    plt.plot(df_roll["datetime_end"], df_roll["MAE"], label="MAE (rolling24)")
    plt.plot(df_roll["datetime_end"], df_roll["RMSE"], label="RMSE (rolling24)")
    plt.title("Metrics over a Moving Window N=24 (comparable to the firmware)  -  Pruned Model (Conv1D Tiny)")
    plt.xlabel("Datetime (window end)")
    plt.ylabel("Value")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_pruned_model_Conv1D_Tiny_metrics_rolling24_mae_rmse.png")
    plt.close()

    # Plot: rolling R²
    plt.figure(figsize=(10, 4))
    plt.plot(df_roll["datetime_end"], df_roll["R2"], label="R² (rolling24)")
    plt.title("R² over a Moving Window N=24 (comparable to the firmware)  -  Pruned Model (Conv1D Tiny)")
    plt.xlabel("Datetime (window end)")
    plt.ylabel("R²")
    plt.ylim([-0.2, 1.05])
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_pruned_model_Conv1D_Tiny_metrics_rolling24_r2.png")
    plt.close()

# ---------------------
# Diagnosis (uses loss/val_loss from the 3 phases: prune + recovery + fine-tune)
# ---------------------
hist_loss = list(history.history['loss']) \
         + list(history_recovery.history['loss']) \
         + list(history_finetune.history['loss'])

hist_vls  = list(history.history.get('val_loss', [])) \
         + list(history_recovery.history.get('val_loss', [])) \
         + list(history_finetune.history.get('val_loss', []))

train_loss = np.array(hist_loss)
val_loss   = np.array(hist_vls)
n = min(5, len(train_loss))
mean_train_loss = float(np.mean(train_loss[-n:])) if n>0 else float('nan')
mean_val_loss   = float(np.mean(val_loss[-n:])) if (n>0 and val_loss.size>0) else float('nan')
gap = abs(mean_val_loss - mean_train_loss) if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss) else float('nan')
gap_pct = (gap / mean_train_loss) * 100 if (np.isfinite(gap) and mean_train_loss>0) else float('nan')

if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss):
    if mean_train_loss > 0.3 and mean_val_loss > 0.3:
        status = "Underfitting detected  (high losses)"
    elif mean_val_loss < mean_train_loss * 0.8:
        status = "Potential underfitting  (validation loss significantly lower than training loss)"
    elif (gap_pct > 50) or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
        status = "Overfitting detected  (large generalization gap or significant divergence)"
    elif gap_pct < 10:
        status = "Well-fitted model  (gap < 10%)"
    elif gap_pct < 30:
        status = "Acceptably fitted model  (gap < 30%)"
    else:
        status = "Mild overfitting  (moderate gap)"
else:
    status = "Insufficient history for fit diagnosis."

print("\n Model diagnostics:")
print(f"- Mean training loss:     {mean_train_loss:.8f}")
print(f"- Mean validation loss:  {mean_val_loss:.8f}")
print(f"- Absolute gap:           {gap:.8f}")
print(f"- Generalization gap:         {gap_pct:.2f}%")
print(f"- Status:                 {status}")

# ---------------------
# UNIQUE PLOT: Training vs Validation
# ---------------------
plt.figure(figsize=(10, 6))
plt.plot(hist_loss, label="Training", linewidth=1.6)
if len(hist_vls) > 0:
    plt.plot(hist_vls, label="Validation", linewidth=1.6)

subtitle = "Well-fitted model  (gap < 10%)" if (np.isfinite(gap_pct) and gap_pct < 10.0) \
           else (f"gap = {gap_pct:.1f}%" if np.isfinite(gap_pct) else "")
plt.title("Loss during training with EarlyStopping\n" + subtitle)

plt.xlabel("Epoch")
plt.ylabel("Loss (Huber / multi-head)")
plt.legend()
plt.grid(True)

all_vals = np.array(list(hist_loss) + (list(hist_vls) if len(hist_vls) > 0 else []))
if all_vals.size > 0:
    y_max = float(all_vals.max())
    if y_max <= 0:
        y_max = 1e-6
    plt.ylim(0.0, y_max * 1.05)

plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_Conv1D_Tiny_training_validation_loss_diagnosis_pruning.png")
plt.close()

# 2) Overall bar chart (joint)
plt.figure(figsize=(8, 6))
labels = ["MSE", "RMSE", "MAE", "R²"]
values = [mse, rmse, mae, r2]
bars = plt.bar(labels, values)
plt.title("Metrics Summary - Pruned Model (T_in + H_in Set)")
plt.grid(axis='y')
for bar in bars:
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width() / 2.0, yval, f"{yval:.8f}", ha='center', va='bottom')
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_Conv1D_Tiny_final_metrics_summary_plot_pruned.png")
plt.close()

# 3) Scatter plots (joint parity)
plt.figure(figsize=(6, 6))
plt.scatter(y_test_orig[:, 0], y_pred_orig[:, 0], alpha=0.5, label="Temperature (T_in)")
plt.scatter(y_test_orig[:, 1], y_pred_orig[:, 1], alpha=0.5, label="Humidity (H_in)")
min_val = float(min(y_test_orig.min(), y_pred_orig.min()))
max_val = float(max(y_test_orig.max(), y_pred_orig.max()))
plt.plot([min_val, max_val], [min_val, max_val], 'k--')
plt.xlabel("Actual Value"); plt.ylabel("Predicted Value")
plt.title("Prediction Scatter Plot - Pruned Model")
plt.legend(); plt.grid(True); plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_Conv1D_Tiny_scatter_predictions_pruned.png"); plt.close()

def parity_plot(y_true, y_pred, label, fname):
    plt.figure(figsize=(6,6))
    plt.scatter(y_true, y_pred, alpha=0.5, label=label)
    mn = float(min(y_true.min(), y_pred.min())); mx = float(max(y_true.max(), y_pred.max()))
    plt.plot([mn, mx], [mn, mx], 'k--')
    plt.xlabel("Actual Value"); plt.ylabel("Predicted Value")
    plt.title(f"Prediction Scatter Plot - {label}")
    plt.legend(); plt.grid(True); plt.tight_layout()
    plt.savefig(metrics_run_dir/fname); plt.close()

parity_plot(y_test_orig[:,0], y_pred_orig[:,0], "Temperature (T_in)", "environment_Conv1D_Tiny_scatter_Tin_pruned.png")
parity_plot(y_test_orig[:,1], y_pred_orig[:,1], "Humidity (H_in)",     "environment_Conv1D_Tiny_scatter_Hin_pruned.png")

# ---------------------
# Metrics table (CSV / Excel)
# ---------------------
# =========================
# (NEW) Firmware-style Rolling(24) summary for metrics_summary export
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)

metrics_dist = {
    "Pruned Model Metrics": [
        # Normalized scale (aggregate over T_in + H_in)
        "MSE (normalized residual joint)", "RMSE (normalized residual joint)", "MAE (normalized residual joint)", "R² (normalized residual joint)",
        # Joint original scale
        "MSE (original joint)", "RMSE (original joint)", "MAE (original joint)", "R² (original joint)",
        # By variable T_in
        "MSE_T (T_in)", "RMSE_T (T_in)", "MAE_T (T_in)", "R²_T (T_in)",
        # By variable H_in
        "MSE_H (H_in)", "RMSE_H (H_in)", "MAE_H (H_in)", "R²_H (H_in)",
        # Rolling(24) firmware-equivalent (window=24, HOUR+Invoke gating)
        "N_ALL (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, last window)", "RMSE (rolling24 firmware, last window)", "R² (rolling24 firmware, last window)",
        "MAE_T (rolling24 firmware, last window)", "RMSE_T (rolling24 firmware, last window)", "R²_T (rolling24 firmware, last window)",
        "MAE_H (rolling24 firmware, last window)", "RMSE_H (rolling24 firmware, last window)", "R²_H (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, test mean)", "RMSE (rolling24 firmware, test mean)", "R² (rolling24 firmware, test mean)",
        "MAE_T (rolling24 firmware, test mean)", "RMSE_T (rolling24 firmware, test mean)", "R²_T (rolling24 firmware, test mean)",
        "MAE_H (rolling24 firmware, test mean)", "RMSE_H (rolling24 firmware, test mean)", "R²_H (rolling24 firmware, test mean)",
        # Sparsity / size
        "Sparsity before strip (%)", "Sparsity after strip (%)", "Model size (KB)",
        # Losss / gap
        "Mean Training Loss", "Mean Validation Loss",
        "Absolute Gap", "Gap Percentage (%)",
        # Times
        "Total Inference Time (ms)", "Inference Time per Sample (ms)",
        # Status
        "Fit Status", "Model Status"
    ],
    "Value": [
        f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
        f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
        f"{mse_T:.4f}",      f"{rmse_T:.4f}",      f"{mae_T:.4f}",      f"{r2_T:.4f}",
        f"{mse_H:.4f}",      f"{rmse_H:.4f}",      f"{mae_H:.4f}",      f"{r2_H:.4f}",
        f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else "",
        f"{fw_last['MAE']:.4f}" if np.isfinite(fw_last['MAE']) else "",
        f"{fw_last['RMSE']:.4f}" if np.isfinite(fw_last['RMSE']) else "",
        f"{fw_last['R2']:.4f}" if np.isfinite(fw_last['R2']) else "",
        f"{fw_last['MAE_T']:.4f}" if np.isfinite(fw_last['MAE_T']) else "",
        f"{fw_last['RMSE_T']:.4f}" if np.isfinite(fw_last['RMSE_T']) else "",
        f"{fw_last['R2_T']:.4f}" if np.isfinite(fw_last['R2_T']) else "",
        f"{fw_last['MAE_H']:.4f}" if np.isfinite(fw_last['MAE_H']) else "",
        f"{fw_last['RMSE_H']:.4f}" if np.isfinite(fw_last['RMSE_H']) else "",
        f"{fw_last['R2_H']:.4f}" if np.isfinite(fw_last['R2_H']) else "",
        f"{fw_mean['MAE']:.4f}" if np.isfinite(fw_mean['MAE']) else "",
        f"{fw_mean['RMSE']:.4f}" if np.isfinite(fw_mean['RMSE']) else "",
        f"{fw_mean['R2']:.4f}" if np.isfinite(fw_mean['R2']) else "",
        f"{fw_mean['MAE_T']:.4f}" if np.isfinite(fw_mean['MAE_T']) else "",
        f"{fw_mean['RMSE_T']:.4f}" if np.isfinite(fw_mean['RMSE_T']) else "",
        f"{fw_mean['R2_T']:.4f}" if np.isfinite(fw_mean['R2_T']) else "",
        f"{fw_mean['MAE_H']:.4f}" if np.isfinite(fw_mean['MAE_H']) else "",
        f"{fw_mean['RMSE_H']:.4f}" if np.isfinite(fw_mean['RMSE_H']) else "",
        f"{fw_mean['R2_H']:.4f}" if np.isfinite(fw_mean['R2_H']) else "",
        f"{sparsity_before * 100:.2f} %", f"{sparsity_after * 100:.2f} %", f"{(0.0 if np.isnan(size_kb) else size_kb):.2f} KB",
        f"{mean_train_loss:.8f}", f"{mean_val_loss:.8f}",
        f"{gap:.8f}",            f"{gap_pct:.2f} %",
        f"{inference_time_total:.2f} ms", f"{inference_time_per_sample:.2f} ms",
        "", ""
    ],
    "Status": [
        # 0-3: normalized metrics → without status
        "", "", "", "",
        # 4-7: aggregated original-scale metrics → use mse_status, rmse_status, ...
        mse_status, rmse_status, mae_status, r2_status,
        # 8-15: per-variable metrics (T_in, H_in) → empty
        "", "", "", "", "", "", "", "",
        # rolling24 firmware (last window + mean) → empty
        "", "", "", "", "", "", "", "", "", "",
        "", "", "", "", "", "", "", "", "",
        # sparsity/size → empty
        "", "", "",
        # losses/gap → empty
        "", "", "", "",
        # timing → empty
        "", "",
        # status final
        status, model_status
    ],
    "Meaning": [
        "Mean squared error on the normalized scale, aggregated over the residuals ΔT_in and ΔH_in.",
        "Root mean squared error on the normalized scale.",
        "Mean absolute error on the normalized scale.",
        "Coefficient of determination on the normalized scale.",
        "Mean squared error on the original scale (°C / %RH), joint over [T_in, H_in].",
        "Root mean squared error on the original scale  -  joint.",
        "Mean absolute error on the original scale  -  joint.",
        "Coefficient of determination on the original scale  -  joint.",
        "Mean squared error on the original scale for T_in.",
        "Root mean squared error on the original scale for T_in.",
        "Mean absolute error on the original scale for T_in.",
        "Coefficient of determination on the original scale for T_in.",
        "Mean squared error on the original scale for H_in.",
        "Root mean squared error on the original scale for H_in.",
        "Mean absolute error on the original scale for H_in.",
        "Coefficient of determination on the original scale for H_in.",
        "Rolling(24): number of samples (HOUR events) effectively present in the final window.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H)  -  MEAN across all test windows.",
        "Rolling(24) aggregate (T+H)  -  MEAN across all test windows.",
        "Rolling(24) aggregate (T+H)  -  MEAN across all test windows.",
        "Rolling(24) T_in  -  MEAN across all test windows.",
        "Rolling(24) T_in  -  MEAN across all test windows.",
        "Rolling(24) T_in  -  MEAN across all test windows.",
        "Rolling(24) H_in  -  MEAN across all test windows.",
        "Rolling(24) H_in  -  MEAN across all test windows.",
        "Rolling(24) H_in  -  MEAN across all test windows.",
        "Percentage of weights equal to zero before removing pruning wrappers.",
        "Percentage of weights equal to zero after removing pruning wrappers.",
        "Size of the model file saved on disk (KB).",
        "Mean training loss over the last epochs (fit diagnosis).",
        "Mean validation loss over the last epochs (fit diagnosis).",
        "Absolute difference between mean losses (val - train).",
        "Percentage difference between mean losses (val - train) / train.",
        "Total time to infer the full test set (ms).",
        "Average latency per sample in the test set (ms).",
        "Diagnosis based on the training vs validation loss gap.",
        "Overall status (aggregated metrics on the original scale)."
    ],
}

dfm = pd.DataFrame(metrics_dist)
dfm.to_csv(metrics_run_dir/"environment_Conv1D_Tiny_pruned_model_metrics_summary.csv", index=False, encoding="utf-8-sig")
print(" File saved: environment_Conv1D_Tiny_pruned_model_metrics_summary.csv")

excel_path = metrics_run_dir/"environment_Conv1D_Tiny_pruned_model_metrics_summary.xlsx"
dfm.to_excel(excel_path, index=False)
wb = load_workbook(excel_path)
ws = wb.active
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
wb.save(excel_path)
print(" File saved: environment_Conv1D_Tiny_pruned_model_metrics_summary.xlsx (with adjusted columns)")


# === Post-execution: update 'latest' and manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Could not update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Could not write manifest.json:", _e)

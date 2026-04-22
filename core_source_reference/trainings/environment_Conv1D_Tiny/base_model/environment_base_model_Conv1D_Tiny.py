"""
Script: environment_base_model_Conv1D_Tiny.py
Module role:
    Train and evaluate the baseline Conv1D Tiny model adopted in the
    LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, trains the baseline residual model, evaluates the
    model in normalized and reconstructed physical domains, and exports metrics,
    figures, and versioned artifacts.

Inputs:
    - environment_dataset_Conv1D_Tiny.csv
    - Project path and versioning utilities from utils.global_utils.paths_Conv1D_Tiny and
      utils.versioning

Outputs:
    - Versioned baseline model artifacts
    - Fitted X and y scalers
    - Tables, figures, and summary spreadsheets for offline evaluation

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Aliases compatible with environments that only expose tf.keras
layers = tf.keras.layers
regularizers = tf.keras.regularizers
models = tf.keras.models

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
rcParams['font.family'] = 'Segoe UI Emoji'

# --- Bootstrap: allows importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
# Paths and helpers (preserved)
from utils.global_utils.paths_Conv1D_Tiny import PROJECT_ROOT, DATASET_ENVIRONMENT, BASE_MODEL, BASE_MODEL_METRICS
from utils.global_utils.versioning import create_versioned_dir, ensure_dir, update_latest, write_manifest
from utils.global_utils.global_seed import set_global_seed
set_global_seed(42)

# === Versioned directories (preserved) ===
run_dir = create_versioned_dir(BASE_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(BASE_MODEL_METRICS / run_dir.name)

# -----------------------------------------------------------------
# Load data
dataset_path = DATASET_ENVIRONMENT / "environment_dataset_Conv1D_Tiny.csv"
df = pd.read_csv(dataset_path)

# Timeral feature engineering (preserved)
df['datetime'] = pd.to_datetime(df['datetime'])
df['hour'] = df['datetime'].dt.hour + df['datetime'].dt.minute / 60.0
df['weekday'] = df['datetime'].dt.weekday
df['month'] = df['datetime'].dt.month

# === (1) Selective causal smoothing on H_in (same as the pruned pipeline) ===
if 'H_in' in df.columns:
    df['H_in'] = pd.to_numeric(df['H_in'], errors='coerce')
    df['H_in'] = df['H_in'].ewm(alpha=0.08, adjust=False).mean()

# Cyclical features
df['sin_hour'] = np.sin(2 * np.pi * df['hour'] / 24)
df['cos_hour'] = np.cos(2 * np.pi * df['hour'] / 24)

# Causal lags (official LiteML contract aligned with MLP)
df['T_in_lag1']  = df['T_in'].shift(1)
df['H_in_lag1']  = df['H_in'].shift(1)
df['T_out_lag1'] = df['T_out'].shift(1)
df['H_out_lag1'] = df['H_out'].shift(1)
df['T_in_lag2']  = df['T_in'].shift(2)
df['H_in_lag2']  = df['H_in'].shift(2)
df.dropna(inplace=True)

# === Timeral split (without shuffling) ===
df = df.sort_values('datetime').reset_index(drop=True)

# 12 fixed features of the official LiteML contract 
features = [
    'T_out', 'H_out',
    'T_in_lag1', 'H_in_lag1',
    'T_out_lag1', 'H_out_lag1',
    'T_in_lag2', 'H_in_lag2',
    'sin_hour', 'cos_hour', 'weekday', 'month'
]

# Pure residual target under the LiteML-Edge contract: ΔT_in, ΔH_in
y_all = np.stack([
    (df['T_in'] - df['T_in_lag1']).values,
    (df['H_in'] - df['H_in_lag1']).values
], axis=1).astype(np.float32)
y_all = y_all.astype(np.float32, copy=False)

# === (2b) Timeral sliding window for Conv1D (24 h of history -> 1 prediction) ===
WINDOW_STEPS = 24  # 24 horas
X_source = df[features].values.astype(np.float32)
N_total = len(df)
if N_total < WINDOW_STEPS:
    raise ValueError("Dataset too short for a 24 h sliding window.")

X_seq = []
y_seq = []
idx_seq = []

for t in range(WINDOW_STEPS - 1, N_total):
    X_seq.append(X_source[t-WINDOW_STEPS+1 : t+1, :])
    y_seq.append(y_all[t])
    idx_seq.append(t)

X_all = np.stack(X_seq, axis=0).astype(np.float32)        # (N_seq, WINDOW_STEPS, n_features)
y_all_seq = np.stack(y_seq, axis=0).astype(np.float32)    # (N_seq, 2)
idx_seq = np.array(idx_seq, dtype=np.int64)
N_seq = X_all.shape[0]

# 60/20/20 temporal split over the generated windows
i1, i2 = int(0.6 * N_seq), int(0.8 * N_seq)
X_train, X_val, X_test = X_all[:i1], X_all[i1:i2], X_all[i2:]
y_train, y_val, y_test = y_all_seq[:i1], y_all_seq[i1:i2], y_all_seq[i2:]

# Preserve NON-normalized references for absolute reconstruction on the test split (mapped by idx_seq)
T_prev_all = df['T_in_lag1'].values.astype(np.float32)
H_prev_all = df['H_in_lag1'].values.astype(np.float32)
y_abs_all  = df[['T_in', 'H_in']].values.astype(np.float32)

T_prev_seq = T_prev_all[idx_seq]
H_prev_seq = H_prev_all[idx_seq]
y_abs_seq  = y_abs_all[idx_seq]

T_prev_test = T_prev_seq[i2:]
H_prev_test = H_prev_seq[i2:]
y_test_abs  = y_abs_seq[i2:]  # ground truth absolute

# Test timestamps (to align with HOUR events in the firmware)
dt_seq = df['datetime'].values[idx_seq]
dt_test = dt_seq[i2:]

# Normalization (features and target)  -  preserve MinMaxScaler
scaler_X = MinMaxScaler()
scaler_y = MinMaxScaler()

# X is in the format (N, WINDOW_STEPS, N_FEATURES); flatten the time axis to normalize per feature
N_train, W_steps, F = X_train.shape
X_train_2d = X_train.reshape(-1, F)
scaler_X.fit(X_train_2d)

# LiteML-Edge (1:1 contract): clamp X in the physical domain before the transform
X_train_2d = np.clip(X_train_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_train = scaler_X.transform(X_train_2d).reshape(N_train, W_steps, F)

N_val = X_val.shape[0]
X_val_2d = X_val.reshape(-1, F)
X_val_2d = np.clip(X_val_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_val = scaler_X.transform(X_val_2d).reshape(N_val, W_steps, F)

N_test = X_test.shape[0]
X_test_2d = X_test.reshape(-1, F)
X_test_2d = np.clip(X_test_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_test = scaler_X.transform(X_test_2d).reshape(N_test, W_steps, F)

y_train = scaler_y.fit_transform(y_train)
y_val   = scaler_y.transform(y_val)
y_test  = scaler_y.transform(y_test)

# --- Separate targets by task (T and H) after normalization ---
y_train_T = y_train[:, 0:1]
y_train_H = y_train[:, 1:1+1]
y_val_T   = y_val[:,   0:1]
y_val_H   = y_val[:,   1:1+1]
y_test_T  = y_test[:,  0:1]
y_test_H  = y_test[:,  1:1+1]

# Reshape for temporal 1D CNN: (N, WINDOW_STEPS, n_features)
# (already in the correct format, but enforced explicitly)
n_features = X_train.shape[-1]
X_train = X_train.reshape((-1, WINDOW_STEPS, n_features))
X_val   = X_val.reshape((-1, WINDOW_STEPS, n_features))
X_test  = X_test.reshape((-1, WINDOW_STEPS, n_features))

from tensorflow.keras import layers, models, regularizers

# -----------------------------------------------------------------
# Smaller Tiny Conv1D multi-head
from tensorflow.keras import layers, models, regularizers

def build_base_model(
    input_shape=(WINDOW_STEPS, None),
    n_outputs=2,
    l2_reg=1e-5,
    dp=0.02,
    noise_std=0.003
):
    assert n_outputs == 2, "This model is designed for 2 outputs (ΔT, ΔH)."

    inputs = layers.Input(shape=input_shape, name="inputs_windowxfeat")

    x = layers.GaussianNoise(noise_std, name="gn_train_only")(inputs)

    x = layers.SeparableConv1D(
        filters=12,
        kernel_size=3,
        padding="same",
        activation="relu",
        depthwise_regularizer=regularizers.l2(l2_reg),
        pointwise_regularizer=regularizers.l2(l2_reg),
        name="sepconv1"
    )(x)

    x = layers.DepthwiseConv1D(
        kernel_size=3,
        padding="same",
        activation="relu",
        depthwise_regularizer=regularizers.l2(l2_reg),
        name="dwconv"
    )(x)

    x = layers.GlobalAveragePooling1D(name="gap")(x)
    x = layers.Dropout(dp, name="drop")(x)

    head_T = layers.Dense(1, kernel_regularizer=regularizers.l2(l2_reg), name="head_T")(x)
    head_H = layers.Dense(1, kernel_regularizer=regularizers.l2(l2_reg), name="head_H")(x)

    model = models.Model(
        inputs=inputs,
        outputs=[head_T, head_H],
        name="cnn1d_window24to2_ultratiny_multihead"
    )
    return model

# Instanciar
model = build_base_model(input_shape=(WINDOW_STEPS, n_features), n_outputs=2)

# ====== EMA smoothing + Light fine-tuning (structure preserved) ======
class EMACallback(tf.keras.callbacks.Callback):
    """Batch-level EMA; at the end, replace the model weights with the smoothed weights."""
    def __init__(self, decay=0.9997):
        super().__init__()
        self.decay = decay
        self.shadow = None
    def on_train_begin(self, logs=None):
        self.shadow = [w.numpy().copy() for w in self.model.weights]
    def on_train_batch_end(self, batch, logs=None):
        for i, w in enumerate(self.model.weights):
            self.shadow[i] = self.decay * self.shadow[i] + (1.0 - self.decay) * w.numpy()
    def on_train_end(self, logs=None):
        self.model.set_weights(self.shadow)

# === (3) Loss multi-tarefa ponderada (foco em H_in) ===
# (the function is preserved for future use, but per-head losses are adopted below)
def make_dual_weighted_huber(delta_T=0.25, delta_H=0.5, w_T=1.2, w_H=1.8, aux_mae=0.1):
    huber_T = tf.keras.losses.Huber(delta=delta_T, reduction='none')
    huber_H = tf.keras.losses.Huber(delta=delta_H, reduction='none')
    mae     = tf.keras.losses.MeanAbsoluteError(reduction='none')

    @tf.function(experimental_relax_shapes=True)
    def loss(y_true, y_pred):
        y_true = tf.cast(y_true, tf.float32)
        y_pred = tf.cast(y_pred, tf.float32)
        yT_true, yH_true = y_true[..., 0], y_true[..., 1]
        yT_pred, yH_pred = y_pred[..., 0], y_pred[..., 1]
        lT = tf.reduce_mean(huber_T(yT_true, yT_pred))
        lH = tf.reduce_mean(huber_H(yH_true, yH_pred))
        mT = tf.reduce_mean(mae(yT_true, yT_pred))
        mH = tf.reduce_mean(mae(yH_true, yH_pred))
        total = w_T*(lT + aux_mae*mT) + w_H*(lH + aux_mae*mH)
        return tf.cast(total, tf.float32)
    return loss

# Per-head loss/weight configuration (used below)
losses = {
    "head_T": tf.keras.losses.Huber(delta=0.25),
    "head_H": tf.keras.losses.Huber(delta=0.5),
}
loss_weights = {
    "head_T": 1.5,
    "head_H": 1.5,
}
metrics_dict = {
    "head_T": [tf.keras.metrics.MeanAbsoluteError(name="mae_T")],
    "head_H": [tf.keras.metrics.MeanAbsoluteError(name="mae_H")],
}

# === (4) Schedules + stable optimizers ===
steps_per_epoch = max(1, int(np.ceil(len(X_train) / 256)))
cosine_schedule_1 = tf.keras.optimizers.schedules.CosineDecayRestarts(
    initial_learning_rate=3e-4, first_decay_steps=5*steps_per_epoch, t_mul=2.0, m_mul=0.8, alpha=1e-5/3e-4
)
cosine_schedule_2 = tf.keras.optimizers.schedules.CosineDecayRestarts(
    initial_learning_rate=1e-4, first_decay_steps=3*steps_per_epoch, t_mul=2.0, m_mul=0.8, alpha=1e-6/1e-4
)

def make_adamw(lr_schedule):
    try:
        return tf.keras.optimizers.AdamW(learning_rate=lr_schedule, weight_decay=1e-5, clipnorm=1.0)
    except Exception:
        try:
            return tf.keras.optimizers.legacy.Adam(learning_rate=lr_schedule, clipnorm=1.0)
        except Exception:
            return tf.keras.optimizers.Adam(learning_rate=lr_schedule, clipnorm=1.0)

# ===== Short warm-up stage (does not alter architecture or scaler) =====
# Compile with Adam at a low learning rate for 2 epochs to stabilize the target response
_warmup_opt = tf.keras.optimizers.Adam(learning_rate=1e-5)
model.compile(
    optimizer=_warmup_opt,
    loss=losses,
    loss_weights=loss_weights,
    metrics=metrics_dict,
)
_ = model.fit(
    X_train, {"head_T": y_train_T, "head_H": y_train_H},
    validation_data=(X_val, {"head_T": y_val_T, "head_H": y_val_H}),
    epochs=2, batch_size=256, verbose=0
)

# ===== Stage 1 =====
optimizer_stage1 = make_adamw(cosine_schedule_1)
model.compile(
    optimizer=optimizer_stage1,
    loss=losses,
    loss_weights=loss_weights,
    metrics=metrics_dict,
)
early_stop = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=30, restore_best_weights=True, verbose=1)
# ---- PATCH: Less aggressive EMA (lower decay) to capture micro-variations
ema_cb_1 = EMACallback(decay=0.995)

history = model.fit(
    X_train, {"head_T": y_train_T, "head_H": y_train_H},
    validation_data=(X_val, {"head_T": y_val_T, "head_H": y_val_H}),
    epochs=1000,
    batch_size=256,
    callbacks=[early_stop, ema_cb_1],
    verbose=1
)

# ===== Stage 2 (short fine-tuning) =====
optimizer_stage2 = make_adamw(cosine_schedule_2)
model.compile(
    optimizer=optimizer_stage2,
    loss=losses,
    loss_weights=loss_weights,
    metrics=metrics_dict,
)
early_stop_ft = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=10, restore_best_weights=True, verbose=1)
# ---- PATCH: Less aggressive EMA also during fine-tuning
ema_cb_2 = EMACallback(decay=0.995)

history_ft = model.fit(
    X_train, {"head_T": y_train_T, "head_H": y_train_H},
    validation_data=(X_val, {"head_T": y_val_T, "head_H": y_val_H}),
    epochs=60,
    batch_size=256,
    callbacks=[early_stop_ft, ema_cb_2],
    verbose=1
)

# -----------------------------------------------------------------
# Save the model and scalers (ONLY the already smoothed configuration)
base_model_path = run_dir/'environment_base_model_Conv1D_Tiny.keras'
model.save(str(base_model_path))  # <- only saved model file
joblib.dump(scaler_X, run_dir/"environment_base_model_Conv1D_Tiny_scaler_X.pkl")
joblib.dump(scaler_y, run_dir/"environment_base_model_Conv1D_Tiny_scaler_y.pkl")

# Inference and latency computation
start_time = time.time()
_ = model.predict(X_test, verbose=0)  # now returns [head_T, head_H], but we preserve the structure
end_time = time.time()
inference_time_total = (end_time - start_time) * 1000  # ms
inference_time_per_sample = (inference_time_total / len(X_test)) if len(X_test) > 0 else float('nan')
print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.2f} ms")

# Model size
saved_model_path = base_model_path
if os.path.isfile(saved_model_path):
    size_kb = os.path.getsize(saved_model_path) / 1024
    print(f"Trained model size: {size_kb:.2f} KB ")
else:
    print(" Model file not found.")
    size_kb = float('nan')

# -----------------------------------------------------------------
# Aggregated evaluation on the NORMALIZED scale (y_test vs y_pred_test_scaled)
y_pred_T_scaled, y_pred_H_scaled = model.predict(X_test, verbose=0)
y_pred_test_scaled = np.concatenate([y_pred_T_scaled, y_pred_H_scaled], axis=1)

mse_scaled = mean_squared_error(y_test, y_pred_test_scaled)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled = mean_absolute_error(y_test, y_pred_test_scaled)
r2_scaled  = r2_score(y_test, y_pred_test_scaled)
print("\n Aggregated results (normalized scale):")
print(f"MSE  (normalized residual) = {mse_scaled:.8f}")
print(f"RMSE (normalized residual) = {rmse_scaled:.8f}")
print(f"MAE  (normalized residual) = {mae_scaled:.8f}")
print(f"R²   (normalized residual) = {r2_scaled:.8f}")

# Aggregated evaluation on the ORIGINAL ABSOLUTE scale (T_in, H_in)
y_pred_delta = scaler_y.inverse_transform(y_pred_test_scaled)
T_pred = T_prev_test + y_pred_delta[:, 0]
H_pred = H_prev_test + y_pred_delta[:, 1]
y_pred_orig = np.stack([T_pred, H_pred], axis=1)

# Global metrics (joint T_in + H_in evaluation)
mse = mean_squared_error(y_test_abs, y_pred_orig)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_abs, y_pred_orig)
r2  = r2_score(y_test_abs, y_pred_orig)

# >>> Per-output metrics (including R²)
mse_T  = mean_squared_error(y_test_abs[:, 0], y_pred_orig[:, 0])
rmse_T = np.sqrt(mse_T)
mae_T  = mean_absolute_error(y_test_abs[:, 0], y_pred_orig[:, 0])
r2_T   = r2_score(y_test_abs[:, 0], y_pred_orig[:, 0])

mse_H  = mean_squared_error(y_test_abs[:, 1], y_pred_orig[:, 1])
rmse_H = np.sqrt(mse_H)
mae_H  = mean_absolute_error(y_test_abs[:, 1], y_pred_orig[:, 1])
r2_H   = r2_score(y_test_abs[:, 1], y_pred_orig[:, 1])

print("\n Per variable (original scale):")
print(f"[T_in] MSE={mse_T:.6f} RMSE={rmse_T:.6f} MAE={mae_T:.6f} R²={r2_T:.6f}")
print(f"[H_in] MSE={mse_H:.6f} RMSE={rmse_H:.6f} MAE={mae_H:.6f} R²={r2_H:.6f}")

print("\n Aggregated results (original scale):")
mse_status  = " MSE within threshold"  if mse  <= 0.1 else " MSE above threshold"
rmse_status = " RMSE within threshold" if rmse <= 0.32 else " RMSE above threshold"
mae_status  = " MAE within threshold"  if mae  <= 0.3 else " MAE above threshold"
r2_status   = " R² within threshold"   if r2   >= 0.8 else " R² below threshold"
print(f"MSE  (original) = {mse:.8f}   {mse_status}")
print(f"RMSE (original) = {rmse:.8f}   {rmse_status}")
print(f"MAE  (original) = {mae:.8f}   {mae_status}")
print(f"R²   (original) = {r2:.8f}   {r2_status}")

# Overall model status
model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied " if model_ok else "Performance thresholds not satisfied "
print("\n Result:", model_status)

# ============================================================
# New: Rolling metrics over a 24-sample moving window (firmware-equivalent)
# ============================================================
ROLLING_N = 24  # N=24 samples (HOUR) as in metrics.cpp

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """Replicate the firmware logic (metrics.cpp):
      R² = 1 - SS_res/SS_tot
      SS_tot = sum(y^2) - (sum(y)^2)/n
      Se n < 2 -> NaN
      Se SS_tot <= 1e-6 -> NaN
    """
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return float("nan")

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))

    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return float("nan")

    return 1.0 - (ss_res / ss_tot)


def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    """
    Returns (mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H)
    seguindo exatamente:
      mae = 0.5*(mae_T + mae_H)
      rmse = sqrt( 0.5*(mse_T + mse_H) )
      r2 = 0.5*(r2_T + r2_H)
    """
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # T
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T * err_T))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = float(_r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0]))

    # H
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H * err_H))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = float(_r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1]))

    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H)
    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

rolling_rows = []
n_test = y_test_abs.shape[0]

# Gating identical to the firmware (offline: all samples count)
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)

if n_test < ROLLING_N:
    print(f"\n Rolling N=24 cannot be computed: the test split has only {n_test} samples.")
else:
    for end in range(ROLLING_N - 1, n_test):
        # Firmware: metrics are updated only when an actual Invoke occurs and the event is HOUR
        if (not invoked_mask[end]) or (not is_rollover_mask[end]):
            rolling_rows.append({
                "window_start": int(end - ROLLING_N + 1),
                "window_end": int(end),
                "datetime_end": pd.to_datetime(dt_test[end]),
                "N": int(ROLLING_N),

                "MAE": float("nan"),
                "RMSE": float("nan"),
                "R2": float("nan"),

                "MAE_T": float("nan"),
                "RMSE_T": float("nan"),
                "R2_T": float("nan"),

                "MAE_H": float("nan"),
                "RMSE_H": float("nan"),
                "R2_H": float("nan"),
            })
            continue
        start = end - ROLLING_N + 1
        yt_w = y_test_abs[start:end+1, :]
        yp_w = y_pred_orig[start:end+1, :]

        mae_w, rmse_w, r2_w, maeT_w, rmseT_w, r2T_w, maeH_w, rmseH_w, r2H_w = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": pd.to_datetime(dt_test[end]),
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": maeT_w,
            "RMSE_T": rmseT_w,
            "R2_T": r2T_w,

            "MAE_H": maeH_w,
            "RMSE_H": rmseH_w,
            "R2_H": r2H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # Useful summaries to compare with the current firmware state
    last = df_roll.iloc[-1]
    mean_roll = df_roll[["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]].mean()

    print("\n Rolling N=24 (firmware-equivalent)  -  LAST window (current state):")
    print(f"MAE={last['MAE']:.4f} RMSE={last['RMSE']:.4f} R²={last['R2']:.4f} | "
          f"T: MAE={last['MAE_T']:.4f} RMSE={last['RMSE_T']:.4f} R²={last['R2_T']:.4f} | "
          f"H: MAE={last['MAE_H']:.4f} RMSE={last['RMSE_H']:.4f} R²={last['R2_H']:.4f}")

    print("\n Rolling N=24  -  MEAN across all TEST windows:")
    print(f"MAE={mean_roll['MAE']:.4f} RMSE={mean_roll['RMSE']:.4f} R²={mean_roll['R2']:.4f} | "
          f"T: MAE={mean_roll['MAE_T']:.4f} RMSE={mean_roll['RMSE_T']:.4f} R²={mean_roll['R2_T']:.4f} | "
          f"H: MAE={mean_roll['MAE_H']:.4f} RMSE={mean_roll['RMSE_H']:.4f} R²={mean_roll['R2_H']:.4f}")

    # Save rolling metrics to CSV/Excel (for 1:1 comparison with firmware logs)
    df_roll = df_roll.round(4)
    df_roll.to_csv(metrics_run_dir / "environment_base_model_Conv1D_Tiny_metrics_rolling24.csv",
                   index=False, encoding="utf-8-sig")
    excel_roll_path = metrics_run_dir / "environment_base_model_Conv1D_Tiny_metrics_rolling24.xlsx"
    df_roll.to_excel(excel_roll_path, index=False)

    # Auto-adjust Excel columns
    wb2 = load_workbook(excel_roll_path)
    ws2 = wb2.active
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb2.save(excel_roll_path)

    print(" Files saved:")
    print(" - environment_base_model_Conv1D_Tiny_metrics_rolling24.csv")
    print(" - environment_base_model_Conv1D_Tiny_metrics_rolling24.xlsx")

    # Plot: rolling (MAE/RMSE/R2)
    plt.figure(figsize=(10, 5))
    plt.plot(df_roll["datetime_end"], df_roll["MAE"], label="MAE (rolling24)")
    plt.plot(df_roll["datetime_end"], df_roll["RMSE"], label="RMSE (rolling24)")
    plt.title("Metrics over a Moving Window N=24 (comparable to the firmware)")
    plt.xlabel("Datetime (window end)")
    plt.ylabel("Value")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_base_model_Conv1D_Tiny_metrics_rolling24_mae_rmse.png")
    plt.close()

    plt.figure(figsize=(10, 4))
    plt.plot(df_roll["datetime_end"], df_roll["R2"], label="R² (rolling24)")
    plt.title("R² over a Moving Window N=24 (comparable to the firmware)")
    plt.xlabel("Datetime (window end)")
    plt.ylabel("R²")
    plt.ylim([-0.2, 1.05])
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_base_model_Conv1D_Tiny_metrics_rolling24_r2.png")
    plt.close()


# -----------------------------------------------------------------
# Overfitting / underfitting diagnosis
train_loss = history.history['loss'] + history_ft.history['loss']
val_loss   = history.history['val_loss'] + history_ft.history['val_loss']
n = min(5, len(train_loss))
mean_train_loss = float(np.mean(train_loss[-n:])) if n > 0 else float('nan')
mean_val_loss   = float(np.mean(val_loss[-n:])) if n > 0 else float('nan')
gap = abs(mean_val_loss - mean_train_loss) if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss) else float('nan')
gap_pct = (gap / mean_train_loss) * 100 if (np.isfinite(gap) and mean_train_loss > 0) else float('nan')

if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss):
    if mean_train_loss > 0.3 and mean_val_loss > 0.3:
        status = "Underfitting detected (high losses)"
    elif mean_val_loss < mean_train_loss * 0.8:
        status = "Potential underfitting  (validation loss significantly lower than training loss)"
    elif (gap_pct > 50) or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
        status = "Overfitting detected  (large generalization gap or significant divergence)"
    elif gap_pct < 10:
        status = "Well-fitted model  (gap < 10%)"
    elif gap_pct < 30:
        status = "Acceptably fitted model  (gap < 30%)"
    else:
        status = "Mild overfitting  (gap moderado)"
else:
    status = "Insufficient history for fit diagnosis."

print("\n Model diagnostics:")
print(f"- Mean training loss:     {mean_train_loss:.8f}")
print(f"- Mean validation loss:  {mean_val_loss:.8f}")
print(f"- Absolute gap:           {gap:.8f}")
print(f"- Generalization gap:         {gap_pct:.2f}%")
print(f"- Status:                 {status}")

# -----------------------------------------------------------------
# Plots
plt.figure(figsize=(8, 5))
plt.plot(train_loss, label='Treinamento')
plt.plot(val_loss, label='Validation')
plt.title("Loss during training with EarlyStopping + EMA\n" + status)
plt.xlabel("Epoch")
plt.ylabel("Loss (Huber)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_base_model_Conv1D_Tiny_training_validation_loss_diagnosis.png")
plt.close()

# Metrics (bar chart) aggregated on the original scale
metric_labels = ['MSE', 'RMSE', 'MAE', 'R²']
metric_values = [mse, rmse, mae, r2]
plt.figure(figsize=(8, 5))
bars = plt.bar(metric_labels, metric_values)
plt.title(f"Aggregated evaluation metrics  -  {'satisfied' if model_ok else 'not satisfied'}")
plt.ylabel("Value")
plt.ylim([0, max(1e-9, max(metric_values)) * 1.2])
for bar in bars:
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width() / 2, yval + (0.01 if np.isfinite(yval) else 0.0), f'{yval:.8f}', ha='center', va='bottom')
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_base_model_Conv1D_Tiny_final_metrics_summary_plot.png")
plt.close()

# Prediction scatter plot (original scale)
plt.figure(figsize=(6, 6))
plt.scatter(y_test_abs[:, 0], y_pred_orig[:, 0], alpha=0.5, label="Temperature (T_in)",)
plt.scatter(y_test_abs[:, 1], y_pred_orig[:, 1], alpha=0.5, label="Humidity (H_in)",)
min_val = float(min(y_test_abs.min(), y_pred_orig.min()))
max_val = float(max(y_test_abs.max(), y_pred_orig.max()))
plt.plot([min_val, max_val], [min_val, max_val], 'k--')
plt.xlabel("Value Real")
plt.ylabel("Value Previsto")
plt.title("Prediction scatter plot (original scale absolute)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_base_model_Conv1D_Tiny_scatter_predictions.png")
plt.close()

# -----------------------------------------------------------------
# Scatter plots by variable (T_in and H_in separately)

# --- Temperature ---
plt.figure(figsize=(6, 6))
plt.scatter(y_test_abs[:, 0], y_pred_orig[:, 0], alpha=0.6, label="Predicted T_in")
min_val_T = float(min(y_test_abs[:, 0].min(), y_pred_orig[:, 0].min()))
max_val_T = float(max(y_test_abs[:, 0].max(), y_pred_orig[:, 0].max()))
plt.plot([min_val_T, max_val_T], [min_val_T, max_val_T], 'k--', label="Linha 1:1")
plt.xlabel("T_in Ground Truth (°C)")
plt.ylabel("T_in Prediction (°C)")
plt.title("Scatter plot  -  Temperature (T_in)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_base_model_Conv1D_Tiny_scatter_T_in.png")
plt.close()

# --- Humidity ---
plt.figure(figsize=(6, 6))
plt.scatter(y_test_abs[:, 1], y_pred_orig[:, 1], alpha=0.6, color='tab:orange', label="Predicted H_in")
min_val_H = float(min(y_test_abs[:, 1].min(), y_pred_orig[:, 1].min()))
max_val_H = float(max(y_test_abs[:, 1].max(), y_pred_orig[:, 1].max()))
plt.plot([min_val_H, max_val_H], [min_val_H, max_val_H], 'k--', label="Linha 1:1")
plt.xlabel("H_in Ground Truth (%)")
plt.ylabel("H_in Prediction (%)")
plt.title("Scatter plot  -  Humidity (H_in)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir/"environment_base_model_Conv1D_Tiny_scatter_H_in.png")
plt.close()

# -----------------------------------------------------------------
# Metrics table (aggregated + separated, WITHOUT "residual" in the label)
# =========================
# New: Firmware-style Rolling(24) summary for metrics_summary export
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)

metrics_dist = {
    "Base Model Metrics": [
        # Normalized (aggregate over T_in + H_in)
        "MSE (normalized residual joint)", "RMSE (normalized residual joint)", "MAE (normalized residual joint)", "R² (normalized residual joint)",
        # Joint original scale
        "MSE (original joint)", "RMSE (original joint)", "MAE (original joint)", "R² (original joint)",
        # Per-variable T_in
        "MSE_T (T_in)", "RMSE_T (T_in)", "MAE_T (T_in)", "R²_T (T_in)",
        # Per-variable H_in
        "MSE_H (H_in)", "RMSE_H (H_in)", "MAE_H (H_in)", "R²_H (H_in)",
        # Rolling(24) firmware-equivalent (window=24, HOUR+Invoke gating)
        "N_ALL (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, last window)", "RMSE (rolling24 firmware, last window)", "R² (rolling24 firmware, last window)",
        "MAE_T (rolling24 firmware, last window)", "RMSE_T (rolling24 firmware, last window)", "R²_T (rolling24 firmware, last window)",
        "MAE_H (rolling24 firmware, last window)", "RMSE_H (rolling24 firmware, last window)", "R²_H (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, test mean)", "RMSE (rolling24 firmware, test mean)", "R² (rolling24 firmware, test mean)",
        "MAE_T (rolling24 firmware, test mean)", "RMSE_T (rolling24 firmware, test mean)", "R²_T (rolling24 firmware, test mean)",
        "MAE_H (rolling24 firmware, test mean)", "RMSE_H (rolling24 firmware, test mean)", "R²_H (rolling24 firmware, test mean)",
        # Outros
        "Model size (KB)",
        "Mean Training Loss", "Mean Validation Loss",
        "Absolute Gap", "Gap Percentage (%)",
        "Total Inference Time (ms)", "Inference Time per Sample (ms)",
        "Fit Status", "Model Status"
    ],
    "Value": [
        f"{mse_scaled:.8f}", f"{rmse_scaled:.8f}", f"{mae_scaled:.8f}", f"{r2_scaled:.8f}",
        f"{mse:.8f}",        f"{rmse:.8f}",        f"{mae:.8f}",        f"{r2:.8f}",
        f"{mse_T:.8f}",      f"{rmse_T:.8f}",      f"{mae_T:.8f}",      f"{r2_T:.8f}",
        f"{mse_H:.8f}",      f"{rmse_H:.8f}",      f"{mae_H:.8f}",      f"{r2_H:.8f}",
        f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else "",
        f"{fw_last['MAE']:.4f}" if np.isfinite(fw_last['MAE']) else "",
        f"{fw_last['RMSE']:.4f}" if np.isfinite(fw_last['RMSE']) else "",
        f"{fw_last['R2']:.4f}" if np.isfinite(fw_last['R2']) else "",
        f"{fw_last['MAE_T']:.4f}" if np.isfinite(fw_last['MAE_T']) else "",
        f"{fw_last['RMSE_T']:.4f}" if np.isfinite(fw_last['RMSE_T']) else "",
        f"{fw_last['R2_T']:.4f}" if np.isfinite(fw_last['R2_T']) else "",
        f"{fw_last['MAE_H']:.4f}" if np.isfinite(fw_last['MAE_H']) else "",
        f"{fw_last['RMSE_H']:.4f}" if np.isfinite(fw_last['RMSE_H']) else "",
        f"{fw_last['R2_H']:.4f}" if np.isfinite(fw_last['R2_H']) else "",
        f"{fw_mean['MAE']:.4f}" if np.isfinite(fw_mean['MAE']) else "",
        f"{fw_mean['RMSE']:.4f}" if np.isfinite(fw_mean['RMSE']) else "",
        f"{fw_mean['R2']:.4f}" if np.isfinite(fw_mean['R2']) else "",
        f"{fw_mean['MAE_T']:.4f}" if np.isfinite(fw_mean['MAE_T']) else "",
        f"{fw_mean['RMSE_T']:.4f}" if np.isfinite(fw_mean['RMSE_T']) else "",
        f"{fw_mean['R2_T']:.4f}" if np.isfinite(fw_mean['R2_T']) else "",
        f"{fw_mean['MAE_H']:.4f}" if np.isfinite(fw_mean['MAE_H']) else "",
        f"{fw_mean['RMSE_H']:.4f}" if np.isfinite(fw_mean['RMSE_H']) else "",
        f"{fw_mean['R2_H']:.4f}" if np.isfinite(fw_mean['R2_H']) else "",
        f"{size_kb:.2f} KB",
        f"{mean_train_loss:.8f}", f"{mean_val_loss:.8f}",
        f"{gap:.8f}",            f"{gap_pct:.2f} %",
        f"{inference_time_total:.2f} ms", f"{inference_time_per_sample:.2f} ms",
        "", ""
    ],
    "Status": [
        "", "", "", "",
        mse_status, rmse_status, mae_status, r2_status,
        "", "", "", "",
        "", "", "", "",
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", status, model_status
    ],
    "Meaning": [
        "Mean squared error on the normalized scale (T_in e H_in em joint).",
        "Root mean squared error on the normalized scale.",
        "Mean absolute error on the normalized scale.",
        "Coefficient of determination on the normalized scale.",
        "Mean squared error on the original scale (for example, °C and %RH)  -  joint [T_in, H_in].",
        "Root mean squared error on the original scale  -  joint.",
        "Mean absolute error on the original scale  -  joint.",
        "Coefficient of determination on the original scale  -  joint.",
        "Mean squared error on the original scale for T_in.",
        "Root mean squared error on the original scale for T_in.",
        "Mean absolute error on the original scale for T_in.",
        "Coefficient of determination on the original scale for T_in.",
        "Mean squared error on the original scale for H_in.",
        "Root mean squared error on the original scale for H_in.",
        "Mean absolute error on the original scale for H_in.",
        "Coefficient of determination on the original scale for H_in.",
        "Rolling(24): number of samples (HOUR events) effectively present in the final window.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",
        "Final model file size.",
        "Mean training loss over the last epochs.",
        "Mean validation loss over the last epochs.",
        "Absolute difference between mean losses.",
        "Generalization gap between losses.",
        "Total inference time in ms.",
        "Average time per sample in ms.",
        "Diagnosis based on the gap.",
        "Overall diagnosis based on the threshold limits."
    ],
    "Expected Values / Thresholds": [
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Ideally > 0.95.",
        "→ < 0.1 excelente (depends on the problem).",
        "→ < 0.32 as a reference.",
        "→ < 0.3 as a reference.",
        "→ > 0.8 desirable.",
        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Ideally > 0.8 (T_in).",
        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Ideally > 0.8 (H_in).",
        "→ Prefer < 256 KB on constrained MCUs.",
        "→ Low (for example, < 0.01).",
        "→ Close to the training loss.",
        "→ < 0.05 is good.",
        "→ < 10% is excellent.",
        "→ Lower is better.",
        "→ < 1 ms is ideal in TinyML.",
        "→ Well-fitted / Overfitting / Underfitting.",
        "→ Satisfactory when all metrics remain within the limits."
    ]
}


# --- Sanity: ensure that all summary columns have the same length (avoids a pandas ValueError) ---
try:
    _max_len = max(len(v) for v in metrics_dist.values() if isinstance(v, list))
    for _k, _v in metrics_dist.items():
        if isinstance(_v, list) and len(_v) < _max_len:
            _v.extend([""] * (_max_len - len(_v)))
except Exception as _e:
    print("[WARN] Could not normalize metrics_dist:", _e)

dfm = pd.DataFrame(metrics_dist)
dfm.to_csv(metrics_run_dir/"environment_base_model_Conv1D_Tiny_metrics_summary.csv", index=False, encoding="utf-8-sig")
print(" File saved: environment_base_model_Conv1D_Tiny_metrics_summary.csv")

excel_path = metrics_run_dir/"environment_base_model_Conv1D_Tiny_metrics_summary.xlsx"
dfm.to_excel(excel_path, index=False)
wb = load_workbook(excel_path)
ws = wb.active
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
wb.save(excel_path)
print(" File saved: environment_base_model_Conv1D_Tiny_metrics_summary.xlsx (with adjusted columns)")

# === Post-execution: update 'latest' e manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Could not update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Could not write manifest.json:", _e)

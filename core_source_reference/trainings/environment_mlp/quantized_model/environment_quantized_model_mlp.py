"""
Script: environment_quantized_model_mlp.py
Module role:
    Train, convert, evaluate, and export the quantized MLP artifacts used in
    the LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, performs quantization-aware training, converts the
    model to TensorFlow Lite INT8, evaluates the resulting artifact in
    normalized and reconstructed physical domains, and exports reference tables,
    figures, and firmware-support artifacts.

Inputs:
    - environment_dataset_mlp.csv
    - Baseline and pruned model artifacts
    - Project path and versioning utilities

Outputs:
    - Versioned quantized TensorFlow Lite artifact
    - Rolling24 evaluation tables and reference spreadsheets
    - Debug workbooks, replay headers, and figure files

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys, random
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.model_selection import train_test_split
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr
import tensorflow_model_optimization as tfmot

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
rcParams["font.family"] = "Segoe UI Emoji"

# --- Bootstrap: allows importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base)
            break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_mlp import (
    PROJECT_ROOT,
    DATASET_ENVIRONMENT_MLP,
    BASE_MODEL,
    PRUNED_MODEL,
    QUANTIZED_MODEL,
    QUANTIZED_MODEL_METRICS,
)
from utils.global_utils.versioning import (
    create_versioned_dir,
    ensure_dir,
    update_latest,
    write_manifest,
    list_runs,
    resolve_run,
    resolve_latest,
)
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Must be called at the top of the script, BEFORE model creation


# Figure export helpers
# Save figure files (PNG, PDF, SVG) to quantization_graphics
def _savefig_pub(stem: str):
    """Save the current matplotlib figure to quantization_graphics as PNG (600 dpi), PDF, and SVG."""
    png_path = graphics_dir / f"{stem}.png"
    pdf_path = graphics_dir / f"{stem}.pdf"
    svg_path = graphics_dir / f"{stem}.svg"
    plt.savefig(png_path, dpi=600, bbox_inches="tight")
    plt.savefig(pdf_path, bbox_inches="tight")
    plt.savefig(svg_path, bbox_inches="tight")
    print(f"[INFO] Figure saved: {png_path}")

# === Versioned directories for the current execution ===
run_dir = create_versioned_dir(QUANTIZED_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(QUANTIZED_MODEL_METRICS / run_dir.name)

# Output subfolders (inside metrics_run_dir) for paper-ready organization
graphics_dir = ensure_dir(metrics_run_dir / "quantization_graphics")
results_dir  = ensure_dir(metrics_run_dir / "quantization_metrics_results")
base_version_path = resolve_latest(BASE_MODEL)
pruned_version_path = resolve_latest(PRUNED_MODEL)

# ---------------------
# Paths
# ---------------------
model_path      = pruned_version_path / "environment_pruned_model_mlp.keras"
scaler_X_path   = base_version_path / "environment_base_model_mlp_scaler_X.pkl"
scaler_y_path   = base_version_path / "environment_base_model_mlp_scaler_y.pkl"
dataset_path    = DATASET_ENVIRONMENT_MLP / "environment_dataset_mlp.csv"
quantized_path  = run_dir / "environment_quantized_model_mlp.tflite"

# ---------------------
# Load dataset
# ---------------------
df_raw = pd.read_csv(dataset_path)

# Sort temporally to preserve a deterministic raw-real timeline
# that can later be used for firmware replay/header exports.
df_raw = df_raw.sort_values("datetime").reset_index(drop=True)
df_raw["__row_id_raw"] = np.arange(len(df_raw), dtype=np.int64)

df_raw["datetime"] = pd.to_datetime(df_raw["datetime"])
for _c in ["T_out", "H_out", "T_in", "H_in"]:
    if _c in df_raw.columns:
        df_raw[_c] = pd.to_numeric(df_raw[_c], errors="coerce")

# Raw-real dataframe preserved for replay/header exports.
df_raw_real = df_raw.copy()

# Processed dataframe used by the offline LiteML-Edge pipeline.
df = df_raw.copy()

# Temporal feature engineering
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in ===
HIN_EMA_ALPHA = 0.08
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=HIN_EMA_ALPHA, adjust=False).mean()

# Cyclical attributes
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lag features (first-order causal dependencies)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)
df_processed_row_ids = df["__row_id_raw"].astype(np.int64).to_numpy()

# Fixed feature order (12 features)
features = [
    "T_out", "H_out",           # current outdoor
    "T_in_lag1", "H_in_lag1",   # indoor lag1
    "T_out_lag1", "H_out_lag1", # outdoor lag1
    "T_in_lag2", "H_in_lag2",   # indoor lag2
    "sin_hour", "cos_hour",     # cyclical hour
    "weekday", "month",         # calendar attributes
]

targets = ["T_in", "H_in"]  # absolute targets retained for reconstruction

# =====================================================================
#        DEFINITION OF ABSOLUTE TARGETS, RESIDUAL TARGETS, AND SWITCH
# =====================================================================
# Training is always performed on residual targets (ΔT_in, ΔH_in) to preserve
# consistency with the firmware-side output contract.
USE_RESIDUAL_TARGETS = True  # retained for logging and documentation

# Full absolute targets (used later for physical-domain reconstruction)
y_full_abs = df[targets].values.astype(np.float32)

# Pure residual targets: ΔT_in, ΔH_in
y_full_res = np.stack(
    [
        df["T_in"].values.astype(np.float32) - df["T_in_lag1"].values.astype(np.float32),
        df["H_in"].values.astype(np.float32) - df["H_in_lag1"].values.astype(np.float32),
    ],
    axis=1,
).astype(np.float32)

# Primary model target (always residual)
y_full_main = y_full_res

# =====================================================================
#            SLIDING WINDOW 24×12 → MLP TARGETS
# =====================================================================
WINDOW   = 24
X_source = df[features].values.astype(np.float32)
N        = len(df)

if N < WINDOW:
    raise ValueError("Insufficient dataset to construct 24-step windows.")

X_win_list = []
y_main_list = []
y_abs_list  = []
idx_list    = []

for t in range(WINDOW - 1, N):
    # Window [t-23, ..., t] → (24, 12)
    window = X_source[t - WINDOW + 1 : t + 1, :]
    X_win_list.append(window)
    y_main_list.append(y_full_main[t])   # residual target used in training
    y_abs_list.append(y_full_abs[t])     # absolute target retained for evaluation
    idx_list.append(t)

# 3D tensor representation: (N_seq, 24, 12)
X_win      = np.stack(X_win_list).astype(np.float32)
y_all_main = np.stack(y_main_list).astype(np.float32)
y_all_abs  = np.stack(y_abs_list).astype(np.float32)
idx_arr    = np.array(idx_list, dtype=np.int64)

# Auxiliary vectors for absolute reconstruction
T_prev_all = df["T_in_lag1"].values.astype(np.float32)[idx_arr]
H_prev_all = df["H_in_lag1"].values.astype(np.float32)[idx_arr]
y_abs_all  = y_all_abs  # semantic alias

# ---------------------
# Load scalers (same as in the base model)
# ---------------------
scaler_X = joblib.load(scaler_X_path)
scaler_y = joblib.load(scaler_y_path)

# ===== CORRECT NORMALIZATION: 12 FEATURES FIRST → THEN FLATTEN 24×12 =====
N_seq, W_steps, F = X_win.shape  # F must be 12
X_2d        = X_win.reshape(-1, F)
X_2d = np.clip(X_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
# LiteML-Edge contract: clamp X in the physical domain to keep minmax_forward within [0,1]
X_2d_scaled = scaler_X.transform(X_2d)
X_scaled    = X_2d_scaled.reshape(N_seq, W_steps * F)

# Normalized residual target
y_scaled = scaler_y.transform(y_all_main)

# =====================================================================
#                 TRAIN / VAL / TEST SPLIT (TEMPORAL)
# =====================================================================
n_total = X_scaled.shape[0]

train_frac = 0.6
val_frac   = 0.2  # the remainder is assigned to the test set

n_train = int(n_total * train_frac)
n_val   = int(n_total * val_frac)
n_test  = n_total - n_train - n_val

if n_train <= 0 or n_val <= 0 or n_test <= 0:
    raise ValueError(
        f"Invalid split with n_total={n_total}, "
        f"n_train={n_train}, n_val={n_val}, n_test={n_test}"
    )

# --- TRAIN: oldest segment ---
X_train     = X_scaled[:n_train]
y_train     = y_scaled[:n_train]
Tprev_train = T_prev_all[:n_train]
Hprev_train = H_prev_all[:n_train]
yabs_train  = y_abs_all[:n_train]

# --- VALIDATION: intermediate segment ---
X_val     = X_scaled[n_train : n_train + n_val]
y_val     = y_scaled[n_train : n_train + n_val]
Tprev_val = T_prev_all[n_train : n_train + n_val]
Hprev_val = H_prev_all[n_train : n_train + n_val]
yabs_val  = y_abs_all[n_train : n_train + n_val]

# --- TEST: most recent segment ---
X_test     = X_scaled[n_train + n_val :]
y_test     = y_scaled[n_train + n_val :]
Tprev_test = T_prev_all[n_train + n_val :]
Hprev_test = H_prev_all[n_train + n_val :]
yabs_test  = y_abs_all[n_train + n_val :]

target_mode = "RESIDUAL (ΔT_in, ΔH_in)" if USE_RESIDUAL_TARGETS else "ABSOLUTE (T_in, H_in)"
print(f"[INFO] Temporal split applied (QAT/INT8 - target {target_mode}):")
print(f"  • Training:   {n_train} samples")
print(f"  • Validation: {n_val} samples")
print(f"  • Test:       {n_test} samples")

# ---------------------
# Load original model (pruned, non-quantized)
# ---------------------
base_model = tf.keras.models.load_model(model_path)

# ---------------------
# Define QAT model (quantization-aware training)
# ---------------------
quantize_model = tfmot.quantization.keras.quantize_model
qat_model = quantize_model(base_model)

qat_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),
    loss=tf.keras.losses.MeanSquaredError(),
    metrics=[tf.keras.metrics.MeanAbsoluteError()],
)

qat_model.summary()

# ---------------------
# Callbacks
# ---------------------
callbacks = [
    tf.keras.callbacks.EarlyStopping(
        monitor="val_loss",
        patience=10,
        restore_best_weights=True,
        verbose=1,
    ),
]

# ---------------------
# Training
# ---------------------
epochs_qat = 20
batch_size = 512

history = qat_model.fit(
    X_train,
    y_train,
    validation_data=(X_val, y_val),
    epochs=epochs_qat,
    batch_size=batch_size,
    shuffle=True,
    callbacks=callbacks,
    verbose=1,
)

# ---------------------
# Convert to TFLite after QAT
# ---------------------
converter = tf.lite.TFLiteConverter.from_keras_model(qat_model)
converter.optimizations = [tf.lite.Optimize.DEFAULT]

def representative_dataset():
    n = min(2000, len(X_train))
    idx = np.linspace(0, len(X_train)-1, n, dtype=int)

    for i in range(0, len(idx), 16):
        batch = X_train[idx[i:i+16]].astype(np.float32)
        yield [batch]

converter.representative_dataset       = representative_dataset
converter.target_spec.supported_ops    = [tf.lite.OpsSet.TFLITE_BUILTINS_INT8]
converter.inference_input_type         = tf.int8
converter.inference_output_type        = tf.int8

tflite_quant_model = converter.convert()

with open(quantized_path, "wb") as f:
    f.write(tflite_quant_model)

print(f"[INFO] Quantized model saved to: {quantized_path}")


DBG_MODEL_RAW_MAX_DIMS = 8
DBG_MODEL_RAW_MAX_BYTES = 32

def _tflite_dtype_code(dtype) -> int:
    _dt = np.dtype(dtype)
    _map = {
        np.dtype(np.float32): 1,
        np.dtype(np.int32): 2,
        np.dtype(np.uint8): 3,
        np.dtype(np.int64): 4,
        np.dtype(np.str_): 5,
        np.dtype(np.bool_): 6,
        np.dtype(np.int16): 7,
        np.dtype(np.complex64): 8,
        np.dtype(np.int8): 9,
        np.dtype(np.float16): 10,
        np.dtype(np.float64): 11,
        np.dtype(np.complex128): 12,
        np.dtype(np.uint64): 13,
        np.dtype(np.uint32): 16,
        np.dtype(np.uint16): 17,
    }
    return int(_map.get(_dt, -1))


def _tflite_dtype_name(dtype) -> str:
    try:
        return np.dtype(dtype).name
    except Exception:
        return str(dtype)


def _capture_tflite_output_raw_records(interpreter, output_details,
                                       max_dims: int = DBG_MODEL_RAW_MAX_DIMS,
                                       max_dump_bytes: int = DBG_MODEL_RAW_MAX_BYTES):
    records = []
    for out_idx, od in enumerate(output_details):
        y_q = interpreter.get_tensor(od["index"])
        y_arr = np.asarray(y_q)
        y_arr_c = np.ascontiguousarray(y_arr)
        raw_bytes = y_arr_c.tobytes(order="C")
        shape = list(y_arr_c.shape)
        bytes_total = int(len(raw_bytes))
        bytes_dumped = int(min(bytes_total, int(max_dump_bytes)))

        rec = {
            "out_idx": int(out_idx),
            "tensor_index": int(od["index"]),
            "type_code": _tflite_dtype_code(y_arr_c.dtype),
            "type_name": _tflite_dtype_name(y_arr_c.dtype),
            "bytes_total": bytes_total,
            "bytes_dumped": bytes_dumped,
            "dims_size": int(len(shape)),
        }

        for di in range(int(max_dims)):
            rec[f"dim{di:02d}"] = int(shape[di]) if di < len(shape) else -1

        for bi in range(int(max_dump_bytes)):
            rec[f"b{bi:02d}_hex"] = f"0x{raw_bytes[bi]:02X}" if bi < bytes_dumped else ""

        records.append(rec)

    return records

# ---------------------
# Auxiliary function for TFLite inference
# ---------------------
def run_tflite_inference(tflite_model_path, X, capture_io: bool = False):
    interpreter = tf.lite.Interpreter(model_path=str(tflite_model_path))
    interpreter.allocate_tensors()

    input_details  = interpreter.get_input_details()
    output_details = interpreter.get_output_details()

    input_info = input_details[0]
    output_info = output_details[0]
    input_scale, input_zero_point   = input_info["quantization"]
    output_scale, output_zero_point = output_info["quantization"]
    input_dtype = input_info["dtype"]
    output_dtype = output_info["dtype"]

    if input_dtype == np.float32:
        X_infer = X.astype(np.float32, copy=False)
    else:
        X_infer = X / input_scale + input_zero_point
        if input_dtype == np.int8:
            X_infer = np.clip(np.round(X_infer), -128, 127).astype(np.int8)
        elif input_dtype == np.uint8:
            X_infer = np.clip(np.round(X_infer), 0, 255).astype(np.uint8)
        else:
            X_infer = np.round(X_infer).astype(input_dtype)

    y_pred_float = []
    captured_input_payload = []
    captured_input_float = []
    captured_output_payload = []
    captured_output_float = []
    captured_output_raw = []

    for i in range(X.shape[0]):
        interpreter.set_tensor(input_info["index"], X_infer[i : i + 1])
        interpreter.invoke()

        output_raw_records = _capture_tflite_output_raw_records(interpreter, output_details)
        y_payload = interpreter.get_tensor(output_info["index"])
        y_payload = np.asarray(y_payload)
        if output_dtype == np.float32:
            y_float = y_payload.astype(np.float32, copy=False)
        else:
            y_float = (y_payload.astype(np.float32) - output_zero_point) * output_scale

        y_pred_float.append(np.asarray(y_float[0], dtype=np.float32))

        if capture_io:
            x_payload = np.asarray(X_infer[i], dtype=np.float32).reshape(-1)

            # IMPORTANT:
            #   - input_payload stores the exact quantized tensor values sent to TFLite.
            #   - input_float stores the original normalized float vector BEFORE quantization.
            # This matches the expected Python-side debug contract for comparison against
            # firmware logs that report both payload (p*) and float-domain input (x*).
            x_float = np.asarray(X[i], dtype=np.float32).reshape(-1)

            captured_input_payload.append(x_payload.astype(np.float32, copy=False))
            captured_input_float.append(np.asarray(x_float, dtype=np.float32))
            captured_output_payload.append(np.asarray(y_payload[0], dtype=np.float32).reshape(-1))
            captured_output_float.append(np.asarray(y_float[0], dtype=np.float32).reshape(-1))
            captured_output_raw.append(output_raw_records)

    y_pred = np.asarray(y_pred_float, dtype=np.float32)
    if not capture_io:
        return y_pred

    return {
        "preds": y_pred,
        "input_payload": np.asarray(captured_input_payload, dtype=np.float32),
        "input_float": np.asarray(captured_input_float, dtype=np.float32),
        "output_payload": np.asarray(captured_output_payload, dtype=np.float32),
        "output_float": np.asarray(captured_output_float, dtype=np.float32),
        "output_raw": captured_output_raw,
    }



def _build_dbg_model_output_raw_reference(sample_epochs, output_raw_records):
    rows = []
    max_dims = int(DBG_MODEL_RAW_MAX_DIMS)
    max_bytes = int(DBG_MODEL_RAW_MAX_BYTES)

    for _idx, (_epoch_dbg, _sample_records) in enumerate(zip(sample_epochs, output_raw_records)):
        for _rec in _sample_records:
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "out_idx": int(_rec.get("out_idx", -1)),
                "tensor_index": int(_rec.get("tensor_index", -1)),
                "type_code": int(_rec.get("type_code", -1)),
                "type_name": str(_rec.get("type_name", "")),
                "bytes_total": int(_rec.get("bytes_total", 0)),
                "bytes_dumped": int(_rec.get("bytes_dumped", 0)),
                "dims_size": int(_rec.get("dims_size", 0)),
            }
            for _di in range(max_dims):
                _row[f"dim{_di:02d}"] = int(_rec.get(f"dim{_di:02d}", -1))
            for _bi in range(max_bytes):
                _row[f"b{_bi:02d}_hex"] = str(_rec.get(f"b{_bi:02d}_hex", ""))
            rows.append(_row)

    if rows:
        return pd.DataFrame(rows)

    _columns = [
        "idx", "epoch", "out_idx", "tensor_index", "type_code", "type_name",
        "bytes_total", "bytes_dumped", "dims_size",
    ]
    _columns += [f"dim{_di:02d}" for _di in range(max_dims)]
    _columns += [f"b{_bi:02d}_hex" for _bi in range(max_bytes)]
    return pd.DataFrame(columns=_columns)



def _dbg_f32(x) -> float:
    return float(np.asarray(x, dtype=np.float32))


def _float32_to_hex_bits(x) -> str:
    _u = np.asarray([x], dtype=np.float32).view(np.uint32)[0]
    return f"0x{int(_u):08X}"


def _build_replay_canonical_processed(raw_replay_df, ema_prev, alpha):
    """Rebuild the replay-processed path with float32 arithmetic from the
    exported raw 2+47 block plus the EMA seed, mirroring the firmware contract.
    """
    _df = raw_replay_df.copy().reset_index(drop=True)

    for _c in ["T_out", "H_out", "T_in", "H_in"]:
        _df[_c] = pd.to_numeric(_df[_c], errors="coerce").astype(np.float32)

    _dt = pd.to_datetime(_df["datetime"])
    _hour = (_dt.dt.hour.astype(np.int64).to_numpy() + (_dt.dt.minute.astype(np.int64).to_numpy() / 60.0)).astype(np.float32)
    _weekday = _dt.dt.weekday.astype(np.int64).to_numpy()
    _month = _dt.dt.month.astype(np.int64).to_numpy()

    _df["hour"] = _hour
    _df["weekday"] = _weekday.astype(np.float32)
    _df["month"] = _month.astype(np.float32)
    _df["sin_hour"] = np.sin(2.0 * np.pi * _hour / 24.0).astype(np.float32)
    _df["cos_hour"] = np.cos(2.0 * np.pi * _hour / 24.0).astype(np.float32)

    _alpha = np.float32(alpha)
    _one_minus = np.float32(1.0) - _alpha
    _prev = np.float32(ema_prev)
    _h_raw = _df["H_in"].to_numpy(dtype=np.float32, copy=True)
    _h_filt = np.empty_like(_h_raw, dtype=np.float32)
    for _i, _raw in enumerate(_h_raw):
        _curr = np.float32(_raw)
        _filt = np.float32(_alpha * _curr + _one_minus * _prev)
        _h_filt[_i] = _filt
        _prev = _filt
    _df["H_in"] = _h_filt

    _df["T_in_lag1"] = _df["T_in"].shift(1).astype(np.float32)
    _df["H_in_lag1"] = _df["H_in"].shift(1).astype(np.float32)
    _df["T_out_lag1"] = _df["T_out"].shift(1).astype(np.float32)
    _df["H_out_lag1"] = _df["H_out"].shift(1).astype(np.float32)
    _df["T_in_lag2"] = _df["T_in"].shift(2).astype(np.float32)
    _df["H_in_lag2"] = _df["H_in"].shift(2).astype(np.float32)

    return _df


def _extract_canonical_replay_debug_inputs(sample_df_indices, replay_processed_df, replay_processed_df_indices, raw_replay_df=None):
    """Build canonical preprocessing windows from the replay-processed source
    mirrored by firmware during DBG_MODEL_IN_CSV, while also exposing explicit
    preprocessing checkpoints for auditability.
    """
    _index_map = {int(_df_idx): int(_pos) for _pos, _df_idx in enumerate(replay_processed_df_indices)}
    _pre_raw_seqs = []
    _pre_smooth_seqs = []
    _pre_lags_seqs = []
    _pre_time_seqs = []
    _pre_phys_seqs = []
    _step_epoch_seqs = []
    _y_abs_rows = []
    _tprev_rows = []
    _hprev_rows = []

    _raw_df = raw_replay_df if raw_replay_df is not None else replay_processed_df

    for _df_idx in sample_df_indices:
        _end = _index_map.get(int(_df_idx))
        if _end is None:
            raise KeyError(f"Replay canonical index not found for df index {_df_idx}")
        if _end < WINDOW - 1:
            raise ValueError(
                f"Replay canonical window too short for df index {_df_idx}: end={_end}, window={WINDOW}"
            )

        _start = _end - WINDOW + 1
        _win_proc = replay_processed_df.iloc[_start : _end + 1]
        _win_raw = _raw_df.iloc[_start : _end + 1]

        if "epoch" in _win_raw.columns:
            _step_epochs = _win_raw["epoch"].to_numpy(dtype=np.int64, copy=True)
        else:
            _step_epochs = (pd.to_datetime(_win_raw["datetime"]).astype("int64") // 10**9).to_numpy(dtype=np.int64, copy=True)

        _pre_raw_seqs.append(
            _win_raw[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_smooth_seqs.append(
            _win_proc[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_lags_seqs.append(
            _win_proc[["T_in_lag1", "H_in_lag1", "T_out_lag1", "H_out_lag1", "T_in_lag2", "H_in_lag2"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_time_seqs.append(
            np.column_stack([
                _step_epochs.astype(np.float32, copy=False),
                _win_proc["sin_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["cos_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["weekday"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["month"].to_numpy(dtype=np.float32, copy=True),
            ]).astype(np.float32, copy=False)
        )
        _pre_phys = _win_proc[features].to_numpy(dtype=np.float32, copy=True)
        _pre_phys_seqs.append(_pre_phys)
        _step_epoch_seqs.append(_step_epochs)

        _row = replay_processed_df.iloc[_end]
        _y_abs_rows.append(np.asarray([_row["T_in"], _row["H_in"]], dtype=np.float32))
        _tprev_rows.append(np.float32(_row["T_in_lag1"]))
        _hprev_rows.append(np.float32(_row["H_in_lag1"]))

    return {
        "pre_raw": np.asarray(_pre_raw_seqs, dtype=np.float32),
        "pre_smooth": np.asarray(_pre_smooth_seqs, dtype=np.float32),
        "pre_lags": np.asarray(_pre_lags_seqs, dtype=np.float32),
        "pre_time": np.asarray(_pre_time_seqs, dtype=np.float32),
        "pre_phys": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "step_epoch": np.asarray(_step_epoch_seqs, dtype=np.int64),
        "x_raw": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "y_abs": np.asarray(_y_abs_rows, dtype=np.float32),
        "tprev": np.asarray(_tprev_rows, dtype=np.float32),
        "hprev": np.asarray(_hprev_rows, dtype=np.float32),
    }


def _build_dbg_model_input_reference(sample_test_positions, sample_df_indices, sample_epochs, x_raw_seqs, x_clip_seqs, x_scaled_seqs, x_tensor_payload_seqs, x_tensor_float_seqs, y_true_abs_rows):
    rows = []
    _n_features = len(features)
    for _idx, (_test_pos, _df_idx, _epoch_dbg) in enumerate(zip(sample_test_positions, sample_df_indices, sample_epochs)):
        _x_raw = np.asarray(x_raw_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _x_clip = np.asarray(x_clip_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _x_scaled = np.asarray(x_scaled_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _x_tensor_payload = np.asarray(x_tensor_payload_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _x_tensor_float = np.asarray(x_tensor_float_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _gt = np.asarray(y_true_abs_rows[_idx], dtype=np.float32).reshape(2)

        for _step in range(WINDOW):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "gt_Tin_true": float(_gt[0]),
                "gt_Hin_true": _dbg_f32(_gt[1]),
                "state_Tout_phys_raw": float(_x_raw[_step, 0]),
                "state_Hout_phys_raw": float(_x_raw[_step, 1]),
                "state_Tin_lag1_phys_raw": float(_x_raw[_step, 2]),
                "state_Hin_lag1_phys_raw": _dbg_f32(_x_raw[_step, 3]),
                "state_Tout_lag1_phys_raw": float(_x_raw[_step, 4]),
                "state_Hout_lag1_phys_raw": float(_x_raw[_step, 5]),
                "state_Tin_lag2_phys_raw": float(_x_raw[_step, 6]),
                "state_Hin_lag2_phys_raw": _dbg_f32(_x_raw[_step, 7]),
                "state_sin_hour": float(_x_raw[_step, 8]),
                "state_cos_hour": float(_x_raw[_step, 9]),
                "state_weekday": float(_x_raw[_step, 10]),
                "state_month": float(_x_raw[_step, 11]),
            }
            for _feat in range(_n_features):
                _value = _x_raw[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_raw"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _value = _x_clip[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_clip"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _row[f"in_f{_feat:02d}_scaled"] = float(_x_scaled[_step, _feat])
            for _feat in range(_n_features):
                # Compatibility alias preserved for existing comparison scripts.
                _row[f"in_x{_feat:02d}_float"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_payload"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_float_prequant"] = float(_x_tensor_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_stage_reference(sample_epochs, step_epoch_seqs, data_seqs, field_names):
    rows = []
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _data = np.asarray(data_seqs[_idx], dtype=np.float32)
        if _data.ndim == 1:
            _data = _data.reshape(WINDOW, 1)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(WINDOW)
        for _step in range(_data.shape[0]):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _j, _name in enumerate(field_names):
                _val = _data[_step, _j]
                _row[_name] = _dbg_f32(_val)
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_sample_reference(raw_df, smooth_df, seed_rows: int = 2):
    """Build PRE_RAW/PRE_SMOOTH at firmware-equivalent logical-sample granularity.

    The firmware emits one row per logical replay sample after the seed rows,
    using idx=0..N-1 and raw_idx pointing to the underlying 2+47 replay row.
    This builder mirrors that exact contract while preserving the detailed
    window-level exports in separate sheets.
    """
    if len(raw_df) != len(smooth_df):
        raise ValueError(
            f"PRE_RAW/PRE_SMOOTH export length mismatch: raw={len(raw_df)} smooth={len(smooth_df)}"
        )
    if len(raw_df) <= seed_rows:
        raise ValueError(
            f"Need more than {seed_rows} rows to build logical PRE_RAW/PRE_SMOOTH exports"
        )

    rows_raw = []
    rows_smooth = []
    for _raw_idx in range(seed_rows, len(raw_df)):
        _idx = _raw_idx - seed_rows
        _epoch = int(raw_df.loc[_raw_idx, "epoch"]) if "epoch" in raw_df.columns else 0
        _row_raw = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_out"]),
            "Hout_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_out"]),
            "Tin_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_in"]),
            "Hin_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_in"]),
        }
        _row_smooth = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_out"]),
            "Hout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_out"]),
            "Tin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_in"]),
            "Hin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_in"]),
        }
        rows_raw.append(_row_raw)
        rows_smooth.append(_row_smooth)

    return pd.DataFrame(rows_raw), pd.DataFrame(rows_smooth)


def _build_preprocess_model_in_reference(sample_epochs, step_epoch_seqs, x_payload_seqs, x_float_seqs):
    rows = []
    _n_features = len(features)
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _x_payload = np.asarray(x_payload_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _x_float = np.asarray(x_float_seqs[_idx], dtype=np.float32).reshape(WINDOW, _n_features)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(WINDOW)
        for _step in range(WINDOW):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _feat in range(_n_features):
                _row[f"x_payload_{_feat:02d}"] = float(_x_payload[_step, _feat])
                _row[f"x_float_prequant_{_feat:02d}"] = float(_x_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_dbg_model_output_reference(sample_test_positions, sample_epochs, output_payload, output_float, tprev_override=None, hprev_override=None):
    _y_scaled = np.asarray(output_float, dtype=np.float32).reshape(-1, 2)
    _d_pred = scaler_y.inverse_transform(_y_scaled).astype(np.float32, copy=False)
    if tprev_override is None:
        _tprev = np.asarray(Tprev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _tprev = np.asarray(tprev_override, dtype=np.float32).reshape(-1)
    if hprev_override is None:
        _hprev = np.asarray(Hprev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _hprev = np.asarray(hprev_override, dtype=np.float32).reshape(-1)

    _dT_pred = np.asarray(_d_pred[:, 0], dtype=np.float32)
    _dH_pred = np.asarray(_d_pred[:, 1], dtype=np.float32)
    _t_pred = (_tprev + _dT_pred).astype(np.float32, copy=False)
    _h_pred = (_hprev + _dH_pred).astype(np.float32, copy=False)

    _out_o0_float = np.asarray(output_float[:, 0], dtype=np.float32)
    _out_o1_float = np.asarray(output_float[:, 1], dtype=np.float32)
    _out_o0_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o0_float]
    _out_o1_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o1_float]

    return pd.DataFrame({
        "idx": np.arange(len(sample_epochs), dtype=np.int64),
        "epoch": np.asarray(sample_epochs, dtype=np.int64),
        "out_o0_tensor": np.asarray(output_payload[:, 0], dtype=np.float32),
        "out_o1_tensor": np.asarray(output_payload[:, 1], dtype=np.float32),
        "out_o0_float": _out_o0_float,
        "out_o1_float": _out_o1_float,
        "out_o0_bits_hex": _out_o0_bits_hex,
        "out_o1_bits_hex": _out_o1_bits_hex,
        "y_T_scaled": _out_o0_float,
        "y_H_scaled": _out_o1_float,
        "d_T_pred": _dT_pred,
        "d_H_pred": _dH_pred,
        "p_Tprev_phys": _tprev,
        "p_Hprev_phys": np.asarray(_hprev, dtype=np.float32),
        "p_T_pred": _t_pred,
        "p_H_pred": _h_pred,
    })


# ---------------------
# TFLite evaluation (VAL and TEST)
# ---------------------
start_time = time.time()
y_val_pred_tflite = run_tflite_inference(quantized_path, X_val)
y_pred_tflite     = run_tflite_inference(quantized_path, X_test)
end_time = time.time()

# Timing metrics
inference_time_total       = (end_time - start_time) * 1000.0
inference_time_per_sample  = inference_time_total / len(X_test)
print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Mean latency per sample: {inference_time_per_sample:.4f} ms")

# ---------------------
# Model sizes
# ---------------------
original_model_size_kb = os.path.getsize(model_path) / 1024
model_size_kb          = os.path.getsize(quantized_path) / 1024
print(f"Original model size (.keras): {original_model_size_kb:.2f} KB")
print(f"Quantized model size: {model_size_kb:.2f} KB")

# ---------------------
# Normalized-domain evaluation (RESIDUAL TRAINING TARGET)
# ---------------------
mse_scaled  = mean_squared_error(y_test, y_pred_tflite)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled  = mean_absolute_error(y_test, y_pred_tflite)
r2_scaled   = r2_score(y_test, y_pred_tflite)

print("\n Results (normalized scale - residual training target):")
print(f"MSE  (norm.) = {mse_scaled:.8f}")
print(f"RMSE (norm.) = {rmse_scaled:.8f}")
print(f"MAE  (norm.) = {mae_scaled:.8f}")
print(f"R²   (norm.) = {r2_scaled:.8f}")

# Evaluation in the ORIGINAL SCALE (absolute reconstruction of T_in, H_in)
if USE_RESIDUAL_TARGETS:
    y_val_orig_res      = scaler_y.inverse_transform(y_val).astype(np.float32, copy=False)
    y_val_pred_orig_res = scaler_y.inverse_transform(y_val_pred_tflite).astype(np.float32, copy=False)

    y_test_orig_res     = scaler_y.inverse_transform(y_test).astype(np.float32, copy=False)
    y_pred_orig_raw_res = scaler_y.inverse_transform(y_pred_tflite).astype(np.float32, copy=False)

    # LiteML-Edge contract: Δ_pred = inverse_transform(scaler_y)
    y_pred_res = y_pred_orig_raw_res.astype(np.float32, copy=False)

    # Ensure prev arrays are float32
    Tprev_test = Tprev_test.astype(np.float32, copy=False)
    Hprev_test = Hprev_test.astype(np.float32, copy=False)

    # Absolute reconstruction
    T_pred = (Tprev_test + y_pred_res[:, 0]).astype(np.float32, copy=False)
    H_pred = (Hprev_test + y_pred_res[:, 1]).astype(np.float32, copy=False)
else:
    # Fallback case: model already outputs absolute values
    T_pred = y_pred_tflite[:, 0].astype(np.float32, copy=False)
    H_pred = y_pred_tflite[:, 1].astype(np.float32, copy=False)

y_pred_orig = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)
y_test_orig = yabs_test.astype(np.float32, copy=False)  # true absolute [T_in, H_in]

mse  = mean_squared_error(y_test_orig, y_pred_orig)
rmse = np.sqrt(mse)
mae  = mean_absolute_error(y_test_orig, y_pred_orig)
r2   = r2_score(y_test_orig, y_pred_orig)

print("\n Results (original scale - joint target set [T_in, H_in]):")
mse_status  = "MSE within threshold"  if mse  <= 0.1  else "MSE above threshold"
rmse_status = "RMSE within threshold" if rmse <= 0.32 else "RMSE above threshold"
mae_status  = "MAE within threshold"  if mae  <= 0.3  else "MAE above threshold"
r2_status   = "R² within threshold"   if r2   >= 0.8  else "R² below threshold"

print(f"MSE  = {mse:.8f}   {mse_status}")
print(f"RMSE = {rmse:.8f}   {rmse_status}")
print(f"MAE  = {mae:.8f}   {mae_status}")
print(f"R²   = {r2:.8f}   {r2_status}")

model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied" if model_ok else "Performance thresholds not satisfied"
print("\n Overall assessment:", model_status)

# ---------------------
# Overfitting / underfitting diagnosis
# ---------------------
train_loss = history.history["loss"]
val_loss   = history.history["val_loss"]
n          = min(5, len(train_loss))
mean_train_loss = float(np.mean(train_loss[-n:]))
mean_val_loss   = float(np.mean(val_loss[-n:]))
gap      = abs(mean_val_loss - mean_train_loss)
gap_pct  = (gap / mean_train_loss) * 100 if mean_train_loss > 0 else 0.0

if mean_train_loss > 0.3 and mean_val_loss > 0.3:
    status = "Underfitting detected (high training and validation losses)"
elif mean_val_loss < mean_train_loss * 0.8:
    status = "Potential underfitting (validation loss significantly lower than training loss)"
elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
    status = "Overfitting detected (large generalization gap or significant divergence)"
elif gap_pct < 10:
    status = "Well-fitted model (generalization gap < 10%)"
elif gap_pct < 30:
    status = "Acceptably fitted model (generalization gap < 30%)"
else:
    status = "Mild overfitting (moderate generalization gap)"

print("\n Model diagnosis:")
print(f"• Mean training loss:    {mean_train_loss:.8f}")
print(f"• Mean validation loss:  {mean_val_loss:.8f}")
print(f"• Absolute gap:          {gap:.8f}")
print(f"• Percentage gap:        {gap_pct:.2f}%")
print(f"• Assessment:            {status}")


# ============================================================
# ROLLING(24) METRICS (FIRMWARE-EQUIVALENT) + EXPORT (paper-ready)
# ============================================================
ROLLING_N = 24              # equals METRICS_WINDOW_SIZE in firmware
ROLLING_EXPORT_ROWS = 24    # export only the last 24 rolling windows

def _r2_like_firmware(y_true_1d, y_pred_1d):
    """Replicate firmware logic (metrics.cpp): R² is NaN when variance is too small."""
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return np.nan

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))
    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return np.nan
    r2v = 1.0 - (ss_res / ss_tot)
    return float(r2v) if np.isfinite(r2v) else np.nan

def _metrics_like_firmware(y_true_2d, y_pred_2d):
    # T_in
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T ** 2))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = _r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0])

    # H_in
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H ** 2))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = _r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1])

    # Aggregate metrics (same semantics as firmware printout)
    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H) if (np.isfinite(r2_T) and np.isfinite(r2_H)) else np.nan

    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

# Test mapping (sequence index -> original dataframe row)
idx_test = idx_arr[n_train + n_val:]  # dataframe indices for TEST windows

rolling_rows = []
n_test_roll = y_test_orig.shape[0]

# --- Gating identical to firmware: the window advances only when Invoke occurred
#     and the event is an HOUR event. Offline evaluation assumes both are true. ---
invoked_mask = np.ones(n_test_roll, dtype=bool)
is_rollover_mask = np.ones(n_test_roll, dtype=bool)

accepted_idxs = [i for i in range(n_test_roll) if (invoked_mask[i] and is_rollover_mask[i])]

if len(accepted_idxs) >= ROLLING_N:
    for k in range(ROLLING_N - 1, len(accepted_idxs)):
        end = accepted_idxs[k]
        start = accepted_idxs[k - ROLLING_N + 1]
        win = accepted_idxs[k - ROLLING_N + 1 : k + 1]

        yt_w = y_test_orig[win]
        yp_w = y_pred_orig[win]

        mae_r, rmse_r, r2_r, mae_T_r, rmse_T_r, r2_T_r, mae_H_r, rmse_H_r, r2_H_r = _metrics_like_firmware(yt_w, yp_w)

        # datetime_end: last sample of the rolling window
        dt_end = df.iloc[int(idx_test[end])]["datetime"] if "datetime" in df.columns else ""

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": str(dt_end),
            "N": int(ROLLING_N),

            "MAE": mae_r,
            "RMSE": rmse_r,
            "R2": r2_r,

            "MAE_T": mae_T_r,
            "RMSE_T": rmse_T_r,
            "R2_T": r2_T_r,

            "MAE_H": mae_H_r,
            "RMSE_H": rmse_H_r,
            "R2_H": r2_H_r,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # --- Rolling(24) distribution summary over the full test set ---
    try:
        _mae_mean = float(df_roll["MAE"].mean())
        _rmse_mean = float(df_roll["RMSE"].mean())
        _mae_p95 = float(df_roll["MAE"].quantile(0.95))
        _rmse_p95 = float(df_roll["RMSE"].quantile(0.95))
        print("\n[INFO] Rolling(24) summary over the full test set:")
        print(f"  • MAE_mean={_mae_mean:.4f} | MAE_p95={_mae_p95:.4f}")
        print(f"  • RMSE_mean={_rmse_mean:.4f} | RMSE_p95={_rmse_p95:.4f}")
    except Exception as _e:
        print("[WARN] Could not summarize the Rolling(24) distribution:", _e)

    df_roll_24 = df_roll.tail(ROLLING_EXPORT_ROWS).copy().reset_index(drop=True)

    # -----------------------------
    # Build the minimal sample stream required to reproduce the exported rolling24 windows
    # -----------------------------
    pos_of = {v: i for i, v in enumerate(accepted_idxs)}
    sample_set = set()
    for _, row in df_roll_24.iterrows():
        end_v = int(row["window_end"])
        pos = pos_of.get(end_v, None)
        if pos is None:
            continue
        win = accepted_idxs[max(0, pos - ROLLING_N + 1):pos + 1]
        for v in win:
            sample_set.add(int(v))

    sample_vals = sorted(sample_set, key=lambda v: pos_of.get(v, 1_000_000_000))
    ROLL24_24_PACK_IDXS = list(sample_vals)
    sample_df_indices = [int(idx_test[v]) for v in sample_vals if 0 <= v < len(idx_test)]

    # -----------------------------
    # Replay 2+47 contract for the raw firmware header
    #   • 47 real rows: exact rows already used in Python for the exported replay
    #   • 2 seed rows : immediately previous processed rows (t-2, t-1)
    # This removes the need for a boot-time mirror in firmware.
    # -----------------------------
    REPLAY_SEED_ROWS = 2
    REPLAY_REAL_ROWS = 47
    REPLAY_TOTAL_RAW_ROWS = REPLAY_SEED_ROWS + REPLAY_REAL_ROWS

    if len(sample_vals) != REPLAY_REAL_ROWS:
        raise ValueError(
            f"Expected exactly {REPLAY_REAL_ROWS} replay rows from Rolling(24), got {len(sample_vals)}"
        )
    if not np.all(np.diff(np.asarray(sample_vals, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must be contiguous in TEST space for the 2+47 export")
    if not np.all(np.diff(np.asarray(sample_df_indices, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must map to contiguous processed dataframe rows for the 2+47 export")

    first_real_df_idx = int(sample_df_indices[0])
    last_real_df_idx  = int(sample_df_indices[-1])

    if first_real_df_idx < REPLAY_SEED_ROWS:
        raise ValueError(
            f"Need {REPLAY_SEED_ROWS} previous processed rows to export replay seeds, got first_real_df_idx={first_real_df_idx}"
        )

    raw_replay_df_indices = list(range(first_real_df_idx - REPLAY_SEED_ROWS, last_real_df_idx + 1))
    if len(raw_replay_df_indices) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw replay rows, got {len(raw_replay_df_indices)}"
        )

    # Map the processed 2+47 block back to the raw-real timeline.
    raw_replay_row_ids = df_processed_row_ids[np.asarray(raw_replay_df_indices, dtype=np.int64)]
    if not np.all(np.diff(raw_replay_row_ids) == 1):
        raise ValueError("Replay 2+47 rows must map to contiguous raw-real dataframe rows")

    raw_replay_start_row_id = int(raw_replay_row_ids[0])
    raw_replay_end_row_id   = int(raw_replay_row_ids[-1])
    raw_replay_real = df_raw_real.iloc[raw_replay_start_row_id:raw_replay_end_row_id + 1].copy().reset_index(drop=True)
    if len(raw_replay_real) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw-real replay rows, got {len(raw_replay_real)}"
        )

    if raw_replay_start_row_id <= 0:
        raise ValueError(
            "Need one previous raw-real row to export the initial H_in EMA state for firmware replay"
        )

    ema_prev_row_id = raw_replay_start_row_id - 1
    ema_prev_raw_row = df_raw_real.iloc[ema_prev_row_id].copy()
    hin_ema_prev = float(df.iloc[first_real_df_idx - REPLAY_SEED_ROWS - 1]["H_in"])

    # 47 real rows preserved for the processed analysis/debug export
    df_samples_roll = df.iloc[sample_df_indices].copy().reset_index(drop=True)

    # 49 raw-real rows (2 seeds + 47 real) used by the firmware replay header and its mirrors
    df_samples_replay_raw = raw_replay_real.copy().reset_index(drop=True)

    # 49 processed rows aligned with the replay block for Python-side filtered/lagged reference
    df_samples_replay_proc = df.iloc[raw_replay_df_indices].copy().reset_index(drop=True)

    # --- Ensure epoch (seconds) for firmware replay/debug exports ---
    for _df_export in (df_samples_roll, df_samples_replay_raw, df_samples_replay_proc):
        if "datetime" in _df_export.columns:
            try:
                _df_export["epoch"] = pd.to_datetime(_df_export["datetime"]).astype("int64") // 10**9
            except Exception:
                _df_export["epoch"] = 0
        else:
            _df_export["epoch"] = 0

    # -----------------------------
    # Export rolling24 metrics (CSV + XLSX) with four-decimal formatting
    # -----------------------------
    df_roll_24_export = df_roll_24.copy()
    for c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]:
        if c in df_roll_24_export.columns:
            df_roll_24_export[c] = pd.to_numeric(df_roll_24_export[c], errors="coerce")

    df_roll_24_export = df_roll_24_export.round(4)

    csv_roll = results_dir / "environment_quantized_metrics_rolling24_mlp.csv"
    xlsx_roll = results_dir / "environment_quantized_metrics_rolling24_mlp.xlsx"
    df_roll_24_export.to_csv(csv_roll, index=False, encoding="utf-8-sig", float_format="%.4f")
    df_roll_24_export.to_excel(xlsx_roll, index=False)

    # Apply four-decimal number format in XLSX
    try:
        wb = load_workbook(xlsx_roll)
        ws = wb.active
        num_fmt = "0.0000"
        num_cols = {c: (list(df_roll_24_export.columns).index(c) + 1) for c in df_roll_24_export.columns if c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
        for _, col_idx in num_cols.items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = num_fmt
        wb.save(xlsx_roll)
    except Exception as _e:
        print("[WARN] Could not apply XLSX number formats for rolling24:", _e)

    print(f"[INFO] rolling24 exported: {csv_roll}")
    print(f"[INFO] rolling24 exported: {xlsx_roll}")

    # -----------------------------
    # Export real replay rows (47, processed analysis only), raw replay rows (2+47),
    # replay reference (2+47 mixed), and firmware header (.h)
    # -----------------------------
    def _autosize_xlsx(_xlsx_path):
        try:
            _wb = load_workbook(_xlsx_path)
            _ws = _wb.active
            for _col_idx, _col_cells in enumerate(_ws.columns, 1):
                _max_length = max(len(str(_cell.value)) if _cell.value is not None else 0 for _cell in _col_cells)
                _ws.column_dimensions[get_column_letter(_col_idx)].width = _max_length + 2
            _wb.save(_xlsx_path)
        except Exception as _e:
            print(f"[WARN] Could not autosize XLSX {_xlsx_path}:", _e)

    samples_csv_path  = results_dir / "environment_quantized_samples_rolling24_mlp.csv"
    samples_xlsx_path = results_dir / "environment_quantized_samples_rolling24_mlp.xlsx"
    if "datetime" in df_samples_roll.columns:
        df_samples_roll["datetime"] = df_samples_roll["datetime"].astype(str)
    df_samples_roll.to_csv(samples_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    df_samples_roll.to_excel(samples_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(samples_xlsx_path)
    print(f"[INFO] rolling24 samples exported: {samples_csv_path} | rows={len(df_samples_roll)}")
    print(f"[INFO] rolling24 samples exported: {samples_xlsx_path} | rows={len(df_samples_roll)}")

    replay_raw_export = df_samples_replay_raw.copy()
    if "datetime" in replay_raw_export.columns:
        replay_raw_export["datetime"] = replay_raw_export["datetime"].astype(str)
    replay_raw_csv_path  = results_dir / "environment_quantized_samples_replay_raw_2plus47_mlp.csv"
    replay_raw_xlsx_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_mlp.xlsx"
    replay_raw_export.to_csv(replay_raw_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    replay_raw_export.to_excel(replay_raw_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(replay_raw_xlsx_path)
    print(f"[INFO] replay raw 2+47 samples exported: {replay_raw_csv_path} | rows={len(replay_raw_export)}")
    print(f"[INFO] replay raw 2+47 samples exported: {replay_raw_xlsx_path} | rows={len(replay_raw_export)}")

    # Replay reference export (mixed: raw + processed reference + Python GT/prediction)
    replay_reference = pd.DataFrame({
        "replay_idx": np.arange(REPLAY_TOTAL_RAW_ROWS, dtype=np.int64),
        "is_seed": [1 if i < REPLAY_SEED_ROWS else 0 for i in range(REPLAY_TOTAL_RAW_ROWS)],
        "epoch": df_samples_replay_raw["epoch"].astype(np.int64),
        "datetime": df_samples_replay_raw["datetime"].astype(str),
        "T_out_raw": df_samples_replay_raw["T_out"].astype(np.float32),
        "H_out_raw": df_samples_replay_raw["H_out"].astype(np.float32),
        "T_in_raw": df_samples_replay_raw["T_in"].astype(np.float32),
        "H_in_raw": df_samples_replay_raw["H_in"].astype(np.float32),
        "T_in_ref": df_samples_replay_proc["T_in"].astype(np.float32),
        "H_in_filt_ref": df_samples_replay_proc["H_in"].astype(np.float32),
        "T_in_lag1_ref": df_samples_replay_proc["T_in_lag1"].astype(np.float32),
        "H_in_lag1_ref": df_samples_replay_proc["H_in_lag1"].astype(np.float32),
        "T_in_lag2_ref": df_samples_replay_proc["T_in_lag2"].astype(np.float32),
        "H_in_lag2_ref": df_samples_replay_proc["H_in_lag2"].astype(np.float32),
    })

    # Python reference outputs only for the 47 real rows.
    replay_reference["T_in_ground truth"] = np.nan
    replay_reference["T_in_pred_python"] = np.nan
    replay_reference["H_in_ground truth"] = np.nan
    replay_reference["H_in_pred_python"] = np.nan

    real_rows_slice = slice(REPLAY_SEED_ROWS, REPLAY_TOTAL_RAW_ROWS)
    y_true_roll_ref = y_test_orig[np.asarray(sample_vals, dtype=np.int64)]
    y_pred_roll_ref = y_pred_orig[np.asarray(sample_vals, dtype=np.int64)]
    replay_reference.loc[real_rows_slice, "T_in_ground truth"] = y_true_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "T_in_pred_python"] = y_pred_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "H_in_ground truth"] = y_true_roll_ref[:, 1]
    replay_reference.loc[real_rows_slice, "H_in_pred_python"] = y_pred_roll_ref[:, 1]

    replay_ref_csv_path  = results_dir / "environment_quantized_replay_reference_2plus47_mlp.csv"
    replay_ref_xlsx_path = results_dir / "environment_quantized_replay_reference_2plus47_mlp.xlsx"
    replay_reference.to_csv(replay_ref_csv_path, index=False, encoding="utf-8-sig", float_format="%.4f")
    replay_reference.to_excel(replay_ref_xlsx_path, index=False, float_format="%.4f")
    _autosize_xlsx(replay_ref_xlsx_path)
    print(f"[INFO] replay reference 2+47 exported: {replay_ref_csv_path} | rows={len(replay_reference)}")
    print(f"[INFO] replay reference 2+47 exported: {replay_ref_xlsx_path} | rows={len(replay_reference)}")

    # -----------------------------
    # model I/O reference (immediate tensors that enter/exit the model)
    # -----------------------------
    dbg_model_valid_idxs = (
        np.asarray(ROLL24_24_PACK_IDXS[-ROLLING_N:], dtype=np.int64)
        if "ROLL24_24_PACK_IDXS" in globals() and len(ROLL24_24_PACK_IDXS) > 0
        else np.arange(max(0, len(X_test) - ROLLING_N), len(X_test), dtype=np.int64)
    )
    if dbg_model_valid_idxs.size > ROLLING_N:
        dbg_model_valid_idxs = dbg_model_valid_idxs[-ROLLING_N:]

    dbg_model_df_indices = [int(idx_test[v]) for v in dbg_model_valid_idxs]
    dbg_model_epochs = []
    for _df_idx in dbg_model_df_indices:
        _dt_dbg = pd.to_datetime(df.iloc[_df_idx]["datetime"])
        dbg_model_epochs.append(int(_dt_dbg.value // 10**9))

    dbg_model_replay_proc = _build_replay_canonical_processed(
        raw_replay_df=df_samples_replay_raw,
        ema_prev=hin_ema_prev,
        alpha=HIN_EMA_ALPHA,
    )

    dbg_model_replay_canonical = _extract_canonical_replay_debug_inputs(
        sample_df_indices=dbg_model_df_indices,
        replay_processed_df=dbg_model_replay_proc,
        replay_processed_df_indices=raw_replay_df_indices,
        raw_replay_df=df_samples_replay_raw,
    )

    dbg_model_capture = run_tflite_inference(quantized_path, X_test[dbg_model_valid_idxs], capture_io=True)
    dbg_model_input_payload = dbg_model_capture["input_payload"]
    dbg_model_input_float = dbg_model_capture["input_float"]
    dbg_model_output_payload = dbg_model_capture["output_payload"]
    dbg_model_output_float = dbg_model_capture["output_float"]
    dbg_model_output_raw = dbg_model_capture["output_raw"]

    dbg_model_pre_clip = np.clip(
        dbg_model_replay_canonical["pre_phys"],
        np.asarray(scaler_X.data_min_, dtype=np.float32),
        np.asarray(scaler_X.data_max_, dtype=np.float32),
    ).astype(np.float32, copy=False)
    dbg_model_pre_scaled = X_test[dbg_model_valid_idxs].reshape(-1, WINDOW, len(features)).astype(np.float32, copy=False)

    dbg_model_input_reference = _build_dbg_model_input_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_df_indices=dbg_model_df_indices,
        sample_epochs=dbg_model_epochs,
        x_raw_seqs=dbg_model_replay_canonical["pre_phys"],
        x_clip_seqs=dbg_model_pre_clip,
        x_scaled_seqs=dbg_model_pre_scaled,
        x_tensor_payload_seqs=dbg_model_input_payload,
        x_tensor_float_seqs=dbg_model_input_float,
        y_true_abs_rows=dbg_model_replay_canonical["y_abs"],
    )
    dbg_model_output_reference = _build_dbg_model_output_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_epochs=dbg_model_epochs,
        output_payload=dbg_model_output_payload,
        output_float=dbg_model_output_float,
        tprev_override=dbg_model_replay_canonical["tprev"],
        hprev_override=dbg_model_replay_canonical["hprev"],
    )
    dbg_model_output_raw_reference = _build_dbg_model_output_raw_reference(
        sample_epochs=dbg_model_epochs,
        output_raw_records=dbg_model_output_raw,
    )

    preprocess_raw_reference, preprocess_smooth_reference = _build_preprocess_sample_reference(
        raw_df=df_samples_replay_raw,
        smooth_df=dbg_model_replay_proc,
        seed_rows=REPLAY_SEED_ROWS,
    )
    preprocess_raw_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_raw"],
        field_names=["Tout_raw", "Hout_raw", "Tin_raw", "Hin_raw"],
    )
    preprocess_smooth_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_smooth"],
        field_names=["Tout_smooth", "Hout_smooth", "Tin_smooth", "Hin_smooth"],
    )
    preprocess_lags_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_lags"],
        field_names=["Tin_lag1", "Hin_lag1", "Tout_lag1", "Hout_lag1", "Tin_lag2", "Hin_lag2"],
    )
    preprocess_time_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_time"],
        field_names=["epoch_ref", "sin_hour", "cos_hour", "weekday", "month"],
    )
    preprocess_phys_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_phys"],
        field_names=[
            "f00_Tout_phys", "f01_Hout_phys", "f02_Tin_lag1_phys", "f03_Hin_lag1_phys",
            "f04_Tout_lag1_phys", "f05_Hout_lag1_phys", "f06_Tin_lag2_phys", "f07_Hin_lag2_phys",
            "f08_sin_hour_phys", "f09_cos_hour_phys", "f10_weekday_phys", "f11_month_phys",
        ],
    )
    preprocess_clip_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_clip,
        field_names=[
            "f00_Tout_clip", "f01_Hout_clip", "f02_Tin_lag1_clip", "f03_Hin_lag1_clip",
            "f04_Tout_lag1_clip", "f05_Hout_lag1_clip", "f06_Tin_lag2_clip", "f07_Hin_lag2_clip",
            "f08_sin_hour_clip", "f09_cos_hour_clip", "f10_weekday_clip", "f11_month_clip",
        ],
    )
    preprocess_scaled_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_scaled,
        field_names=[
            "f00_Tout_scaled", "f01_Hout_scaled", "f02_Tin_lag1_scaled", "f03_Hin_lag1_scaled",
            "f04_Tout_lag1_scaled", "f05_Hout_lag1_scaled", "f06_Tin_lag2_scaled", "f07_Hin_lag2_scaled",
            "f08_sin_hour_scaled", "f09_cos_hour_scaled", "f10_weekday_scaled", "f11_month_scaled",
        ],
    )
    preprocess_model_in_reference = _build_preprocess_model_in_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        x_payload_seqs=dbg_model_input_payload,
        x_float_seqs=dbg_model_input_float,
    )

    dbg_model_input_csv_path = results_dir / "environment_quantized_dbg_model_input_reference_mlp.csv"
    dbg_model_input_xlsx_path = results_dir / "environment_quantized_dbg_model_input_reference_mlp.xlsx"
    dbg_model_output_csv_path = results_dir / "environment_quantized_dbg_model_output_reference_mlp.csv"
    dbg_model_output_xlsx_path = results_dir / "environment_quantized_dbg_model_output_reference_mlp.xlsx"
    dbg_model_output_raw_csv_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_mlp.csv"
    dbg_model_output_raw_xlsx_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_mlp.xlsx"
    dbg_model_workbook_path = results_dir / "environment_quantized_model_io_debug_reference_mlp.xlsx"
    preprocess_workbook_path = results_dir / "environment_quantized_preprocess_debug_reference_mlp.xlsx"

    dbg_model_input_reference.to_csv(dbg_model_input_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_input_reference.to_excel(dbg_model_input_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_reference.to_csv(dbg_model_output_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_output_reference.to_excel(dbg_model_output_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_raw_reference.to_csv(dbg_model_output_raw_csv_path, index=False, encoding="utf-8-sig")
    dbg_model_output_raw_reference.to_excel(dbg_model_output_raw_xlsx_path, index=False)
    with pd.ExcelWriter(preprocess_workbook_path, engine="openpyxl") as _writer_pre_dbg:
        preprocess_raw_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
    with pd.ExcelWriter(dbg_model_workbook_path, engine="openpyxl") as _writer_model_dbg:
        preprocess_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
        dbg_model_input_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_IN_CSV", float_format="%.8f")
        dbg_model_output_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_CSV", float_format="%.8f")
        dbg_model_output_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_RAW_CSV")

    _autosize_xlsx(dbg_model_input_xlsx_path)
    _autosize_xlsx(dbg_model_output_xlsx_path)
    _autosize_xlsx(dbg_model_output_raw_xlsx_path)
    _autosize_xlsx(preprocess_workbook_path)
    _autosize_xlsx(dbg_model_workbook_path)

    print(f"[INFO] model input reference exported: {dbg_model_input_csv_path} | rows={len(dbg_model_input_reference)}")
    print(f"[INFO] model input reference exported: {dbg_model_input_xlsx_path} | rows={len(dbg_model_input_reference)}")
    print(f"[INFO] model output reference exported: {dbg_model_output_csv_path} | rows={len(dbg_model_output_reference)}")
    print(f"[INFO] model output reference exported: {dbg_model_output_xlsx_path} | rows={len(dbg_model_output_reference)}")
    print(f"[INFO] model output RAW reference exported: {dbg_model_output_raw_csv_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f"[INFO] model output RAW reference exported: {dbg_model_output_raw_xlsx_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f"[INFO] preprocessing debug workbook exported: {preprocess_workbook_path}")
    print(f"[INFO] model I/O debug workbook exported: {dbg_model_workbook_path}")

    hdr_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_mlp.h"

    def _fmt_f(x):
        try:
            xf = float(x)
        except Exception:
            xf = float("nan")
        if not np.isfinite(xf):
            return "NAN"
        if abs(xf) < 0.0005:
            xf = 0.0
        # Keep higher precision in replay header to preserve feature fidelity in MCU validation
        return f"{xf:.6f}f"

    # Robust column lookup
    def _get_col(df_, names):
        for n in names:
            if n in df_.columns:
                return n
        return None

    c_tin  = _get_col(df_samples_replay_raw, ["Tin","T_in","tin","T_IN"])
    c_hin  = _get_col(df_samples_replay_raw, ["Hin","H_in","hin","H_IN"])
    c_tout = _get_col(df_samples_replay_raw, ["Tout","T_out","tout","T_OUT"])
    c_hout = _get_col(df_samples_replay_raw, ["Hout","H_out","hout","H_OUT"])

    with open(hdr_path, "w", encoding="utf-8") as hf:
        hf.write("#pragma once\n")
        hf.write("#include <stdint.h>\n\n")
        hf.write("typedef struct {\n")
        hf.write("  uint32_t epoch;\n")
        hf.write("  float T_out; float H_out; float T_in; float H_in_raw;\n")
        hf.write("} liteml_sample_raw_t;\n\n")

        hf.write(f"static const uint16_t LITEML_REPLAY_SEED_ROWS = {REPLAY_SEED_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_REAL_ROWS = {REPLAY_REAL_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_TOTAL_RAW_ROWS = {REPLAY_TOTAL_RAW_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_ROLL24_24_N_SAMPLES = {len(df_samples_replay_raw)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_ALPHA = {_fmt_f(HIN_EMA_ALPHA)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_PREV = {_fmt_f(hin_ema_prev)};\n")
        hf.write("static const liteml_sample_raw_t LITEML_ROLL24_24_SAMPLES[] = {\n")

        if not all([c_tin, c_hin, c_tout, c_hout]):
            hf.write("  // [WARN] Expected columns were not found in the raw-real dataset.\n")
        else:
            for i in range(len(df_samples_replay_raw)):
                epoch = int(df_samples_replay_raw.loc[i, "epoch"]) if "epoch" in df_samples_replay_raw.columns else 0
                Tout = df_samples_replay_raw.loc[i, c_tout]
                Hout = df_samples_replay_raw.loc[i, c_hout]
                Tin  = df_samples_replay_raw.loc[i, c_tin]
                Hin_raw  = df_samples_replay_raw.loc[i, c_hin]

                if i == 0:
                    comment = " // seed prev2 (t-2)"
                elif i == 1:
                    comment = " // seed prev  (t-1)"
                else:
                    comment = f" // real row {i - REPLAY_SEED_ROWS:02d}"

                hf.write(
                    f"  {{ {epoch}u, {_fmt_f(Tout)}, {_fmt_f(Hout)}, {_fmt_f(Tin)}, {_fmt_f(Hin_raw)} }},{comment}\n"
                )

        hf.write("};\n")
    print(f"[INFO] rolling24 firmware header exported (2+47 raw replay + EMA seed): {hdr_path}")

    # -----------------------------
    # Rolling24 plots (paper-ready, English)
    # -----------------------------
    try:
        roll_end_idxs = df_roll_24["window_end"].astype(int).to_list()

        yT_true_r = y_test_orig[roll_end_idxs, 0]
        yT_pred_r = y_pred_orig[roll_end_idxs, 0]
        yH_true_r = y_test_orig[roll_end_idxs, 1]
        yH_pred_r = y_pred_orig[roll_end_idxs, 1]

        # Scatter: T_in (rolling24)
        plt.figure(figsize=(6, 6))
        plt.scatter(yT_true_r, yT_pred_r, s=14, alpha=0.65)
        mn = float(min(yT_true_r.min(), yT_pred_r.min()))
        mx = float(max(yT_true_r.max(), yT_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], "k--", linewidth=1.0)
        plt.gca().set_aspect("equal", adjustable="box")
        mae_t_r = float(np.mean(np.abs(yT_pred_r - yT_true_r)))
        rmse_t_r = float(np.sqrt(np.mean((yT_pred_r - yT_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yT_true_r)}\nMAE={mae_t_r:.3f} °C\nRMSE={rmse_t_r:.3f} °C",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (T_in, °C)")
        plt.ylabel("MLP prediction (T_in, °C)")
        plt.title("Rolling(24): ground truth vs prediction (T_in)")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_T_in_rolling24_mlp")
        plt.close()

        # Scatter: H_in (rolling24)
        plt.figure(figsize=(6, 6))
        plt.scatter(yH_true_r, yH_pred_r, s=14, alpha=0.65)
        mn = float(min(yH_true_r.min(), yH_pred_r.min()))
        mx = float(max(yH_true_r.max(), yH_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], "k--", linewidth=1.0)
        plt.gca().set_aspect("equal", adjustable="box")
        mae_h_r = float(np.mean(np.abs(yH_pred_r - yH_true_r)))
        rmse_h_r = float(np.sqrt(np.mean((yH_pred_r - yH_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yH_true_r)}\nMAE={mae_h_r:.3f} %\nRMSE={rmse_h_r:.3f} %",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (H_in, %)")
        plt.ylabel("MLP prediction (H_in, %)")
        plt.title("Rolling(24): ground truth vs prediction (H_in)")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_H_in_rolling24_mlp")
        plt.close()

        # Rolling metrics over exported windows (MAE/RMSE)
        plt.figure(figsize=(8, 4))
        plt.plot(df_roll_24_export["MAE"].values, label="MAE (aggregate)")
        plt.plot(df_roll_24_export["RMSE"].values, label="RMSE (aggregate)")
        plt.xlabel("Rolling window index (last 24)")
        plt.ylabel("Error")
        plt.title("Rolling(24) metrics over the last 24 windows")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_rolling24_metrics_last24_mlp")
        plt.close()

        # Time series: Rolling(24) endpoints (T_in)
        plt.figure(figsize=(10, 4))
        try:
            x_dt = pd.to_datetime(df_roll_24["datetime_end"], errors="coerce")
            if x_dt.isna().all():
                x_dt = np.arange(len(df_roll_24))
                plt.xlabel("Rolling window index (last 24)")
            else:
                plt.xlabel("Datetime (window end)")
        except Exception:
            x_dt = np.arange(len(df_roll_24))
            plt.xlabel("Rolling window index (last 24)")

        plt.plot(x_dt, yT_true_r, label="Ground truth (T_in)")
        plt.plot(x_dt, yT_pred_r, label="Prediction (T_in)")
        plt.ylabel("Temperature (°C)")
        plt.title("Rolling(24) time series at window endpoints: T_in")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_timeseries_T_in_rolling24_mlp")
        plt.close()

        # Time series: Rolling(24) endpoints (H_in)
        plt.figure(figsize=(10, 4))
        try:
            x_dt = pd.to_datetime(df_roll_24["datetime_end"], errors="coerce")
            if x_dt.isna().all():
                x_dt = np.arange(len(df_roll_24))
                plt.xlabel("Rolling window index (last 24)")
            else:
                plt.xlabel("Datetime (window end)")
        except Exception:
            x_dt = np.arange(len(df_roll_24))
            plt.xlabel("Rolling window index (last 24)")

        plt.plot(x_dt, yH_true_r, label="Ground truth (H_in)")
        plt.plot(x_dt, yH_pred_r, label="Prediction (H_in)")
        plt.ylabel("Relative humidity (%)")
        plt.title("Rolling(24) time series at window endpoints: H_in")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.legend()
        plt.tight_layout()
        _savefig_pub("environment_timeseries_H_in_rolling24_mlp")
        plt.close()

    except Exception as _e:
        print("[WARN] Could not generate rolling24 plots:", _e)

else:
    print("Warning: Not enough samples to compute Rolling(24) metrics.")

# ---------------------
# Loss plot (training vs validation)
# ---------------------
plt.figure(figsize=(8, 5))
plt.plot(train_loss, label="Training (QAT)")
plt.plot(val_loss,   label="Validation")
plt.title("QAT training and validation loss\n" + status)
plt.xlabel("Epoch")
plt.ylabel("MSE loss")
plt.legend()
plt.grid(True, linestyle="--", alpha=0.5)
plt.tight_layout()
_savefig_pub("environment_quantized_training_validation_loss_diagnosis_mlp")
plt.close()

# ---------------------
# Individual metrics for T_in and H_in
# ---------------------
T_true     = y_test_orig[:, 0]
H_true     = y_test_orig[:, 1]
T_pred_abs = y_pred_orig[:, 0]
H_pred_abs = y_pred_orig[:, 1]

mse_T  = mean_squared_error(T_true, T_pred_abs)
rmse_T = np.sqrt(mse_T)
mae_T  = mean_absolute_error(T_true, T_pred_abs)
r2_T   = r2_score(T_true, T_pred_abs)

mse_H  = mean_squared_error(H_true, H_pred_abs)
rmse_H = np.sqrt(mse_H)
mae_H  = mean_absolute_error(H_true, H_pred_abs)
r2_H   = r2_score(H_true, H_pred_abs)

print("\n Individual metrics - Temperature (T_in):")
print(f"MSE_T  = {mse_T:.8f}")
print(f"RMSE_T = {rmse_T:.8f}")
print(f"MAE_T  = {mae_T:.8f}")
print(f"R²_T   = {r2_T:.8f}")

print("\n Individual metrics - Humidity (H_in):")
print(f"MSE_H  = {mse_H:.8f}")
print(f"RMSE_H = {rmse_H:.8f}")
print(f"MAE_H  = {mae_H:.8f}")
print(f"R²_H   = {r2_H:.8f}")

# ============================================================
# Joint scatter plot  -  T_in and H_in (original scale)
# ============================================================
plt.figure(figsize=(6, 6))
plt.scatter(y_test_orig[:, 0], y_pred_orig[:, 0], alpha=0.5, label="Temperature (T_in)")
plt.scatter(y_test_orig[:, 1], y_pred_orig[:, 1], alpha=0.5, label="Humidity (H_in)")
min_val = min(y_test_orig.min(), y_pred_orig.min())
max_val = max(y_test_orig.max(), y_pred_orig.max())
plt.plot([min_val, max_val], [min_val, max_val], "k--")
plt.xlabel("Ground truth (original scale)")
plt.ylabel("Prediction (original scale)")
plt.title("Joint scatter: ground truth vs prediction (quantized MLP)")
plt.legend()
plt.grid(True, linestyle="--", alpha=0.5)
plt.tight_layout()
_savefig_pub("environment_quantized_model_scatter_predictions_mlp")
plt.close()

# ============================================================
# Separate scatter plot  -  T_in
# ============================================================
plt.figure(figsize=(6, 6))
plt.scatter(T_true, T_pred_abs, alpha=0.5, label="T_in")
min_val_T = min(T_true.min(), T_pred_abs.min())
max_val_T = max(T_true.max(), T_pred_abs.max())
plt.plot([min_val_T, max_val_T], [min_val_T, max_val_T], "k--")
plt.xlabel("T_in ground truth (°C)")
plt.ylabel("T_in prediction (°C)")
plt.title("Scatter: ground truth vs prediction (T_in)")
plt.grid(True, linestyle="--", alpha=0.5)
plt.tight_layout()
_savefig_pub("environment_quantized_model_scatter_Tin_mlp")
plt.close()

# ============================================================
# Separate scatter plot  -  H_in
# ============================================================
plt.figure(figsize=(6, 6))
plt.scatter(H_true, H_pred_abs, alpha=0.5, label="H_in")
min_val_H = min(H_true.min(), H_pred_abs.min())
max_val_H = max(H_true.max(), H_pred_abs.max())
plt.plot([min_val_H, max_val_H], [min_val_H, max_val_H], "k--")
plt.xlabel("H_in ground truth (%)")
plt.ylabel("H_in prediction (%)")
plt.title("Scatter: ground truth vs prediction (H_in)")
plt.grid(True, linestyle="--", alpha=0.5)
plt.tight_layout()
_savefig_pub("environment_quantized_model_scatter_Hin_mlp")
plt.close()

# ============================================================
# Bar plot of the main metrics (joint)
# ============================================================
plt.figure(figsize=(8, 6))
labels = ["MSE", "RMSE", "MAE", "R²"]
values = [mse, rmse, mae, r2]
bars = plt.bar(labels, values)
plt.title("Metrics summary (original scale): quantized model")
plt.ylabel("Metric value")
plt.grid(axis="y", linestyle="--", alpha=0.5)
for bar in bars:
    yval = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2.0,
        yval,
        f"{yval:.8f}",
        ha="center",
        va="bottom",
    )
plt.tight_layout()
_savefig_pub("environment_quantized_model_final_metrics_summary_plot_mlp")
plt.close()

# ---------------------
corr_T = np.corrcoef(T_true, T_pred_abs)[0, 1]
corr_H = np.corrcoef(H_true, H_pred_abs)[0, 1]

print("\n[INFO] Correlation (Temperature):", f"{corr_T:.4f}")
print("[INFO] Correlation (Humidity):   ", f"{corr_H:.4f}")

if corr_T >= 0.9:
    print("[INFO] Temperature diagnosis: very strong correlation (r ≥ 0.9)")
elif corr_T >= 0.75:
    print("[INFO] Temperature diagnosis: strong correlation (r ≥ 0.75)")
elif corr_T >= 0.5:
    print("[INFO] Temperature diagnosis: moderate correlation (r ≥ 0.5)")
else:
    print("[INFO] Temperature diagnosis: weak or negligible correlation (r < 0.5)")

if corr_H >= 0.9:
    print("[INFO] Humidity diagnosis: very strong correlation (r ≥ 0.9)")
elif corr_H >= 0.75:
    print("[INFO] Humidity diagnosis: strong correlation (r ≥ 0.75)")
elif corr_H >= 0.5:
    print("[INFO] Humidity diagnosis: moderate correlation (r ≥ 0.5)")
else:
    print("[INFO] Humidity diagnosis: weak or negligible correlation (r < 0.5)")

# ---------------------
# Pearson correlation + p-value
# ---------------------
pearson_t, pval_t = pearsonr(y_test_orig[:, 0], y_pred_orig[:, 0])
if pval_t < 0.05:
    print(
        f"[INFO] Pearson correlation (Temperature): r = {pearson_t:.4f} (p = {pval_t:.4e}) Statistically significant"
    )
else:
    print(
        f"[INFO] Pearson correlation (Temperature): r = {pearson_t:.4f} (p = {pval_t:.4e}) Not significant at the 5% level"
    )

pearson_h, pval_h = pearsonr(y_test_orig[:, 1], y_pred_orig[:, 1])
if pval_h < 0.05:
    print(
        f"[INFO] Pearson correlation (Humidity): r = {pearson_h:.4f} (p = {pval_h:.4e}) Statistically significant"
    )
else:
    print(
        f"[INFO] Pearson correlation (Humidity): r = {pearson_h:.4f} (p = {pval_h:.4e}) Not significant at the 5% level"
    )

# ========================= Metrics table (CSV / Excel) =========================

# =========================
# Firmware-style Rolling(24) summary for metrics_summary export
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols if k in _last.index}
        fw_mean = {k: float(df_roll[_cols].mean(numeric_only=True)[k]) for k in _cols if k in df_roll.columns}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract the Rolling(24) summary:", _e)

# --- Compatibility block (quantized): sparsity/strip metrics in the same style as pruned_model ---
# In this script, the "original" comparison model is the pruned .keras model (model_path).
# Sparsity is estimated here by counting weights exactly equal to zero.
def _sparsity_fraction(model) -> float:
    try:
        total, zeros = 0, 0
        for layer in getattr(model, "layers", []):
            for w in layer.get_weights():
                arr = np.asarray(w)
                total += arr.size
                zeros += int(np.sum(arr == 0))
        return (zeros / total) if total > 0 else float("nan")
    except Exception:
        return float("nan")

sparsity_before = float("nan")
sparsity_after  = float("nan")
try:
    # Before: as stored in the pruned .keras model
    sparsity_before = _sparsity_fraction(base_model)

    # After strip_pruning: if pruning wrappers exist, remove them before measurement
    try:
        from tensorflow_model_optimization.sparsity.keras import strip_pruning
        _stripped = strip_pruning(base_model)
        sparsity_after = _sparsity_fraction(_stripped)
    except Exception:
        sparsity_after = sparsity_before
except Exception:
    pass

# Quantized model size (tflite) and original model size (.keras)
model_size_kb = float(model_size_kb) if 'model_size_kb' in globals() else float("nan")
original_model_size_kb = float(original_model_size_kb) if 'original_model_size_kb' in globals() else float("nan")

# --- Compatibility block (quantized): fit/model status in the same style as pruned_model ---
try:
    status_diag = status_diag
except Exception:
    try:
        status_diag = status
    except Exception:
        status_diag = ""

try:
    status = status_diag
except Exception:
    status = ""

try:
    model_status = "Performance thresholds satisfied" if bool(model_ok) else "Performance thresholds not satisfied"
except Exception:
    model_status = ""

# Structure identical to the pruned_model metrics_summary
_metrics_names = [
    'MSE (normalized residual, joint)',
    'RMSE (normalized residual, joint)',
    'MAE (normalized residual, joint)',
    'R² (normalized residual, joint)',
    'MSE (joint original)',
    'RMSE (joint original)',
    'MAE (joint original)',
    'R² (joint original)',
    'MSE_T (T_in)',
    'RMSE_T (T_in)',
    'MAE_T (T_in)',
    'R²_T (T_in)',
    'MSE_H (H_in)',
    'RMSE_H (H_in)',
    'MAE_H (H_in)',
    'R²_H (H_in)',
    'N_ALL (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, last window)',
    'RMSE (rolling24 firmware, last window)',
    'R² (rolling24 firmware, last window)',
    'MAE_T (rolling24 firmware, last window)',
    'RMSE_T (rolling24 firmware, last window)',
    'R²_T (rolling24 firmware, last window)',
    'MAE_H (rolling24 firmware, last window)',
    'RMSE_H (rolling24 firmware, last window)',
    'R²_H (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, test mean)',
    'RMSE (rolling24 firmware, test mean)',
    'R² (rolling24 firmware, test mean)',
    'MAE_T (rolling24 firmware, test mean)',
    'RMSE_T (rolling24 firmware, test mean)',
    'R²_T (rolling24 firmware, test mean)',
    'MAE_H (rolling24 firmware, test mean)',
    'RMSE_H (rolling24 firmware, test mean)',
    'R²_H (rolling24 firmware, test mean)',
    'Sparsity before strip (%)',
    'Sparsity after strip (%)',
    'Model size (KB)',
    'Original model size (KB)',
    'Mean training loss',
    'Mean validation loss',
    'Absolute gap',
    'Percentage gap (%)',
    'Total inference time (ms)',
    'Inference time per sample (ms)',
    'Fit status',
    'Model status',
]
_significados  = [
    'Mean squared error in the normalized scale for residual targets (ΔT_in, ΔH_in), jointly evaluated.',
    'Root mean squared error in the normalized scale, jointly evaluated.',
    'Mean absolute error in the normalized scale, jointly evaluated.',
    'Coefficient of determination in the normalized scale, jointly evaluated.',
    'Mean squared error in the original physical scale (°C / %RH) for the joint target set [T_in, H_in].',
    'Root mean squared error in the original physical scale for the joint target set.',
    'Mean absolute error in the original physical scale for the joint target set.',
    'Coefficient of determination in the original physical scale for the joint target set.',
    'Mean squared error in the original physical scale for T_in.',
    'Root mean squared error in the original physical scale for T_in.',
    'Mean absolute error in the original physical scale for T_in.',
    'Coefficient of determination in the original physical scale for T_in.',
    'Mean squared error in the original physical scale for H_in.',
    'Root mean squared error in the original physical scale for H_in.',
    'Mean absolute error in the original physical scale for H_in.',
    'Coefficient of determination in the original physical scale for H_in.',
    'Rolling(24): number of effective HOUR-event samples in the final window.',
    'Aggregated Rolling(24) MAE in the LAST window, directly comparable with the firmware log.',
    'Aggregated Rolling(24) RMSE in the LAST window, directly comparable with the firmware log.',
    'Aggregated Rolling(24) R² in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) T_in MAE in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) T_in RMSE in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) T_in R² in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) H_in MAE in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) H_in RMSE in the LAST window, directly comparable with the firmware log.',
    'Rolling(24) H_in R² in the LAST window, directly comparable with the firmware log.',
    'Mean aggregated Rolling(24) MAE across the full test set.',
    'Mean aggregated Rolling(24) RMSE across the full test set.',
    'Mean aggregated Rolling(24) R² across the full test set.',
    'Mean Rolling(24) MAE for T_in across the full test set.',
    'Mean Rolling(24) RMSE for T_in across the full test set.',
    'Mean Rolling(24) R² for T_in across the full test set.',
    'Mean Rolling(24) MAE for H_in across the full test set.',
    'Mean Rolling(24) RMSE for H_in across the full test set.',
    'Mean Rolling(24) R² for H_in across the full test set.',
    'Percentage of zero-valued weights before applying strip_pruning.',
    'Percentage of zero-valued weights after removing pruning operations.',
    'Final model file size in kilobytes (KB).',
    'Original model file size (.keras) in kilobytes (KB).',
    'Mean training loss over the last training epochs.',
    'Mean validation loss over the last validation epochs.',
    'Absolute difference between mean losses.',
    'Percentage gap between validation and training losses.',
    'Total time required to infer all test samples.',
    'Mean time required to infer one sample.',
    'Diagnosis based on the gap/loss pattern: fitted, overfitting, or underfitting.',
    'Overall diagnosis considering predefined MSE/RMSE/MAE/R² limits.',
]
_thresholds      = [
    '→ Lower is better.',
    '→ Lower is better.',
    '→ Lower is better.',
    '→ Ideally > 0.95.',
    '→ < 0.1 is excellent, depending on the problem.',
    '→ < 0.32 as a reference threshold.',
    '→ < 0.3 as a reference threshold.',
    '→ > 0.8 is desirable.',
    '→ Lower is better (T_in).',
    '→ Lower is better (T_in).',
    '→ Lower is better (T_in).',
    '→ Ideally > 0.8 (T_in).',
    '→ Lower is better (H_in).',
    '→ Lower is better (H_in).',
    '→ Lower is better (H_in).',
    '→ Ideally > 0.8 (H_in).',
    '→ Must reach 24 when the rolling window is fully populated.',
    '→ Lower is better (rolling24).',
    '→ Lower is better (rolling24).',
    '→ May be NaN under low-variance conditions, consistent with firmware behavior.',
    '→ Lower is better (rolling24 T_in).',
    '→ Lower is better (rolling24 T_in).',
    '→ May be NaN under low-variance conditions.',
    '→ Lower is better (rolling24 H_in).',
    '→ Lower is better (rolling24 H_in).',
    '→ May be NaN under low-variance conditions.',
    '→ Lower is better (mean across windows).',
    '→ Lower is better (mean across windows).',
    '→ May be NaN under low-variance conditions.',
    '→ Lower is better (mean across windows for T_in).',
    '→ Lower is better (mean across windows for T_in).',
    '→ May be NaN under low-variance conditions.',
    '→ Lower is better (mean across windows for H_in).',
    '→ Lower is better (mean across windows for H_in).',
    '→ May be NaN under low-variance conditions.',
    '→ In general, > 50% supports meaningful compression gains.',
    '→ Ideally close to the target final sparsity.',
    '→ Preferably < 256 KB on constrained MCUs.',
    '→ Reference value for comparison with the quantized model.',
    '→ Low value preferred (e.g., < 0.01).',
    '→ Should remain close to the training loss.',
    '→ < 0.05 is good.',
    '→ < 10% is excellent.',
    '→ Lower is better.',
    '→ < 1 ms is ideal in TinyML scenarios.',
    "→ 'Fitted' when the gap is low and losses are stable.",
    "→ 'Performance thresholds satisfied' when the model satisfies the predefined limits.",
]

_valores = [
    f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
    f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
    f"{mse_T:.4f}",      f"{rmse_T:.4f}",      f"{mae_T:.4f}",      f"{r2_T:.4f}",
    f"{mse_H:.4f}",      f"{rmse_H:.4f}",      f"{mae_H:.4f}",      f"{r2_H:.4f}",
    (f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else ""),
    (f"{fw_last.get('MAE', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE', float('nan'))) else ""),
    (f"{fw_last.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE', float('nan'))) else ""),
    (f"{fw_last.get('R2', float('nan')):.4f}" if np.isfinite(fw_last.get('R2', float('nan'))) else ""),
    (f"{fw_last.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_T', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_last.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_T', float('nan'))) else ""),
    (f"{fw_last.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_H', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_last.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_H', float('nan'))) else ""),
    (f"{fw_mean.get('MAE', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE', float('nan'))) else ""),
    (f"{fw_mean.get('R2', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_T', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_mean.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_T', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_H', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_mean.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_H', float('nan'))) else ""),
    (f"{sparsity_before * 100:.2f} %" if np.isfinite(sparsity_before) else ""),
    (f"{sparsity_after * 100:.2f} %" if np.isfinite(sparsity_after) else ""),
    (f"{model_size_kb:.2f} KB" if np.isfinite(model_size_kb) else ""),
    (f"{original_model_size_kb:.2f} KB" if np.isfinite(original_model_size_kb) else ""),
    (f"{mean_train_loss:.4f}" if 'mean_train_loss' in globals() and np.isfinite(mean_train_loss) else ""),
    (f"{mean_val_loss:.4f}"   if 'mean_val_loss' in globals() and np.isfinite(mean_val_loss) else ""),
    (f"{gap:.4f}"             if 'gap' in globals() and np.isfinite(gap) else ""),
    (f"{gap_pct:.2f} %"       if 'gap_pct' in globals() and np.isfinite(gap_pct) else ""),
    f"{inference_time_total:.2f} ms",
    f"{inference_time_per_sample:.2f} ms",
    "",
    ""
]

_statuses = [
    "", "", "", "",
    mse_status, rmse_status, mae_status, r2_status,
    "", "", "", "",
    "", "", "", "",
    "", "", "", "", "", "", "", "", "", "",
    "", "", "", "", "", "", "", "", "",
    "", "", "", "",
    "", "", "", "",
    "", "",
    status,
    model_status
]

# --- FIX: guarantee identical length across all columns ---
_max_len = max(len(_metrics_names), len(_valores), len(_statuses), len(_significados), len(_thresholds))
def _pad(lst):
    lst = list(lst)
    if len(lst) < _max_len:
        lst += [""] * (_max_len - len(lst))
    return lst

metrics_dist = {
    "Quantized Model Metrics": _pad(_metrics_names),
    "Value": _pad(_valores),
    "Status": _pad(_statuses),
    "Meaning": _pad(_significados),
    "Expected Values / Thresholds": _pad(_thresholds),
}

df_metrics = pd.DataFrame(metrics_dist)

excel_path_metrics = results_dir/"environment_quantized_model_metrics_summary_mlp.xlsx"
df_metrics.to_csv(results_dir/"environment_quantized_model_metrics_summary_mlp.csv", index=False, encoding="utf-8-sig")
df_metrics.to_excel(excel_path_metrics, index=False)

# Adjust column widths
wb = load_workbook(excel_path_metrics); ws = wb.active
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2
wb.save(excel_path_metrics)
print(f"[INFO] File saved successfully: {excel_path_metrics}")

# ============================================================
# Rolling(24) prediction export table
#   • window_start/window_end in TEST space
#   • datetime_end via idx_test -> df
#   • exported features correspond to the LAST step of the window (12 columns)
# ============================================================
try:
    _PRED_N = 47

    if "ROLL24_24_PACK_IDXS" in globals() and isinstance(ROLL24_24_PACK_IDXS, (list, tuple, np.ndarray)) and len(ROLL24_24_PACK_IDXS) > 0:
        _pred_sel = np.asarray(ROLL24_24_PACK_IDXS, dtype=np.int64)
        if _pred_sel.size >= _PRED_N:
            _pred_sel = _pred_sel[-_PRED_N:]
        else:
            _pad = np.arange(max(0, len(X_test) - (_PRED_N - _pred_sel.size)), len(X_test), dtype=np.int64)
            _pred_sel = np.unique(np.concatenate([_pred_sel, _pad]))[-_PRED_N:]
    else:
        _pred_sel = np.arange(max(0, len(X_test) - _PRED_N), len(X_test), dtype=np.int64)

    # Select samples in TEST space
    X_pred_sel = X_test[_pred_sel]
    y_true_sel = y_test_orig[_pred_sel]
    y_pred_sel = y_pred_orig[_pred_sel]

    # Window metadata (in TEST space)
    _window_end   = _pred_sel.astype(np.int64)
    _window_start = (_window_end - (24 - 1)).astype(np.int64)
    _pred_df_indices = [int(idx_test[int(_i)]) for _i in _window_end]

    _datetime_end = []
    for _df_idx in _pred_df_indices:
        _datetime_end.append(str(df.iloc[_df_idx]["datetime"]))

    # Base dataframe with features taken DIRECTLY from the engineered raw timeline.
    # This guarantees that the spreadsheet shows the same real rows that feed the processed analysis exports.
    df_preds_source = df.iloc[_pred_df_indices].copy().reset_index(drop=True)
    _missing_feature_cols = [c for c in features if c not in df_preds_source.columns]
    if _missing_feature_cols:
        raise ValueError(f"[export] Missing raw feature columns in df_preds_source: {_missing_feature_cols}")
    df_preds = df_preds_source[features].copy()

    # Provenance / alignment check against the model-side last-step tensor.
    # This check is diagnostic only; the spreadsheet export uses the engineered dataframe.
    X_last_scaled = X_pred_sel[:, -len(features):]
    X_last_orig   = scaler_X.inverse_transform(X_last_scaled)
    _raw_feature_values = df_preds[features].values.astype(np.float32)
    _feature_abs_diff = np.abs(_raw_feature_values - X_last_orig.astype(np.float32))
    _feature_max_abs_diff = float(np.nanmax(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
    _feature_mean_abs_diff = float(np.nanmean(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
    print(
        f"Feature provenance check (engineered df vs inverse-transformed model input): "
        f"max_abs_diff={_feature_max_abs_diff:.6f} | mean_abs_diff={_feature_mean_abs_diff:.6f}"
    )

    df_preds["T_in_ground truth"] = y_true_sel[:, 0]
    df_preds["T_in_pred"]         = y_pred_sel[:, 0]
    df_preds["H_in_ground truth"] = y_true_sel[:, 1]
    df_preds["H_in_pred"]         = y_pred_sel[:, 1]

    df_preds.insert(0, "datetime_end", _datetime_end)
    df_preds.insert(0, "window_end", _window_end)
    df_preds.insert(0, "window_start", _window_start)

    _export_cols = (
        ["window_start", "window_end", "datetime_end"]
        + list(features)
        + ["T_in_ground truth", "T_in_pred", "H_in_ground truth", "H_in_pred"]
    )

    _missing_cols = [c for c in _export_cols if c not in df_preds.columns]
    if _missing_cols:
        raise ValueError(f"[export] Missing columns in df_preds: {_missing_cols}")

    df_preds_export = df_preds[_export_cols].copy()

    # Export rounding (presentation only)
    num_cols = df_preds_export.select_dtypes(include=[np.number]).columns
    df_preds_export[num_cols] = (np.round(df_preds_export[num_cols].astype(np.float32) * 100.0) / 100.0)

    # Export using the original filenames
    excel_path_pred = results_dir / "environment_quantized_predictions_rolling24_mlp.xlsx"
    csv_path_pred   = results_dir / "environment_quantized_predictions_rolling24_mlp.csv"

    df_preds_export.to_excel(excel_path_pred, index=False, float_format="%.2f")
    df_preds_export.to_csv(csv_path_pred, index=False, encoding="utf-8-sig", float_format="%.2f")

    print(f"[INFO] Files saved: {excel_path_pred.name} | {csv_path_pred.name} | rows={len(df_preds_export)}")

    # Adjust column widths in XLSX
    wb = load_workbook(excel_path_pred)
    ws = wb.active
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        wb.save(excel_path_pred)

except Exception as _e:
    print("[WARN] Could not export the Rolling(24) predictions table:", _e)

# === Post-execution: update 'latest' and manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Unable to update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Unable to write manifest.json:", _e)

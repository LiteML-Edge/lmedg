"""
Script: environment_base_model_mlp.py
Module role:
    Train and evaluate the baseline MLP model adopted in the LiteML-Edge
    environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, trains the baseline residual model, evaluates the
    model in normalized and reconstructed physical domains, and exports metrics,
    figures, and versioned artifacts.

Inputs:
    - environment_dataset_mlp.csv
    - Project path and versioning utilities from utils.global_utils.paths_mlp and
      utils.global_utils.versioning.

Outputs:
    - Versioned baseline model artifacts
    - Fitted X and y scalers
    - Tables, figures, and summary spreadsheets for offline evaluation

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys  
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
rcParams["font.family"] = "Segoe UI Emoji"
# Figure export settings
rcParams["figure.dpi"] = 120
rcParams["savefig.dpi"] = 600
rcParams["font.size"] = 11
rcParams["axes.titlesize"] = 13
rcParams["axes.labelsize"] = 12
rcParams["legend.fontsize"] = 10
rcParams["xtick.labelsize"] = 10
rcParams["ytick.labelsize"] = 10
rcParams["axes.grid"] = True
rcParams["grid.alpha"] = 0.25
rcParams["grid.linestyle"] = "--"
rcParams["axes.spines.top"] = False
rcParams["axes.spines.right"] = False
# ================================================
# --- Bootstrap: allow importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_mlp import PROJECT_ROOT, DATASET_ENVIRONMENT_MLP, BASE_MODEL, BASE_MODEL_METRICS  # use project root in a stable way
from utils.global_utils.versioning import create_versioned_dir, ensure_dir, update_latest, write_manifest
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Call at the top of the script, BEFORE model creation

# === Versioned directories per execution ===
run_dir = create_versioned_dir(BASE_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(BASE_MODEL_METRICS / run_dir.name)
# ===========================================

# =====================================================================
#                           DATA PREPARATION
# =====================================================================
dataset_path = DATASET_ENVIRONMENT_MLP / "environment_dataset_mlp.csv"
df = pd.read_csv(dataset_path)

# Sort temporally to construct sliding windows
df = df.sort_values("datetime").reset_index(drop=True)

# Temporal feature engineering
df["datetime"] = pd.to_datetime(df["datetime"])
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (same idea as Conv1D Tiny) ===
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=0.08, adjust=False).mean()

# Cyclical features
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lag features (first!)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Add lag2 only for indoor variables to keep 12 + weekday/month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)

# Separate features (12 features) 
features = [
    "T_out", "H_out",          # current outdoor
    "T_in_lag1", "H_in_lag1",  # indoor lag1
    "T_out_lag1", "H_out_lag1",# outdoor lag1
    "T_in_lag2", "H_in_lag2",  # indoor lag2
    "sin_hour", "cos_hour",    # cyclical hour
    "weekday", "month",        # calendar
]

# =====================================================================
#          ABSOLUTE AND RESIDUAL TARGETS (ΔT_in, ΔH_in) + CLIPPING
# =====================================================================
# Full absolute targets (for reconstruction and metrics)
targets_abs = ["T_in", "H_in"]
y_full_abs = df[targets_abs].values.astype(np.float32)

# Raw residual targets: ΔT_in, ΔH_in
y_full_res = np.stack(
    [
        df["T_in"].values.astype(np.float32) - df["T_in_lag1"].values.astype(np.float32),
        df["H_in"].values.astype(np.float32) - df["H_in_lag1"].values.astype(np.float32),
    ],
    axis=1,
).astype(np.float32)

# === Residual Δ = abs - lag1 (LiteML-Edge contract: NO clamp on Δ) ===
# (Firmware reconstructs abs = lag1 + Δ without clamp; keep Δ pure here)
y_full_res = y_full_res.astype(np.float32, copy=False)

# =====================================================================
#            SLIDING WINDOW 24×12 → MLP (RESIDUAL TARGET)
#   (NOW IN 3D TO NORMALIZE 12 FEATURES BEFORE FLATTENING)
# =====================================================================
WINDOW = 24
X_source = df[features].values.astype(np.float32)
N = len(df)

if N < WINDOW:
    raise ValueError("Dataset is insufficient to build 24-step windows.")

X_win_list = []   # will store (24, 12)
y_res_list = []
y_abs_list = []
idx_list = []

for t in range(WINDOW - 1, N):
    # window [t-23, ..., t] → (24, 12)
    window = X_source[t - WINDOW + 1 : t + 1, :]
    X_win_list.append(window)
    y_res_list.append(y_full_res[t])   # residual target ΔT_in, ΔH_in
    y_abs_list.append(y_full_abs[t])   # absolute target T_in, H_in (for metrics)
    idx_list.append(t)

# X in 3D (N_seq, 24, 12)
X_win = np.stack(X_win_list).astype(np.float32)
y_res_all = np.stack(y_res_list).astype(np.float32)
y_abs_all_idx = np.stack(y_abs_list).astype(np.float32)
idx_arr = np.array(idx_list, dtype=np.int64)

# === Main training target: ALWAYS residual (ΔT_in, ΔH_in) ===
y_all_res = y_res_all

# Auxiliary vectors for ABSOLUTE reconstruction
T_prev_all = df["T_in_lag1"].values.astype(np.float32)[idx_arr]
H_prev_all = df["H_in_lag1"].values.astype(np.float32)[idx_arr]

# =====================================================================
#                      TRAIN / VAL / TEST SPLIT (TEMPORAL)
#                (SPLIT ON TOP OF 3D X_win, AS IN Conv1D)
# =====================================================================
n_total = X_win.shape[0]

train_frac = 0.6
val_frac   = 0.2  # the remainder is assigned to test

n_train = int(n_total * train_frac)
n_val   = int(n_total * val_frac)
n_test  = n_total - n_train - n_val

if n_train <= 0 or n_val <= 0 or n_test <= 0:
    raise ValueError(
        f"Invalid split with n_total={n_total}, "
        f"n_train={n_train}, n_val={n_val}, n_test={n_test}"
    )

# --- TRAIN: oldest segment ---
X_train_win   = X_win[:n_train]                 # (n_train, 24, 12)
y_train_res   = y_all_res[:n_train]
Tprev_train   = T_prev_all[:n_train]
Hprev_train   = H_prev_all[:n_train]
yabs_train    = y_abs_all_idx[:n_train]

# --- VALIDATION: middle segment ---
X_val_win     = X_win[n_train:n_train + n_val]  # (n_val, 24, 12)
y_val_res     = y_all_res[n_train:n_train + n_val]
Tprev_val     = T_prev_all[n_train:n_train + n_val]
Hprev_val     = H_prev_all[n_train:n_train + n_val]
yabs_val      = y_abs_all_idx[n_train:n_train + n_val]

# --- TEST: most recent segment (real future scenario) ---
X_test_win    = X_win[n_train + n_val:]         # (n_test, 24, 12)
y_test_res    = y_all_res[n_train + n_val:]
Tprev_test    = T_prev_all[n_train + n_val:]
Hprev_test    = H_prev_all[n_train + n_val:]
yabs_test     = y_abs_all_idx[n_train + n_val:]
 
# TEST timestamps (to align with EVENT HOUR in firmware)
dt_seq = df["datetime"].values[idx_arr]
dt_test = dt_seq[n_train + n_val:]


print("[INFO] Temporal split applied:")
print(f"  • Training:   {n_train} samples")
print(f"  • Validation: {n_val} samples")
print(f"  • Test:       {n_test} samples")

# =====================================================================
#                             NORMALIZATION
#        >>> KEY STEP: NORMALIZE 12 FEATURES, THEN FLATTEN
# =====================================================================
scaler_X = MinMaxScaler()
scaler_y = MinMaxScaler()

# --- X normalization (12 features), as in Conv1D ---
N_train, W_steps, F = X_train_win.shape  # F must be 12
# Flatten only to fit the scaler per feature
X_train_2d = X_train_win.reshape(-1, F)           # (N_train * 24, 12)
scaler_X.fit(X_train_2d)

# LiteML-Edge (1:1 contract): clamp X in the physical domain to keep minmax_forward in [0,1]
# (equivalent to clamp01 in firmware, but applied BEFORE transform)
X_train_2d = np.clip(X_train_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)

# Apply the scaler and flatten windows to 288 features (24*12) for the MLP
X_train = scaler_X.transform(X_train_2d).reshape(N_train, W_steps * F)

# X_val
N_val = X_val_win.shape[0]
X_val_2d = X_val_win.reshape(-1, F)
X_val_2d = np.clip(X_val_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_val = scaler_X.transform(X_val_2d).reshape(N_val, W_steps * F)

# X_test
N_test = X_test_win.shape[0]
X_test_2d = X_test_win.reshape(-1, F)
X_test_2d = np.clip(X_test_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_test = scaler_X.transform(X_test_2d).reshape(N_test, W_steps * F)

# --- y normalization (residuals ΔT_in, ΔH_in) ---
y_train = scaler_y.fit_transform(y_train_res)
y_val   = scaler_y.transform(y_val_res)
y_test  = scaler_y.transform(y_test_res)

# =====================================================================
#                        MLP MODEL (SAME STRUCTURE)
# =====================================================================
# MLP model
model = tf.keras.Sequential([
    tf.keras.layers.Input(shape=(X_train.shape[1],)),
    tf.keras.layers.Dense(16, activation="relu", kernel_initializer="he_normal"),
    tf.keras.layers.Dense(2, activation="linear"),
])

model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=3e-3),
    loss="mse",
    metrics=[tf.keras.metrics.MeanAbsoluteError()],
)

callbacks = [
    tf.keras.callbacks.ReduceLROnPlateau(
        monitor="val_loss", factor=0.5, patience=10, min_delta=1e-4, min_lr=1e-6, verbose=1
    ),
    tf.keras.callbacks.EarlyStopping(
        monitor="val_loss", patience=15, restore_best_weights=True, verbose=1
    ),
]

history = model.fit(
    X_train, y_train,
    validation_data=(X_val, y_val),
    epochs=500,
    batch_size=576,
    shuffle=True,
    callbacks=callbacks,
    verbose=1,
)
# =====================================================================
#                    SAVE MODEL AND NORMALIZERS
# =====================================================================
model.save(str(run_dir / "environment_base_model_mlp.keras"))
joblib.dump(scaler_X, run_dir / "environment_base_model_mlp_scaler_X.pkl")
joblib.dump(scaler_y, run_dir / "environment_base_model_mlp_scaler_y.pkl")

# =====================================================================
#                       INFERENCE AND LATENCY
# =====================================================================
start_time = time.time()
y_pred_test_scaled = model.predict(X_test)
end_time = time.time()
inference_time_total = (end_time - start_time) * 1000  # in milliseconds
inference_time_per_sample = inference_time_total / len(X_test) if len(X_test) > 0 else float("nan")

print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.2f} ms")

# =====================================================================
#     DENORMALIZE RESIDUALS AND RECONSTRUCT ABSOLUTE VALUES
# =====================================================================

# Return to the original residual domain ΔT_in, ΔH_in
y_pred_delta = scaler_y.inverse_transform(y_pred_test_scaled).astype(np.float32, copy=False)

# Ensure prev values are float32
Tprev_test = Tprev_test.astype(np.float32, copy=False)
Hprev_test = Hprev_test.astype(np.float32, copy=False)

# Absolute reconstruction using T_in_lag1 / H_in_lag1
T_pred = (Tprev_test + y_pred_delta[:, 0]).astype(np.float32, copy=False)
H_pred = (Hprev_test + y_pred_delta[:, 1]).astype(np.float32, copy=False)

# Final vector also in float32
y_pred_abs = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)

# Absolute ground truth also in float32
y_test_abs = yabs_test.astype(np.float32, copy=False)  # [T_in, H_in]

# =====================================================================
#                        METRICS (ORIGINAL SCALE)
# =====================================================================
mse = mean_squared_error(y_test_abs, y_pred_abs)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_abs, y_pred_abs)
r2 = r2_score(y_test_abs, y_pred_abs)

print("\n Results (original scale - set [T_in, H_in]):")
mse_status = "MSE within threshold" if mse <= 0.1 else "MSE above threshold"
rmse_status = "RMSE within threshold" if rmse <= 0.32 else "RMSE above threshold"
mae_status = "MAE within threshold" if mae <= 0.3 else "MAE above threshold"
r2_status = "R² within threshold" if r2 >= 0.8 else "R² below threshold"

print(f"MSE  = {mse:.8f}   {mse_status}")
print(f"RMSE = {rmse:.8f}   {rmse_status}")
print(f"MAE  = {mae:.8f}   {mae_status}")
print(f"R²   = {r2:.8f}   {r2_status}")

# =====================================================================
#   METRICS IN THE NORMALIZED SPACE (RESIDUALS ΔT_in, ΔH_in)
# =====================================================================
mse_scaled = mean_squared_error(y_test, y_pred_test_scaled)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled = mean_absolute_error(y_test, y_pred_test_scaled)
r2_scaled = r2_score(y_test, y_pred_test_scaled)

print("\n Results (normalized scale - residuals [ΔT_in, ΔH_in]):")
print(f"MSE  (norm. residual) = {mse_scaled:.8f}")
print(f"RMSE (norm. residual) = {rmse_scaled:.8f}")
print(f"MAE  (norm. residual) = {mae_scaled:.8f}")
print(f"R²   (norm. residual) = {r2_scaled:.8f}")

# ========= Individual metrics for T_in and H_in (original scale) =========
T_true = y_test_abs[:, 0]
T_pred_ind = y_pred_abs[:, 0]
H_true = y_test_abs[:, 1]
H_pred_ind = y_pred_abs[:, 1]

mse_T = mean_squared_error(T_true, T_pred_ind)
rmse_T = np.sqrt(mse_T)
mae_T = mean_absolute_error(T_true, T_pred_ind)
r2_T = r2_score(T_true, T_pred_ind)

mse_H = mean_squared_error(H_true, H_pred_ind)
rmse_H = np.sqrt(mse_H)
mae_H = mean_absolute_error(H_true, H_pred_ind)
r2_H = r2_score(H_true, H_pred_ind)

print("\n Individual metrics - Temperature (T_in):")
print(f"MSE_T  = {mse_T:.8f}")
print(f"RMSE_T = {rmse_T:.8f}")
print(f"MAE_T  = {mae_T:.8f}")
print(f"R²_T   = {r2_T:.8f}")

print("\n Individual metrics - Humidity (H_in):")
print(f"MSE_H  = {mse_H:.8f}")
print(f"RMSE_H = {rmse_H:.8f}")
print(f"MAE_H  = {mae_H:.8f}")
print(f"R²_H   = {r2_H:.8f}")
# =========================================================================

# Rolling-window metrics with 24 samples, aligned with firmware semantics
# ============================================================
ROLLING_N = 24  # N=24 samples (HOUR), as in metrics.cpp

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """Replicates the firmware logic (metrics.cpp):
      R² = 1 - SS_res/SS_tot
      SS_tot = sum(y^2) - (sum(y)^2)/n
      If n < 2 -> NaN
      If SS_tot <= 1e-6 -> NaN
    """
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return float("nan")

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))

    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return float("nan")

    return 1.0 - (ss_res / ss_tot)


def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    """
    Returns (mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H)
    following exactly:
      mae = 0.5*(mae_T + mae_H)
      rmse = sqrt( 0.5*(mse_T + mse_H) )
      r2 = 0.5*(r2_T + r2_H)
    """
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # T
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T * err_T))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = float(_r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0]))

    # H
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H * err_H))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = float(_r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1]))

    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H)
    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

rolling_rows = []
n_test = y_test_abs.shape[0]

# Gating equal to firmware (offline: all samples count)
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)

if n_test < ROLLING_N:
    print(f"\n Rolling N=24 cannot be computed: test set contains only {n_test} samples.")
else:
    for end in range(ROLLING_N - 1, n_test):
        # Firmware only updates metrics if there was a real Invoke and EVENT=HOUR
        if (not invoked_mask[end]) or (not is_rollover_mask[end]):
            rolling_rows.append({
                "window_start": int(end - ROLLING_N + 1),
                "window_end": int(end),
                "datetime_end": pd.to_datetime(dt_test[end]),
                "N": int(ROLLING_N),

                "MAE": float("nan"),
                "RMSE": float("nan"),
                "R2": float("nan"),

                "MAE_T": float("nan"),
                "RMSE_T": float("nan"),
                "R2_T": float("nan"),

                "MAE_H": float("nan"),
                "RMSE_H": float("nan"),
                "R2_H": float("nan"),
            })
            continue
        start = end - ROLLING_N + 1
        yt_w = y_test_abs[start:end+1, :]
        yp_w = y_pred_abs[start:end+1, :]

        mae_w, rmse_w, r2_w, maeT_w, rmseT_w, r2T_w, maeH_w, rmseH_w, r2H_w = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": pd.to_datetime(dt_test[end]),
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": maeT_w,
            "RMSE_T": rmseT_w,
            "R2_T": r2T_w,

            "MAE_H": maeH_w,
            "RMSE_H": rmseH_w,
            "R2_H": r2H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # Useful summaries for comparison with the firmware current state
    last = df_roll.iloc[-1]
    mean_roll = df_roll[["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]].mean()

    print("\n[INFO] Rolling N=24 (firmware-equivalent), last window (current state):")
    print(f"MAE={last['MAE']:.4f} RMSE={last['RMSE']:.4f} R²={last['R2']:.4f} | "
          f"T: MAE={last['MAE_T']:.4f} RMSE={last['RMSE_T']:.4f} R²={last['R2_T']:.4f} | "
          f"H: MAE={last['MAE_H']:.4f} RMSE={last['RMSE_H']:.4f} R²={last['R2_H']:.4f}")

    print("\n[INFO] Rolling N=24 mean over all test windows:")
    print(f"MAE={mean_roll['MAE']:.4f} RMSE={mean_roll['RMSE']:.4f} R²={mean_roll['R2']:.4f} | "
          f"T: MAE={mean_roll['MAE_T']:.4f} RMSE={mean_roll['RMSE_T']:.4f} R²={mean_roll['R2_T']:.4f} | "
          f"H: MAE={mean_roll['MAE_H']:.4f} RMSE={mean_roll['RMSE_H']:.4f} R²={mean_roll['R2_H']:.4f}")

    # Save rolling metrics to CSV/Excel (for 1:1 comparison with firmware logs)
    df_roll = df_roll.round(4)
    df_roll.to_csv(metrics_run_dir / "environment_base_model_metrics_rolling24_mlp.csv",
                   index=False, encoding="utf-8-sig")
    excel_roll_path = metrics_run_dir / "environment_base_model_metrics_rolling24_mlp.xlsx"
    df_roll.to_excel(excel_roll_path, index=False)

    # Auto-adjust column widths in Excel
    wb2 = load_workbook(excel_roll_path)
    ws2 = wb2.active
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb2.save(excel_roll_path)

    print("[INFO] Files saved:")
    print(" - environment_base_model_metrics_rolling24_mlp.csv")
    print(" - environment_base_model_metrics_rolling24_mlp.xlsx")

    # Plot: rolling MAE and RMSE
    plt.figure(figsize=(10, 4.8))
    plt.plot(df_roll["datetime_end"], df_roll["MAE"], label="MAE (Rolling-24)")
    plt.plot(df_roll["datetime_end"], df_roll["RMSE"], label="RMSE (Rolling-24)")
    plt.title("Rolling-Window Performance (N=24)  -  Firmware-Comparable")
    plt.xlabel("Time (window end)")
    plt.ylabel("Error (absolute scale)")
    plt.legend(frameon=False, ncol=2, loc="upper right")
    plt.grid(True, which="major")
    plt.tight_layout()
    plt.savefig(
        metrics_run_dir / "environment_base_model_metrics_rolling24_mae_rmse_mlp.png",
        dpi=600, bbox_inches="tight"
    )
    plt.close()
    
    plt.figure(figsize=(10, 4.2))
    plt.plot(df_roll["datetime_end"], df_roll["R2"], label="R² (Rolling-24)")
    plt.title("Rolling-Window Coefficient of Determination (N=24)  -  Firmware-Comparable")
    plt.xlabel("Time (window end)")
    plt.ylabel("R²")
    plt.ylim([-0.2, 1.05])
    plt.legend(frameon=False, loc="lower right")
    plt.grid(True, which="major")
    plt.tight_layout()
    plt.savefig(
        metrics_run_dir / "environment_base_model_metrics_rolling24_r2_mlp.png",
        dpi=600, bbox_inches="tight"
    )
    plt.close()
# =========================
# (NEW) Firmware-style Rolling(24) summary for metrics_summary export
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)


# Overall model status (full set)
model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied" if model_ok else "Performance thresholds not satisfied"
print("\n Overall assessment:", model_status)

# Overfitting / underfitting diagnosis - TinyML aware
train_loss = history.history["loss"]
val_loss = history.history["val_loss"]
n = min(5, len(train_loss))
mean_train_loss = float(np.mean(train_loss[-n:])) if n > 0 else float("nan")
mean_val_loss = float(np.mean(val_loss[-n:])) if n > 0 else float("nan")
gap = abs(mean_val_loss - mean_train_loss) if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss) else float("nan")
gap_pct = (gap / mean_train_loss) * 100 if (np.isfinite(gap) and mean_train_loss > 0) else float("nan")

if np.isfinite(mean_train_loss) and np.isfinite(mean_val_loss):
    if mean_train_loss > 0.3 and mean_val_loss > 0.3:
        status = "Underfitting detected (high training and validation losses)"
    elif mean_val_loss < mean_train_loss * 0.8:
        status = "Potential underfitting (validation loss significantly lower than training loss)"
    elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
        status = "Overfitting detected (large generalization gap or significant divergence)"
    elif gap_pct < 10:
        status = "Well-fitted model (generalization gap < 10%)"
    elif gap_pct < 30:
        status = "Acceptably fitted model (generalization gap < 30%)"
    else:
        status = "Mild overfitting (moderate generalization gap)"
else:
    status = "Undefined (insufficient training history for reliable assessment)"

print("\n Model diagnosis:")
print(f"• Mean training loss:     {mean_train_loss:.8f}")
print(f"• Mean validation loss:   {mean_val_loss:.8f}")
print(f"• Absolute gap:           {gap:.8f}")
print(f"• Gap percentage:         {gap_pct:.2f}%")
print(f"• Status:                 {status}")

# Plots
plt.figure(figsize=(8.2, 5.0))
plt.plot(train_loss, label="Training")
plt.plot(val_loss, label="Validation")
plt.title("Training Dynamics with Early Stopping" + status)
plt.xlabel("Epoch")
plt.ylabel("Loss (MSE)")
plt.legend(frameon=False, loc="upper right")
plt.grid(True, which="major")
plt.tight_layout()
plt.savefig(
    metrics_run_dir / "environment_base_model__training_validation_loss_diagnosis_mlp.png",
    dpi=600, bbox_inches="tight"
)
plt.close()
# Metrics bar chart in original scale
metric_labels = ["MSE", "RMSE", "MAE", "R²"]
metric_values = [mse, rmse, mae, r2]
plt.figure(figsize=(7.8, 4.8))
bars = plt.bar(metric_labels, metric_values)
plt.title(f"Evaluation Metrics (Original Scale)  -  {'Thresholds satisfied' if model_ok else 'Thresholds not satisfied'}")
plt.ylabel("Value")
plt.ylim([0, max(1e-9, max(metric_values)) * 1.25])
for bar in bars:
    yval = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2,
        yval + 0.01 * max(metric_values),
        f"{yval:.4f}",
        ha="center",
        va="bottom",
        fontsize=10
    )
plt.grid(axis="y", linestyle="--", alpha=0.25)
plt.tight_layout()
plt.savefig(
    metrics_run_dir / "environment_base_model__final_metrics_summary_plot_mlp.png",
    dpi=600, bbox_inches="tight"
)
plt.close()
# Predicted versus ground truth in original scale
plt.figure(figsize=(6.3, 6.3))
plt.scatter(y_test_abs[:, 0], y_pred_abs[:, 0], alpha=0.55, label="Temperature (T_in)")
plt.scatter(y_test_abs[:, 1], y_pred_abs[:, 1], alpha=0.55, label="Humidity (H_in)")
min_val = float(min(y_test_abs.min(), y_pred_abs.min()))
max_val = float(max(y_test_abs.max(), y_pred_abs.max()))
plt.plot([min_val, max_val], [min_val, max_val], "k--", linewidth=1.2, label="Ideal (y=x)")
plt.xlabel("Ground truth")
plt.ylabel("Prediction")
plt.title("Predicted vs. Ground Truth (Original Scale)")
try:
    plt.text(
        0.02, 0.98,
        f"N={len(y_test_abs)}\nMAE={mae:.4f}\nRMSE={rmse:.4f}\nR²={r2:.4f}",
        transform=plt.gca().transAxes,
        va="top",
        fontsize=10,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85, edgecolor="0.8")
    )
except Exception:
    pass
plt.legend(frameon=False, loc="lower right")
plt.grid(True, which="major")
plt.tight_layout()
plt.savefig(
    metrics_run_dir / "environment_base_model__scatter_predictions_mlp.png",
    dpi=600, bbox_inches="tight"
)
plt.close()
# ============================================================
# Scatter plot: temperature (T_in)
# ============================================================
plt.figure(figsize=(6.2, 6.2))
plt.scatter(T_true, T_pred_ind, alpha=0.6, label="T_in")
min_val_T = float(min(T_true.min(), T_pred_ind.min()))
max_val_T = float(max(T_true.max(), T_pred_ind.max()))
plt.plot([min_val_T, max_val_T], [min_val_T, max_val_T], "k--", linewidth=1.2, label="Ideal (y=x)")
plt.xlabel("Ground truth T_in (°C)")
plt.ylabel("Predicted T_in (°C)")
plt.title("Predicted vs. Ground Truth  -  Temperature (T_in)")
try:
    plt.text(
        0.02, 0.98,
        f"N={len(T_true)}\nMAE={mae_T:.4f} °C\nRMSE={rmse_T:.4f} °C\nR²={r2_T:.4f}",
        transform=plt.gca().transAxes,
        va="top",
        fontsize=10,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85, edgecolor="0.8")
    )
except Exception:
    pass
plt.legend(frameon=False, loc="lower right")
plt.grid(True, which="major")
plt.tight_layout()
plt.savefig(
    metrics_run_dir / "environment_base_model__scatter_Tin_mlp.png",
    dpi=600, bbox_inches="tight"
)
plt.close()
# ============================================================
# Scatter plot: humidity (H_in)
# ============================================================
plt.figure(figsize=(6.2, 6.2))
plt.scatter(H_true, H_pred_ind, alpha=0.6, label="H_in")
min_val_H = float(min(H_true.min(), H_pred_ind.min()))
max_val_H = float(max(H_true.max(), H_pred_ind.max()))
plt.plot([min_val_H, max_val_H], [min_val_H, max_val_H], "k--", linewidth=1.2, label="Ideal (y=x)")
plt.xlabel("Ground truth H_in (%)")
plt.ylabel("Predicted H_in (%)")
plt.title("Predicted vs. Ground Truth  -  Humidity (H_in)")
try:
    plt.text(
        0.02, 0.98,
        f"N={len(H_true)}\nMAE={mae_H:.4f} %\nRMSE={rmse_H:.4f} %\nR²={r2_H:.4f}",
        transform=plt.gca().transAxes,
        va="top",
        fontsize=10,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85, edgecolor="0.8")
    )
except Exception:
    pass
plt.legend(frameon=False, loc="lower right")
plt.grid(True, which="major")
plt.tight_layout()
plt.savefig(
    metrics_run_dir / "environment_base_model__scatter_Hin_mlp.png",
    dpi=600, bbox_inches="tight"
)
plt.close()
# =========================
# (NEW) Firmware-style Rolling(24) summary for metrics_summary export
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)

# ================== Metrics table ==================
metrics_dist = {
    "Base Model Metrics": [
        # Normalized (aggregated over T_in + H_in)
        "MSE (normalized residual, joint)",
        "RMSE (normalized residual, joint)",
        "MAE (normalized residual, joint)",
        "R² (normalized residual, joint)",

        # Joint original scale
        "MSE (joint original scale)",
        "RMSE (joint original scale)",
        "MAE (joint original scale)",
        "R² (joint original scale)",

        # Per-variable T_in
        "MSE_T (T_in)",
        "RMSE_T (T_in)",
        "MAE_T (T_in)",
        "R²_T (T_in)",

        # Per-variable H_in
        "MSE_H (H_in)",
        "RMSE_H (H_in)",
        "MAE_H (H_in)",
        "R²_H (H_in)",

        # Rolling(24) matching the firmware (window=24, HOUR+Invoke gating)
        "N_ALL (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, last window)",
        "RMSE (rolling24 firmware, last window)",
        "R² (rolling24 firmware, last window)",
        "MAE_T (rolling24 firmware, last window)",
        "RMSE_T (rolling24 firmware, last window)",
        "R²_T (rolling24 firmware, last window)",
        "MAE_H (rolling24 firmware, last window)",
        "RMSE_H (rolling24 firmware, last window)",
        "R²_H (rolling24 firmware, last window)",

        "MAE (rolling24 firmware, test mean)",
        "RMSE (rolling24 firmware, test mean)",
        "R² (rolling24 firmware, test mean)",
        "MAE_T (rolling24 firmware, test mean)",
        "RMSE_T (rolling24 firmware, test mean)",
        "R²_T (rolling24 firmware, test mean)",
        "MAE_H (rolling24 firmware, test mean)",
        "RMSE_H (rolling24 firmware, test mean)",
        "R²_H (rolling24 firmware, test mean)",

        # Size
        "Model size (KB)",

        # Losses / gap
        "Mean Training Loss",
        "Mean Validation Loss",
        "Absolute Gap",
        "Gap Percentage (%)",

        # Times
        "Total Inference Time (ms)",
        "Inference Time per Sample (ms)",

        # Status
        "Fit Status",
        "Model Status",
    ],
    "Value": [
        f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
        f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
        f"{mse_T:.4f}",      f"{rmse_T:.4f}",      f"{mae_T:.4f}",      f"{r2_T:.4f}",
        f"{mse_H:.4f}",      f"{rmse_H:.4f}",      f"{mae_H:.4f}",      f"{r2_H:.4f}",

        f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else "",
        f"{fw_last['MAE']:.4f}" if np.isfinite(fw_last['MAE']) else "",
        f"{fw_last['RMSE']:.4f}" if np.isfinite(fw_last['RMSE']) else "",
        f"{fw_last['R2']:.4f}" if np.isfinite(fw_last['R2']) else "",
        f"{fw_last['MAE_T']:.4f}" if np.isfinite(fw_last['MAE_T']) else "",
        f"{fw_last['RMSE_T']:.4f}" if np.isfinite(fw_last['RMSE_T']) else "",
        f"{fw_last['R2_T']:.4f}" if np.isfinite(fw_last['R2_T']) else "",
        f"{fw_last['MAE_H']:.4f}" if np.isfinite(fw_last['MAE_H']) else "",
        f"{fw_last['RMSE_H']:.4f}" if np.isfinite(fw_last['RMSE_H']) else "",
        f"{fw_last['R2_H']:.4f}" if np.isfinite(fw_last['R2_H']) else "",

        f"{fw_mean['MAE']:.4f}" if np.isfinite(fw_mean['MAE']) else "",
        f"{fw_mean['RMSE']:.4f}" if np.isfinite(fw_mean['RMSE']) else "",
        f"{fw_mean['R2']:.4f}" if np.isfinite(fw_mean['R2']) else "",
        f"{fw_mean['MAE_T']:.4f}" if np.isfinite(fw_mean['MAE_T']) else "",
        f"{fw_mean['RMSE_T']:.4f}" if np.isfinite(fw_mean['RMSE_T']) else "",
        f"{fw_mean['R2_T']:.4f}" if np.isfinite(fw_mean['R2_T']) else "",
        f"{fw_mean['MAE_H']:.4f}" if np.isfinite(fw_mean['MAE_H']) else "",
        f"{fw_mean['RMSE_H']:.4f}" if np.isfinite(fw_mean['RMSE_H']) else "",
        f"{fw_mean['R2_H']:.4f}" if np.isfinite(fw_mean['R2_H']) else "",

        "",  # filled after reading the file

        f"{mean_train_loss:.4f}",
        f"{mean_val_loss:.4f}",
        f"{gap:.4f}",
        f"{gap_pct:.2f}",

        f"{inference_time_total:.2f}",
        f"{inference_time_per_sample:.2f}",

        "",
        "",
    ],
    "Status": [
        "", "", "", "",                      # normalized joint
        mse_status, rmse_status, mae_status, r2_status,  # joint original
        "", "", "", "",                      # T_in
        "", "", "", "",                      # H_in

        "", "", "", "", "", "", "", "", "", "",          # rolling last window (10)
        "", "", "", "", "", "", "", "", "",             # rolling test mean (9)

        "",                                   # size
        "", "", "", "",                        # losses/gap
        "", "",                                # tempos

        status,                                # Fit Status
        model_status,                         # Model Status
    ],
    "Meaning": [
        "Mean squared error in the normalized residual space (ΔT_in, ΔH_in), aggregated over T_in + H_in.",
        "Root mean squared error in the normalized residual space, aggregated over T_in + H_in.",
        "Mean absolute error in the normalized residual space, aggregated over T_in + H_in.",
        "Coefficient of determination in the normalized residual space, aggregated over T_in + H_in.",

        "Joint mean squared error on the original absolute scale (T_in + H_in).",
        "Joint root mean squared error on the original absolute scale (T_in + H_in).",
        "Joint mean absolute error on the original absolute scale (T_in + H_in).",
        "Joint coefficient of determination on the original absolute scale (T_in + H_in).",

        "MSE of indoor temperature (T_in) only, on the original absolute scale.",
        "RMSE of indoor temperature (T_in) only, on the original absolute scale.",
        "MAE of indoor temperature (T_in) only, on the original absolute scale.",
        "R² of indoor temperature (T_in) only, on the original absolute scale.",

        "MSE of indoor humidity (H_in) only, on the original absolute scale.",
        "RMSE of indoor humidity (H_in) only, on the original absolute scale.",
        "MAE of indoor humidity (H_in) only, on the original absolute scale.",
        "R² of indoor humidity (H_in) only, on the original absolute scale.",

        "Rolling(24): number of samples (HOUR events) effectively present in the final window.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",

        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) aggregate (T+H) MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) T_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",
        "Rolling(24) H_in MEAN over the test set (mean across windows).",

        "Size of the model file saved on disk (proxy for FLASH/storage).",

        "Mean of the last training losses (convergence indicator).",
        "Mean of the last validation losses (generalization indicator).",
        "Absolute difference between mean losses (validation vs training).",
        "Percentage difference between mean losses (validation vs training).",

        "Total time spent inferring the test set (ms).",
        "Average latency per sample in the test set (ms/sample).",

        "Fit diagnosis (underfitting/overfitting) based on the training/validation gap.",
        "Overall model status based on MSE/RMSE/MAE/R² thresholds (original scale).",
    ],
    "Expected Values": [
        "→ Lower is better (normalized).",
        "→ Lower is better (normalized).",
        "→ Lower is better (normalized).",
        "→ The closer to 1, the better (normalized).",

        "→ Lower is better (original scale).",
        "→ Lower is better (original scale).",
        "→ Lower is better (original scale).",
        "→ Ideally ≥ 0.8 (may be NaN if variance is too low).",

        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Ideally ≥ 0.8 (may be NaN if variance is too low).",

        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Ideally ≥ 0.8 (may be NaN if variance is too low).",

        "→ Should reach 24 when the window is full (after warm-up).",
        "→ Lower is better (rolling24).",
        "→ Lower is better (rolling24).",
        "→ May be NaN if variance is low (same as firmware).",
        "→ Lower is better (rolling24 T_in).",
        "→ Lower is better (rolling24 T_in).",
        "→ May be NaN if variance is low (same as firmware).",
        "→ Lower is better (rolling24 H_in).",
        "→ Lower is better (rolling24 H_in).",
        "→ May be NaN if variance is low (same as firmware).",

        "→ Lower is better (window mean).",
        "→ Lower is better (window mean).",
        "→ May be NaN if variance is low (same as firmware).",
        "→ Lower is better (window mean T_in).",
        "→ Lower is better (window mean T_in).",
        "→ May be NaN if variance is low (same as firmware).",
        "→ Lower is better (window mean H_in).",
        "→ Lower is better (window mean H_in).",
        "→ May be NaN if variance is low (same as firmware).",

        "→ Smaller is better (TinyML: smaller KB values ease deployment).",

        "→ Lower and more stable is better.",
        "→ Lower and more stable is better.",
        "→ Ideal: low (small gap).",
        "→ Ideal: <10% (well-fitted) | <30% (acceptable fit) | >50% (overfitting).",

        "→ Lower is better (execution).",
        "→ Lower is better (latency).",

        "→ Ideal: 'Well-fitted Model' or 'Possibly well-fitted Model'.",
        "→ Ideal: 'Performance thresholds satisfied'.",
    ],
}

# Model size (fill now)
saved_model_path = run_dir / "environment_base_model_mlp.keras"
if os.path.isfile(saved_model_path):
    model_size_kb = os.path.getsize(saved_model_path) / 1024
    print(f"Trained model size: {model_size_kb:.2f} KB")
    metrics_dist["Value"][35] = f"{model_size_kb:.2f} KB"
else:
    print("Model file not found.")
    metrics_dist["Value"][35] = ""

df_metrics = pd.DataFrame(metrics_dist)
df_metrics.to_csv(metrics_run_dir / "environment_base_model_metrics_summary_mlp.csv", index=False, encoding="utf-8-sig")
print("[INFO] File saved: environment_base_model_metrics_summary_mlp.csv")

excel_path = metrics_run_dir / "environment_base_model_metrics_summary_mlp.xlsx"
df_metrics.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

wb.save(excel_path)
print("[INFO] File saved: environment_base_model_metrics_summary_mlp.xlsx (with adjusted columns)")

# === Post-execution: update 'latest' and manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Could not update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Could not write manifest.json:", _e)

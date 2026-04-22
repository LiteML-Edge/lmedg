"""
Script: environment_pruned_model_mlp.py
Module role:
    Train, prune, fine-tune, and evaluate the pruned MLP model adopted in the
    LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, performs pruning-aware training and fine-tuning,
    evaluates the resulting model in normalized and reconstructed physical
    domains, and exports metrics, figures, and versioned artifacts.

Inputs:
    - environment_dataset_mlp.csv
    - Baseline model path utilities and versioning helpers

Outputs:
    - Versioned pruned-model artifacts
    - Evaluation tables, figures, and summary spreadsheets

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os,sys 
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import joblib
import tensorflow_model_optimization as tfmot
from tensorflow_model_optimization.sparsity import keras as sparsity
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"

rcParams["font.family"] = "Segoe UI Emoji"
# Figure export settings
rcParams["figure.dpi"] = 120
rcParams["savefig.dpi"] = 600
rcParams["font.size"] = 11
rcParams["axes.titlesize"] = 13
rcParams["axes.labelsize"] = 12
rcParams["legend.fontsize"] = 10
rcParams["xtick.labelsize"] = 10
rcParams["ytick.labelsize"] = 10
rcParams["axes.grid"] = True
rcParams["grid.alpha"] = 0.25
rcParams["grid.linestyle"] = "--"
rcParams["axes.spines.top"] = False
rcParams["axes.spines.right"] = False
# ================================================

rcParams['font.family'] = 'Segoe UI Emoji'

# --- Bootstrap: allows importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_mlp import (
    PROJECT_ROOT,
    DATASET_ENVIRONMENT_MLP,
    BASE_MODEL,
    PRUNED_MODEL,
    PRUNED_MODEL_METRICS
)
from utils.global_utils.versioning import (
    create_versioned_dir,
    ensure_dir,
    update_latest,
    write_manifest,
    list_runs,
    resolve_run,
    resolve_latest,
)
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Must be called at the top of the script, BEFORE model creation

# === Versioned directories for the current execution ===
run_dir = create_versioned_dir(PRUNED_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(PRUNED_MODEL_METRICS / run_dir.name)
version_path = resolve_latest(BASE_MODEL)
# ===========================================

# -------------------------------------------------------------------------
# LiteML-Edge Artifact Contract
# -------------------------------------------------------------------------
# 1. Feature order must remain identical across:
#      - dataset generation
#      - scaler export
#      - Python-side evaluation
#      - embedded firmware inference
#
# 2. Targets are residual:
#      ΔT_in = T_in(t) − T_in_lag1
#      ΔH_in = H_in(t) − H_in_lag1
#
# 3. Absolute reconstruction:
#      T_in = T_in_lag1 + ΔT_in
#      H_in = H_in_lag1 + ΔH_in
#
# 4. Input normalization:
#      X_clamped = clip(X, scaler_X.data_min_, scaler_X.data_max_)
#      X_norm = scaler_X.transform(X_clamped)
#
# 5. Rolling(24) metrics are firmware-equivalent:
#      - window length = 24
#      - same aggregation rules
#      - same R² formulation
#      - direct comparability with firmware logs
# -------------------------------------------------------------------------

# ---------------------
# Paths
# ---------------------
model_path     = version_path / "environment_base_model_mlp.keras"
scaler_X_path  = version_path / "environment_base_model_mlp_scaler_X.pkl"
scaler_y_path  = version_path / "environment_base_model_mlp_scaler_y.pkl"
dataset_path   = DATASET_ENVIRONMENT_MLP / "environment_dataset_mlp.csv"

# =====================================================================
#                           DATA PREPARATION
# =====================================================================
df = pd.read_csv(dataset_path)

# Sort temporally to build causal windows
df = df.sort_values("datetime").reset_index(drop=True)

# Temporal feature engineering
df["datetime"] = pd.to_datetime(df["datetime"])
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (consistent with Conv1D Tiny) ===
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=0.08, adjust=False).mean()

# Cyclical attributes
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lag features (first-order causal dependencies)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Fixed feature option: no diff terms; indoor lag2 retained to preserve the 12-feature specification plus weekday and month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)

# Fixed feature order (12 features)  -  Option 3 (without T_diff/H_diff) + weekday/month
features = [
    "T_out", "H_out",           # current outdoor
    "T_in_lag1", "H_in_lag1",   # indoor lag1
    "T_out_lag1", "H_out_lag1", # outdoor lag1
    "T_in_lag2", "H_in_lag2",   # indoor lag2
    "sin_hour", "cos_hour",     # cyclical hour
    "weekday", "month",         # calendar
]

# =====================================================================
#          ABSOLUTE AND RESIDUAL TARGETS (ΔT_in, ΔH_in)  -  RESIDUAL ONLY
# =====================================================================

# Absolute targets retained for reconstruction and reporting
targets_abs = ["T_in", "H_in"]
y_full_abs = df[targets_abs].values.astype(np.float32)

# Raw residual targets: ΔT_in, ΔH_in
y_full_res = np.stack(
    [
        df["T_in"].values.astype(np.float32) - df["T_in_lag1"].values.astype(np.float32),
        df["H_in"].values.astype(np.float32) - df["H_in_lag1"].values.astype(np.float32),
    ],
    axis=1,
).astype(np.float32)

y_full_res = y_full_res.astype(np.float32)

# The training target is exclusively residual from this point onward
y_full_main = y_full_res  # RESIDUAL ONLY

# =====================================================================
#      SLIDING WINDOW 24×12 → MLP TARGETS  -  3D TENSOR AS IN Conv1D
#      NORMALIZE THE 12 FEATURES FIRST, THEN FLATTEN TO 288
# =====================================================================
WINDOW = 24
X_source = df[features].values.astype(np.float32)
N = len(df)

if N < WINDOW:
    raise ValueError("Insufficient dataset to construct 24-step windows.")

X_win_list   = []  # (24, 12)
y_main_list  = []
y_abs_list   = []
idx_list     = []

for t in range(WINDOW - 1, N):
    # window [t-23, ..., t] → (24, 12)
    window = X_source[t - WINDOW + 1 : t + 1, :]
    X_win_list.append(window)
    y_main_list.append(y_full_main[t])   # residual target used during training
    y_abs_list.append(y_full_abs[t])     # absolute targets retained for reporting
    idx_list.append(t)

# 3D representation for feature-wise normalization
X_win      = np.stack(X_win_list).astype(np.float32)   # (N_seq, 24, 12)
y_all_main = np.stack(y_main_list).astype(np.float32)  # (N_seq, 2) residual
y_all_abs  = np.stack(y_abs_list).astype(np.float32)   # (N_seq, 2) absolute
idx_arr    = np.array(idx_list, dtype=np.int64)

# Auxiliary vectors for absolute reconstruction
T_prev_all = df["T_in_lag1"].values.astype(np.float32)[idx_arr]
H_prev_all = df["H_in_lag1"].values.astype(np.float32)[idx_arr]

# =====================================================================
#                    TRAIN / VAL / TEST SPLIT (TEMPORAL)
#              SPLIT APPLIED ON THE 3D TENSOR, AS IN Conv1D
# =====================================================================
n_total = X_win.shape[0]

train_frac = 0.6
val_frac   = 0.2  # the test set receives the remainder

n_train = int(n_total * train_frac)
n_val   = int(n_total * val_frac)
n_test  = n_total - n_train - n_val

if n_train <= 0 or n_val <= 0 or n_test <= 0:
    raise ValueError(
        f"Invalid split with n_total={n_total}, "
        f"n_train={n_train}, n_val={n_val}, n_test={n_test}"
    )

# --- TRAIN: oldest segment ---
X_train_win     = X_win[:n_train]                 # (n_train, 24, 12)
y_train_raw     = y_all_main[:n_train]           # residual
Tprev_train     = T_prev_all[:n_train]
Hprev_train     = H_prev_all[:n_train]
yabs_train_full = y_all_abs[:n_train]

# --- VALIDATION: intermediate segment ---
X_val_win       = X_win[n_train : n_train + n_val]
y_val_raw       = y_all_main[n_train : n_train + n_val]
Tprev_val       = T_prev_all[n_train : n_train + n_val]
Hprev_val       = H_prev_all[n_train : n_train + n_val]
yabs_val_full   = y_all_abs[n_train : n_train + n_val]

# --- TEST: most recent segment ---
X_test_win      = X_win[n_train + n_val :]
y_test_raw      = y_all_main[n_train + n_val :]
Tprev_test      = T_prev_all[n_train + n_val :]
Hprev_test      = H_prev_all[n_train + n_val :]
yabs_test_full  = y_all_abs[n_train + n_val :]

print("[INFO] Temporal split applied (pruned residual MLP):")
print(f"  • Training:   {n_train} samples")
print(f"  • Validation: {n_val} samples")
print(f"  • Test:       {n_test} samples")

# ---------------------
# Load scalers trained in BASE_MODEL
# ---------------------
scaler_X = joblib.load(scaler_X_path)
scaler_y = joblib.load(scaler_y_path)

# =====================================================================
#            X NORMALIZATION ACROSS THE 12 FEATURES → THEN FLATTEN 24×12
#            SAME LOGIC AS Conv1D/BASE_MODEL FOR SCALER_X
# =====================================================================
N_train, W_steps, F = X_train_win.shape   # F must be 12

# X_train
X_train_2d = X_train_win.reshape(-1, F)   # (N_train * 24, 12)
X_train_2d = np.clip(X_train_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
# LiteML-Edge contract: clamping preserves scaler-domain compatibility and
# keeps min-max forward values within [0,1] for deployment-consistent inference.
X_train_2d_norm = scaler_X.transform(X_train_2d)
X_train = X_train_2d_norm.reshape(N_train, W_steps * F)  # (N_train, 288)

# X_val
N_val = X_val_win.shape[0]
X_val_2d = X_val_win.reshape(-1, F)
X_val_2d = np.clip(X_val_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
# LiteML-Edge contract: clamping preserves scaler-domain compatibility and
# keeps min-max forward values within [0,1] for deployment-consistent inference.
X_val_2d_norm = scaler_X.transform(X_val_2d)
X_val = X_val_2d_norm.reshape(N_val, W_steps * F)

# X_test
N_test = X_test_win.shape[0]
X_test_2d = X_test_win.reshape(-1, F)
X_test_2d = np.clip(X_test_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
# LiteML-Edge contract: clamping preserves scaler-domain compatibility and
# keeps min-max forward values within [0,1] for deployment-consistent inference.
X_test_2d_norm = scaler_X.transform(X_test_2d)
X_test = X_test_2d_norm.reshape(N_test, W_steps * F)

# =====================================================================
#           y NORMALIZATION WITH THE SAME scaler_y FROM BASE_MODEL
#           y IS ALWAYS RESIDUAL HERE: ΔT_in, ΔH_in
# =====================================================================
y_train = scaler_y.transform(y_train_raw)
y_val   = scaler_y.transform(y_val_raw)
y_test  = scaler_y.transform(y_test_raw)

# ---------------------
# Load base model (same MLP architecture 288→2)
# ---------------------
base_model = tf.keras.models.load_model(model_path)

# ---------------------
# Define pruning schedule (constant sparsity)
# ---------------------
batch_size      = 512
epochs_prune    = 100
steps_per_epoch = int(np.ceil(len(X_train) / batch_size))

pruning_schedule = sparsity.ConstantSparsity(
    target_sparsity=0.50,
    begin_step=0,
    frequency=steps_per_epoch,
)

pruned_model = sparsity.prune_low_magnitude(base_model, pruning_schedule=pruning_schedule)

# ---------------------
# Compile (reduced learning rate for stability)
# ---------------------
pruned_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-3),
    loss=tf.keras.losses.MeanSquaredError(),
    metrics=[tf.keras.metrics.MeanAbsoluteError()],
)

# ---------------------
# Callbacks
# ---------------------
callbacks = [
    sparsity.UpdatePruningStep(),
    tf.keras.callbacks.EarlyStopping(
        monitor="val_loss",
        patience=10,
        restore_best_weights=True,
        verbose=1,
    ),
]

# ---------------------
# PHASE 1  -  Pruning-aware training
# ---------------------
history = pruned_model.fit(
    X_train,
    y_train,
    validation_data=(X_val, y_val),
    batch_size=batch_size,
    epochs=epochs_prune,
    shuffle=True,  # acceptable because temporal order was already respected in the split
    callbacks=callbacks,
    verbose=1,
)

# ---------------------
# Remove pruning wrapper and create final deployment model
# ---------------------
final_model = sparsity.strip_pruning(pruned_model)

# ---------------------
# PHASE 2  -  Post-pruning fine-tuning
# ---------------------
fine_tune_epochs = 8
fine_tune_lr     = 1e-5

final_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=fine_tune_lr),
    loss=tf.keras.losses.MeanSquaredError(),
    metrics=[tf.keras.metrics.MeanAbsoluteError()],
)

history_ft = final_model.fit(
    X_train,
    y_train,
    validation_data=(X_val, y_val),
    batch_size=batch_size,
    epochs=fine_tune_epochs,
    shuffle=True,
    verbose=1,
)

# Concatenate training histories
for key, values in history_ft.history.items():
    if key in history.history:
        history.history[key] = history.history[key] + values
    else:
        history.history[key] = values

# Save fine-tuned final model
final_model_path = run_dir / "environment_pruned_model_mlp.keras"
final_model.save(final_model_path)

# ---------------------
# Function for actual sparsity measurement
# ---------------------
def calc_sparsity(m):
    zeros = 0
    total = 0
    for w in m.get_weights():
        arr = np.asarray(w)
        zeros += np.sum(arr == 0)
        total += arr.size
    return zeros / total if total else 0.0

sparsity_before = calc_sparsity(pruned_model)
sparsity_after  = calc_sparsity(final_model)
print(f"\n Sparsity (before strip): {sparsity_before:.2%}")
print(f"Sparsity (after strip):  {sparsity_after:.2%}")

# ---------------------
# Inference and latency measurement
# ---------------------
start_time = time.time()
y_pred_test_scaled = final_model.predict(X_test)
end_time = time.time()
inference_time_total = (end_time - start_time) * 1000.0  # ms
inference_time_per_sample = inference_time_total / len(X_test)

print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Mean latency per sample: {inference_time_per_sample:.4f} ms")

# Model size in bytes
if os.path.isfile(final_model_path):
    model_size_kb = os.path.getsize(final_model_path) / 1024
    print(f"Pruned model size: {model_size_kb:.2f} KB")
else:
    print("Model file not found.")

# =====================================================================
#     DENORMALIZE AND RECONSTRUCT (RESIDUAL → ABSOLUTE) - TEST SET
# =====================================================================
# Return to the original domain of the training target (always residual ΔT_in, ΔH_in)
y_pred_delta = scaler_y.inverse_transform(y_pred_test_scaled).astype(np.float32, copy=False)

# Ensure prev arrays are float32
Tprev_test = Tprev_test.astype(np.float32, copy=False)
Hprev_test = Hprev_test.astype(np.float32, copy=False)

# y_pred_delta = [ΔT_in, ΔH_in] → absolute reconstruction
T_pred = (Tprev_test + y_pred_delta[:, 0]).astype(np.float32, copy=False)
H_pred = (Hprev_test + y_pred_delta[:, 1]).astype(np.float32, copy=False)

y_pred_abs = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)
y_test_abs = yabs_test_full.astype(np.float32, copy=False)  # true absolute [T_in, H_in]

# =====================================================================
#                        METRICS (ORIGINAL SCALE)
# =====================================================================
mse = mean_squared_error(y_test_abs, y_pred_abs)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_abs, y_pred_abs)
r2  = r2_score(y_test_abs, y_pred_abs)

print("\n Results (original scale - joint target set [T_in, H_in]):")
mse_status  = "MSE within threshold"  if mse  <= 0.1  else "MSE above threshold"
rmse_status = "RMSE within threshold" if rmse <= 0.32 else "RMSE above threshold"
mae_status  = "MAE within threshold"  if mae  <= 0.3  else "MAE above threshold"
r2_status   = "R² within threshold"   if r2   >= 0.8  else "R² below threshold"

print(f"MSE  = {mse:.4f}   {mse_status}")
print(f"RMSE = {rmse:.4f}   {rmse_status}")
print(f"MAE  = {mae:.4f}   {mae_status}")
print(f"R²   = {r2:.4f}   {r2_status}")

# =====================================================================
#     METRICS IN THE TRAINING TARGET SPACE (NORMALIZED RESIDUAL)
# =====================================================================
mse_scaled  = mean_squared_error(y_test, y_pred_test_scaled)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled  = mean_absolute_error(y_test, y_pred_test_scaled)
r2_scaled   = r2_score(y_test, y_pred_test_scaled)

print("\n Results (normalized scale - residual training target):")
print(f"MSE  (norm.) = {mse_scaled:.4f}")
print(f"RMSE (norm.) = {rmse_scaled:.4f}")
print(f"MAE  (norm.) = {mae_scaled:.4f}")
print(f"R²   (norm.) = {r2_scaled:.4f}")

# ========= Individual metrics for T_in and H_in (original scale) =========
T_true = y_test_abs[:, 0]
T_pred_ind = y_pred_abs[:, 0]
H_true = y_test_abs[:, 1]
H_pred_ind = y_pred_abs[:, 1]

mse_T  = mean_squared_error(T_true, T_pred_ind)
rmse_T = np.sqrt(mse_T)
mae_T  = mean_absolute_error(T_true, T_pred_ind)
r2_T   = r2_score(T_true, T_pred_ind)

mse_H  = mean_squared_error(H_true, H_pred_ind)
rmse_H = np.sqrt(mse_H)
mae_H  = mean_absolute_error(H_true, H_pred_ind)
r2_H   = r2_score(H_true, H_pred_ind)

print("\n Individual metrics - Temperature (T_in):")
print(f"MSE_T  = {mse_T:.4f}")
print(f"RMSE_T = {rmse_T:.4f}")
print(f"MAE_T  = {mae_T:.4f}")
print(f"R²_T   = {r2_T:.4f}")

print("\n Individual metrics - Humidity (H_in):")
print(f"MSE_H  = {mse_H:.4f}")
print(f"RMSE_H = {rmse_H:.4f}")
print(f"MAE_H  = {mae_H:.4f}")
print(f"R²_H   = {r2_H:.4f}")

# Overall model status (joint)
model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied" if model_ok else "Performance thresholds not satisfied"
print("\n Overall assessment:", model_status)

# ---------------------
# Generalization assessment (using concatenated history)
# ---------------------
train_loss = history.history["loss"]
val_loss   = history.history["val_loss"]
n = min(5, len(train_loss))
mean_train_loss = float(np.mean(train_loss[-n:]))
mean_val_loss   = float(np.mean(val_loss[-n:]))
gap = abs(mean_val_loss - mean_train_loss)
gap_pct = (gap / mean_train_loss) * 100 if mean_train_loss > 0 else 0.0

if mean_train_loss > 0.3 and mean_val_loss > 0.3:
    status = "Underfitting detected (high training and validation losses)"
elif mean_val_loss < mean_train_loss * 0.8:
    status = "Potential underfitting (validation loss significantly lower than training loss)"
elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
    status = "Overfitting detected (large generalization gap or significant divergence)"
elif gap_pct < 10:
    status = "Well-fitted model (generalization gap < 10%)"
elif gap_pct < 30:
    status = "Acceptably fitted model (generalization gap < 30%)"
else:
    status = "Mild overfitting (moderate generalization gap)"

print("\n Model diagnosis:")
print(f"• Mean training loss:    {mean_train_loss:.4f}")
print(f"• Mean validation loss:  {mean_val_loss:.4f}")
print(f"• Absolute gap:          {gap:.4f}")
print(f"• Percentage gap:        {gap_pct:.2f}%")
print(f"• Assessment:            {status}")

# Training convergence plots
plt.figure(figsize=(8, 5))
plt.plot(train_loss, label="Training (pruning + fine-tuning)")
plt.plot(val_loss,   label="Validation")
plt.title("Training Convergence (Pruning + Fine-Tuning)\n" + status)
plt.xlabel("Epoch")
plt.ylabel("Loss (MSE)")
plt.legend()
plt.grid()
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_pruned_model_training_validation_loss_diagnosis_mlp.png", dpi=600, bbox_inches="tight")
plt.close()

# Prediction scatter (original scale) - joint
plt.figure(figsize=(6, 6))
plt.scatter(y_test_abs[:, 0], y_pred_abs[:, 0], alpha=0.5, label="Temperature (T_in)")
plt.scatter(y_test_abs[:, 1], y_pred_abs[:, 1], alpha=0.5, label="Humidity (H_in)")
min_val = min(y_test_abs.min(), y_pred_abs.min())
max_val = max(y_test_abs.max(), y_pred_abs.max())
plt.plot([min_val, max_val], [min_val, max_val], "k--")
plt.xlabel("Ground truth")
plt.ylabel("Prediction")
plt.title("Predicted vs. Ground Truth (Pruned MLP, Original Scale)")
plt.legend()
plt.grid()
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_pruned_model_scatter_predictions_mlp.png", dpi=600, bbox_inches="tight")
plt.close()

# =====================================================================
# Firmware-equivalent Rolling(24) evaluation (window=24, HOUR+Invoke gating)
# =====================================================================

# -------------------------------
# Datetime aligned with sequences (for rolling plots)
# -------------------------------
try:
    dt_seq = df["datetime"].values[idx_arr]
    dt_test = dt_seq[n_train + n_val :]
except Exception:
    dt_test = pd.to_datetime(np.arange(y_test_abs.shape[0]))

ROLLING_N = 24  # N=24 samples (HOUR), consistent with metrics.cpp

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """Replicates the firmware logic (metrics.cpp):
      R² = 1 - SS_res/SS_tot
      SS_tot = sum(y^2) - (sum(y)^2)/n
      If n < 2 -> NaN
      If SS_tot <= 1e-6 -> NaN
    """
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return float("nan")

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))

    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return float("nan")

    return 1.0 - (ss_res / ss_tot)


def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    """
    Returns (mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H)
    following exactly:
      mae = 0.5*(mae_T + mae_H)
      rmse = sqrt(0.5*(mse_T + mse_H))
      r2 = 0.5*(r2_T + r2_H)
    """
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # Temperature channel
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T * err_T))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = float(_r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0]))

    # Humidity channel
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H * err_H))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = float(_r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1]))

    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H)
    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

rolling_rows = []
n_test = y_test_abs.shape[0]

# Offline gating equivalent to firmware semantics: all test samples are valid
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)

if n_test < ROLLING_N:
    print(f"\n Rolling N=24 cannot be computed: the test set has only {n_test} samples.")
else:
    for end in range(ROLLING_N - 1, n_test):
        # Firmware semantics: metrics are updated only after real Invoke and HOUR event
        if (not invoked_mask[end]) or (not is_rollover_mask[end]):
            rolling_rows.append({
                "window_start": int(end - ROLLING_N + 1),
                "window_end": int(end),
                "datetime_end": pd.to_datetime(dt_test[end]),
                "N": int(ROLLING_N),

                "MAE": float("nan"),
                "RMSE": float("nan"),
                "R2": float("nan"),

                "MAE_T": float("nan"),
                "RMSE_T": float("nan"),
                "R2_T": float("nan"),

                "MAE_H": float("nan"),
                "RMSE_H": float("nan"),
                "R2_H": float("nan"),
            })
            continue
        start = end - ROLLING_N + 1
        yt_w = y_test_abs[start:end+1, :]
        yp_w = y_pred_abs[start:end+1, :]

        mae_w, rmse_w, r2_w, maeT_w, rmseT_w, r2T_w, maeH_w, rmseH_w, r2H_w = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": pd.to_datetime(dt_test[end]),
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": maeT_w,
            "RMSE_T": rmseT_w,
            "R2_T": r2T_w,

            "MAE_H": maeH_w,
            "RMSE_H": rmseH_w,
            "R2_H": r2H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # Summaries for direct comparison with the firmware current-state view
    last = df_roll.iloc[-1]
    mean_roll = df_roll[["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]].mean()

    print("\n[INFO] Rolling N=24 (firmware-equivalent), last window:")
    print(f"MAE={last['MAE']:.4f} RMSE={last['RMSE']:.4f} R²={last['R2']:.4f} | "
          f"T: MAE={last['MAE_T']:.4f} RMSE={last['RMSE_T']:.4f} R²={last['R2_T']:.4f} | "
          f"H: MAE={last['MAE_H']:.4f} RMSE={last['RMSE_H']:.4f} R²={last['R2_H']:.4f}")

    print("\n[INFO] Rolling N=24 mean across all test windows:")
    print(f"MAE={mean_roll['MAE']:.4f} RMSE={mean_roll['RMSE']:.4f} R²={mean_roll['R2']:.4f} | "
          f"T: MAE={mean_roll['MAE_T']:.4f} RMSE={mean_roll['RMSE_T']:.4f} R²={mean_roll['R2_T']:.4f} | "
          f"H: MAE={mean_roll['MAE_H']:.4f} RMSE={mean_roll['RMSE_H']:.4f} R²={mean_roll['R2_H']:.4f}")

    # Save rolling metrics as CSV/Excel for direct firmware-log comparison
    df_roll = df_roll.round(4)
    df_roll.to_csv(metrics_run_dir / "environment_pruned_model_metrics_rolling24_mlp.csv",
                   index=False, encoding="utf-8-sig")
    excel_roll_path = metrics_run_dir / "environment_pruned_model_metrics_rolling24_mlp.xlsx"
    df_roll.to_excel(excel_roll_path, index=False)

    # Auto-adjust column widths in Excel
    wb2 = load_workbook(excel_roll_path)
    ws2 = wb2.active
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb2.save(excel_roll_path)

    print("[INFO] Files saved:")
    print(" - environment_pruned_model_metrics_rolling24_mlp.csv")
    print(" - environment_pruned_model_metrics_rolling24_mlp.xlsx")

    # Plot: rolling MAE and RMSE
    plt.figure(figsize=(10, 4.8))
    plt.plot(df_roll["datetime_end"], df_roll["MAE"], label="MAE (Rolling-24)")
    plt.plot(df_roll["datetime_end"], df_roll["RMSE"], label="RMSE (Rolling-24)")
    plt.title("Firmware-Equivalent Rolling-Window Performance (N=24)")
    plt.xlabel("Time (window end)")
    plt.ylabel("Error (absolute scale)")
    plt.legend(frameon=False, ncol=2, loc="upper right")
    plt.grid(True, which="major")
    plt.tight_layout()
    plt.savefig(
        metrics_run_dir / "environment_pruned_model_metrics_rolling24_mae_rmse_mlp.png",
        dpi=600, bbox_inches="tight"
    )
    plt.close()

    plt.figure(figsize=(10, 4.2))
    plt.plot(df_roll["datetime_end"], df_roll["R2"], label="R² (Rolling-24)")
    plt.title("Firmware-Equivalent Rolling Coefficient of Determination (N=24)")
    plt.xlabel("Time (window end)")
    plt.ylabel("R²")
    plt.ylim([-0.2, 1.05])
    plt.legend(frameon=False, loc="lower right")
    plt.grid(True, which="major")
    plt.tight_layout()
    plt.savefig(
        metrics_run_dir / "environment_pruned_model_metrics_rolling24_r2_mlp.png",
        dpi=600, bbox_inches="tight"
    )
    plt.close()

# =========================
# (NEW) Firmware-style Rolling(24) summary for inclusion in metrics_summary
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract Rolling(24) summary:", _e)

# ============================================================
# Separate scatter plot  -  T_in
# ============================================================
plt.figure(figsize=(6, 6))
plt.scatter(T_true, T_pred_ind, alpha=0.5, label="T_in")
min_val_T = min(T_true.min(), T_pred_ind.min())
max_val_T = max(T_true.max(), T_pred_ind.max())
plt.plot([min_val_T, max_val_T], [min_val_T, max_val_T], "k--")
plt.xlabel("Ground truth T_in (°C)")
plt.ylabel("Predicted T_in (°C)")
plt.title("Predicted vs. Ground Truth  -  Temperature (T_in)")
plt.grid()
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_pruned_model_scatter_Tin_mlp.png", dpi=600, bbox_inches="tight")
plt.close()

# ============================================================
# Separate scatter plot  -  H_in
# ============================================================
plt.figure(figsize=(6, 6))
plt.scatter(H_true, H_pred_ind, alpha=0.5, label="H_in")
min_val_H = min(H_true.min(), H_pred_ind.min())
max_val_H = max(H_true.max(), H_pred_ind.max())
plt.plot([min_val_H, max_val_H], [min_val_H, max_val_H], "k--")
plt.xlabel("Ground truth H_in (%)")
plt.ylabel("Predicted H_in (%)")
plt.title("Predicted vs. Ground Truth  -  Humidity (H_in)")
plt.grid()
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_pruned_model_scatter_Hin_mlp.png", dpi=600, bbox_inches="tight")
plt.close()

# ============================================================
# Bar plot of the main metrics (joint)
# ============================================================
plt.figure(figsize=(8, 6))
labels = ["MSE", "RMSE", "MAE", "R²"]
values = [mse, rmse, mae, r2]
bars = plt.bar(labels, values)
plt.title("Evaluation Metrics Summary  -  Pruned Model (Original Scale)")
plt.grid(axis="y")
for bar in bars:
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width() / 2.0, yval, f"{yval:.4f}", ha="center", va="bottom")
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_pruned_model_final_metrics_summary_plot_mlp.png", dpi=600, bbox_inches="tight")
plt.close()

# =====================================================================
# Metrics table (CSV / XLSX)
# =====================================================================
try:
    size_str = f"{model_size_kb:.2f}"
except NameError:
    size_str = ""

metrics_dist = {
    "Metric": [
        "MSE (normalized residual)",
        "RMSE (normalized residual)",
        "MAE (normalized residual)",
        "R² (normalized residual)",
        "MSE (joint original)",
        "RMSE (joint original)",
        "MAE (joint original)",
        "R² (joint original)",
        "MSE_T (T_in)",
        "RMSE_T (T_in)",
        "MAE_T (T_in)",
        "R²_T (T_in)",
        "MSE_H (H_in)",
        "RMSE_H (H_in)",
        "MAE_H (H_in)",
        "R²_H (H_in)",

        # Firmware-equivalent Rolling(24) summary
        "N_ALL (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, last window)",
        "RMSE (rolling24 firmware, last window)",
        "R² (rolling24 firmware, last window)",
        "MAE_T (rolling24 firmware, last window)",
        "RMSE_T (rolling24 firmware, last window)",
        "R²_T (rolling24 firmware, last window)",
        "MAE_H (rolling24 firmware, last window)",
        "RMSE_H (rolling24 firmware, last window)",
        "R²_H (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, test mean)",
        "RMSE (rolling24 firmware, test mean)",
        "R² (rolling24 firmware, test mean)",
        "MAE_T (rolling24 firmware, test mean)",
        "RMSE_T (rolling24 firmware, test mean)",
        "R²_T (rolling24 firmware, test mean)",
        "MAE_H (rolling24 firmware, test mean)",
        "RMSE_H (rolling24 firmware, test mean)",
        "R²_H (rolling24 firmware, test mean)",
        "Sparsity before strip (%)",
        "Sparsity after strip (%)",
        "Model size (KB)",
        "Mean training loss",
        "Mean validation loss",
        "Absolute gap",
        "Percentage gap (%)",
        "Total inference time (ms)",
        "Inference time per sample (ms)",
        "Generalization assessment",
        "Overall model status",
    ],
    "Value": [
        f"{mse_scaled:.4f}",
        f"{rmse_scaled:.4f}",
        f"{mae_scaled:.4f}",
        f"{r2_scaled:.4f}",
        f"{mse:.4f}",
        f"{rmse:.4f}",
        f"{mae:.4f}",
        f"{r2:.4f}",
        f"{mse_T:.4f}",
        f"{rmse_T:.4f}",
        f"{mae_T:.4f}",
        f"{r2_T:.4f}",
        f"{mse_H:.4f}",
        f"{rmse_H:.4f}",
        f"{mae_H:.4f}",
        f"{r2_H:.4f}",

        f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else "",
        f"{fw_last['MAE']:.4f}" if np.isfinite(fw_last['MAE']) else "",
        f"{fw_last['RMSE']:.4f}" if np.isfinite(fw_last['RMSE']) else "",
        f"{fw_last['R2']:.4f}" if np.isfinite(fw_last['R2']) else "",
        f"{fw_last['MAE_T']:.4f}" if np.isfinite(fw_last['MAE_T']) else "",
        f"{fw_last['RMSE_T']:.4f}" if np.isfinite(fw_last['RMSE_T']) else "",
        f"{fw_last['R2_T']:.4f}" if np.isfinite(fw_last['R2_T']) else "",
        f"{fw_last['MAE_H']:.4f}" if np.isfinite(fw_last['MAE_H']) else "",
        f"{fw_last['RMSE_H']:.4f}" if np.isfinite(fw_last['RMSE_H']) else "",
        f"{fw_last['R2_H']:.4f}" if np.isfinite(fw_last['R2_H']) else "",
        f"{fw_mean['MAE']:.4f}" if np.isfinite(fw_mean['MAE']) else "",
        f"{fw_mean['RMSE']:.4f}" if np.isfinite(fw_mean['RMSE']) else "",
        f"{fw_mean['R2']:.4f}" if np.isfinite(fw_mean['R2']) else "",
        f"{fw_mean['MAE_T']:.4f}" if np.isfinite(fw_mean['MAE_T']) else "",
        f"{fw_mean['RMSE_T']:.4f}" if np.isfinite(fw_mean['RMSE_T']) else "",
        f"{fw_mean['R2_T']:.4f}" if np.isfinite(fw_mean['R2_T']) else "",
        f"{fw_mean['MAE_H']:.4f}" if np.isfinite(fw_mean['MAE_H']) else "",
        f"{fw_mean['RMSE_H']:.4f}" if np.isfinite(fw_mean['RMSE_H']) else "",
        f"{fw_mean['R2_H']:.4f}" if np.isfinite(fw_mean['R2_H']) else "",
        f"{sparsity_before * 100:.2f}",
        f"{sparsity_after * 100:.2f}",
        size_str,
        f"{mean_train_loss:.4f}",
        f"{mean_val_loss:.4f}",
        f"{gap:.4f}",
        f"{gap_pct:.2f}",
        f"{inference_time_total:.2f}",
        f"{inference_time_per_sample:.2f}",
        "",
        "",
    ],
    "Status": [
        "",  # MSE (normalized residual)
        "",  # RMSE (normalized residual)
        "",  # MAE (normalized residual)
        "",  # R² (normalized residual)

        mse_status,   # MSE (joint original)
        rmse_status,  # RMSE (joint original)
        mae_status,   # MAE (joint original)
        r2_status,    # R² (joint original)

        "",  # MSE_T
        "",  # RMSE_T
        "",  # MAE_T
        "",  # R²_T
        "",  # MSE_H
        "",  # RMSE_H
        "",  # MAE_H
        "",  # R²_H

        # Rolling(24) is primarily comparison-oriented; fields intentionally left blank
        "", "", "", "", "", "", "", "", "", "",
        "", "", "", "", "", "", "", "", "",

        "",  # Sparsity before strip (%)
        "",  # Sparsity after strip (%)
        "",  # Model size (KB)
        "",  # Mean training loss
        "",  # Mean validation loss
        "",  # Absolute gap
        "",  # Percentage gap (%)
        "",  # Total inference time (ms)
        "",  # Inference time per sample (ms)
        status,         # Generalization assessment
        model_status,  # Overall model status
    ],

    "Meaning": [
        "Mean squared error in the normalized residual scale (ΔT_in, ΔH_in).",
        "Root mean squared error in the normalized residual scale.",
        "Mean absolute error in the normalized residual scale.",
        "Coefficient of determination in the normalized residual scale.",
        "Joint mean squared error in the original scale (T_in + H_in).",
        "Joint root mean squared error in the original scale.",
        "Joint mean absolute error in the original scale.",
        "Joint coefficient of determination in the original scale.",
        "Mean squared error for indoor temperature only (T_in).",
        "Root mean squared error for indoor temperature.",
        "Mean absolute error for indoor temperature.",
        "Coefficient of determination for indoor temperature.",
        "Mean squared error for indoor humidity only (H_in).",
        "Root mean squared error for indoor humidity.",
        "Mean absolute error for indoor humidity.",
        "Coefficient of determination for indoor humidity.",

        "Number of samples considered in the last N=24 window (must reach 24 after warm-up).",
        "Aggregated MAE (mean of MAE_T and MAE_H) in the last N=24 window, consistent with firmware semantics.",
        "Aggregated RMSE in the last N=24 window, consistent with firmware semantics.",
        "Aggregated R² (mean of R²_T and R²_H) in the last N=24 window; may be NaN under low-variance conditions.",
        "T_in MAE in the last N=24 window, consistent with firmware semantics.",
        "T_in RMSE in the last N=24 window, consistent with firmware semantics.",
        "T_in R² in the last N=24 window; may be NaN under low-variance conditions.",
        "H_in MAE in the last N=24 window, consistent with firmware semantics.",
        "H_in RMSE in the last N=24 window, consistent with firmware semantics.",
        "H_in R² in the last N=24 window; may be NaN under low-variance conditions.",
        "Mean aggregated MAE across the test set (mean over N=24 windows).",
        "Mean aggregated RMSE across the test set (mean over N=24 windows).",
        "Mean aggregated R² across the test set (mean over N=24 windows).",
        "Mean MAE_T across the test set (mean over N=24 windows).",
        "Mean RMSE_T across the test set (mean over N=24 windows).",
        "Mean R²_T across the test set (mean over N=24 windows).",
        "Mean MAE_H across the test set (mean over N=24 windows).",
        "Mean RMSE_H across the test set (mean over N=24 windows).",
        "Mean R²_H across the test set (mean over N=24 windows).",
        "Percentage of weights equal to zero before strip_pruning (active pruning stage).",
        "Percentage of weights equal to zero after strip_pruning (final deployment model).",
        "Final size of the pruned model file in kilobytes.",
        "Mean loss over the last training epochs (MSE).",
        "Mean loss over the last validation epochs (MSE).",
        "Absolute difference between mean validation and training losses.",
        "Percentage gap between validation and training losses.",
        "Total inference time on the test set.",
        "Mean inference time per sample.",
        "Automatic generalization diagnosis based on training and validation behavior.",
        "Overall model diagnosis based on the main decision thresholds.",
    ],
    "Expected Values / Thresholds": [
        "→ Lower is better (e.g., ≤ 0.001 is excellent).",
        "→ Lower is better (ideal < 0.05).",
        "→ Lower is better (ideal < 0.05).",
        "→ Ideally > 0.95.",
        "→ < 0.1 is considered excellent, depending on the variable.",
        "→ Ideal < 0.32.",
        "→ Ideal < 0.3.",
        "→ Desirable ≥ 0.8.",
        "→ Lower is better; compare with joint MSE.",
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Ideally close to 1.",
        "→ Lower is better; compare with T_in metrics.",
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Ideally close to 1.",
        "→ High sparsity indicates good compression provided that predictive metrics are preserved.",
        "→ High post-strip sparsity indicates a genuinely sparse deployment model.",
        "→ Ideal < 256 KB for TinyML; < 100 KB is excellent.",
        "→ Ideal < 0.01.",
        "→ Should remain close to training loss to avoid overfitting.",
        "→ < 0.05 is good; > 0.1 may indicate overfitting.",
        "→ < 10% is excellent; > 30% may indicate possible overfitting.",
        "→ Lower is better; ideal < 250 ms total for TinyML scenarios.",
        "→ Ideal < 1 ms per sample; < 10 ms is acceptable for many applications.",
        "→ Interpreted from the loss-behavior diagnostic text.",
        "→ Interpreted from the main evaluation metrics (MSE, RMSE, MAE, R²).",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ],
}

df_metrics = pd.DataFrame(metrics_dist)
csv_path = metrics_run_dir / "environment_pruned_model_metrics_summary_mlp.csv"
xlsx_path = metrics_run_dir / "environment_pruned_model_metrics_summary_mlp.xlsx"

df_metrics.to_csv(csv_path, index=False, encoding="utf-8-sig")
print(f"[INFO] File saved: {csv_path.name}")

df_metrics.to_excel(xlsx_path, index=False)
wb = load_workbook(xlsx_path)
ws = wb.active
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
wb.save(xlsx_path)
print(f"[INFO] File saved: {xlsx_path.name} (with adjusted columns)")

# === Post-execution: update 'latest' and manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Unable to update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Unable to write manifest.json:", _e)

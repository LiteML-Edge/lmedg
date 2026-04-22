"""
Script: environment_base_model_lstm.py
Module role:
    Train and evaluate the baseline LSTM model adopted in the LiteML-Edge
    environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, trains the baseline residual model, evaluates the
    model in normalized and reconstructed physical domains, and exports metrics,
    figures, and versioned artifacts.

Inputs:
    - environment_dataset_lstm.csv
    - Project path and versioning utilities from utils.global_utils.paths_lstm and
      utils.global_utils.versioning.

Outputs:
    - Versioned baseline model artifacts
    - Fitted X and y scalers
    - Tables, figures, and summary spreadsheets for offline evaluation

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys 
from pathlib import Path
import time
import numpy as np
import pandas as pd

os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"

import tensorflow as tf
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

rcParams['font.family'] = 'Segoe UI Emoji'

# --- bootstrap: allow importing utils/ locally and in the runner ---
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)
# -----------------------------------------------------------------
from utils.global_utils.paths_lstm import PROJECT_ROOT, DATASET_ENVIRONMENT_LSTM, BASE_MODEL, BASE_MODEL_METRICS
from utils.global_utils.versioning import create_versioned_dir, ensure_dir, update_latest, write_manifest
from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Call before model creation

# === Versioned directories per execution ===
run_dir = create_versioned_dir(BASE_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(BASE_MODEL_METRICS / run_dir.name)
# ===========================================

# =========================
# Load data and features
# =========================
dataset_path = DATASET_ENVIRONMENT_LSTM / "environment_dataset_lstm.csv"
df = pd.read_csv(dataset_path)

# Check expected columns
cols_req = {"datetime", "T_out", "T_in", "H_out", "H_in"}
missing = cols_req - set(df.columns)
if missing:
    raise ValueError(
        f"Missing columns in the dataset: {missing}. "
        f"Expected: {sorted(cols_req)}"
    )

# Sort temporally to construct sliding windows
df = df.sort_values("datetime").reset_index(drop=True)

# Temporal feature engineering
df["datetime"] = pd.to_datetime(df["datetime"])
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (same idea as Conv1D Tiny) ===
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=0.08, adjust=False).mean()

# Cyclical features
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lag features (first!)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Add lag2 only for indoor variables to keep 12 + weekday/month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)

# Separate features (12 features) 
features = [
    "T_out", "H_out",          # outdoor atual
    "T_in_lag1", "H_in_lag1",  # indoor lag1
    "T_out_lag1", "H_out_lag1",# outdoor lag1
    "T_in_lag2", "H_in_lag2",  # indoor lag2
    "sin_hour", "cos_hour",    # cyclical hour
    "weekday", "month",        # calendar features
]

# Internal residual target: ΔT_in, ΔH_in
y_all = np.stack([
    (df['T_in'] - df['T_in_lag1']).values,
    (df['H_in'] - df['H_in_lag1']).values
], axis=1).astype(np.float32)

# LiteML-Edge contract: do not apply clamp/clip to the residual Δ.
# Δ must remain exactly abs − lag1 (without clipping) for 1:1 compatibility with the firmware.
y_all = y_all.astype(np.float32, copy=False)

# Temporal sliding window (24 steps)
WINDOW_STEPS = 24  # 24 horas
X_source = df[features].values.astype(np.float32)
N_total = len(df)
if N_total < WINDOW_STEPS:
    raise ValueError("Dataset too short for the 24-step sliding window.")

X_seq = []
y_seq = []
idx_seq = []

for t in range(WINDOW_STEPS - 1, N_total):
    X_seq.append(X_source[t-WINDOW_STEPS+1 : t+1, :])
    y_seq.append(y_all[t])
    idx_seq.append(t)

X_all = np.stack(X_seq, axis=0).astype(np.float32)        # (N_seq, WINDOW_STEPS, n_features)
y_all_seq = np.stack(y_seq, axis=0).astype(np.float32)    # (N_seq, 2)
idx_seq = np.array(idx_seq, dtype=np.int64)
N_seq, SEQ_LEN, K_NUM_FEATURES = X_all.shape

# Temporal 60/20/20 split over the windows (same logic as Conv1D Tiny)
i1, i2 = int(0.6 * N_seq), int(0.8 * N_seq)
X_train, X_val, X_test = X_all[:i1], X_all[i1:i2], X_all[i2:]
y_train, y_val, y_test = y_all_seq[:i1], y_all_seq[i1:i2], y_all_seq[i2:]

# Keep non-normalized references for absolute reconstruction on the test split
T_prev_all = df['T_in_lag1'].values.astype(np.float32)
H_prev_all = df['H_in_lag1'].values.astype(np.float32)
y_abs_all  = df[['T_in', 'H_in']].values.astype(np.float32)

T_prev_seq = T_prev_all[idx_seq]
H_prev_seq = H_prev_all[idx_seq]
y_abs_seq  = y_abs_all[idx_seq]

T_prev_test = T_prev_seq[i2:]
H_prev_test = H_prev_seq[i2:]
y_test_abs  = y_abs_seq[i2:]  # absolute ground truth (T_in, H_in)

# TEST timestamps (to align with EVENT HOUR in firmware)
dt_seq = df['datetime'].values[idx_seq]
dt_test = dt_seq[i2:]

# =========================
# Normalization (fit on training split)
# =========================
scaler_X = MinMaxScaler()
scaler_y = MinMaxScaler()

# X in 3D: flatten the temporal axis to normalize per feature
N_train = X_train.shape[0]
F = K_NUM_FEATURES
X_train_2d = X_train.reshape(-1, F)
scaler_X.fit(X_train_2d)
# LiteML-Edge contract: clamp X in the physical domain to keep minmax_forward in [0,1]
X_train_2d = np.clip(X_train_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_train = scaler_X.transform(X_train_2d).reshape(N_train, WINDOW_STEPS, F)

N_val = X_val.shape[0]
X_val_2d = X_val.reshape(-1, F)
X_val_2d = np.clip(X_val_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_val = scaler_X.transform(X_val_2d).reshape(N_val, WINDOW_STEPS, F)
N_test = X_test.shape[0]
X_test_2d = X_test.reshape(-1, F)
X_test_2d = np.clip(X_test_2d, scaler_X.data_min_, scaler_X.data_max_).astype(np.float32)
X_test = scaler_X.transform(X_test_2d).reshape(N_test, WINDOW_STEPS, F)
# y: residual targets ΔT_in and ΔH_in
y_train = scaler_y.fit_transform(y_train)
y_val   = scaler_y.transform(y_val)
y_test  = scaler_y.transform(y_test)

# =========================
# LSTM model (3D input, residual output ΔT_in, ΔH_in)
# =========================
model = tf.keras.Sequential([
    tf.keras.layers.Input(shape=(SEQ_LEN, K_NUM_FEATURES)),
    tf.keras.layers.LSTM(24, return_sequences=False),
    tf.keras.layers.Dense(16, activation='relu'),
    tf.keras.layers.Dense(2)
])

model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-4, clipnorm=1.0),
    loss=tf.keras.losses.MeanSquaredError(),
    metrics=[tf.keras.metrics.MeanAbsoluteError()]
)

callbacks = [
    tf.keras.callbacks.ReduceLROnPlateau(
        monitor="val_loss", factor=0.5, patience=8, min_lr=1e-6, verbose=1
    ),
    tf.keras.callbacks.EarlyStopping(
        monitor="val_loss", patience=10, restore_best_weights=True, verbose=1
    )
]

history = model.fit(
    X_train, y_train,
    validation_data=(X_val, y_val),
    epochs=50,
    batch_size=220,
    callbacks=callbacks,
    verbose=1
)

# =========================
# Save model and scalers
# =========================
model.save(str(run_dir / 'environment_base_model_lstm.keras'))
joblib.dump(scaler_X, run_dir / "environment_base_model_lstm_scaler_X.pkl")
joblib.dump(scaler_y, run_dir / "environment_base_model_lstm_scaler_y.pkl")

# =========================
# Inference and latency calculation
# =========================
start_time = time.time()
_ = model.predict(X_test, verbose=0)
end_time = time.time()
inference_time_total = (end_time - start_time) * 1000  # ms
inference_time_per_sample = (inference_time_total / len(X_test)) if len(X_test) > 0 else float('nan')

print(f"Total inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.2f} ms")

# =========================
# Model size
# =========================
saved_model_path = run_dir / "environment_base_model_lstm.keras"
if os.path.isfile(saved_model_path):
    model_size_kb = os.path.getsize(saved_model_path) / 1024
    print(f"Trained model size: {model_size_kb:.2f} KB ")
else:
    print("Model file not found.")
    model_size_kb = float('nan')

# =========================
# Evaluation: normalized scale (residual target space)
# =========================
y_pred_test_scaled = model.predict(X_test, verbose=0)
mse_scaled = mean_squared_error(y_test, y_pred_test_scaled)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled = mean_absolute_error(y_test, y_pred_test_scaled)
r2_scaled  = r2_score(y_test, y_pred_test_scaled)

print("\n Results (normalized scale - residual training target):")
print(f"MSE  = {mse_scaled:.4f}")
print(f"RMSE = {rmse_scaled:.4f}")
print(f"MAE  = {mae_scaled:.4f}")
print(f"R²   = {r2_scaled:.4f}")

# =========================
# Evaluation: original absolute scale (T_in, H_in)
# =========================

# Denormalize residuals and ensure float32
y_pred_delta = scaler_y.inverse_transform(y_pred_test_scaled).astype(np.float32, copy=False)

# Ensure the previous values (lag1) are also float32
T_prev_test = T_prev_test.astype(np.float32, copy=False)
H_prev_test = H_prev_test.astype(np.float32, copy=False)

# Absolute reconstruction
T_pred = (T_prev_test + y_pred_delta[:, 0]).astype(np.float32, copy=False)
H_pred = (H_prev_test + y_pred_delta[:, 1]).astype(np.float32, copy=False)

# Final float32 vector
y_pred_orig = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)

# Ground truth also in float32
y_test_orig = y_test_abs.astype(np.float32, copy=False)

mse = mean_squared_error(y_test_orig, y_pred_orig)
rmse = np.sqrt(mse)
mae = mean_absolute_error(y_test_orig, y_pred_orig)
r2  = r2_score(y_test_orig, y_pred_orig)

print("\n Results (original scale - joint target set [T_in, H_in]):")
mse_status = "MSE within threshold" if mse <= 0.1 else "MSE above threshold"
rmse_status = "RMSE within threshold" if rmse <= 0.32 else "RMSE above threshold"
mae_status = "MAE within threshold" if mae <= 0.3 else "MAE above threshold"
r2_status = "R² within threshold" if r2 >= 0.8 else "R² below threshold"

print(f"MSE  = {mse:.4f}   {mse_status}")
print(f"RMSE = {rmse:.4f}   {rmse_status}")
print(f"MAE  = {mae:.4f}   {mae_status}")
print(f"R²   = {r2:.4f}   {r2_status}")
# >>> Per-variable metrics (original absolute scale)
mse_T  = mean_squared_error(y_test_orig[:, 0], y_pred_orig[:, 0])
rmse_T = np.sqrt(mse_T)
mae_T  = mean_absolute_error(y_test_orig[:, 0], y_pred_orig[:, 0])
r2_T   = r2_score(y_test_orig[:, 0], y_pred_orig[:, 0])

mse_H  = mean_squared_error(y_test_orig[:, 1], y_pred_orig[:, 1])
rmse_H = np.sqrt(mse_H)
mae_H  = mean_absolute_error(y_test_orig[:, 1], y_pred_orig[:, 1])
r2_H   = r2_score(y_test_orig[:, 1], y_pred_orig[:, 1])

print("\n Results (by variable - original scale):")
print(f"[T_in] MSE={mse_T:.4f} RMSE={rmse_T:.4f} MAE={mae_T:.4f} R²={r2_T:.4f}")
print(f"[H_in] MSE={mse_H:.4f} RMSE={rmse_H:.4f} MAE={mae_H:.4f} R²={r2_H:.4f}")

# Overall model status
model_ok = all([mse <= 0.1, rmse <= 0.32, mae <= 0.3, r2 >= 0.8])
model_status = "Performance thresholds satisfied" if model_ok else "Performance thresholds not satisfied"
print("\n Overall assessment:", model_status)

# ============================================================
# NEW: Rolling-window metrics with N=24 (firmware-equivalent)
# ============================================================
ROLLING_N = 24  # N=24 samples (HOUR) as in metrics.cpp

def _r2_like_firmware(y_true_1d: np.ndarray, y_pred_1d: np.ndarray) -> float:
    """Replicate the firmware logic (metrics.cpp):
      R² = 1 - SS_res/SS_tot
      SS_tot = sum(y^2) - (sum(y)^2)/n
      Se n < 2 -> NaN
      Se SS_tot <= 1e-6 -> NaN
    """
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return float("nan")

    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))

    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    if ss_tot <= 1e-6:
        return float("nan")

    return 1.0 - (ss_res / ss_tot)


def _metrics_like_firmware(y_true_2d: np.ndarray, y_pred_2d: np.ndarray):
    """
    Returns (mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H)
    seguindo exatamente:
      mae = 0.5*(mae_T + mae_H)
      rmse = sqrt( 0.5*(mse_T + mse_H) )
      r2 = 0.5*(r2_T + r2_H)
    """
    y_true_2d = np.asarray(y_true_2d, dtype=np.float32)
    y_pred_2d = np.asarray(y_pred_2d, dtype=np.float32)

    # T
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = float(np.mean(np.abs(err_T)))
    mse_T = float(np.mean(err_T * err_T))
    rmse_T = float(np.sqrt(mse_T))
    r2_T = float(_r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0]))

    # H
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = float(np.mean(np.abs(err_H)))
    mse_H = float(np.mean(err_H * err_H))
    rmse_H = float(np.sqrt(mse_H))
    r2_H = float(_r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1]))

    mae = 0.5 * (mae_T + mae_H)
    rmse = float(np.sqrt(0.5 * (mse_T + mse_H)))
    r2 = 0.5 * (r2_T + r2_H)
    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

rolling_rows = []
n_test = y_test_orig.shape[0]

# Firmware-equivalent gating (offline: all samples count)
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)

if n_test < ROLLING_N:
    print(f"\n Rolling N=24 cannot be computed: test set contains only {n_test} samples.")
else:
    for end in range(ROLLING_N - 1, n_test):
        # Firmware: update metrics only when a real Invoke occurred and EVENT=HOUR
        if (not invoked_mask[end]) or (not is_rollover_mask[end]):
            rolling_rows.append({
                "window_start": int(end - ROLLING_N + 1),
                "window_end": int(end),
                "datetime_end": pd.to_datetime(dt_test[end]),
                "N": int(ROLLING_N),

                "MAE": float("nan"),
                "RMSE": float("nan"),
                "R2": float("nan"),

                "MAE_T": float("nan"),
                "RMSE_T": float("nan"),
                "R2_T": float("nan"),

                "MAE_H": float("nan"),
                "RMSE_H": float("nan"),
                "R2_H": float("nan"),
            })
            continue
        start = end - ROLLING_N + 1
        yt_w = y_test_orig[start:end+1, :]
        yp_w = y_pred_orig[start:end+1, :]

        mae_w, rmse_w, r2_w, maeT_w, rmseT_w, r2T_w, maeH_w, rmseH_w, r2H_w = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": pd.to_datetime(dt_test[end]),
            "N": int(ROLLING_N),

            "MAE": mae_w,
            "RMSE": rmse_w,
            "R2": r2_w,

            "MAE_T": maeT_w,
            "RMSE_T": rmseT_w,
            "R2_T": r2T_w,

            "MAE_H": maeH_w,
            "RMSE_H": rmseH_w,
            "R2_H": r2H_w,
        })

    df_roll = pd.DataFrame(rolling_rows)

    # Useful summaries for comparison with the firmware current state
    last = df_roll.iloc[-1]
    mean_roll = df_roll[["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]].mean()

    print("\n[INFO] Rolling N=24 (firmware-equivalent), last window (current state):")
    print(f"MAE={last['MAE']:.4f} RMSE={last['RMSE']:.4f} R²={last['R2']:.4f} | "
          f"T: MAE={last['MAE_T']:.4f} RMSE={last['RMSE_T']:.4f} R²={last['R2_T']:.4f} | "
          f"H: MAE={last['MAE_H']:.4f} RMSE={last['RMSE_H']:.4f} R²={last['R2_H']:.4f}")

    print("\n[INFO] Rolling N=24 mean over all test windows:")
    print(f"MAE={mean_roll['MAE']:.4f} RMSE={mean_roll['RMSE']:.4f} R²={mean_roll['R2']:.4f} | "
          f"T: MAE={mean_roll['MAE_T']:.4f} RMSE={mean_roll['RMSE_T']:.4f} R²={mean_roll['R2_T']:.4f} | "
          f"H: MAE={mean_roll['MAE_H']:.4f} RMSE={mean_roll['RMSE_H']:.4f} R²={mean_roll['R2_H']:.4f}")

    # Save rolling metrics to CSV/Excel (for 1:1 comparison with firmware logs)
    df_roll = df_roll.round(4)
    df_roll.to_csv(metrics_run_dir / "environment_base_model_lstm_metrics_rolling24.csv",
                   index=False, encoding="utf-8-sig")
    excel_roll_path = metrics_run_dir / "environment_base_model_lstm_metrics_rolling24.xlsx"
    df_roll.to_excel(excel_roll_path, index=False)

    # Auto-adjust column widths in Excel
    wb2 = load_workbook(excel_roll_path)
    ws2 = wb2.active
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb2.save(excel_roll_path)

    print("[INFO] Files saved:")
    print(" - environment_base_model_lstm_metrics_rolling24.csv")
    print(" - environment_base_model_lstm_metrics_rolling24.xlsx")

    # Plot: rolling (MAE/RMSE/R2)
    plt.figure(figsize=(10, 5))
    plt.plot(df_roll["datetime_end"], df_roll["MAE"], label="MAE (rolling24)")
    plt.plot(df_roll["datetime_end"], df_roll["RMSE"], label="RMSE (rolling24)")
    plt.title("Rolling-Window Performance (N=24)  -  Firmware-Comparable")
    plt.xlabel("Time (window end)")
    plt.ylabel("Value")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_base_model_lstm_metrics_rolling24_mae_rmse.png")
    plt.close()

    plt.figure(figsize=(10, 4))
    plt.plot(df_roll["datetime_end"], df_roll["R2"], label="R² (rolling24)")
    plt.title("Rolling-Window Coefficient of Determination (N=24)  -  Firmware-Comparable")
    plt.xlabel("Time (window end)")
    plt.ylabel("R²")
    plt.ylim([-0.2, 1.05])
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(metrics_run_dir / "environment_base_model_lstm_metrics_rolling24_r2.png")
    plt.close()

# =========================
# Generalization diagnosis
# =========================
train_loss = history.history['loss']
val_loss   = history.history['val_loss']
n = min(5, len(train_loss))
mean_train_loss = np.mean(train_loss[-n:])
mean_val_loss   = np.mean(val_loss[-n:])
gap = abs(mean_val_loss - mean_train_loss)
gap_pct = (gap / mean_train_loss) * 100 if mean_train_loss > 0 else 0

if mean_train_loss > 0.3 and mean_val_loss > 0.3:
    status = "Underfitting detected (high training and validation losses)"
elif mean_val_loss < mean_train_loss * 0.8:
    status = "Potential underfitting (validation loss significantly lower than training loss)"
elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
    status = "Overfitting detected (large generalization gap or significant divergence)"
elif gap_pct < 10:
    status = "Well-fitted model (generalization gap < 10%)"
elif gap_pct < 30:
    status = "Acceptably fitted model (generalization gap < 30%)"
else:
    status = "Mild overfitting (moderate generalization gap)"

print("\n Model diagnostics:")
print(f"• Mean training loss:     {mean_train_loss:.4f}")
print(f"• Mean validation loss:   {mean_val_loss:.4f}")
print(f"• Absolute gap:           {gap:.4f}")
print(f"• Generalization gap:     {gap_pct:.2f}%")
print(f"• Status:                 {status}")

# =========================
# Plots / Reports
# =========================
plt.figure(figsize=(8, 5))
plt.plot(train_loss, label='Training')
plt.plot(val_loss, label='Validation')
plt.title("Training Dynamics with Early Stopping\n" + status)
plt.xlabel("Epoch")
plt.ylabel("MSE")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_base_model_lstm_training_validation_loss_diagnosis.png")
plt.close()

# Metrics bar chart  -  original absolute scale
metric_labels = ['MSE', 'RMSE', 'MAE', 'R²']
metric_values = [mse, rmse, mae, r2]
plt.figure(figsize=(8, 5))
bars = plt.bar(metric_labels, metric_values, color='skyblue')
plt.title(f"Evaluation metrics - {'thresholds satisfied' if model_ok else 'thresholds not satisfied'}")
plt.ylabel("Value")
plt.ylim([0, max(metric_values) * 1.2 if np.isfinite(max(metric_values)) else 1.0])
for bar in bars:
    yval = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2,
        yval + 0.01,
        f'{yval:.4f}',
        ha='center',
        va='bottom'
    )
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_base_model_lstm_final_metrics_summary_plot.png")
plt.close()

# Prediction scatter  -  original absolute scale
plt.figure(figsize=(6, 6))
plt.scatter(
    y_test_orig[:, 0], y_pred_orig[:, 0],
    alpha=0.5, label="Temperature (T_in)", color="blue"
)
plt.scatter(
    y_test_orig[:, 1], y_pred_orig[:, 1],
    alpha=0.5, label="Humidity (H_in)", color="green"
)
min_val = min(y_test_orig.min(), y_pred_orig.min())
max_val = max(y_test_orig.max(), y_pred_orig.max())
plt.plot([min_val, max_val], [min_val, max_val], 'k--')
plt.xlabel("Ground-truth value")
plt.ylabel("Predicted value")
plt.title("Prediction scatter (original scale)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_base_model_lstm_scatter_predictions.png")
plt.close()

# Per-variable scatter plots (T_in and H_in separated)

# --- Temperatura ---
plt.figure(figsize=(6, 6))
plt.scatter(y_test_orig[:, 0], y_pred_orig[:, 0], alpha=0.6, label="Predicted T_in")
min_val_T = float(min(y_test_orig[:, 0].min(), y_pred_orig[:, 0].min()))
max_val_T = float(max(y_test_orig[:, 0].max(), y_pred_orig[:, 0].max()))
plt.plot([min_val_T, max_val_T], [min_val_T, max_val_T], 'k--', label="Linha 1:1")
plt.xlabel("Ground truth T_in (°C)")
plt.ylabel("Predicted T_in (°C)")
plt.title("Predicted vs. Ground Truth  -  Temperature (T_in)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_base_model_lstm_scatter_T_in.png")
plt.close()

# --- Umidade ---
plt.figure(figsize=(6, 6))
plt.scatter(y_test_orig[:, 1], y_pred_orig[:, 1], alpha=0.6, label="Predicted H_in")
min_val_H = float(min(y_test_orig[:, 1].min(), y_pred_orig[:, 1].min()))
max_val_H = float(max(y_test_orig[:, 1].max(), y_pred_orig[:, 1].max()))
plt.plot([min_val_H, max_val_H], [min_val_H, max_val_H], 'k--', label="Linha 1:1")
plt.xlabel("Ground truth H_in (%)")
plt.ylabel("Predicted H_in (%)")
plt.title("Predicted vs. Ground Truth  -  Humidity (H_in)")
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.savefig(metrics_run_dir / "environment_base_model_lstm_scatter_H_in.png")
plt.close()

# =========================
# Rolling(24) firmware-style summary to store in metrics_summary
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols}
        fw_mean = {k: float(df_roll[_cols].mean()[k]) for k in _cols}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)

metrics_dist = {
    "Baseline Model Metrics": [
        # Normalized (joint T_in + H_in)
        "MSE (normalized joint residual)", "RMSE (normalized joint residual)", "MAE (normalized joint residual)", "R² (normalized joint residual)",
        # Joint original scale
        "MSE (joint original scale)", "RMSE (joint original scale)", "MAE (joint original scale)", "R² (joint original scale)",
        # Per-variable T_in
        "MSE_T (T_in)", "RMSE_T (T_in)", "MAE_T (T_in)", "R²_T (T_in)",
        # Per-variable H_in
        "MSE_H (H_in)", "RMSE_H (H_in)", "MAE_H (H_in)", "R²_H (H_in)",
        # Rolling(24) firmware-equivalent (window=24, HOUR+Invoke gating)
        "N_ALL (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, last window)", "RMSE (rolling24 firmware, last window)", "R² (rolling24 firmware, last window)",
        "MAE_T (rolling24 firmware, last window)", "RMSE_T (rolling24 firmware, last window)", "R²_T (rolling24 firmware, last window)",
        "MAE_H (rolling24 firmware, last window)", "RMSE_H (rolling24 firmware, last window)", "R²_H (rolling24 firmware, last window)",
        "MAE (rolling24 firmware, test mean)", "RMSE (rolling24 firmware, test mean)", "R² (rolling24 firmware, test mean)",
        "MAE_T (rolling24 firmware, test mean)", "RMSE_T (rolling24 firmware, test mean)", "R²_T (rolling24 firmware, test mean)",
        "MAE_H (rolling24 firmware, test mean)", "RMSE_H (rolling24 firmware, test mean)", "R²_H (rolling24 firmware, test mean)",
        # Sparsity / size
        "Model size (KB)",
        # Loss / gap
        "Mean training loss", "Mean validation loss",
        "Absolute gap", "Percentage gap (%)",
        # Timing
        "Total inference time (ms)", "Inference time per sample (ms)",
        # Status
        "Fit Status", "Model Status"
    ],
    "Value": [
        f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
        f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
        f"{mse_T:.4f}",      f"{rmse_T:.4f}",      f"{mae_T:.4f}",      f"{r2_T:.4f}",
        f"{mse_H:.4f}",      f"{rmse_H:.4f}",      f"{mae_H:.4f}",      f"{r2_H:.4f}",
        f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else "",
        f"{fw_last['MAE']:.4f}" if np.isfinite(fw_last['MAE']) else "",
        f"{fw_last['RMSE']:.4f}" if np.isfinite(fw_last['RMSE']) else "",
        f"{fw_last['R2']:.4f}" if np.isfinite(fw_last['R2']) else "",
        f"{fw_last['MAE_T']:.4f}" if np.isfinite(fw_last['MAE_T']) else "",
        f"{fw_last['RMSE_T']:.4f}" if np.isfinite(fw_last['RMSE_T']) else "",
        f"{fw_last['R2_T']:.4f}" if np.isfinite(fw_last['R2_T']) else "",
        f"{fw_last['MAE_H']:.4f}" if np.isfinite(fw_last['MAE_H']) else "",
        f"{fw_last['RMSE_H']:.4f}" if np.isfinite(fw_last['RMSE_H']) else "",
        f"{fw_last['R2_H']:.4f}" if np.isfinite(fw_last['R2_H']) else "",
        f"{fw_mean['MAE']:.4f}" if np.isfinite(fw_mean['MAE']) else "",
        f"{fw_mean['RMSE']:.4f}" if np.isfinite(fw_mean['RMSE']) else "",
        f"{fw_mean['R2']:.4f}" if np.isfinite(fw_mean['R2']) else "",
        f"{fw_mean['MAE_T']:.4f}" if np.isfinite(fw_mean['MAE_T']) else "",
        f"{fw_mean['RMSE_T']:.4f}" if np.isfinite(fw_mean['RMSE_T']) else "",
        f"{fw_mean['R2_T']:.4f}" if np.isfinite(fw_mean['R2_T']) else "",
        f"{fw_mean['MAE_H']:.4f}" if np.isfinite(fw_mean['MAE_H']) else "",
        f"{fw_mean['RMSE_H']:.4f}" if np.isfinite(fw_mean['RMSE_H']) else "",
        f"{fw_mean['R2_H']:.4f}" if np.isfinite(fw_mean['R2_H']) else "",
        f"{model_size_kb:.2f} KB",
        f"{mean_train_loss:.4f}", f"{mean_val_loss:.4f}",
        f"{gap:.4f}",            f"{gap_pct:.2f} %",
        f"{inference_time_total:.2f} ms", f"{inference_time_per_sample:.2f} ms",
        "", ""
    ],
    "Status": [
        # 0-3: normalized metrics → no status
        "", "", "", "",
        # 4-7: original-scale joint metrics → use mse_status, rmse_status, ...
        mse_status, rmse_status, mae_status, r2_status,
        # 8-15: per-variable metrics (T_in, H_in) → kept empty
        "", "", "", "", "", "", "", "",
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",  # rolling24 firmware (last window + mean)
        # 16-18: sparsity / size → empty
        "",
        # 19-22: losses / gaps → empty
        "", "", "", "",
        # 23-24: inference timing → empty
        "", "",
        # 25-26: rows “Fit Status” and “Model Status”
        status, model_status
    ],
    "Significado": [
        "Mean squared error in the normalized scale for the joint residual target (ΔT_in, ΔH_in).",
        "Root mean squared error in the normalized scale.",
        "Mean absolute error in the normalized scale.",
        "Coefficient of determination in the normalized scale.",
        "Mean squared error in the original scale (°C / %RH) for the joint [T_in, H_in] target.",
        "Root mean squared error in the original scale for the joint target.",
        "Mean absolute error in the original scale for the joint target.",
        "Coefficient of determination in the original scale for the joint target.",
        "Mean squared error in the original scale for T_in.",
        "Root mean squared error in the original scale for T_in.",
        "Mean absolute error in the original scale for T_in.",
        "Coefficient of determination in the original scale for T_in.",
        "Mean squared error in the original scale for H_in.",
        "Root mean squared error in the original scale for H_in.",
        "Mean absolute error in the original scale for H_in.",
        "Coefficient of determination in the original scale for H_in.",
        "Rolling(24): number of samples (HOUR events) effectively present in the final window.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) T_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) H_in in the LAST window  -  comparable to the firmware log.",
        "Rolling(24) aggregate (T+H) MEAN across the test set (window average).",
        "Rolling(24) aggregate (T+H) MEAN across the test set (window average).",
        "Rolling(24) aggregate (T+H) MEAN across the test set (window average).",
        "Rolling(24) T_in MEAN across the test set (window average).",
        "Rolling(24) T_in MEAN across the test set (window average).",
        "Rolling(24) T_in MEAN across the test set (window average).",
        "Rolling(24) H_in MEAN across the test set (window average).",
        "Rolling(24) H_in MEAN across the test set (window average).",
        "Rolling(24) H_in MEAN across the test set (window average).",
        "Final model file size in kilobytes (KB).",
        "Mean loss over the last training epochs.",
        "Mean loss over the last validation epochs.",
        "Absolute difference between mean losses.",
        "Percentage gap between losses (validation vs training).",
        "Total time to infer all test samples.",
        "Mean time to infer a single sample.",
        "Fit diagnosis based on losses and generalization gap.",
        "Overall diagnosis based on the predefined MSE/RMSE/MAE/R² thresholds."
    ],
    "Expected Values / Thresholds": [
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Lower is better.",
        "→ Ideally > 0.95.",
        "→ < 0.1 is excellent, depending on the problem.",
        "→ < 0.32 as a reference.",
        "→ < 0.30 as a reference.",
        "→ > 0.8 is desirable.",
        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Lower is better (T_in).",
        "→ Ideally > 0.8 (T_in).",
        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Lower is better (H_in).",
        "→ Ideally > 0.8 (H_in).",
        "→ Should reach 24 when the window is full (after warm-up).",
        "→ Lower is better (rolling24).",
        "→ Lower is better (rolling24).",
        "→ May be NaN when variance is low (same as firmware).",
        "→ Lower is better (rolling24 T_in).",
        "→ Lower is better (rolling24 T_in).",
        "→ May be NaN when variance is low.",
        "→ Lower is better (rolling24 H_in).",
        "→ Lower is better (rolling24 H_in).",
        "→ May be NaN when variance is low.",
        "→ Lower is better (window mean).",
        "→ Lower is better (window mean).",
        "→ May be NaN when variance is low.",
        "→ Lower is better (T_in window mean).",
        "→ Lower is better (T_in window mean).",
        "→ May be NaN when variance is low.",
        "→ Lower is better (H_in window mean).",
        "→ Lower is better (H_in window mean).",
        "→ May be NaN when variance is low.",
        "→ Prefer < 256 KB on constrained MCUs.",
        "→ Low (for example, < 0.01).",
        "→ Close to the training loss.",
        "→ < 0.05 is good.",
        "→ < 10% is excellent.",
        "→ Lower is better.",
        "→ < 1 ms is ideal in TinyML.",
        "→ 'Well-fitted model' when the generalization gap is low and losses remain stable.",
        "→ 'Performance thresholds satisfied' when the model satisfies the predefined limits."
    ]
}

# Create DataFrame
# --- FIX: ensure that all columns have the same length (required by pandas) ---
_max_len = max(len(v) for v in metrics_dist.values())
for _k, _v in list(metrics_dist.items()):
    if len(_v) < _max_len:
        metrics_dist[_k] = list(_v) + [""] * (_max_len - len(_v))
# -------------------------------------------------------------------------------

dfm = pd.DataFrame(metrics_dist)

# Save as CSV
dfm.to_csv(
    metrics_run_dir / "environment_base_model_lstm_metrics_summary.csv",
    index=False,
    encoding="utf-8-sig"
)
print("[INFO] File saved: environment_base_model_lstm_metrics_summary.csv")

# Save as Excel with automatic column-width adjustment
excel_path = metrics_run_dir / "environment_base_model_lstm_metrics_summary.xlsx"
dfm.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
wb.save(excel_path)
print("[INFO] File saved: environment_base_model_lstm_metrics_summary.xlsx (with adjusted columns)")

# === Post-execution: update 'latest' and manifest ===
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Unable to update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Unable to write manifest.json:", _e)

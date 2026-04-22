"""
Script: environment_quantized_model_lstm.py
Module role:
    Train, convert, evaluate, and export the quantized LSTM artifacts used in
    the LiteML-Edge environment pipeline.

Technical summary:
    This script prepares the time-ordered dataset, applies the fixed
    preprocessing contract, performs quantization-aware training, converts the
    model to TensorFlow Lite, evaluates the resulting artifact in normalized
    and reconstructed physical domains, and exports reference tables, figures,
    and firmware-support artifacts.

Inputs:
    - environment_dataset_lstm.csv
    - Baseline and pruned model artifacts
    - Project path and versioning utilities

Outputs:
    - Versioned quantized TensorFlow Lite artifact
    - Rolling24 evaluation tables and reference spreadsheets
    - Debug workbooks, replay headers, and figure files

Notes:
    This script assumes the repository project structure and the referenced
    utility modules. The computational logic, numerical procedures, and
    execution flow are preserved.
"""
import os, sys, random
import struct
from pathlib import Path
import time
import numpy as np
import pandas as pd
import tensorflow as tf
from tensorflow import lite as tflite
import matplotlib.pyplot as plt
from matplotlib import rcParams
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import joblib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tensorflow_model_optimization as tfmot
from tensorflow_model_optimization.python.core.quantization.keras import quantize_config

# ========================= Global settings =========================
os.environ["TF_USE_LEGACY_KERAS"] = "False"
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "2")
rcParams['font.family'] = 'Segoe UI Emoji'

def set_global_seed(seed: int = 42):
    random.seed(seed); np.random.seed(seed); tf.random.set_seed(seed)
set_global_seed(42)

try:
    tf.get_logger().setLevel("ERROR")
    try:
        tf.autograph.set_verbosity(0, alsologtostderr=False)
    except Exception:
        pass
except Exception:
    pass

# ========================= Local bootstrap utilities =========================
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base); break
if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from utils.global_utils.paths_lstm import (
    PROJECT_ROOT, DATASET_ENVIRONMENT_LSTM, BASE_MODEL, PRUNED_MODEL, QUANTIZED_MODEL, QUANTIZED_MODEL_METRICS
)
from utils.global_utils.versioning import (
    create_versioned_dir, ensure_dir, update_latest, write_manifest, resolve_latest
)

from utils.global_utils.global_seed import set_global_seed

set_global_seed(42)  # Call at the top of the script, BEFORE model creation

# ========================= QuantizeConfig no-op =========================
class NoOpQuantizeConfig(quantize_config.QuantizeConfig):
    def get_weights_and_quantizers(self, layer): return []
    def get_activations_and_quantizers(self, layer): return []
    def set_quantize_weights(self, layer, quantize_weights): pass
    def set_quantize_activations(self, layer, quantize_activations): pass
    def get_output_quantizers(self, layer): return []
    def get_config(self): return {}

# ========================= Hyperparameters =========================
SEQ_LEN = 24
WINDOW = SEQ_LEN
USE_WEIGHTED_HUBER = True
H_WEIGHT = 1.4

EPOCHS = 300
BATCH_SIZE = 256
LR = 1e-5
PATIENCE_ES = 15
PATIENCE_RLR = 10
MIN_LR = 1e-6
REP_SAMPLES = 512

# ========================= Directories =========================
run_dir = create_versioned_dir(QUANTIZED_MODEL, strategy="counter")
metrics_run_dir = ensure_dir(QUANTIZED_MODEL_METRICS / run_dir.name)

# Output subfolders (inside metrics_run_dir) for organization (paper-ready)
graphics_dir = ensure_dir(metrics_run_dir / "quantization_graphics")
results_dir  = ensure_dir(metrics_run_dir / "quantization_metrics_results")

base_version_path   = resolve_latest(BASE_MODEL)
pruned_version_path = resolve_latest(PRUNED_MODEL)

# ========================= Paths =========================
model_path     = pruned_version_path / "environment_pruned_model_lstm.keras"
scaler_X_path  = base_version_path / "environment_base_model_lstm_scaler_X.pkl"
scaler_y_path  = base_version_path / "environment_base_model_lstm_scaler_y.pkl"
dataset_path   = DATASET_ENVIRONMENT_LSTM / "environment_dataset_lstm.csv"
quantized_path = run_dir / "environment_quantized_model_lstm.tflite"

# =========================
# Load data and features
# (harmonized with base/pruned: residual target + lags + 24-step window)
# =========================
df_raw = pd.read_csv(dataset_path)

# Check expected columns
cols_req = {"datetime","T_out","T_in","H_out","H_in"}
missing = cols_req - set(df_raw.columns)
if missing:
    raise ValueError(
        f"Missing columns in the dataset: {missing}. "
        f"Expected: {sorted(cols_req)}"
    )

# Sort temporally to preserve a deterministic raw_real timeline.
df_raw = df_raw.sort_values("datetime").reset_index(drop=True)
df_raw["__row_id_raw"] = np.arange(len(df_raw), dtype=np.int64)

df_raw["datetime"] = pd.to_datetime(df_raw["datetime"])
for _c in ["T_out", "H_out", "T_in", "H_in"]:
    if _c in df_raw.columns:
        df_raw[_c] = pd.to_numeric(df_raw[_c], errors="coerce")

# Preserve raw_real for replay header/mirror exports.
df_raw_real = df_raw.copy()

# Processed DataFrame used in the offline LiteML-Edge pipeline.
df = df_raw.copy()

# Temporal feature engineering
df["hour"] = df["datetime"].dt.hour + df["datetime"].dt.minute / 60.0
df["weekday"] = df["datetime"].dt.weekday
df["month"] = df["datetime"].dt.month

# === Selective causal smoothing on H_in (same idea as Conv1D Tiny) ===
HIN_EMA_ALPHA = 0.08
if "H_in" in df.columns:
    df["H_in"] = pd.to_numeric(df["H_in"], errors="coerce")
    df["H_in"] = df["H_in"].ewm(alpha=HIN_EMA_ALPHA, adjust=False).mean()

# Cyclical features
df["sin_hour"] = np.sin(2 * np.pi * df["hour"] / 24)
df["cos_hour"] = np.cos(2 * np.pi * df["hour"] / 24)

# Lag features (first!)
df["T_in_lag1"]  = df["T_in"].shift(1)
df["H_in_lag1"]  = df["H_in"].shift(1)
df["T_out_lag1"] = df["T_out"].shift(1)
df["H_out_lag1"] = df["H_out"].shift(1)

# Add lag2 only for indoor variables to keep 12 + weekday/month
df["T_in_lag2"]  = df["T_in"].shift(2)
df["H_in_lag2"]  = df["H_in"].shift(2)

df.dropna(inplace=True)
df_processed_row_ids = df["__row_id_raw"].astype(np.int64).to_numpy()

# Separate features (12 features) 
features = [
    "T_out", "H_out",          # current outdoor state
    "T_in_lag1", "H_in_lag1",  # indoor lag1
    "T_out_lag1", "H_out_lag1",# outdoor lag1
    "T_in_lag2", "H_in_lag2",  # indoor lag2
    "sin_hour", "cos_hour",    # cyclical hour
    "weekday", "month",        # calendar features
]

# Internal residual target: ΔT_in, ΔH_in
y_all = np.stack([
    (df['T_in'] - df['T_in_lag1']).values,
    (df['H_in'] - df['H_in_lag1']).values
], axis=1).astype(np.float32)

# LiteML-Edge contract: DO NOT apply clamp/clip to the residual Δ (only to X, in the physical domain).
# (we keep the residual pure: Δ = abs - lag1)
# Original scale feature matrix
X_source = df[features].values.astype(np.float32)

# ===== Temporal sliding window (24 steps) =====
def build_sequences(X2D, Y2D, seq_len):
    Xs, Ys, idxs = [], [], []
    for i in range(seq_len - 1, len(X2D)):
        Xs.append(X2D[i - seq_len + 1 : i + 1, :])  # (seq_len, n_features)
        Ys.append(Y2D[i])                           # (2,)
        idxs.append(i)
    return (
        np.asarray(Xs, dtype=np.float32),
        np.asarray(Ys, dtype=np.float32),
        np.asarray(idxs, dtype=np.int64),
    )

X_seq, y_seq, idx_seq = build_sequences(X_source, y_all, SEQ_LEN)
N_seq, _, K_NUM_FEATURES = X_seq.shape  # K_NUM_FEATURES = 12

# ===== Temporal split (60/20/20) in sequences =====
n_train = int(0.6 * N_seq)
n_val   = int(0.2 * N_seq)

X_train, y_train = X_seq[:n_train],           y_seq[:n_train]
X_val,   y_val   = X_seq[n_train:n_train+n_val], y_seq[n_train:n_train+n_val]
X_test,  y_test  = X_seq[n_train+n_val:],     y_seq[n_train+n_val:]

# Preserve the original physical sequences for debug/replay exports.
X_train_raw = X_train.astype(np.float32, copy=True)
X_val_raw   = X_val.astype(np.float32, copy=True)
X_test_raw  = X_test.astype(np.float32, copy=True)

idx_train = idx_seq[:n_train]
idx_val   = idx_seq[n_train:n_train+n_val]
idx_test  = idx_seq[n_train+n_val:]

# ===== Absolute references (for T_in, H_in reconstruction) =====
T_prev_all = df['T_in_lag1'].values.astype(np.float32)
H_prev_all = df['H_in_lag1'].values.astype(np.float32)
y_abs_all  = df[['T_in', 'H_in']].values.astype(np.float32)

T_prev_train = T_prev_all[idx_train]
H_prev_train = H_prev_all[idx_train]
y_train_abs  = y_abs_all[idx_train]

T_prev_val = T_prev_all[idx_val]
H_prev_val = H_prev_all[idx_val]
y_val_abs  = y_abs_all[idx_val]

T_prev_test = T_prev_all[idx_test]
H_prev_test = H_prev_all[idx_test]
y_test_abs  = y_abs_all[idx_test]

# ========================= Pretrained scalers (residual) =========================
scaler_X = joblib.load(scaler_X_path)
scaler_y = joblib.load(scaler_y_path)

# X in 3D flatten to normalize per feature
def clip_sequences(X3D, scaler):
    n, s, f = X3D.shape
    X2D = X3D.reshape(-1, f)
    X2D_clip = np.clip(X2D, scaler.data_min_, scaler.data_max_).astype(np.float32)
    return X2D_clip.reshape(n, s, f)

def scale_sequences(X3D, scaler):
    n, s, f = X3D.shape
    X2D = X3D.reshape(-1, f)
    # LiteML-Edge contract: clamp X in the physical domain to keep minmax_forward in [0,1]
    X2D = np.clip(X2D, scaler.data_min_, scaler.data_max_).astype(np.float32)
    X2D_scaled = scaler.transform(X2D)
    return X2D_scaled.reshape(n, s, f)

X_train_clip = clip_sequences(X_train_raw, scaler_X)
X_val_clip   = clip_sequences(X_val_raw,   scaler_X)
X_test_clip  = clip_sequences(X_test_raw,  scaler_X)

X_train = scale_sequences(X_train_raw, scaler_X)
X_val   = scale_sequences(X_val_raw,   scaler_X)
X_test  = scale_sequences(X_test_raw,  scaler_X)

# y: residual targets ΔT_in and ΔH_in normalized
y_train = scaler_y.transform(y_train)
y_val   = scaler_y.transform(y_val)
y_test  = scaler_y.transform(y_test)

# ========================= Load pruned model =========================
base_model = tf.keras.models.load_model(model_path)

# ========================= QAT only on Dense =========================
# Ensure that LSTM/BiLSTM layers are NOT quantized; they remain float
def is_lstm_like(layer):
    return isinstance(layer, (tf.keras.layers.LSTM, tf.keras.layers.Bidirectional))

def _annotate_clone(layer):
    # LSTM/Bidirectional and normalization layers: NoOp (preserve float)
    if is_lstm_like(layer) or isinstance(layer, (tf.keras.layers.BatchNormalization, tf.keras.layers.LayerNormalization)):
        cfg = layer.get_config()
        new_layer = layer.__class__.from_config(cfg)
        return tfmot.quantization.keras.quantize_annotate_layer(new_layer, NoOpQuantizeConfig())
    # Dense with QAT (INT8)
    if isinstance(layer, tf.keras.layers.Dense):
        cfg = layer.get_config()
        new_layer = layer.__class__.from_config(cfg)
        return tfmot.quantization.keras.quantize_annotate_layer(new_layer)
    return layer

annotate_model = tf.keras.models.clone_model(base_model, clone_function=_annotate_clone)

with tfmot.quantization.keras.quantize_scope({'NoOpQuantizeConfig': NoOpQuantizeConfig}):
    qat_model = tfmot.quantization.keras.quantize_apply(annotate_model)

# ========================= Losses =========================
def weighted_huber(y_true, y_pred, delta=1.0):
    error = y_true - y_pred
    abs_err = tf.abs(error)
    quadratic = tf.minimum(abs_err, delta)
    linear = abs_err - quadratic
    hub = 0.5 * tf.square(quadratic) + delta * linear
    w = tf.constant([1.0, H_WEIGHT], dtype=hub.dtype)
    hub_w = hub * w
    return tf.reduce_mean(hub_w)

loss_fn = weighted_huber if USE_WEIGHTED_HUBER else tf.keras.losses.MeanSquaredError()

qat_model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=LR),
    loss=loss_fn,
    metrics=[tf.keras.metrics.MeanAbsoluteError()]
)

reduce_lr = tf.keras.callbacks.ReduceLROnPlateau(
    monitor='val_loss', factor=0.5, patience=PATIENCE_RLR, min_lr=MIN_LR, verbose=1
)
early_stop = tf.keras.callbacks.EarlyStopping(
    monitor='val_loss', patience=PATIENCE_ES, restore_best_weights=True, verbose=1
)

print("[INFO] Training with QAT (INT8 in Dense layers; LSTM kept in float)...")
history = qat_model.fit(
    X_train, y_train,
    validation_data=(X_val, y_val),
    epochs=EPOCHS,
    batch_size=BATCH_SIZE,
    callbacks=[early_stop, reduce_lr],
    verbose=1
)

# ========================= Sparsity (informational only) =========================
def print_model_sparsity(model):
    total_params, zero_params = 0, 0
    for layer in model.layers:
        for w in layer.get_weights():
            total_params += w.size
            zero_params  += np.sum(w == 0)
    sparsity = (zero_params / total_params) * 100 if total_params > 0 else 0
    print(f"Sparsity: {sparsity:.2f}% ({zero_params} / {total_params})")

print_model_sparsity(qat_model)
qat_model.save(run_dir / "environment_qat_model_lstm.keras")

# ========================= Representative dataset =========================
X_calib = X_train.astype(np.float32, copy=False)

def representative_dataset(n_samples=REP_SAMPLES):
    for i in range(min(n_samples, len(X_calib))):
        yield [X_calib[i:i+1].astype(np.float32)]

# ========================= ConcreteFunction p/ TFLite =========================
def _get_concrete_function(keras_model, batch_size=1, timesteps=24, features=12):
    @tf.function(input_signature=[tf.TensorSpec(shape=[batch_size, timesteps, features], dtype=tf.float32)])
    def serve(x):
        return keras_model(x, training=False)
    return serve.get_concrete_function()

# ========================= TFLite conversion =========================
def unwrap_layer(l):
    # try to unwrap TFMOT quantization wrappers
    try:
        from tensorflow_model_optimization.python.core.quantization.keras.quantize_wrapper import QuantizeWrapper, QuantizeWrapperV2
        while isinstance(l, (QuantizeWrapper, QuantizeWrapperV2)):
            l = l.layer
    except Exception:
        pass
    return l

def is_lstm_deep(layer):
    import tensorflow as tf
    LSTM = tf.keras.layers.LSTM
    RNN  = tf.keras.layers.RNN
    Bi   = tf.keras.layers.Bidirectional
    LSTMCell = tf.keras.layers.LSTMCell

    l = unwrap_layer(layer)

    if isinstance(l, LSTM):
        return True

    if isinstance(l, Bi):
        inner = getattr(l, "layer", None) or getattr(l, "forward_layer", None) or None
        if inner is not None:
            return is_lstm_deep(inner)

    if isinstance(l, RNN):
        cell = getattr(l, "cell", None)
        if isinstance(cell, LSTMCell):
            return True
        if hasattr(cell, "cells"):
            return any(isinstance(c, LSTMCell) for c in cell.cells)

    return False

HAS_LSTM = any(is_lstm_deep(m) for m in qat_model.submodules)
print(f"[INFO] (Keras) LSTM present? {HAS_LSTM}  (note: this check may fail; confirm in the .tflite artifact)")

try:
    concrete_func = _get_concrete_function(qat_model, batch_size=1, timesteps=SEQ_LEN, features=K_NUM_FEATURES)
    converter = tf.lite.TFLiteConverter.from_concrete_functions([concrete_func], qat_model)
except Exception as _e:
    print("[WARN] from_concrete_functions failed, using from_keras_model:", _e)
    converter = tf.lite.TFLiteConverter.from_keras_model(qat_model)

converter.optimizations = [tf.lite.Optimize.DEFAULT]
converter.representative_dataset = representative_dataset
converter.target_spec.supported_ops = [
    tf.lite.OpsSet.TFLITE_BUILTINS_INT8,  # int8 for Dense
    tf.lite.OpsSet.TFLITE_BUILTINS,       # float ops
    tf.lite.OpsSet.SELECT_TF_OPS          # required for Keras LSTM
]

# critical to preserve TensorList/control-flow behavior for LSTM through Select TF Ops
try:
    converter._experimental_lower_tensor_list_ops = False
except Exception:
    pass

# force float I/O when LSTM is present (hybrid: float input/output, mixed internal path)
if HAS_LSTM:
    converter.inference_input_type  = tf.float32
    converter.inference_output_type = tf.float32
else:
    converter.inference_input_type  = tf.int8
    converter.inference_output_type = tf.int8

try:
    quantized_model = converter.convert()
except Exception as e:
    print("[WARN] Hybrid conversion failed; falling back to BUILTINS+SELECT_TF_OPS (float I/O):", e)
    converter.optimizations = [tf.lite.Optimize.DEFAULT]
    converter.target_spec.supported_ops = [tf.lite.OpsSet.TFLITE_BUILTINS, tf.lite.OpsSet.SELECT_TF_OPS]
    converter.inference_input_type  = tf.float32
    converter.inference_output_type = tf.float32
    quantized_model = converter.convert()

with open(quantized_path, "wb") as f:
    f.write(quantized_model)
print("[INFO] Quantized model saved at:", quantized_path)

# ===== Verification: was the LSTM preserved in the .tflite model? =====
def assert_lstm_preserved(tflite_path: str,
                          must_have_any: tuple = (
                              "While",
                              "BatchMatMulV2", "MatMul",
                              "TensorList",
                              "UnidirectionalSequenceLSTM",
                              "BidirectionalSequenceLSTM",
                              "LSTM"
                          )) -> None:
    import numpy as np
    import tensorflow as tf
    try:
        intr = tf.lite.Interpreter(model_path=tflite_path)
        intr.allocate_tensors()
    except Exception as e:
        raise RuntimeError(f"[check] Failed to open TFLite artifact: {e}")

    try:
        ops = intr._get_ops_details()
        op_names = [op.get("op_name", "") for op in ops]
    except Exception as e:
        raise RuntimeError(f"[check] Unable to inspect operators: {e}")

    print("[check] Operators detected in the .tflite artifact:")
    for name in op_names:
        print("    -", name)

    found = any(any(token.lower() in name.lower() for token in must_have_any) for name in op_names)

    if not found:
        tips = [
            "Confirm that 'SELECT_TF_OPS' is included in converter.target_spec.supported_ops.",
            "Keep converter._experimental_lower_tensor_list_ops = False.",
            "Ensure that LSTM/Bidirectional layers were annotated with NoOpQuantizeConfig (without QAT).",
            "If needed, force float I/O: converter.inference_input_type/output_type = tf.float32.",
        ]
        msg = "\n".join([
            "[WARN] Verification failed: no LSTM or control-flow pattern was identified in the .tflite artifact.",
            "This suggests that the LSTM may have been removed/fused or the conversion fell back to a purely int8 path.",
            "Suggested checks:",
            *[f" - {t}" for t in tips]
        ])
        raise SystemExit(msg)

    print("[INFO] Verification passed: LSTM or control-flow patterns were detected in the .tflite artifact.")

assert_lstm_preserved(str(quantized_path))

# ========================= TFLite inference =========================
interpreter = tf.lite.Interpreter(model_path=str(quantized_path))
interpreter.allocate_tensors()
input_details  = interpreter.get_input_details()
output_details = interpreter.get_output_details()

try:
    ops = interpreter._get_ops_details()
    print("[ops] Operators in the TFLite artifact:")
    for op in ops:
        print(" -", op.get("op_name"))
except Exception:
    pass

in_info  = input_details[0]
out_info = output_details[0]
in_scale,  in_zp  = in_info["quantization"]
out_scale, out_zp = out_info["quantization"]

print(f"[quant] Input dtype={in_info['dtype']}  scale={in_scale}  zp={in_zp}  shape={in_info['shape']}")
print(f"[quant] Output dtype={out_info['dtype']} scale={out_scale} zp={out_zp} shape={out_info['shape']}")

DBG_MODEL_RAW_MAX_DIMS = 8
DBG_MODEL_RAW_MAX_BYTES = 32

def quantize(x_f32: np.ndarray, scale: float, zp: int, dtype) -> np.ndarray:
    if scale == 0: return x_f32.astype(dtype)
    x_q = np.round(x_f32 / scale + zp)
    if dtype == np.int8:    x_q = np.clip(x_q, -128, 127)
    elif dtype == np.uint8: x_q = np.clip(x_q, 0, 255)
    return x_q.astype(dtype)

def dequantize(x_q: np.ndarray, scale: float, zp: int) -> np.ndarray:
    if scale == 0: return x_q.astype(np.float32)
    return (x_q.astype(np.float32) - zp) * scale

def _tflite_dtype_code(dtype) -> int:
    _dt = np.dtype(dtype)
    _map = {
        np.dtype(np.float32): 1,
        np.dtype(np.int32): 2,
        np.dtype(np.uint8): 3,
        np.dtype(np.int64): 4,
        np.dtype(np.str_): 5,
        np.dtype(np.bool_): 6,
        np.dtype(np.int16): 7,
        np.dtype(np.complex64): 8,
        np.dtype(np.int8): 9,
        np.dtype(np.float16): 10,
        np.dtype(np.float64): 11,
        np.dtype(np.complex128): 12,
        np.dtype(np.uint64): 13,
        np.dtype(np.uint32): 16,
        np.dtype(np.uint16): 17,
    }
    return int(_map.get(_dt, -1))


def _tflite_dtype_name(dtype) -> str:
    try:
        return np.dtype(dtype).name
    except Exception:
        return str(dtype)


def _capture_tflite_output_raw_records(max_dims: int = DBG_MODEL_RAW_MAX_DIMS,
                                       max_dump_bytes: int = DBG_MODEL_RAW_MAX_BYTES):
    records = []
    for out_idx, od in enumerate(output_details):
        y_q = interpreter.get_tensor(od["index"])
        y_arr = np.asarray(y_q)
        y_arr_c = np.ascontiguousarray(y_arr)
        raw_bytes = y_arr_c.tobytes(order="C")
        shape = list(y_arr_c.shape)
        bytes_total = int(len(raw_bytes))
        bytes_dumped = int(min(bytes_total, int(max_dump_bytes)))

        rec = {
            "out_idx": int(out_idx),
            "tensor_index": int(od["index"]),
            "type_code": _tflite_dtype_code(y_arr_c.dtype),
            "type_name": _tflite_dtype_name(y_arr_c.dtype),
            "bytes_total": bytes_total,
            "bytes_dumped": bytes_dumped,
            "dims_size": int(len(shape)),
        }

        for di in range(int(max_dims)):
            rec[f"dim{di:02d}"] = int(shape[di]) if di < len(shape) else -1

        for bi in range(int(max_dump_bytes)):
            rec[f"b{bi:02d}_hex"] = f"0x{raw_bytes[bi]:02X}" if bi < bytes_dumped else ""

        records.append(rec)

    return records


def _read_tflite_output_payload_and_float():
    payload_parts = []
    float_parts = []
    for od in output_details:
        y_q = interpreter.get_tensor(od["index"])
        y_q_arr = np.asarray(y_q)
        payload_parts.append(y_q_arr)

        od_scale, od_zp = od["quantization"]
        if od["dtype"] == np.float32:
            y_f_arr = y_q_arr.astype(np.float32, copy=False)
        else:
            y_f_arr = dequantize(y_q_arr, od_scale, od_zp)
        float_parts.append(np.asarray(y_f_arr, dtype=np.float32))

    if len(payload_parts) == 1:
        payload_flat = np.asarray(payload_parts[0][0], dtype=np.float32).reshape(-1)
        float_flat = np.asarray(float_parts[0][0], dtype=np.float32).reshape(-1)
    else:
        payload_flat = np.asarray([
            np.asarray(part[0]).reshape(-1)[0] for part in payload_parts
        ], dtype=np.float32)
        float_flat = np.asarray([
            np.asarray(part[0], dtype=np.float32).reshape(-1)[0] for part in float_parts
        ], dtype=np.float32)

    return payload_flat.astype(np.float32, copy=False), float_flat.astype(np.float32, copy=False)


def _predict_tflite(X3D, capture_io: bool = False):
    preds = []
    captured_input_payload = []
    captured_input_float = []
    captured_output_payload = []
    captured_output_float = []
    captured_output_raw = []

    input_dtype = in_info["dtype"]
    for x in X3D:
        x_f32 = x.reshape(1, SEQ_LEN, X3D.shape[-1]).astype(np.float32)
        x_in = x_f32 if input_dtype == np.float32 else quantize(x_f32, in_scale, in_zp, input_dtype)
        interpreter.set_tensor(in_info["index"], x_in)
        interpreter.invoke()

        output_raw_records = _capture_tflite_output_raw_records()
        y_payload, y_float = _read_tflite_output_payload_and_float()
        preds.append(y_float)

        if capture_io:
            x_payload_flat = np.asarray(x_in[0], dtype=np.float32).reshape(-1)
            # Keep DBG_MODEL_IN_CSV aligned with the firmware contract:
            #   in_x*_float = actual tensor payload converted to float
            #   in_f*_scaled = normalized float input BEFORE the tensor write
            x_float_flat = np.asarray(x_f32[0], dtype=np.float32).reshape(-1)
            captured_input_payload.append(x_payload_flat.astype(np.float32, copy=False))
            captured_input_float.append(np.asarray(x_float_flat, dtype=np.float32))
            captured_output_payload.append(np.asarray(y_payload, dtype=np.float32))
            captured_output_float.append(np.asarray(y_float, dtype=np.float32))
            captured_output_raw.append(output_raw_records)

    preds = np.asarray(preds, dtype=np.float32)
    if not capture_io:
        return preds

    return {
        "preds": preds,
        "input_payload": np.asarray(captured_input_payload, dtype=np.float32),
        "input_float": np.asarray(captured_input_float, dtype=np.float32),
        "output_payload": np.asarray(captured_output_payload, dtype=np.float32),
        "output_float": np.asarray(captured_output_float, dtype=np.float32),
        "output_raw": captured_output_raw,
    }


def _dbg_f32(value):
    return float(np.float32(value))


def _float32_to_hex_bits(value):
    _v = np.float32(value)
    return f"0x{struct.unpack('<I', struct.pack('<f', float(_v)))[0]:08X}"


def _build_replay_canonical_processed(raw_replay_df, ema_prev, alpha):
    """Rebuild the replay-processed path with float32 arithmetic, mirroring the
    firmware contract from the exported raw 2+47 block plus the EMA seed.
    """
    _df = raw_replay_df.copy().reset_index(drop=True)

    for _c in ["T_out", "H_out", "T_in", "H_in"]:
        _df[_c] = pd.to_numeric(_df[_c], errors="coerce").astype(np.float32)

    _dt = pd.to_datetime(_df["datetime"])
    _hour = _dt.dt.hour.astype(np.int64).to_numpy()
    _weekday = _dt.dt.weekday.astype(np.int64).to_numpy()
    _month = _dt.dt.month.astype(np.int64).to_numpy()

    _df["hour"] = _hour
    _df["weekday"] = _weekday.astype(np.float32)
    _df["month"] = _month.astype(np.float32)
    _df["sin_hour"] = np.sin(2.0 * np.pi * _hour / 24.0).astype(np.float32)
    _df["cos_hour"] = np.cos(2.0 * np.pi * _hour / 24.0).astype(np.float32)

    _alpha = np.float32(alpha)
    _one_minus = np.float32(1.0) - _alpha
    _prev = np.float32(ema_prev)
    _h_raw = _df["H_in"].to_numpy(dtype=np.float32, copy=True)
    _h_filt = np.empty_like(_h_raw, dtype=np.float32)
    for _i, _raw in enumerate(_h_raw):
        _curr = np.float32(_raw)
        _filt = np.float32(_alpha * _curr + _one_minus * _prev)
        _h_filt[_i] = _filt
        _prev = _filt
    _df["H_in"] = _h_filt

    _df["T_in_lag1"] = _df["T_in"].shift(1).astype(np.float32)
    _df["H_in_lag1"] = _df["H_in"].shift(1).astype(np.float32)
    _df["T_out_lag1"] = _df["T_out"].shift(1).astype(np.float32)
    _df["H_out_lag1"] = _df["H_out"].shift(1).astype(np.float32)
    _df["T_in_lag2"] = _df["T_in"].shift(2).astype(np.float32)
    _df["H_in_lag2"] = _df["H_in"].shift(2).astype(np.float32)

    return _df


def _extract_canonical_replay_debug_inputs(sample_df_indices, replay_processed_df, replay_processed_df_indices, raw_replay_df=None):
    """Build canonical preprocessing windows from the replay-processed source
    mirrored by firmware during DBG_MODEL_IN_CSV, while also exposing explicit
    preprocessing checkpoints for auditability.
    """
    _index_map = {int(_df_idx): int(_pos) for _pos, _df_idx in enumerate(replay_processed_df_indices)}
    _pre_raw_seqs = []
    _pre_smooth_seqs = []
    _pre_lags_seqs = []
    _pre_time_seqs = []
    _pre_phys_seqs = []
    _step_epoch_seqs = []
    _y_abs_rows = []
    _tprev_rows = []
    _hprev_rows = []

    _raw_df = raw_replay_df if raw_replay_df is not None else replay_processed_df

    for _df_idx in sample_df_indices:
        _end = _index_map.get(int(_df_idx))
        if _end is None:
            raise KeyError(f"Replay canonical index not found for df index {_df_idx}")
        if _end < SEQ_LEN - 1:
            raise ValueError(
                f"Replay canonical window too short for df index {_df_idx}: end={_end}, seq_len={SEQ_LEN}"
            )

        _start = _end - SEQ_LEN + 1
        _win_proc = replay_processed_df.iloc[_start : _end + 1]
        _win_raw = _raw_df.iloc[_start : _end + 1]

        if "epoch" in _win_raw.columns:
            _step_epochs = _win_raw["epoch"].to_numpy(dtype=np.int64, copy=True)
        else:
            _step_epochs = (pd.to_datetime(_win_raw["datetime"]).astype("int64") // 10**9).to_numpy(dtype=np.int64, copy=True)

        _pre_raw_seqs.append(
            _win_raw[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_smooth_seqs.append(
            _win_proc[["T_out", "H_out", "T_in", "H_in"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_lags_seqs.append(
            _win_proc[["T_in_lag1", "H_in_lag1", "T_out_lag1", "H_out_lag1", "T_in_lag2", "H_in_lag2"]].to_numpy(dtype=np.float32, copy=True)
        )
        _pre_time_seqs.append(
            np.column_stack([
                _step_epochs.astype(np.float32, copy=False),
                _win_proc["sin_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["cos_hour"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["weekday"].to_numpy(dtype=np.float32, copy=True),
                _win_proc["month"].to_numpy(dtype=np.float32, copy=True),
            ]).astype(np.float32, copy=False)
        )
        _pre_phys = _win_proc[features].to_numpy(dtype=np.float32, copy=True)
        _pre_phys_seqs.append(_pre_phys)
        _step_epoch_seqs.append(_step_epochs)

        _row = replay_processed_df.iloc[_end]
        _y_abs_rows.append(np.asarray([_row["T_in"], _row["H_in"]], dtype=np.float32))
        _tprev_rows.append(np.float32(_row["T_in_lag1"]))
        _hprev_rows.append(np.float32(_row["H_in_lag1"]))

    return {
        "pre_raw": np.asarray(_pre_raw_seqs, dtype=np.float32),
        "pre_smooth": np.asarray(_pre_smooth_seqs, dtype=np.float32),
        "pre_lags": np.asarray(_pre_lags_seqs, dtype=np.float32),
        "pre_time": np.asarray(_pre_time_seqs, dtype=np.float32),
        "pre_phys": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "step_epoch": np.asarray(_step_epoch_seqs, dtype=np.int64),
        "x_raw": np.asarray(_pre_phys_seqs, dtype=np.float32),
        "y_abs": np.asarray(_y_abs_rows, dtype=np.float32),
        "tprev": np.asarray(_tprev_rows, dtype=np.float32),
        "hprev": np.asarray(_hprev_rows, dtype=np.float32),
    }


def _build_dbg_model_input_reference(sample_test_positions, sample_df_indices, sample_epochs, x_raw_seqs, x_clip_seqs, x_scaled_seqs, x_tensor_payload_seqs, x_tensor_float_seqs, y_true_abs_rows):
    rows = []
    _n_features = len(features)
    for _idx, (_test_pos, _df_idx, _epoch_dbg) in enumerate(zip(sample_test_positions, sample_df_indices, sample_epochs)):
        _x_raw = np.asarray(x_raw_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _x_clip = np.asarray(x_clip_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _x_scaled = np.asarray(x_scaled_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _x_tensor_payload = np.asarray(x_tensor_payload_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _x_tensor_float = np.asarray(x_tensor_float_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _gt = np.asarray(y_true_abs_rows[_idx], dtype=np.float32).reshape(2)

        for _step in range(SEQ_LEN):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "gt_Tin_true": float(_gt[0]),
                "gt_Hin_true": _dbg_f32(_gt[1]),
                "state_Tout_phys_raw": float(_x_raw[_step, 0]),
                "state_Hout_phys_raw": float(_x_raw[_step, 1]),
                "state_Tin_lag1_phys_raw": float(_x_raw[_step, 2]),
                "state_Hin_lag1_phys_raw": _dbg_f32(_x_raw[_step, 3]),
                "state_Tout_lag1_phys_raw": float(_x_raw[_step, 4]),
                "state_Hout_lag1_phys_raw": float(_x_raw[_step, 5]),
                "state_Tin_lag2_phys_raw": float(_x_raw[_step, 6]),
                "state_Hin_lag2_phys_raw": _dbg_f32(_x_raw[_step, 7]),
                "state_sin_hour": float(_x_raw[_step, 8]),
                "state_cos_hour": float(_x_raw[_step, 9]),
                "state_weekday": float(_x_raw[_step, 10]),
                "state_month": float(_x_raw[_step, 11]),
            }
            for _feat in range(_n_features):
                _value = _x_raw[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_raw"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _value = _x_clip[_step, _feat]
                _row[f"in_f{_feat:02d}_phys_clip"] = _dbg_f32(_value) if _feat in (3, 7) else float(_value)
            for _feat in range(_n_features):
                _row[f"in_f{_feat:02d}_scaled"] = float(_x_scaled[_step, _feat])
            for _feat in range(_n_features):
                _row[f"in_x{_feat:02d}_float"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_payload"] = float(_x_tensor_payload[_step, _feat])
                _row[f"in_x{_feat:02d}_float_prequant"] = float(_x_tensor_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_stage_reference(sample_epochs, step_epoch_seqs, data_seqs, field_names):
    rows = []
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _data = np.asarray(data_seqs[_idx], dtype=np.float32)
        if _data.ndim == 1:
            _data = _data.reshape(SEQ_LEN, 1)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(SEQ_LEN)
        for _step in range(_data.shape[0]):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _j, _name in enumerate(field_names):
                _val = _data[_step, _j]
                _row[_name] = _dbg_f32(_val)
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_preprocess_sample_reference(raw_df, smooth_df, seed_rows: int = 2):
    """Build PRE_RAW/PRE_SMOOTH at firmware-equivalent logical-sample granularity.

    The firmware emits one row per logical replay sample after the seed rows,
    using idx=0..N-1 and raw_idx pointing to the underlying 2+47 replay row.
    This builder mirrors that exact contract while preserving the detailed
    window-level exports in separate sheets.
    """
    if len(raw_df) != len(smooth_df):
        raise ValueError(
            f"PRE_RAW/PRE_SMOOTH export length mismatch: raw={len(raw_df)} smooth={len(smooth_df)}"
        )
    if len(raw_df) <= seed_rows:
        raise ValueError(
            f"Need more than {seed_rows} rows to build logical PRE_RAW/PRE_SMOOTH exports"
        )

    rows_raw = []
    rows_smooth = []
    for _raw_idx in range(seed_rows, len(raw_df)):
        _idx = _raw_idx - seed_rows
        _epoch = int(raw_df.loc[_raw_idx, "epoch"]) if "epoch" in raw_df.columns else 0
        _row_raw = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_out"]),
            "Hout_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_out"]),
            "Tin_raw": _dbg_f32(raw_df.loc[_raw_idx, "T_in"]),
            "Hin_raw": _dbg_f32(raw_df.loc[_raw_idx, "H_in"]),
        }
        _row_smooth = {
            "idx": int(_idx),
            "mode": "REPLAY",
            "raw_idx": int(_raw_idx),
            "epoch": int(_epoch),
            "Tout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_out"]),
            "Hout_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_out"]),
            "Tin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "T_in"]),
            "Hin_smooth": _dbg_f32(smooth_df.loc[_raw_idx, "H_in"]),
        }
        rows_raw.append(_row_raw)
        rows_smooth.append(_row_smooth)

    return pd.DataFrame(rows_raw), pd.DataFrame(rows_smooth)


def _build_preprocess_model_in_reference(sample_epochs, step_epoch_seqs, x_payload_seqs, x_float_seqs):
    rows = []
    _n_features = len(features)
    for _idx, _epoch_dbg in enumerate(sample_epochs):
        _x_payload = np.asarray(x_payload_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _x_float = np.asarray(x_float_seqs[_idx], dtype=np.float32).reshape(SEQ_LEN, _n_features)
        _step_epochs = np.asarray(step_epoch_seqs[_idx], dtype=np.int64).reshape(SEQ_LEN)
        for _step in range(SEQ_LEN):
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "step": int(_step),
                "step_epoch": int(_step_epochs[_step]),
            }
            for _feat in range(_n_features):
                _row[f"x_payload_{_feat:02d}"] = float(_x_payload[_step, _feat])
                _row[f"x_float_prequant_{_feat:02d}"] = float(_x_float[_step, _feat])
            rows.append(_row)
    return pd.DataFrame(rows)


def _build_dbg_model_output_reference(sample_test_positions, sample_epochs, output_payload, output_float, tprev_override=None, hprev_override=None):
    _y_scaled = np.asarray(output_float, dtype=np.float32).reshape(-1, 2)
    _d_pred = scaler_y.inverse_transform(_y_scaled).astype(np.float32, copy=False)
    if tprev_override is None:
        _tprev = np.asarray(T_prev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _tprev = np.asarray(tprev_override, dtype=np.float32).reshape(-1)
    if hprev_override is None:
        _hprev = np.asarray(H_prev_test[np.asarray(sample_test_positions, dtype=np.int64)], dtype=np.float32)
    else:
        _hprev = np.asarray(hprev_override, dtype=np.float32).reshape(-1)
    _t_pred = (_tprev + _d_pred[:, 0]).astype(np.float32, copy=False)
    _h_pred = (_hprev + _d_pred[:, 1]).astype(np.float32, copy=False)

    _out_o0_float = np.asarray(output_float[:, 0], dtype=np.float32)
    _out_o1_float = np.asarray(output_float[:, 1], dtype=np.float32)
    _out_o0_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o0_float]
    _out_o1_bits_hex = [_float32_to_hex_bits(_v) for _v in _out_o1_float]

    return pd.DataFrame({
        "idx": np.arange(len(sample_epochs), dtype=np.int64),
        "epoch": np.asarray(sample_epochs, dtype=np.int64),
        "out_o0_tensor": np.asarray(output_payload[:, 0], dtype=np.float32),
        "out_o1_tensor": np.asarray(output_payload[:, 1], dtype=np.float32),
        "out_o0_float": _out_o0_float,
        "out_o1_float": _out_o1_float,
        "out_o0_bits_hex": _out_o0_bits_hex,
        "out_o1_bits_hex": _out_o1_bits_hex,
        "y_T_scaled": _out_o0_float,
        "y_H_scaled": _out_o1_float,
        "d_T_pred": np.asarray(_d_pred[:, 0], dtype=np.float32),
        "d_H_pred": np.asarray(_d_pred[:, 1], dtype=np.float32),
        "p_Tprev_phys": _tprev,
        "p_Hprev_phys": np.asarray(_hprev, dtype=np.float32),
        "p_T_pred": _t_pred,
        "p_H_pred": _h_pred,
    })


def _build_dbg_model_output_raw_reference(sample_epochs, output_raw_records):
    rows = []
    max_dims = int(DBG_MODEL_RAW_MAX_DIMS)
    max_bytes = int(DBG_MODEL_RAW_MAX_BYTES)

    for _idx, (_epoch_dbg, _sample_records) in enumerate(zip(sample_epochs, output_raw_records)):
        for _rec in _sample_records:
            _row = {
                "idx": int(_idx),
                "epoch": int(_epoch_dbg),
                "out_idx": int(_rec.get("out_idx", -1)),
                "tensor_index": int(_rec.get("tensor_index", -1)),
                "type_code": int(_rec.get("type_code", -1)),
                "type_name": str(_rec.get("type_name", "")),
                "bytes_total": int(_rec.get("bytes_total", 0)),
                "bytes_dumped": int(_rec.get("bytes_dumped", 0)),
                "dims_size": int(_rec.get("dims_size", 0)),
            }
            for _di in range(max_dims):
                _row[f"dim{_di:02d}"] = int(_rec.get(f"dim{_di:02d}", -1))
            for _bi in range(max_bytes):
                _row[f"b{_bi:02d}_hex"] = str(_rec.get(f"b{_bi:02d}_hex", ""))
            rows.append(_row)

    if rows:
        return pd.DataFrame(rows)

    _columns = [
        "idx", "epoch", "out_idx", "tensor_index", "type_code", "type_name",
        "bytes_total", "bytes_dumped", "dims_size",
    ]
    _columns += [f"dim{_di:02d}" for _di in range(max_dims)]
    _columns += [f"b{_bi:02d}_hex" for _bi in range(max_bytes)]
    return pd.DataFrame(columns=_columns)


def _build_dbg_replay_reference(sample_test_positions, replay_epochs, canonical_rows=None, tprev_override=None, hprev_override=None, y_true_abs_override=None):
    _sel = np.asarray(sample_test_positions, dtype=np.int64)
    if canonical_rows is None:
        _df_rows = df.iloc[idx_test[_sel]].copy().reset_index(drop=True)
    else:
        _df_rows = canonical_rows.copy().reset_index(drop=True)
    _y_scaled = np.asarray(y_pred_tflite[_sel], dtype=np.float32)
    _d_pred = np.asarray(y_pred_res[_sel], dtype=np.float32)
    if y_true_abs_override is None:
        _y_true_abs = np.asarray(y_test_abs_final[_sel], dtype=np.float32)
    else:
        _y_true_abs = np.asarray(y_true_abs_override, dtype=np.float32).reshape(-1, 2)
    if tprev_override is None:
        _tprev = np.asarray(T_prev_test[_sel], dtype=np.float32)
    else:
        _tprev = np.asarray(tprev_override, dtype=np.float32).reshape(-1)
    if hprev_override is None:
        _hprev = np.asarray(H_prev_test[_sel], dtype=np.float32)
    else:
        _hprev = np.asarray(hprev_override, dtype=np.float32).reshape(-1)
    _t_pred = (_tprev + _d_pred[:, 0]).astype(np.float32, copy=False)
    _h_pred = (_hprev + _d_pred[:, 1]).astype(np.float32, copy=False)

    return pd.DataFrame({
        "idx": np.arange(len(_sel), dtype=np.int64),
        "epoch": np.asarray(replay_epochs, dtype=np.int64),
        "gt_Tin_true": _y_true_abs[:, 0],
        "gt_Hin_true": np.asarray(_y_true_abs[:, 1], dtype=np.float32),
        "state_Tout_phys_raw": _df_rows["T_out"].astype(np.float32),
        "state_Hout_phys_raw": _df_rows["H_out"].astype(np.float32),
        "state_Tin_lag1_phys_raw": _df_rows["T_in_lag1"].astype(np.float32),
        "state_Hin_lag1_phys_raw": np.asarray(_df_rows["H_in_lag1"], dtype=np.float32),
        "state_Tout_lag1_phys_raw": _df_rows["T_out_lag1"].astype(np.float32),
        "state_Hout_lag1_phys_raw": _df_rows["H_out_lag1"].astype(np.float32),
        "state_Tin_lag2_phys_raw": _df_rows["T_in_lag2"].astype(np.float32),
        "state_Hin_lag2_phys_raw": np.asarray(_df_rows["H_in_lag2"], dtype=np.float32),
        "state_sin_hour": _df_rows["sin_hour"].astype(np.float32),
        "state_cos_hour": _df_rows["cos_hour"].astype(np.float32),
        "state_weekday": _df_rows["weekday"].astype(np.float32),
        "state_month": _df_rows["month"].astype(np.float32),
        "y_T_scaled": _y_scaled[:, 0],
        "y_H_scaled": _y_scaled[:, 1],
        "d_T_pred": _d_pred[:, 0],
        "d_H_pred": _d_pred[:, 1],
        "p_Tprev_phys": _tprev,
        "p_Hprev_phys": np.asarray(_hprev, dtype=np.float32),
        "p_T_pred": _t_pred,
        "p_H_pred": _h_pred,
    })

start_time = time.time()
y_val_pred_tflite = _predict_tflite(X_val)
y_pred_tflite     = _predict_tflite(X_test)
end_time = time.time()

# ========================= Metrics in normalized scale (residual) =========================
mse_scaled  = mean_squared_error(y_test, y_pred_tflite)
rmse_scaled = np.sqrt(mse_scaled)
mae_scaled  = mean_absolute_error(y_test, y_pred_tflite)
r2_scaled   = r2_score(y_test, y_pred_tflite)

print("\n Results (normalized scale - residual training target):")
print(f"MSE  = {mse_scaled:.4f}")
print(f"RMSE = {rmse_scaled:.4f}")
print(f"MAE  = {mae_scaled:.4f}")
print(f"R²   = {r2_scaled:.4f}")

# ========================= Original residual scale =========================
y_val_orig_res      = scaler_y.inverse_transform(y_val).astype(np.float32, copy=False)
y_val_pred_orig_res = scaler_y.inverse_transform(y_val_pred_tflite).astype(np.float32, copy=False)
y_test_orig_res     = scaler_y.inverse_transform(y_test).astype(np.float32, copy=False)
y_pred_orig_raw_res = scaler_y.inverse_transform(y_pred_tflite).astype(np.float32, copy=False)

# ------------------------------------------------------------
# Contract enforcement (LiteML-Edge):
#   - Residual-only output (ΔT_in, ΔH_in) in original units
#   - Absolute reconstruction happens later: Tin = Tin_lag1 + ΔT, Hin = Hin_lag1 + ΔH
# ------------------------------------------------------------
y_pred_res = np.asarray(y_pred_orig_raw_res, dtype=np.float32)

# ========================= T_in, H_in =========================
T_prev_test = T_prev_test.astype(np.float32, copy=False)
H_prev_test = H_prev_test.astype(np.float32, copy=False)

T_pred = (T_prev_test + y_pred_res[:, 0]).astype(np.float32, copy=False)
H_pred = (H_prev_test + y_pred_res[:, 1]).astype(np.float32, copy=False)

y_pred_abs = np.stack([T_pred, H_pred], axis=1).astype(np.float32, copy=False)

# Ground truth also in float32
y_test_abs_final = y_test_abs.astype(np.float32, copy=False)  # (T_in_ground truth, H_in_ground truth)

# ============================================================
# ROLLING-WINDOW METRICS WITH N=24 (FIRMWARE-EQUIVALENT)
# ============================================================

ROLLING_N = 24  # same as METRICS_WINDOW_SIZE in the firmware
ROLLING_EXPORT_ROWS = 24  # export 24 rolling24 windows and generate replay 2+47

def _r2_like_firmware(y_true_1d, y_pred_1d):
    # Replicate the firmware logic (metrics.cpp): R² is undefined when variance is too low.
    y_true_1d = np.asarray(y_true_1d, dtype=np.float32).reshape(-1)
    y_pred_1d = np.asarray(y_pred_1d, dtype=np.float32).reshape(-1)
    n = int(y_true_1d.size)
    if n < 2:
        return np.nan

    # SS_res / SS_tot using accumulations (same as firmware)
    err = y_pred_1d - y_true_1d
    ss_res = float(np.sum(err * err))
    sum_y = float(np.sum(y_true_1d))
    sum_y_sq = float(np.sum(y_true_1d * y_true_1d))
    ss_tot = sum_y_sq - (sum_y * sum_y) / float(n)

    # The firmware uses a 1e-6 threshold
    if ss_tot <= 1e-6:
        return np.nan

    r2 = 1.0 - (ss_res / ss_tot)
    if not np.isfinite(r2):
        return np.nan
    return float(r2)

def _metrics_like_firmware(y_true_2d, y_pred_2d):
    # T_in
    err_T = y_pred_2d[:, 0] - y_true_2d[:, 0]
    mae_T = np.mean(np.abs(err_T))
    mse_T = np.mean(err_T ** 2)
    rmse_T = np.sqrt(mse_T)
    r2_T = _r2_like_firmware(y_true_2d[:, 0], y_pred_2d[:, 0])

    # H_in
    err_H = y_pred_2d[:, 1] - y_true_2d[:, 1]
    mae_H = np.mean(np.abs(err_H))
    mse_H = np.mean(err_H ** 2)
    rmse_H = np.sqrt(mse_H)
    r2_H = _r2_like_firmware(y_true_2d[:, 1], y_pred_2d[:, 1])

    mae = 0.5 * (mae_T + mae_H)
    rmse = np.sqrt(0.5 * (mse_T + mse_H))
    r2 = 0.5 * (r2_T + r2_H) if (np.isfinite(r2_T) and np.isfinite(r2_H)) else np.nan

    return mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H

rolling_rows = []
n_test = y_test_abs_final.shape[0]

# --- Gating identical to the firmware (metrics.cpp) ---
invoked_mask = np.ones(n_test, dtype=bool)
is_rollover_mask = np.ones(n_test, dtype=bool)
accepted_idxs = [i for i in range(n_test) if (invoked_mask[i] and is_rollover_mask[i])]

if len(accepted_idxs) >= ROLLING_N:
    for k in range(ROLLING_N - 1, len(accepted_idxs)):
        end = accepted_idxs[k]
        start = accepted_idxs[k - ROLLING_N + 1]
        win = accepted_idxs[k - ROLLING_N + 1 : k + 1]

        yt_w = y_test_abs_final[win]
        yp_w = y_pred_abs[win]

        mae, rmse, r2, mae_T, rmse_T, r2_T, mae_H, rmse_H, r2_H = _metrics_like_firmware(yt_w, yp_w)

        rolling_rows.append({
            "window_start": int(start),
            "window_end": int(end),
            "datetime_end": df.iloc[idx_test[end]]["datetime"],
            "N": int(ROLLING_N),
            "MAE": mae,
            "RMSE": rmse,
            "R2": r2,
            "MAE_T": mae_T,
            "RMSE_T": rmse_T,
            "R2_T": r2_T,
            "MAE_H": mae_H,
            "RMSE_H": rmse_H,
            "R2_H": r2_H,
        })

    df_roll = pd.DataFrame(rolling_rows)
    df_roll_24 = df_roll.tail(ROLLING_EXPORT_ROWS).copy().reset_index(drop=True)

    # Minimum stream of samples in the test space to reproduce the 24 rolling24 windows.
    pos_of = {v: i for i, v in enumerate(accepted_idxs)}
    sample_set = set()
    for _, row in df_roll_24.iterrows():
        end_v = int(row["window_end"])
        pos = pos_of.get(end_v, None)
        if pos is None:
            continue
        win = accepted_idxs[max(0, pos - ROLLING_N + 1):pos + 1]
        for v in win:
            sample_set.add(int(v))

    sample_vals = sorted(sample_set, key=lambda v: pos_of.get(v, 1_000_000_000))
    ROLL24_24_PACK_IDXS = list(sample_vals)
    DBG_MODEL_VALID_IDXS = list(ROLL24_24_PACK_IDXS[-ROLLING_N:])
    sample_df_indices = [int(idx_test[v]) for v in sample_vals if 0 <= v < len(idx_test)]

    # Replay 2+47 aligned with the MLP: 2 seed rows + 47 real rows from the rolling24 stream.
    REPLAY_SEED_ROWS = 2
    REPLAY_REAL_ROWS = 47
    REPLAY_TOTAL_RAW_ROWS = REPLAY_SEED_ROWS + REPLAY_REAL_ROWS

    if len(sample_vals) != REPLAY_REAL_ROWS:
        raise ValueError(
            f"Expected exactly {REPLAY_REAL_ROWS} Rolling(24) replay rows, obtained {len(sample_vals)}"
        )
    if not np.all(np.diff(np.asarray(sample_vals, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must be contiguous in the test space to export 2+47")
    if not np.all(np.diff(np.asarray(sample_df_indices, dtype=np.int64)) == 1):
        raise ValueError("Rolling(24) replay rows must map to contiguous rows in the processed dataframe")

    first_real_df_idx = int(sample_df_indices[0])
    last_real_df_idx  = int(sample_df_indices[-1])

    if first_real_df_idx < REPLAY_SEED_ROWS:
        raise ValueError(
            f"{REPLAY_SEED_ROWS} previous rows are required to export the replay seeds, first_real_df_idx={first_real_df_idx}"
        )

    raw_replay_df_indices = list(range(first_real_df_idx - REPLAY_SEED_ROWS, last_real_df_idx + 1))
    if len(raw_replay_df_indices) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw replay rows, got {len(raw_replay_df_indices)}"
        )

    # Map the processed 2+47 block back to the raw_real timeline.
    raw_replay_row_ids = df_processed_row_ids[np.asarray(raw_replay_df_indices, dtype=np.int64)]
    if not np.all(np.diff(raw_replay_row_ids) == 1):
        raise ValueError("The replay 2+47 rows must map to contiguous rows in raw_real")

    raw_replay_start_row_id = int(raw_replay_row_ids[0])
    raw_replay_end_row_id   = int(raw_replay_row_ids[-1])
    raw_replay_real = df_raw_real.iloc[raw_replay_start_row_id:raw_replay_end_row_id + 1].copy().reset_index(drop=True)
    if len(raw_replay_real) != REPLAY_TOTAL_RAW_ROWS:
        raise ValueError(
            f"Expected {REPLAY_TOTAL_RAW_ROWS} raw-real replay rows, got {len(raw_replay_real)}"
        )

    if raw_replay_start_row_id <= 0:
        raise ValueError("A previous raw_real row is required before the replay block to export the initial EMA state of H_in")

    hin_ema_prev = float(df.iloc[first_real_df_idx - REPLAY_SEED_ROWS - 1]["H_in"])

    # 47 processed real rows used in the analysis
    df_samples_roll = df.iloc[sample_df_indices].copy().reset_index(drop=True)

    # 49 raw_real rows (2 seeds + 47 real rows) for header/mirror export
    df_samples_replay_raw = raw_replay_real.copy().reset_index(drop=True)

    # 49 canonical processed replay rows: reconstructed in float32 from
    # the raw 2+47 block plus the EMA seed, to mirror the effective firmware path.
    df_samples_replay_proc = _build_replay_canonical_processed(
        raw_replay_df=df_samples_replay_raw,
        ema_prev=hin_ema_prev,
        alpha=HIN_EMA_ALPHA,
    )

    # epoch for replay/analysis exports
    for _df_export in (df_samples_roll, df_samples_replay_raw, df_samples_replay_proc):
        if "datetime" in _df_export.columns:
            try:
                _df_export["epoch"] = pd.to_datetime(_df_export["datetime"]).astype('int64') // 10**9
            except Exception:
                _df_export["epoch"] = 0
        else:
            _df_export["epoch"] = 0

    # -----------------------------
    # Export rolling24 metrics (CSV + XLSX)
    # -----------------------------
    df_roll_24_export = df_roll_24.copy()
    for c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]:
        if c in df_roll_24_export.columns:
            df_roll_24_export[c] = pd.to_numeric(df_roll_24_export[c], errors="coerce")

    df_roll_24_export = df_roll_24_export.round(4)

    csv_roll = results_dir / "environment_quantized_metrics_rolling24_lstm.csv"
    xlsx_roll = results_dir / "environment_quantized_metrics_rolling24_lstm.xlsx"
    df_roll_24_export.to_csv(csv_roll, index=False, encoding="utf-8-sig", float_format="%.4f")
    df_roll_24_export.to_excel(xlsx_roll, index=False)

    try:
        wb = load_workbook(xlsx_roll)
        ws = wb.active
        num_fmt = "0.0000"
        num_cols = {c: (list(df_roll_24_export.columns).index(c) + 1) for c in df_roll_24_export.columns if c in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
        for _, col_idx in num_cols.items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = num_fmt
        wb.save(xlsx_roll)
    except Exception as _e:
        print("[WARN] Could not apply XLSX formatting to rolling24:", _e)

    print(f"[INFO] rolling24 exported: {csv_roll}")
    print(f"[INFO] rolling24 exported: {xlsx_roll}")

    def _autosize_xlsx(_xlsx_path):
        try:
            _wb = load_workbook(_xlsx_path)
            _ws = _wb.active
            for _col_idx, _col_cells in enumerate(_ws.columns, 1):
                _max_length = max(len(str(_cell.value)) if _cell.value is not None else 0 for _cell in _col_cells)
                _ws.column_dimensions[get_column_letter(_col_idx)].width = _max_length + 2
            _wb.save(_xlsx_path)
        except Exception as _e:
            print(f"[WARN] Could not auto-adjust XLSX {_xlsx_path}:", _e)

    # -----------------------------
    # samples_rolling24 (not raw; analysis)
    # -----------------------------
    samples_csv_path  = results_dir / "environment_quantized_samples_rolling24_lstm.csv"
    samples_xlsx_path = results_dir / "environment_quantized_samples_rolling24_lstm.xlsx"
    if "datetime" in df_samples_roll.columns:
        df_samples_roll["datetime"] = df_samples_roll["datetime"].astype(str)
    df_samples_roll.to_csv(samples_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    df_samples_roll.to_excel(samples_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(samples_xlsx_path)
    print(f"[INFO] rolling24 samples exported: {samples_csv_path} | rows={len(df_samples_roll)}")
    print(f"[INFO] rolling24 samples exported: {samples_xlsx_path} | rows={len(df_samples_roll)}")

    # -----------------------------
    # replay raw 2+47 (header + raw mirror)
    # -----------------------------
    replay_raw_export = df_samples_replay_raw.copy()
    if "datetime" in replay_raw_export.columns:
        replay_raw_export["datetime"] = replay_raw_export["datetime"].astype(str)

    replay_raw_csv_path  = results_dir / "environment_quantized_samples_replay_raw_2plus47_lstm.csv"
    replay_raw_xlsx_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_lstm.xlsx"
    replay_raw_export.to_csv(replay_raw_csv_path, index=False, encoding="utf-8-sig", float_format="%.2f")
    replay_raw_export.to_excel(replay_raw_xlsx_path, index=False, float_format="%.2f")
    _autosize_xlsx(replay_raw_xlsx_path)
    print(f"[INFO] replay raw 2+47 exported: {replay_raw_csv_path} | rows={len(replay_raw_export)}")
    print(f"[INFO] replay raw 2+47 exported: {replay_raw_xlsx_path} | rows={len(replay_raw_export)}")

    # -----------------------------
    # replay reference 2+47 (mixed: raw data + Python references)
    # -----------------------------
    replay_reference = pd.DataFrame({
        "replay_idx": np.arange(REPLAY_TOTAL_RAW_ROWS, dtype=np.int64),
        "is_seed": [1 if i < REPLAY_SEED_ROWS else 0 for i in range(REPLAY_TOTAL_RAW_ROWS)],
        "epoch": df_samples_replay_raw["epoch"].astype(np.int64),
        "datetime": df_samples_replay_raw["datetime"].astype(str),
        "T_out_raw": df_samples_replay_raw["T_out"].astype(np.float32),
        "H_out_raw": df_samples_replay_raw["H_out"].astype(np.float32),
        "T_in_raw": df_samples_replay_raw["T_in"].astype(np.float32),
        "H_in_raw": df_samples_replay_raw["H_in"].astype(np.float32),
        "T_in_ref": df_samples_replay_proc["T_in"].astype(np.float32),
        "H_in_filt_ref": df_samples_replay_proc["H_in"].astype(np.float32),
        "T_in_lag1_ref": df_samples_replay_proc["T_in_lag1"].astype(np.float32),
        "H_in_lag1_ref": df_samples_replay_proc["H_in_lag1"].astype(np.float32),
        "T_in_lag2_ref": df_samples_replay_proc["T_in_lag2"].astype(np.float32),
        "H_in_lag2_ref": df_samples_replay_proc["H_in_lag2"].astype(np.float32),
    })

    replay_reference["T_in_ground truth"] = np.nan
    replay_reference["T_in_pred_python"] = np.nan
    replay_reference["H_in_ground truth"] = np.nan
    replay_reference["H_in_pred_python"] = np.nan

    real_rows_slice = slice(REPLAY_SEED_ROWS, REPLAY_TOTAL_RAW_ROWS)
    y_true_roll_ref = y_test_abs_final[np.asarray(sample_vals, dtype=np.int64)]
    y_pred_roll_ref = y_pred_abs[np.asarray(sample_vals, dtype=np.int64)]
    replay_reference.loc[real_rows_slice, "T_in_ground truth"] = y_true_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "T_in_pred_python"] = y_pred_roll_ref[:, 0]
    replay_reference.loc[real_rows_slice, "H_in_ground truth"] = y_true_roll_ref[:, 1]
    replay_reference.loc[real_rows_slice, "H_in_pred_python"] = y_pred_roll_ref[:, 1]

    replay_ref_csv_path  = results_dir / "environment_quantized_replay_reference_2plus47_lstm.csv"
    replay_ref_xlsx_path = results_dir / "environment_quantized_replay_reference_2plus47_lstm.xlsx"
    replay_reference.to_csv(replay_ref_csv_path, index=False, encoding="utf-8-sig", float_format="%.4f")
    replay_reference.to_excel(replay_ref_xlsx_path, index=False, float_format="%.4f")
    _autosize_xlsx(replay_ref_xlsx_path)
    print(f"[INFO] replay reference 2+47 exported: {replay_ref_csv_path} | rows={len(replay_reference)}")
    print(f"[INFO] replay reference 2+47 exported: {replay_ref_xlsx_path} | rows={len(replay_reference)}")

    # -----------------------------
    # DBG_REPLAY_CSV reference (stage-wise standardized semantics)
    # -----------------------------
    dbg_replay_epochs = df_samples_roll["epoch"].astype(np.int64).tolist()
    dbg_replay_canonical_rows = df_samples_replay_proc.iloc[REPLAY_SEED_ROWS:].copy().reset_index(drop=True)
    dbg_replay_reference = _build_dbg_replay_reference(
        sample_vals,
        dbg_replay_epochs,
        canonical_rows=dbg_replay_canonical_rows,
        tprev_override=np.asarray(dbg_replay_canonical_rows["T_in_lag1"], dtype=np.float32),
        hprev_override=np.asarray(dbg_replay_canonical_rows["H_in_lag1"], dtype=np.float32),
        y_true_abs_override=np.asarray(dbg_replay_canonical_rows[["T_in", "H_in"]], dtype=np.float32),
    )

    dbg_replay_csv_path = results_dir / "environment_quantized_dbg_replay_reference_lstm.csv"
    dbg_replay_xlsx_path = results_dir / "environment_quantized_dbg_replay_reference_lstm.xlsx"
    dbg_replay_reference.to_csv(dbg_replay_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_replay_reference.to_excel(dbg_replay_xlsx_path, index=False, float_format="%.8f")
    _autosize_xlsx(dbg_replay_xlsx_path)
    print(f"[INFO] DBG_REPLAY_CSV reference exported: {dbg_replay_csv_path} | rows={len(dbg_replay_reference)}")
    print(f"[INFO] DBG_REPLAY_CSV reference exported: {dbg_replay_xlsx_path} | rows={len(dbg_replay_reference)}")

    # -----------------------------
    # model I/O reference (immediate tensors that enter/exit the model)
    # -----------------------------
    dbg_model_valid_idxs = np.asarray(DBG_MODEL_VALID_IDXS, dtype=np.int64) if "DBG_MODEL_VALID_IDXS" in globals() else np.arange(max(0, len(X_test) - ROLLING_N), len(X_test), dtype=np.int64)
    if dbg_model_valid_idxs.size > ROLLING_N:
        dbg_model_valid_idxs = dbg_model_valid_idxs[-ROLLING_N:]

    dbg_model_df_indices = [int(idx_test[v]) for v in dbg_model_valid_idxs]
    dbg_model_epochs = []
    for _df_idx in dbg_model_df_indices:
        _dt_dbg = pd.to_datetime(df.iloc[_df_idx]["datetime"])
        dbg_model_epochs.append(int(_dt_dbg.value // 10**9))

    dbg_model_replay_proc = _build_replay_canonical_processed(
        raw_replay_df=df_samples_replay_raw,
        ema_prev=hin_ema_prev,
        alpha=HIN_EMA_ALPHA,
    )

    dbg_model_replay_canonical = _extract_canonical_replay_debug_inputs(
        sample_df_indices=dbg_model_df_indices,
        replay_processed_df=dbg_model_replay_proc,
        replay_processed_df_indices=raw_replay_df_indices,
        raw_replay_df=df_samples_replay_raw,
    )

    dbg_model_capture = _predict_tflite(X_test[dbg_model_valid_idxs], capture_io=True)
    dbg_model_input_payload = dbg_model_capture["input_payload"]
    dbg_model_input_float = dbg_model_capture["input_float"]
    dbg_model_output_payload = dbg_model_capture["output_payload"]
    dbg_model_output_float = dbg_model_capture["output_float"]
    dbg_model_output_raw = dbg_model_capture["output_raw"]

    dbg_model_pre_clip = np.clip(
        dbg_model_replay_canonical["pre_phys"],
        np.asarray(scaler_X.data_min_, dtype=np.float32),
        np.asarray(scaler_X.data_max_, dtype=np.float32),
    ).astype(np.float32, copy=False)
    dbg_model_pre_scaled = np.asarray(X_test[dbg_model_valid_idxs], dtype=np.float32).reshape(-1, SEQ_LEN, len(features))

    dbg_model_input_reference = _build_dbg_model_input_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_df_indices=dbg_model_df_indices,
        sample_epochs=dbg_model_epochs,
        x_raw_seqs=dbg_model_replay_canonical["pre_phys"],
        x_clip_seqs=dbg_model_pre_clip,
        x_scaled_seqs=dbg_model_pre_scaled,
        x_tensor_payload_seqs=dbg_model_input_payload,
        x_tensor_float_seqs=dbg_model_input_float,
        y_true_abs_rows=dbg_model_replay_canonical["y_abs"],
    )

    dbg_model_output_reference = _build_dbg_model_output_reference(
        sample_test_positions=dbg_model_valid_idxs,
        sample_epochs=dbg_model_epochs,
        output_payload=dbg_model_output_payload,
        output_float=dbg_model_output_float,
        tprev_override=dbg_model_replay_canonical["tprev"],
        hprev_override=dbg_model_replay_canonical["hprev"],
    )
    dbg_model_output_raw_reference = _build_dbg_model_output_raw_reference(
        sample_epochs=dbg_model_epochs,
        output_raw_records=dbg_model_output_raw,
    )

    preprocess_raw_reference, preprocess_smooth_reference = _build_preprocess_sample_reference(
        raw_df=df_samples_replay_raw,
        smooth_df=dbg_model_replay_proc,
        seed_rows=REPLAY_SEED_ROWS,
    )
    preprocess_raw_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_raw"],
        field_names=["Tout_raw", "Hout_raw", "Tin_raw", "Hin_raw"],
    )
    preprocess_smooth_window_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_smooth"],
        field_names=["Tout_smooth", "Hout_smooth", "Tin_smooth", "Hin_smooth"],
    )
    preprocess_lags_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_lags"],
        field_names=["Tin_lag1", "Hin_lag1", "Tout_lag1", "Hout_lag1", "Tin_lag2", "Hin_lag2"],
    )
    preprocess_time_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_time"],
        field_names=["epoch_ref", "sin_hour", "cos_hour", "weekday", "month"],
    )
    preprocess_phys_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_replay_canonical["pre_phys"],
        field_names=[
            "f00_Tout_phys", "f01_Hout_phys", "f02_Tin_lag1_phys", "f03_Hin_lag1_phys",
            "f04_Tout_lag1_phys", "f05_Hout_lag1_phys", "f06_Tin_lag2_phys", "f07_Hin_lag2_phys",
            "f08_sin_hour_phys", "f09_cos_hour_phys", "f10_weekday_phys", "f11_month_phys",
        ],
    )
    preprocess_clip_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_clip,
        field_names=[
            "f00_Tout_clip", "f01_Hout_clip", "f02_Tin_lag1_clip", "f03_Hin_lag1_clip",
            "f04_Tout_lag1_clip", "f05_Hout_lag1_clip", "f06_Tin_lag2_clip", "f07_Hin_lag2_clip",
            "f08_sin_hour_clip", "f09_cos_hour_clip", "f10_weekday_clip", "f11_month_clip",
        ],
    )
    preprocess_scaled_reference = _build_preprocess_stage_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        data_seqs=dbg_model_pre_scaled,
        field_names=[
            "f00_Tout_scaled", "f01_Hout_scaled", "f02_Tin_lag1_scaled", "f03_Hin_lag1_scaled",
            "f04_Tout_lag1_scaled", "f05_Hout_lag1_scaled", "f06_Tin_lag2_scaled", "f07_Hin_lag2_scaled",
            "f08_sin_hour_scaled", "f09_cos_hour_scaled", "f10_weekday_scaled", "f11_month_scaled",
        ],
    )
    preprocess_model_in_reference = _build_preprocess_model_in_reference(
        sample_epochs=dbg_model_epochs,
        step_epoch_seqs=dbg_model_replay_canonical["step_epoch"],
        x_payload_seqs=dbg_model_input_payload,
        x_float_seqs=dbg_model_input_float,
    )

    dbg_model_input_csv_path = results_dir / "environment_quantized_dbg_model_input_reference_lstm.csv"
    dbg_model_input_xlsx_path = results_dir / "environment_quantized_dbg_model_input_reference_lstm.xlsx"
    dbg_model_output_csv_path = results_dir / "environment_quantized_dbg_model_output_reference_lstm.csv"
    dbg_model_output_xlsx_path = results_dir / "environment_quantized_dbg_model_output_reference_lstm.xlsx"
    dbg_model_output_raw_csv_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_lstm.csv"
    dbg_model_output_raw_xlsx_path = results_dir / "environment_quantized_dbg_model_output_raw_reference_lstm.xlsx"
    dbg_model_workbook_path = results_dir / "environment_quantized_model_io_debug_reference_lstm.xlsx"
    preprocess_workbook_path = results_dir / "environment_quantized_preprocess_debug_reference_lstm.xlsx"

    dbg_model_input_reference.to_csv(dbg_model_input_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_input_reference.to_excel(dbg_model_input_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_reference.to_csv(dbg_model_output_csv_path, index=False, encoding="utf-8-sig", float_format="%.8f")
    dbg_model_output_reference.to_excel(dbg_model_output_xlsx_path, index=False, float_format="%.8f")
    dbg_model_output_raw_reference.to_csv(dbg_model_output_raw_csv_path, index=False, encoding="utf-8-sig")
    dbg_model_output_raw_reference.to_excel(dbg_model_output_raw_xlsx_path, index=False)
    with pd.ExcelWriter(preprocess_workbook_path, engine="openpyxl") as _writer_pre_dbg:
        preprocess_raw_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_pre_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
    with pd.ExcelWriter(dbg_model_workbook_path, engine="openpyxl") as _writer_model_dbg:
        preprocess_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_CSV", float_format="%.8f")
        preprocess_smooth_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_CSV", float_format="%.8f")
        preprocess_raw_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_RAW_WINDOW_CSV", float_format="%.8f")
        preprocess_smooth_window_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SMOOTH_WINDOW_CSV", float_format="%.8f")
        preprocess_lags_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_LAGS_CSV", float_format="%.8f")
        preprocess_time_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_TIME_CSV", float_format="%.8f")
        preprocess_phys_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_PHYS_CSV", float_format="%.8f")
        preprocess_clip_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_CLIP_CSV", float_format="%.8f")
        preprocess_scaled_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_SCALED_CSV", float_format="%.8f")
        preprocess_model_in_reference.to_excel(_writer_model_dbg, index=False, sheet_name="PRE_MODEL_IN_CSV", float_format="%.8f")
        dbg_replay_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_REPLAY_CSV", float_format="%.8f")
        dbg_model_input_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_IN_CSV", float_format="%.8f")
        dbg_model_output_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_CSV", float_format="%.8f")
        dbg_model_output_raw_reference.to_excel(_writer_model_dbg, index=False, sheet_name="DBG_MODEL_OUT_RAW_CSV")

    _autosize_xlsx(dbg_model_input_xlsx_path)
    _autosize_xlsx(dbg_model_output_xlsx_path)
    _autosize_xlsx(dbg_model_output_raw_xlsx_path)
    _autosize_xlsx(preprocess_workbook_path)
    _autosize_xlsx(dbg_model_workbook_path)

    print(f"[INFO] model input reference exported: {dbg_model_input_csv_path} | rows={len(dbg_model_input_reference)}")
    print(f"[INFO] model input reference exported: {dbg_model_input_xlsx_path} | rows={len(dbg_model_input_reference)}")
    print(f"[INFO] model output reference exported: {dbg_model_output_csv_path} | rows={len(dbg_model_output_reference)}")
    print(f"[INFO] model output reference exported: {dbg_model_output_xlsx_path} | rows={len(dbg_model_output_reference)}")
    print(f"[INFO] model output RAW reference exported: {dbg_model_output_raw_csv_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f"[INFO] model output RAW reference exported: {dbg_model_output_raw_xlsx_path} | rows={len(dbg_model_output_raw_reference)}")
    print(f"[INFO] preprocess workbook exported: {preprocess_workbook_path}")
    print(f"[INFO] model I/O workbook exported: {dbg_model_workbook_path}")

    hdr_path = results_dir / "environment_quantized_samples_replay_raw_2plus47_lstm.h"

    def _fmt_f(x):
        try:
            xf = float(x)
        except Exception:
            xf = float('nan')
        if not np.isfinite(xf):
            return "NAN"
        if abs(xf) < 0.0005:
            xf = 0.0
        return f"{xf:.6f}f"

    def _get_col(df_, names):
        for n in names:
            if n in df_.columns:
                return n
        return None

    c_tin  = _get_col(df_samples_replay_raw, ["Tin", "T_in", "tin", "T_IN"])
    c_hin  = _get_col(df_samples_replay_raw, ["Hin", "H_in", "hin", "H_IN"])
    c_tout = _get_col(df_samples_replay_raw, ["Tout", "T_out", "tout", "T_OUT"])
    c_hout = _get_col(df_samples_replay_raw, ["Hout", "H_out", "hout", "H_OUT"])

    with open(hdr_path, "w", encoding="utf-8") as hf:
        hf.write("#pragma once\n")
        hf.write("#include <stdint.h>\n\n")
        hf.write("typedef struct {\n")
        hf.write("  uint32_t epoch;\n")
        hf.write("  float T_out; float H_out; float T_in; float H_in_raw;\n")
        hf.write("} liteml_sample_raw_t;\n\n")

        hf.write(f"static const uint16_t LITEML_REPLAY_SEED_ROWS = {REPLAY_SEED_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_REAL_ROWS = {REPLAY_REAL_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_REPLAY_TOTAL_RAW_ROWS = {REPLAY_TOTAL_RAW_ROWS};\n")
        hf.write(f"static const uint16_t LITEML_ROLL24_24_N_SAMPLES = {len(df_samples_replay_raw)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_ALPHA = {_fmt_f(HIN_EMA_ALPHA)};\n")
        hf.write(f"static const float LITEML_HIN_EMA_PREV = {_fmt_f(hin_ema_prev)};\n")
        hf.write("static const liteml_sample_raw_t LITEML_ROLL24_24_SAMPLES[] = {\n")

        if not all([c_tin, c_hin, c_tout, c_hout]):
            hf.write("  // [WARN] Expected columns were not found in raw_real.\n")
        else:
            for i in range(len(df_samples_replay_raw)):
                epoch = int(df_samples_replay_raw.loc[i, "epoch"]) if "epoch" in df_samples_replay_raw.columns else 0
                Tout = df_samples_replay_raw.loc[i, c_tout]
                Hout = df_samples_replay_raw.loc[i, c_hout]
                Tin  = df_samples_replay_raw.loc[i, c_tin]
                Hin  = df_samples_replay_raw.loc[i, c_hin]

                if i == 0:
                    comment = " // seed prev2 (t-2)"
                elif i == 1:
                    comment = " // seed prev  (t-1)"
                else:
                    comment = f" // real row {i - REPLAY_SEED_ROWS:02d}"

                hf.write(
                    f"  {{ {epoch}u, {_fmt_f(Tout)}, {_fmt_f(Hout)}, {_fmt_f(Tin)}, {_fmt_f(Hin)} }},{comment}\n"
                )

        hf.write("};\n")

    print(f"[INFO] firmware header exported (2+47 raw replay + EMA seed): {hdr_path}")
else:
    print("[WARN] There are not enough samples to compute rolling N=24.")

# ========================= ABSOLUTE METRICS (aggregate and per variable) =========================
mse  = mean_squared_error(y_test_abs_final, y_pred_abs)
rmse = np.sqrt(mse)
mae  = mean_absolute_error(y_test_abs_final, y_pred_abs)
r2   = r2_score(y_test_abs_final, y_pred_abs)

mse_t  = mean_squared_error(y_test_abs_final[:,0], y_pred_abs[:,0])
rmse_t = np.sqrt(mse_t)
mae_t  = mean_absolute_error(y_test_abs_final[:,0], y_pred_abs[:,0])
r2_t   = r2_score(y_test_abs_final[:,0], y_pred_abs[:,0])

mse_h  = mean_squared_error(y_test_abs_final[:,1], y_pred_abs[:,1])
rmse_h = np.sqrt(mse_h)
mae_h  = mean_absolute_error(y_test_abs_final[:,1], y_pred_abs[:,1])
r2_h   = r2_score(y_test_abs_final[:,1], y_pred_abs[:,1])

print("\n Results (original scale - aggregate):")
mse_status = "MSE within threshold" if mse <= 2.125 else "MSE above threshold"
rmse_status = "RMSE within threshold" if rmse <= 1.458 else "RMSE above threshold"
mae_status = "MAE within threshold" if mae <= 0.925 else "MAE above threshold"
r2_status = "R² within threshold" if r2 >= 0.8 else "R² below threshold"
print(f"MSE  = {mse:.4f}   {mse_status}")
print(f"RMSE = {rmse:.4f}   {rmse_status}")
print(f"MAE  = {mae:.4f}   {mae_status}")
print(f"R²   = {r2:.4f}   {r2_status}")

print("\n Results (by variable):")
print(f"[T_in]  MSE={mse_t:.4f} RMSE={rmse_t:.4f} MAE={mae_t:.4f} R²={r2_t:.4f}")
print(f"[H_in]  MSE={mse_h:.4f} RMSE={rmse_h:.4f} MAE={mae_h:.4f} R²={r2_h:.4f}")

# ========================= Thresholds e status =========================
THRESH = {
    "T":    {"rmse": 0.50, "mae": 0.35, "r2": 0.88},
    "H":    {"rmse": 2.00, "mae": 1.50, "r2": 0.80},
    "AGGR": {"r2": 0.80}
}
def ok(b): return "within threshold" if b else "outside threshold"

status_T = {"RMSE": ok(rmse_t <= THRESH["T"]["rmse"]),
            "MAE" : ok(mae_t  <= THRESH["T"]["mae"]),
            "R2"  : ok(r2_t   >= THRESH["T"]["r2"])}
status_H = {"RMSE": ok(rmse_h <= THRESH["H"]["rmse"]),
            "MAE" : ok(mae_h  <= THRESH["H"]["mae"]),
            "R2"  : ok(r2_h   >= THRESH["H"]["r2"])}
status_AGGR = {"R2": ok(r2 >= THRESH["AGGR"]["r2"])}

print("\n Results (original scale - by variable):")
print(f"[T_in]  MSE={mse_t:.4f}  RMSE={rmse_t:.3f}°C {status_T['RMSE']}  "
      f"MAE={mae_t:.3f}°C {status_T['MAE']}  R²={r2_t:.3f} {status_T['R2']}")
print(f"[H_in]  MSE={mse_h:.4f}  RMSE={rmse_h:.3f} p.p. {status_H['RMSE']}  "
      f"MAE={mae_h:.3f} p.p. {status_H['MAE']}  R²={r2_h:.3f} {status_H['R2']}")

print("\n Results (original scale - aggregate T+H):")
print(f"MSE={mse:.4f}  RMSE={rmse:.3f}  MAE={mae:.3f}  R²={r2:.3f}  {status_AGGR['R2']}")

model_ok = all([
    rmse_t <= THRESH["T"]["rmse"],  mae_t <= THRESH["T"]["mae"],  r2_t >= THRESH["T"]["r2"],
    rmse_h <= THRESH["H"]["rmse"],  mae_h <= THRESH["H"]["mae"],  r2_h >= THRESH["H"]["r2"],
    r2 >= THRESH["AGGR"]["r2"]
])
print("\n Overall assessment:", "Performance thresholds satisfied" if model_ok else "Performance thresholds not satisfied")

# ========================= Time and size =========================
inference_time_total = (end_time - start_time) * 1000.0
inference_time_per_sample = inference_time_total / max(1, len(X_test))
original_model_size_kb = os.path.getsize(model_path) / 1024
model_size_kb          = os.path.getsize(quantized_path) / 1024

print(f"\nTotal inference time: {inference_time_total:.2f} ms")
print(f"Average latency per sample: {inference_time_per_sample:.4f} ms")
print(f"Original model size (.keras): {original_model_size_kb:.2f} KB")
print(f"Quantized model size: {model_size_kb:.2f} KB")

# ========================= Overfitting/underfitting diagnosis =========================
train_loss = history.history.get('loss', [])
val_loss   = history.history.get('val_loss', [])

if len(train_loss) > 0 and len(val_loss) > 0:
    n = min(5, len(train_loss))
    mean_train_loss = float(np.mean(train_loss[-n:]))
    mean_val_loss   = float(np.mean(val_loss[-n:]))
    gap = abs(mean_val_loss - mean_train_loss)
    gap_pct = (gap / mean_train_loss) * 100 if mean_train_loss > 0 else 0
    if mean_train_loss > 0.3 and mean_val_loss > 0.3: status_diag = "Underfitting detected (high training and validation losses)"
    elif mean_val_loss < mean_train_loss * 0.8:        status_diag = "Potential underfitting (validation loss significantly lower than training loss)"
    elif gap_pct > 50 or (mean_val_loss > mean_train_loss * 1.2 and gap > 0.05):
        status_diag = "Overfitting detected (large generalization gap or significant divergence)"
    elif gap_pct < 10:                                  status_diag = "Well-fitted model (generalization gap < 10%)" 
    elif gap_pct < 30:                                  status_diag = "Acceptably fitted model (generalization gap < 30%)"
    else:                                               status_diag = "Mild overfitting (moderate generalization gap)"
    print("\n Model diagnostics:")
    print(f"  Mean training loss:     {mean_train_loss:.4f}")
    print(f"  Mean validation loss:   {mean_val_loss:.4f}")
    print(f"  Absolute gap:           {gap:.4f}")
    print(f"  Generalization gap:     {gap_pct:.2f}%")
    print(f"  Status:                 {status_diag}")
else:
    status_diag = "Insufficient history for fit diagnosis."
    print("\n Model diagnostics: (not available)")

# --- Publication-grade figure saving (PNG+PDF+SVG) into quantization_graphics/ ---
def _savefig_pub(stem: str):
    """Save current matplotlib figure to quantization_graphics as PNG (600 dpi) + PDF + SVG."""
    png_path = graphics_dir / f"{stem}.png"
    pdf_path = graphics_dir / f"{stem}.pdf"
    svg_path = graphics_dir / f"{stem}.svg"
    plt.savefig(png_path, dpi=600, bbox_inches="tight")
    plt.savefig(pdf_path, bbox_inches="tight")
    plt.savefig(svg_path, bbox_inches="tight")
    print(f"[INFO] Figure saved: {png_path}")

# ========================= Plots =========================
if len(train_loss) > 0 and len(val_loss) > 0:
    plt.figure(figsize=(8, 5))
    plt.plot(train_loss, label='Training')
    plt.plot(val_loss, label='Validation')
    plt.title("Training and validation loss during QAT (LSTM)\n" + status_diag)
    plt.xlabel("Epoch"); plt.ylabel("Loss")
    plt.legend(); plt.grid(True); plt.tight_layout()
    _savefig_pub("environment_base_model_training_validation_loss_diagnosis_lstm")
    plt.close()

plt.figure(figsize=(8, 5))
labels = ["MSE", "RMSE", "MAE", "R²"]
values = [mse, rmse, mae, r2]
bars = plt.bar(labels, values)
plt.title("Evaluation metrics - INT8-quantized LSTM model")
plt.ylabel("Value"); plt.grid(axis='y', linestyle='--', alpha=0.7)
for bar in bars:
    h = bar.get_height()
    plt.text(bar.get_x() + bar.get_width() / 2, h, f"{h:.4f}", ha='center', va='bottom', fontsize=9)
plt.tight_layout(); _savefig_pub("environment_bar_metrics_quantized_lstm"); plt.close()

# ========================= Joint scatter plot (T_in + H_in) =========================
plt.figure(figsize=(6, 6))
plt.scatter(
    y_test_abs[:, 0], y_pred_abs[:, 0],
    alpha=0.5, label="Temperature (T_in)", color="blue"
)
plt.scatter(
    y_test_abs[:, 1], y_pred_abs[:, 1],
    alpha=0.5, label="Relative humidity (H_in)", color="green"
)

min_val = min(y_test_abs.min(), y_pred_abs.min())
max_val = max(y_test_abs.max(), y_pred_abs.max())
plt.plot([min_val, max_val], [min_val, max_val], 'k--')

plt.xlabel("Ground truth value")
plt.ylabel("Predicted value")
plt.title("Joint prediction scatter - quantized LSTM model")
plt.legend()
plt.grid(True)
plt.tight_layout()
_savefig_pub("environment_scatter_predictions_quantized_lstm")
plt.close()

# === Scatter plots (ground truth vs prediction) ===
plt.figure(figsize=(6,6))
plt.scatter(y_test_abs_final[:,0], y_pred_abs[:,0], s=12, alpha=0.6)
mn = float(min(y_test_abs_final[:,0].min(), y_pred_abs[:,0].min()))
mx = float(max(y_test_abs_final[:,0].max(), y_pred_abs[:,0].max()))
plt.plot([mn, mx], [mn, mx])
plt.xlabel("T_in ground truth"); plt.ylabel("T_in prediction")
plt.title(f"T_in scatter (R²={r2_t:.3f})")
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout(); _savefig_pub("environment_scatter_T_in_lstm"); plt.close()

plt.figure(figsize=(6,6))
plt.scatter(y_test_abs_final[:,1], y_pred_abs[:,1], s=12, alpha=0.6)
mn = float(min(y_test_abs_final[:,1].min(), y_pred_abs[:,1].min()))
mx = float(max(y_test_abs_final[:,1].max(), y_pred_abs[:,1].max()))
plt.plot([mn, mx], [mn, mx])
plt.xlabel("H_in ground truth"); plt.ylabel("H_in prediction")
plt.title(f"H_in scatter (R²={r2_h:.3f})")
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout(); _savefig_pub("environment_scatter_H_in_lstm"); plt.close()

# ========================= Rolling24 plots (paper-ready; same style as global) =========================
try:
    if 'df_roll_24' in globals() and isinstance(df_roll_24, pd.DataFrame) and len(df_roll_24) > 0:
        roll_end_idxs = df_roll_24["window_end"].astype(int).to_list()

        yT_true_r = y_test_abs_final[roll_end_idxs, 0]
        yT_pred_r = y_pred_abs[roll_end_idxs, 0]
        yH_true_r = y_test_abs_final[roll_end_idxs, 1]
        yH_pred_r = y_pred_abs[roll_end_idxs, 1]

        plt.figure(figsize=(6, 6))
        plt.scatter(yT_true_r, yT_pred_r, s=14, alpha=0.65)
        mn = float(min(yT_true_r.min(), yT_pred_r.min()))
        mx = float(max(yT_true_r.max(), yT_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], 'k--', linewidth=1.0)
        plt.gca().set_aspect('equal', adjustable='box')
        mae_t_r = float(np.mean(np.abs(yT_pred_r - yT_true_r)))
        rmse_t_r = float(np.sqrt(np.mean((yT_pred_r - yT_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yT_true_r)}\nMAE={mae_t_r:.3f} °C\nRMSE={rmse_t_r:.3f} °C",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (T_in, °C)")
        plt.ylabel("LSTM prediction (T_in, °C)")
        plt.title("Rolling(24): ground truth vs prediction (T_in)")
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_T_in_rolling24_lstm")
        plt.close()

        plt.figure(figsize=(6, 6))
        plt.scatter(yH_true_r, yH_pred_r, s=14, alpha=0.65)
        mn = float(min(yH_true_r.min(), yH_pred_r.min()))
        mx = float(max(yH_true_r.max(), yH_pred_r.max()))
        plt.plot([mn, mx], [mn, mx], 'k--', linewidth=1.0)
        plt.gca().set_aspect('equal', adjustable='box')
        mae_h_r = float(np.mean(np.abs(yH_pred_r - yH_true_r)))
        rmse_h_r = float(np.sqrt(np.mean((yH_pred_r - yH_true_r) ** 2)))
        plt.text(0.02, 0.98, f"N={len(yH_true_r)}\nMAE={mae_h_r:.3f} %\nRMSE={rmse_h_r:.3f} %",
                 transform=plt.gca().transAxes, va="top")
        plt.xlabel("Ground truth (H_in, %)")
        plt.ylabel("LSTM prediction (H_in, %)")
        plt.title("Rolling(24): ground truth vs prediction (H_in)")
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_scatter_H_in_rolling24_lstm")
        plt.close()

        try:
            t_end = pd.to_datetime(df_roll_24["datetime_end"]).to_list()
        except Exception:
            t_end = list(range(len(roll_end_idxs)))

        plt.figure(figsize=(10, 4))
        plt.plot(t_end, yT_true_r, label="Ground truth (dataset)")
        plt.plot(t_end, yT_pred_r, label="LSTM prediction", linestyle="--")
        plt.xlabel("Window end (time)")
        plt.ylabel("T_in (°C)")
        plt.title("Rolling(24): T_in time series (window ends)")
        plt.legend()
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_timeseries_T_in_rolling24_lstm")
        plt.close()

        plt.figure(figsize=(8, 5))
        plt.plot(t_end, yH_true_r, label="Ground truth (dataset)")
        plt.plot(t_end, yH_pred_r, label="LSTM prediction", linestyle="--")
        plt.xlabel("Window end (time)")
        plt.ylabel("H_in (%)")
        plt.title("Rolling(24): H_in time series (window ends)")
        plt.legend()
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.tight_layout()
        _savefig_pub("environment_timeseries_H_in_rolling24_lstm")
        plt.close()

except Exception as _e:
    print("[WARN] Rolling24 plots could not be generated:", _e)

# ========================= Exports =========================
X_test_laststep = X_test[:, -1, :]
X_test_orig = scaler_X.inverse_transform(X_test_laststep)

df_preds = pd.DataFrame(
    X_test_orig,
    columns=features,
)
df_preds['T_in_ground truth'] = y_test_abs_final[:, 0]
df_preds['T_in_pred'] = y_pred_abs[:, 0]
df_preds['H_in_ground truth'] = y_test_abs_final[:, 1]
df_preds['H_in_pred'] = y_pred_abs[:, 1]

corr_t = np.corrcoef(df_preds['T_in_ground truth'], df_preds['T_in_pred'])[0,1]
corr_h = np.corrcoef(df_preds['H_in_ground truth'], df_preds['H_in_pred'])[0,1]
diag_t = ("Strong positive correlation" if corr_t >= 0.9 else
          "Moderate positive correlation" if corr_t >= 0.75 else
          "Weak or negative correlation")
diag_h = ("Strong positive correlation" if corr_h >= 0.9 else
          "Moderate positive correlation" if corr_h >= 0.75 else
          "Weak or negative correlation")

# ========================= Exports =========================
_PRED_N = 47

if "ROLL24_24_PACK_IDXS" in globals() and isinstance(ROLL24_24_PACK_IDXS, (list, tuple, np.ndarray)) and len(ROLL24_24_PACK_IDXS) > 0:
    _pred_sel = np.asarray(ROLL24_24_PACK_IDXS, dtype=np.int64)
    if _pred_sel.size >= _PRED_N:
        _pred_sel = _pred_sel[-_PRED_N:]
    else:
        _pad = np.arange(max(0, len(X_test) - (_PRED_N - _pred_sel.size)), len(X_test), dtype=np.int64)
        _pred_sel = np.unique(np.concatenate([_pred_sel, _pad]))[-_PRED_N:]
else:
    _pred_sel = np.arange(max(0, len(X_test) - _PRED_N), len(X_test), dtype=np.int64)

X_pred_sel = X_test[_pred_sel]
y_true_sel = y_test_abs_final[_pred_sel]
y_pred_sel = y_pred_abs[_pred_sel]

_window_end = _pred_sel.astype(np.int64)
_window_start = (_window_end - (SEQ_LEN - 1)).astype(np.int64)
_pred_df_indices = [int(idx_test[int(_i)]) for _i in _window_end]

_datetime_end = []
for _df_idx in _pred_df_indices:
    _datetime_end.append(str(df.iloc[_df_idx]["datetime"]))

# Base dataframe with features taken DIRECTLY from the engineered timeline.
# This keeps the predictions spreadsheet under the same 1:1 provenance contract as the MLP export.
df_preds_source = df.iloc[_pred_df_indices].copy().reset_index(drop=True)
_missing_feature_cols = [c for c in features if c not in df_preds_source.columns]
if _missing_feature_cols:
    raise ValueError(f"[export] Missing columns in df_preds_source: {_missing_feature_cols}")
df_preds = df_preds_source[features].copy()

# Provenance / alignment check against the model-side last-step tensor.
# Diagnostic only; the spreadsheet export uses the engineered dataframe rows.
X_last_scaled = X_pred_sel[:, -1, :]
X_last_orig = scaler_X.inverse_transform(X_last_scaled)
_raw_feature_values = df_preds[features].values.astype(np.float32)
_feature_abs_diff = np.abs(_raw_feature_values - X_last_orig.astype(np.float32))
_feature_max_abs_diff = float(np.nanmax(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
_feature_mean_abs_diff = float(np.nanmean(_feature_abs_diff)) if _feature_abs_diff.size else 0.0
print(
    f"Feature provenance check (engineered dataframe vs. inverse-transformed model input): "
    f"max_abs_diff={_feature_max_abs_diff:.6f} | mean_abs_diff={_feature_mean_abs_diff:.6f}"
)

df_preds["T_in_ground truth"] = y_true_sel[:, 0]
df_preds["T_in_pred"] = y_pred_sel[:, 0]
df_preds["H_in_ground truth"] = y_true_sel[:, 1]
df_preds["H_in_pred"] = y_pred_sel[:, 1]

# Compatibility aliases kept for downstream diagnostics.
df_preds["T_in_real"] = df_preds["T_in_ground truth"]
df_preds["H_in_real"] = df_preds["H_in_ground truth"]

df_preds.insert(0, "datetime_end", _datetime_end)
df_preds.insert(0, "window_end", _window_end)
df_preds.insert(0, "window_start", _window_start)

_export_cols = (
    ["window_start", "window_end", "datetime_end"]
    + list(features)
    + ["T_in_ground truth", "T_in_pred", "H_in_ground truth", "H_in_pred"])

_missing_cols = [c for c in _export_cols if c not in df_preds.columns]
if _missing_cols:
    raise ValueError(f"[export] Missing columns in df_preds: {_missing_cols}")

df_preds_export = df_preds[_export_cols].copy()

num_cols = df_preds_export.select_dtypes(include=[np.number]).columns
df_preds_export[num_cols] = (np.round(df_preds_export[num_cols].astype(np.float32) * 100.0) / 100.0)

excel_path_pred = results_dir / "environment_quantized_predictions_rolling24_lstm.xlsx"
csv_path_pred   = results_dir / "environment_quantized_predictions_rolling24_lstm.csv"

df_preds_export.to_excel(excel_path_pred, index=False, float_format="%.2f")
df_preds_export.to_csv(csv_path_pred, index=False, encoding="utf-8-sig", float_format="%.2f")

print(f"[INFO] Files saved: {excel_path_pred.name} | {csv_path_pred.name} | rows={len(df_preds_export)}")

wb = load_workbook(excel_path_pred)
ws = wb.active
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
wb.save(excel_path_pred)

print(f"Temperature diagnostic: {diag_t} (r = {corr_t:.2f})")
print(f"Humidity diagnostic: {diag_h} (r = {corr_h:.2f})")

# ========================= Metrics table (CSV / Excel, pruned_model style) =========================

# =========================
# Rolling(24) firmware-style summary to store in metrics_summary
# =========================
fw_last = {k: float("nan") for k in ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]}
fw_mean = fw_last.copy()
fw_N_ALL = float("nan")
try:
    if 'df_roll' in globals() and isinstance(df_roll, pd.DataFrame) and len(df_roll) > 0:
        _cols = ["MAE","RMSE","R2","MAE_T","RMSE_T","R2_T","MAE_H","RMSE_H","R2_H"]
        _last = df_roll.iloc[-1]
        fw_last = {k: float(_last[k]) for k in _cols if k in _last.index}
        fw_mean = {k: float(df_roll[_cols].mean(numeric_only=True)[k]) for k in _cols if k in df_roll.columns}
        fw_N_ALL = float(_last.get("N", float("nan")))
except Exception as _e:
    print("[WARN] Failed to extract rolling(24) summary:", _e)

# --- Compatibility (quantized): sparsity/strip metrics (pruned_model style) ---
def _sparsity_fraction(model) -> float:
    try:
        total, zeros = 0, 0
        for layer in getattr(model, "layers", []):
            for w in layer.get_weights():
                arr = np.asarray(w)
                total += arr.size
                zeros += int(np.sum(arr == 0))
        return (zeros / total) if total > 0 else float("nan")
    except Exception:
        return float("nan")

sparsity_before = float("nan")
sparsity_after  = float("nan")
try:
    sparsity_before = _sparsity_fraction(base_model)
    try:
        from tensorflow_model_optimization.sparsity.keras import strip_pruning
        _stripped = strip_pruning(base_model)
        sparsity_after = _sparsity_fraction(_stripped)
    except Exception:
        sparsity_after = sparsity_before
except Exception:
    pass

model_size_kb = float(model_size_kb) if 'model_size_kb' in globals() else float("nan")
original_model_size_kb = float(original_model_size_kb) if 'original_model_size_kb' in globals() else float("nan")

try:
    status = status_diag
except Exception:
    status = ""

try:
    model_status = "Performance thresholds satisfied" if bool(model_ok) else "Performance thresholds not satisfied"
except Exception:
    model_status = ""

_metrics_names = [
    'MSE (normalized joint residual)',
    'RMSE (normalized joint residual)',
    'MAE (normalized joint residual)',
    'R² (normalized joint residual)',
    'MSE (original joint target)',
    'RMSE (original joint target)',
    'MAE (original joint target)',
    'R² (original joint target)',
    'MSE_T (T_in)',
    'RMSE_T (T_in)',
    'MAE_T (T_in)',
    'R²_T (T_in)',
    'MSE_H (H_in)',
    'RMSE_H (H_in)',
    'MAE_H (H_in)',
    'R²_H (H_in)',
    'N_ALL (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, last window)',
    'RMSE (rolling24 firmware, last window)',
    'R² (rolling24 firmware, last window)',
    'MAE_T (rolling24 firmware, last window)',
    'RMSE_T (rolling24 firmware, last window)',
    'R²_T (rolling24 firmware, last window)',
    'MAE_H (rolling24 firmware, last window)',
    'RMSE_H (rolling24 firmware, last window)',
    'R²_H (rolling24 firmware, last window)',
    'MAE (rolling24 firmware, test mean)',
    'RMSE (rolling24 firmware, test mean)',
    'R² (rolling24 firmware, test mean)',
    'MAE_T (rolling24 firmware, test mean)',
    'RMSE_T (rolling24 firmware, test mean)',
    'R²_T (rolling24 firmware, test mean)',
    'MAE_H (rolling24 firmware, test mean)',
    'RMSE_H (rolling24 firmware, test mean)',
    'R²_H (rolling24 firmware, test mean)',
    'Sparsity before strip (%)',
    'Sparsity after strip (%)',
    'Model size (KB)',
    'Original model size (KB)',
    'Mean training loss',
    'Mean validation loss',
    'Absolute gap',
    'Gap percentage (%)',
    'Total inference time (ms)',
    'Inference time per sample (ms)',
    'Fit Status',
    'Model Status',
]
_significados  = [
    'Mean squared error in the normalized scale for the joint residual target (ΔT_in, ΔH_in).',
    'Root mean squared error in the normalized scale.',
    'Mean absolute error in the normalized scale.',
    'Coefficient of determination in the normalized scale.',
    'Mean squared error in the original scale (°C / %RH) for the joint [T_in, H_in] target.',
    'Root mean squared error in the original scale for the joint target.',
    'Mean absolute error in the original scale for the joint target.',
    'Coefficient of determination in the original scale for the joint target.',
    'Mean squared error in the original scale for T_in.',
    'Root mean squared error in the original scale for T_in.',
    'Mean absolute error in the original scale for T_in.',
    'Coefficient of determination in the original scale for T_in.',
    'Mean squared error in the original scale for H_in.',
    'Root mean squared error in the original scale for H_in.',
    'Mean absolute error in the original scale for H_in.',
    'Coefficient of determination in the original scale for H_in.',
    'Rolling(24): number of samples (HOUR events) effectively present in the final window.',
    'Rolling(24) aggregate (T+H) in the LAST window - comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) in the LAST window - comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) in the LAST window - comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) T_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) H_in in the LAST window - comparable to the firmware log.',
    'Rolling(24) aggregate (T+H) MEAN across the test set (window average).',
    'Rolling(24) aggregate (T+H) MEAN across the test set (window average).',
    'Rolling(24) aggregate (T+H) MEAN across the test set (window average).',
    'Rolling(24) T_in MEAN across the test set (window average).',
    'Rolling(24) T_in MEAN across the test set (window average).',
    'Rolling(24) T_in MEAN across the test set (window average).',
    'Rolling(24) H_in MEAN across the test set (window average).',
    'Rolling(24) H_in MEAN across the test set (window average).',
    'Rolling(24) H_in MEAN across the test set (window average).',
    'Percentage of zero weights before applying the `strip_pruning` function.',
    'Percentage of zero weights after removing pruning operations.',
    'Final model file size in kilobytes (KB).',
    'Original model (.keras) size in kilobytes (KB).',
    'Mean loss over the last training epochs.',
    'Mean loss over the last validation epochs.',
    'Absolute difference between mean losses.',
    'Percentage gap between losses (validation vs training).',
    'Total time to infer all test samples.',
    'Mean time to infer a single sample.',
    'Fit diagnosis based on losses and generalization gap.',
    'Overall diagnosis based on the predefined MSE/RMSE/MAE/R² thresholds.',
]
_thresholds      = [
    'Lower is better.',
    'Lower is better.',
    'Lower is better.',
    'Ideally > 0.95.',
    '< 0.1 is excellent, depending on the problem.',
    '< 0.32 as a reference.',
    '< 0.30 as a reference.',
    '> 0.8 is desirable.',
    'Lower is better (T_in).',
    'Lower is better (T_in).',
    'Lower is better (T_in).',
    'Ideally > 0.8 (T_in).',
    'Lower is better (H_in).',
    'Lower is better (H_in).',
    'Lower is better (H_in).',
    'Ideally > 0.8 (H_in).',
    'Should reach 24 when the window is full (after warm-up).',
    'Lower is better (rolling24).',
    'Lower is better (rolling24).',
    'May be NaN when variance is low (same as firmware).',
    'Lower is better (rolling24 T_in).',
    'Lower is better (rolling24 T_in).',
    'May be NaN when variance is low.',
    'Lower is better (rolling24 H_in).',
    'Lower is better (rolling24 H_in).',
    'May be NaN when variance is low.',
    'Lower is better (window mean).',
    'Lower is better (window mean).',
    'May be NaN when variance is low.',
    'Lower is better (T_in window mean).',
    'Lower is better (T_in window mean).',
    'May be NaN when variance is low.',
    'Lower is better (H_in window mean).',
    'Lower is better (H_in window mean).',
    'May be NaN when variance is low.',
    'In general, > 50% for good compression gains.',
    'Ideally close to the final target sparsity.',
    'Prefer < 256 KB on constrained MCUs.',
    'Reference value for comparison with the original model.',
    'Low (for example, < 0.01).',
    'Close to the training loss.',
    '< 0.05 is good.',
    '< 10% is excellent.',
    'Lower is better.',
    '< 1 ms is ideal in TinyML.',
    "'Well-fitted model' when the generalization gap is low and losses remain stable.",
    "'Performance thresholds satisfied' when the model satisfies the predefined limits.",
]

_valores = [
    f"{mse_scaled:.4f}", f"{rmse_scaled:.4f}", f"{mae_scaled:.4f}", f"{r2_scaled:.4f}",
    f"{mse:.4f}",        f"{rmse:.4f}",        f"{mae:.4f}",        f"{r2:.4f}",
    f"{mse_t:.4f}",      f"{rmse_t:.4f}",      f"{mae_t:.4f}",      f"{r2_t:.4f}",
    f"{mse_h:.4f}",      f"{rmse_h:.4f}",      f"{mae_h:.4f}",      f"{r2_h:.4f}",
    (f"{fw_N_ALL:.0f}" if np.isfinite(fw_N_ALL) else ""),
    (f"{fw_last.get('MAE', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE', float('nan'))) else ""),
    (f"{fw_last.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE', float('nan'))) else ""),
    (f"{fw_last.get('R2', float('nan')):.4f}" if np.isfinite(fw_last.get('R2', float('nan'))) else ""),
    (f"{fw_last.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_T', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_last.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_T', float('nan'))) else ""),
    (f"{fw_last.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('MAE_H', float('nan'))) else ""),
    (f"{fw_last.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_last.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_last.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_last.get('R2_H', float('nan'))) else ""),
    (f"{fw_mean.get('MAE', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE', float('nan'))) else ""),
    (f"{fw_mean.get('R2', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_T', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_T', float('nan'))) else ""),
    (f"{fw_mean.get('R2_T', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_T', float('nan'))) else ""),
    (f"{fw_mean.get('MAE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('MAE_H', float('nan'))) else ""),
    (f"{fw_mean.get('RMSE_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('RMSE_H', float('nan'))) else ""),
    (f"{fw_mean.get('R2_H', float('nan')):.4f}" if np.isfinite(fw_mean.get('R2_H', float('nan'))) else ""),
    (f"{sparsity_before * 100:.2f} %" if np.isfinite(sparsity_before) else ""),
    (f"{sparsity_after * 100:.2f} %" if np.isfinite(sparsity_after) else ""),
    (f"{model_size_kb:.2f} KB" if np.isfinite(model_size_kb) else ""),
    (f"{original_model_size_kb:.2f} KB" if np.isfinite(original_model_size_kb) else ""),
    (f"{mean_train_loss:.4f}" if 'mean_train_loss' in globals() and np.isfinite(mean_train_loss) else ""),
    (f"{mean_val_loss:.4f}"   if 'mean_val_loss' in globals() and np.isfinite(mean_val_loss) else ""),
    (f"{gap:.4f}"             if 'gap' in globals() and np.isfinite(gap) else ""),
    (f"{gap_pct:.2f} %"       if 'gap_pct' in globals() and np.isfinite(gap_pct) else ""),
    f"{inference_time_total:.2f} ms",
    f"{inference_time_per_sample:.2f} ms",
    "",
    ""
]

_statuses = [
    "", "", "", "",
    mse_status, rmse_status, mae_status, r2_status,
    *[""] * 37,
    status,
    model_status
]

_max_len = max(len(_metrics_names), len(_valores), len(_statuses), len(_significados), len(_thresholds))
def _pad(lst):
    lst = list(lst)
    if len(lst) < _max_len:
        lst += [""] * (_max_len - len(lst))
    return lst

metrics_dist = {
    "Quantized Model Metrics": _pad(_metrics_names),
    "Value": _pad(_valores),
    "Status": _pad(_statuses),
    "Meaning": _pad(_significados),
    "Expected Values / Thresholds": _pad(_thresholds),
}

df_metrics = pd.DataFrame(metrics_dist)

excel_path_metrics = results_dir/"environment_quantized_model_metrics_summary_lstm.xlsx"
df_metrics.to_csv(results_dir/"environment_quantized_model_metrics_summary_lstm.csv", index=False, encoding="utf-8-sig")
df_metrics.to_excel(excel_path_metrics, index=False)

wb = load_workbook(excel_path_metrics); ws = wb.active
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2
wb.save(excel_path_metrics)
print(f"[INFO] File saved successfully: {excel_path_metrics}")

# ========================= Post-execution =========================
try:
    update_latest(run_dir)
except Exception as _e:
    print("[WARN] Unable to update 'latest':", _e)
try:
    write_manifest(run_dir, run=str(run_dir))
except Exception as _e:
    print("[WARN] Unable to write manifest.json:", _e)

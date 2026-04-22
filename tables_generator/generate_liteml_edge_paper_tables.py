#!/usr/bin/env python3
"""
Generate paper-ready LiteML-Edge tables from workbook folders and firmware logs.

Default execution is bootstrap-oriented and does not require path arguments when
this script is placed under:

    LiteML/utils/table_generator/

Expected default artifact locations relative to the LiteML repository root:
- utils/workbook_mlp
- utils/workbook_lstm
- utils/workbook_Conv1D_Tiny
- firmwares/environment_mlp/PlatfIO_ESP32_Wemos_mlp/logs
- firmwares/environment_mlp/PlatfIO_ESP32_Wemos_lstm/logs
- firmwares/environment_mlp/PlatfIO_ESP32_Wemos_Con1D_Tiny/logs

The output files are aligned with the manuscript structure:
- replay conformance table from workbook comparison artifacts
- energy/latency table with separate Replay and Field columns
- memory table with separate minimum free heap values for Replay and Field
- IDLE table with separate Replay and Field baseline values

The generated LaTeX fragments use booktabs-style rules and manuscript-oriented
tabular layouts. The parent manuscript should load the packages:
- booktabs
- multirow
- siunitx

Calculation summary
-------------------
Prediction/model-I/O values are read from the Excel workbooks.
Replay and Field deployment values are parsed from firmware logs.
For each model and mode, the newest valid log is selected.

Per-log aggregates are computed as follows:
- Invoke energy mean: arithmetic mean of E_inference_window(ΔE_total)
- Event energy mean: arithmetic mean of E_inference_pipeline(ΔE_total)
- Invoke time mean: arithmetic mean of t_inference
- Event time mean: arithmetic mean of t_inference_pipeline
- IDLE voltage/current/power means: arithmetic means over all [BENCH] IDLE rows
- Flash/Arena/Total footprint: values from the last [MEM] row in the log
- Minimum free heap: minimum free-heap value parsed from matching [BENCH] rows
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

MODEL_ORDER = ["MLP", "LSTM", "Conv1D Tiny"]
MODE_ORDER = ["Replay", "Field"]

MODEL_CANONICAL = {
    "mlp": "MLP",
    "lstm": "LSTM",
    "conv1d_tiny": "Conv1D Tiny",
    "conv1d tiny": "Conv1D Tiny",
    "conv1d": "Conv1D Tiny",
    "con1d_tiny": "Conv1D Tiny",
    "con1d tiny": "Conv1D Tiny",
    "con1d": "Conv1D Tiny",
}

PREDICTION_PREFIX = "predictions_metrics_vs_log_comparison"
MODEL_IO_PREFIX = "model_io_comparison"

PWR_RE = re.compile(
    r"\[PWR\] infer \| E_inference_window\(ΔE_total\)=([0-9.]+)µWh .*?"
    r"\| E_inference_pipeline\(ΔE_total\)=([0-9.]+)µWh .*?"
    r"\| t_inference=([0-9.]+)ms \| t_inference_pipeline=([0-9.]+)ms",
    re.UNICODE,
)
IDLE_RE = re.compile(
    r"\[BENCH\] IDLE .*?\| V_bus=([0-9.]+)V I_bus=([0-9.]+)mA P_bus=([0-9.]+)mW",
    re.UNICODE,
)
MEM_RE = re.compile(
    r"\[MEM\] Model=([0-9.]+)kB \(FLASH\) \| Arena=([0-9.]+)kB \(RAM\) \| Total≈([0-9.]+)kB",
    re.UNICODE,
)
HEAP_RE = re.compile(
    r"\[BENCH\] .*?heap=([0-9.]+)kB/([0-9.]+)kB \(min=([0-9.]+)kB, biggest=([0-9.]+)kB\) \| arena=([0-9.]+)kB",
    re.UNICODE,
)
MODE_TAG_RE = re.compile(r"LITEML_MODE\s*=\s*([0-9]+)")


@dataclass
class BundleRecord:
    model: str
    workbook_root: Path | None
    log_root: Path | None
    prediction_workbook: Path | None
    model_io_workbook: Path | None
    log_files: list[Path]


class ExtractionError(RuntimeError):
    pass


def normalize_model_name(text: str) -> str | None:
    """Return the canonical model label inferred from a path or filename string."""
    key = text.lower().replace("-", "_")
    for token, canonical in MODEL_CANONICAL.items():
        if token in key:
            return canonical
    return None


def parse_log_order_key(path: Path) -> tuple[int, str]:
    """Build a sortable key from the timestamp-like suffix embedded in a log filename."""
    match = re.search(r"(\d{6}-\d{6})", path.name)
    if match:
        return (int(match.group(1).replace("-", "")), path.name)
    return (0, path.name)


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ExtractionError(message)


def script_repo_root() -> Path:
    """Infer the LiteML repository root from the script location."""
    script_path = Path(__file__).resolve()
    try:
        return script_path.parents[2]
    except IndexError as exc:
        raise ExtractionError(
            "Unable to infer the LiteML repository root from the script location."
        ) from exc


def default_workbook_dirs(repo_root: Path) -> dict[str, Path]:
    """Return the default workbook folders expected for each supported model."""
    return {
        "MLP": repo_root / "utils" / "workbook_mlp",
        "LSTM": repo_root / "utils" / "workbook_lstm",
        "Conv1D Tiny": repo_root / "utils" / "workbook_Conv1D_Tiny",
    }


def default_log_dirs(repo_root: Path) -> dict[str, list[Path]]:
    """Return the default firmware log folders expected for each supported model."""
    return {
        "MLP": [
            repo_root / "firmwares" / "environment_mlp" / "PlatfIO_ESP32_Wemos_mlp" / "logs",
        ],
        "LSTM": [
            repo_root / "firmwares" / "environment_mlp" / "PlatfIO_ESP32_Wemos_lstm" / "logs",
        ],
        "Conv1D Tiny": [
            repo_root / "firmwares" / "environment_mlp" / "PlatfIO_ESP32_Wemos_Con1D_Tiny" / "logs",
        ],
    }


def scan_model_workbooks(model: str, workbook_root: Path) -> tuple[Path | None, Path | None]:
    """Find the newest prediction workbook and model-I/O workbook under one model folder."""
    if not workbook_root.exists():
        return None, None

    prediction_workbook: Path | None = None
    model_io_workbook: Path | None = None

    for path in workbook_root.rglob("*.xlsx"):
        lower_name = path.name.lower()
        if PREDICTION_PREFIX in lower_name:
            if prediction_workbook is None or path.name > prediction_workbook.name:
                prediction_workbook = path
        elif MODEL_IO_PREFIX in lower_name:
            if model_io_workbook is None or path.name > model_io_workbook.name:
                model_io_workbook = path

    return prediction_workbook, model_io_workbook


def scan_model_logs(log_roots: list[Path]) -> list[Path]:
    """Collect every .log file available under the configured log folders."""
    log_files: list[Path] = []
    for log_root in log_roots:
        if not log_root.exists():
            continue
        for path in log_root.rglob("*.log"):
            if path not in log_files:
                log_files.append(path)
    return log_files


def discover_default_bundles(repo_root: Path) -> dict[str, BundleRecord]:
    """Build model records from the standard LiteML workbook and log directories."""
    bundles: dict[str, BundleRecord] = {}
    workbook_dirs = default_workbook_dirs(repo_root)
    log_dirs = default_log_dirs(repo_root)

    for model in MODEL_ORDER:
        workbook_root = workbook_dirs[model]
        prediction_workbook, model_io_workbook = scan_model_workbooks(model, workbook_root)
        log_root_candidates = log_dirs[model]
        log_files = scan_model_logs(log_root_candidates)

        bundles[model] = BundleRecord(
            model=model,
            workbook_root=workbook_root,
            log_root=next((p for p in log_root_candidates if p.exists()), log_root_candidates[0]),
            prediction_workbook=prediction_workbook,
            model_io_workbook=model_io_workbook,
            log_files=log_files,
        )

    return bundles


def discover_fallback_bundles(search_root: Path) -> dict[str, BundleRecord]:
    """Search the repository tree for additional matching workbooks and logs.

    This fallback is used only to fill missing artifacts when the default folder map
    does not provide a complete set for a given model.
    """
    bundles: dict[str, BundleRecord] = {}

    for path in search_root.rglob("*"):
        if not path.is_file():
            continue

        model = normalize_model_name(str(path))
        if model is None:
            continue

        record = bundles.get(model)
        if record is None:
            record = BundleRecord(
                model=model,
                workbook_root=path.parent,
                log_root=path.parent,
                prediction_workbook=None,
                model_io_workbook=None,
                log_files=[],
            )
            bundles[model] = record

        lower_name = path.name.lower()
        if lower_name.endswith(".xlsx"):
            if PREDICTION_PREFIX in lower_name and record.prediction_workbook is None:
                record.prediction_workbook = path
            elif MODEL_IO_PREFIX in lower_name and record.model_io_workbook is None:
                record.model_io_workbook = path
        elif lower_name.endswith(".log") and path not in record.log_files:
            record.log_files.append(path)

    return bundles


def merge_bundle_maps(primary: dict[str, BundleRecord], fallback: dict[str, BundleRecord]) -> dict[str, BundleRecord]:
    """Merge the default artifact map with the fallback discoveries."""
    merged = dict(primary)
    for model, fallback_record in fallback.items():
        if model not in merged:
            merged[model] = fallback_record
            continue

        record = merged[model]
        if record.prediction_workbook is None:
            record.prediction_workbook = fallback_record.prediction_workbook
        if record.model_io_workbook is None:
            record.model_io_workbook = fallback_record.model_io_workbook
        for log_path in fallback_record.log_files:
            if log_path not in record.log_files:
                record.log_files.append(log_path)

    return merged


def read_rows(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    """Read all rows from one workbook sheet using data-only values."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    return list(worksheet.iter_rows(values_only=True))


def read_key_value_sheet(path: Path, sheet_name: str) -> dict[str, Any]:
    """Read a two-column Metric/Value sheet into a dictionary."""
    rows = read_rows(path, sheet_name)
    if not rows:
        return {}
    header = rows[0]
    if len(header) < 2 or header[0] != "Metric" or header[1] != "Value":
        return {}
    output: dict[str, Any] = {}
    for row in rows[1:]:
        if not row or row[0] in (None, ""):
            continue
        output[str(row[0]).strip()] = row[1]
    return output


def read_first_table(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read the first non-empty tabular region from a sheet into row dictionaries."""
    rows = read_rows(path, sheet_name)
    if not rows:
        return []
    header: list[str] | None = None
    start_idx = 0
    for idx, row in enumerate(rows):
        if row and row[0] not in (None, ""):
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            start_idx = idx + 1
            break
    if header is None:
        return []
    table: list[dict[str, Any]] = []
    for row in rows[start_idx:]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        item: dict[str, Any] = {}
        for idx, key in enumerate(header):
            if not key:
                continue
            item[key] = row[idx] if idx < len(row) else None
        table.append(item)
    return table


def detect_log_mode(path: Path) -> str | None:
    """Classify one log as Replay or Field.

    Detection priority:
    1. explicit mode tokens in the filename
    2. LITEML_MODE tag inside the log body
       - 1 => Replay
       - 2 => Field
    """
    lower_name = path.name.lower()
    if "replay" in lower_name:
        return "Replay"
    if "field" in lower_name:
        return "Field"

    text = path.read_text(encoding="utf-8", errors="ignore")
    match = MODE_TAG_RE.search(text)
    if match:
        mode_value = match.group(1)
        if mode_value == "1":
            return "Replay"
        if mode_value == "2":
            return "Field"
    return None


def choose_latest_log_by_mode(log_files: list[Path]) -> dict[str, Path]:
    """Select the newest valid log for Replay and Field independently."""
    by_mode: dict[str, list[Path]] = {"Replay": [], "Field": []}
    for path in log_files:
        mode = detect_log_mode(path)
        if mode in by_mode:
            by_mode[mode].append(path)

    selected: dict[str, Path] = {}
    for mode in MODE_ORDER:
        candidates = by_mode[mode]
        if candidates:
            selected[mode] = sorted(candidates, key=parse_log_order_key)[-1]
    return selected


def parse_prediction_artifacts(workbook: Path) -> dict[str, Any]:
    """Extract prediction agreement fields from the prediction workbook.

    Returned values are sourced as follows:
    - summary: direct key-value export from the Summary sheet
    - latest_match: the Overview row named
      "Latest valid rolling24 metrics at 4 decimals", with Summary fallback
    - prediction_line: the Overview row named
      "Prediction agreement at 2 decimals"
    """
    summary = read_key_value_sheet(workbook, "Summary")
    latest_metrics = read_first_table(workbook, "Latest_Metrics")
    overview = read_first_table(workbook, "Overview")

    latest_match = None
    prediction_line = None
    for row in overview:
        block = row.get("Block")
        if block == "Latest valid rolling24 metrics at 4 decimals":
            latest_match = row.get("Matched_Rows")
        if block == "Prediction agreement at 2 decimals":
            prediction_line = row.get("Matched_Rows")

    return {
        "summary": summary,
        "latest_metrics": latest_metrics,
        "latest_match": latest_match or summary.get("Latest metrics rounded matches"),
        "prediction_line": prediction_line,
    }


def parse_model_io_artifacts(workbook: Path) -> dict[str, Any]:
    """Extract stage-wise comparison status rows and tolerance metadata from the model-I/O workbook."""
    overview = read_first_table(workbook, "Overview")
    tolerance = read_first_table(workbook, "Tolerance_Protocol")
    by_block = {str(row.get("Block")): row for row in overview if row.get("Block")}
    return {"overview": overview, "tolerance": tolerance, "by_block": by_block}


def parse_log_metrics(path: Path) -> dict[str, Any]:
    """Parse aggregate energy, timing, memory, heap, and IDLE statistics from one firmware log."""
    text = path.read_text(encoding="utf-8", errors="ignore")
    power_rows = PWR_RE.findall(text)
    idle_rows = IDLE_RE.findall(text)
    memory_rows = MEM_RE.findall(text)
    heap_rows = HEAP_RE.findall(text)

    require(power_rows, f"No [PWR] inference rows found in {path.name}.")
    require(idle_rows, f"No [BENCH] IDLE rows found in {path.name}.")
    require(memory_rows, f"No [MEM] rows found in {path.name}.")
    require(heap_rows, f"No heap rows found in {path.name}.")

    invoke_energy_uwh = [float(row[0]) for row in power_rows]
    pipeline_energy_uwh = [float(row[1]) for row in power_rows]
    invoke_time_ms = [float(row[2]) for row in power_rows]
    pipeline_time_ms = [float(row[3]) for row in power_rows]

    idle_voltage_v = [float(row[0]) for row in idle_rows]
    idle_current_ma = [float(row[1]) for row in idle_rows]
    idle_power_mw = [float(row[2]) for row in idle_rows]

    model_flash_kb, arena_kb, total_kb = [float(value) for value in memory_rows[-1]]
    free_heap_kb = [float(row[2]) for row in heap_rows]
    reserved_heap_kb = [float(row[1]) for row in heap_rows]

    def mean(values: Iterable[float]) -> float:
        values_list = list(values)
        return sum(values_list) / len(values_list)

    return {
        "log_file": path.name,
        "invoke_events": len(invoke_energy_uwh),
        "idle_samples": len(idle_voltage_v),
        "invoke_energy_uwh_mean": mean(invoke_energy_uwh),
        "pipeline_energy_uwh_mean": mean(pipeline_energy_uwh),
        "invoke_time_ms_mean": mean(invoke_time_ms),
        "pipeline_time_ms_mean": mean(pipeline_time_ms),
        "idle_voltage_v_mean": mean(idle_voltage_v),
        "idle_current_ma_mean": mean(idle_current_ma),
        "idle_power_mw_mean": mean(idle_power_mw),
        "model_flash_kb": model_flash_kb,
        "arena_kb": arena_kb,
        "total_footprint_kb": total_kb,
        "min_free_heap_kb": min(free_heap_kb),
        "max_heap_pool_kb": max(reserved_heap_kb),
    }


def normalize_prediction_agreement(value: Any) -> str:
    """Normalize prediction agreement text to the manuscript style."""
    text = str(value).strip()
    return re.sub(r"\s*\|\s*", "; ", text)

def latex_escape(text: Any) -> str:
    return (
        str(text)
        .replace("\\", r"\textbackslash{}")
        .replace("&", r"\&")
        .replace("%", r"\%")
        .replace("_", r"\_")
        .replace("#", r"\#")
    )


def format_float(value: Any, digits: int) -> str:
    return f"{float(value):.{digits}f}"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name[:31])
        if not rows:
            worksheet["A1"] = "No rows"
            continue
        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column, value=header)
        for row_idx, row in enumerate(rows, start=2):
            for column, header in enumerate(headers, start=1):
                worksheet.cell(row=row_idx, column=column, value=row.get(header))
    workbook.save(path)


def sort_model_rows(model_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {name: idx for idx, name in enumerate(MODEL_ORDER)}
    return sorted(model_rows, key=lambda row: order.get(str(row.get("Model")), 999))


def build_replay_conformance_rows(model_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in sort_model_rows(model_rows):
        rows.append(
            {
                "Model": item["Model"],
                "Shared data pipeline": "MATCH",
                "Critical input (x*, p*)": item["Critical_Input_Status"],
                "Raw tensor dump": item["Raw_Tensor_Status"],
                "Immediate raw output (o_raw)": item["Decoded_Raw_Status"],
                "Final prediction (y*)": item["Postprocessed_Status"],
                "Predictions (2 decimals)": item["Predictions_2dp"],
                "Latest metrics (4 decimals)": item["Latest_Metrics_4dp"],
            }
        )
    return rows


def build_energy_rows(model_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in sort_model_rows(model_rows):
        rows.append(
            {
                "Model": item["Model"],
                "Replay energy (uWh)": round(item["Replay_Invoke_Energy_uWh_Mean"], 3),
                "Replay time (ms)": round(item["Replay_Invoke_Time_ms_Mean"], 2),
                "Field energy (uWh)": round(item["Field_Invoke_Energy_uWh_Mean"], 3),
                "Field time (ms)": round(item["Field_Invoke_Time_ms_Mean"], 2),
                "Replay log": item["Replay_Log_File"],
                "Field log": item["Field_Log_File"],
            }
        )
    return rows


def build_memory_rows(model_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in sort_model_rows(model_rows):
        rows.append(
            {
                "Model": item["Model"],
                "Model flash (kB)": round(item["Model_Flash_kB"], 2),
                "Tensor arena (kB)": round(item["Tensor_Arena_kB"], 2),
                "Min free heap Replay (kB)": round(item["Replay_Minimum_Free_Heap_kB"], 1),
                "Min free heap Field (kB)": round(item["Field_Minimum_Free_Heap_kB"], 1),
                "Replay log": item["Replay_Log_File"],
                "Field log": item["Field_Log_File"],
            }
        )
    return rows


def build_idle_rows(model_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in sort_model_rows(model_rows):
        rows.append(
            {
                "Model": item["Model"],
                "Replay idle current (mA)": round(item["Replay_Idle_Current_mA_Mean"], 2),
                "Replay idle power (mW)": round(item["Replay_Idle_Power_mW_Mean"], 2),
                "Field idle current (mA)": round(item["Field_Idle_Current_mA_Mean"], 2),
                "Field idle power (mW)": round(item["Field_Idle_Power_mW_Mean"], 2),
                "Replay log": item["Replay_Log_File"],
                "Field log": item["Field_Log_File"],
            }
        )
    return rows


def render_replay_conformance_tex(rows: list[dict[str, Any]]) -> str:
    """Render the replay conformance table using the manuscript tabular style."""
    lines = [
        r"\begin{table*}[!t]",
        r"\centering",
        r"\caption{Replay stage-wise conformance summary under Rolling-24 ($n{=}24$).}",
        r"\label{tab:equiv}",
        r"\scriptsize",
        r"\setlength{\tabcolsep}{4pt}",
        r"\renewcommand{\arraystretch}{1.08}",
        r"\begin{tabular}{lccccccc}",
        r"\toprule",
        r"\textbf{Model} & \textbf{Shared data} & \textbf{Critical input} & \textbf{Raw tensor} & \textbf{Decoded raw} & \textbf{Post-processed} & \textbf{Predictions} & \textbf{Latest metrics} \\",
        r" & \textbf{pipeline} & \textbf{($x^*$, $p^*$)} & \textbf{dump} & \textbf{output ($o_{\mathrm{raw}}$)} & \textbf{prediction ($y^*$)} & \textbf{(2 decimals)} & \textbf{(4 decimals)} \\",
        r"\midrule",
    ]
    for row in rows:
        lines.append(
            f"{latex_escape(row['Model'])} & {row['Shared data pipeline']} & {row['Critical input (x*, p*)']} & {row['Raw tensor dump']} & {row['Immediate raw output (o_raw)']} & {row['Final prediction (y*)']} & {latex_escape(row['Predictions (2 decimals)'])} & {latex_escape(row['Latest metrics (4 decimals)'])} \\\\"
        )
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table*}"])
    return "\n".join(lines) + "\n"

def render_energy_tex(rows: list[dict[str, Any]]) -> str:
    """Render the per-inference energy table using the manuscript tabular style."""
    lines = [
        r"\begin{table}[!t]",
        r"\centering",
        r"\begin{minipage}{0.84\columnwidth}",
        r"\caption{Per-inference energy and latency\\ ($n{=}24$) under Rolling-24.}",
        r"\label{tab:infer_energy}",
        r"\centering",
        r"\scriptsize",
        r"\setlength{\tabcolsep}{3.2pt}",
        r"\renewcommand{\arraystretch}{1.05}",
        r"\begin{tabular}{l l",
        r"S[table-format=1.3]",
        r"S[table-format=2.2]}",
        r"\toprule",
        r"\textbf{Model} & \textbf{Mode} &",
        r"{$\boldsymbol{\overline{\Delta E}}$ \textbf{($\mu$Wh)}} &",
        r"{$\boldsymbol{\overline{t}}$ \textbf{(ms)}} \\",
        r"\midrule",
    ]
    for idx, row in enumerate(rows):
        lines.append(rf"\multirow{{2}}{{*}}{{{latex_escape(row['Model'])}}}")
        lines.append(f"& Replay & {format_float(row['Replay energy (uWh)'], 3)} & {format_float(row['Replay time (ms)'], 2)} \\\\")
        lines.append(f"& Field  & {format_float(row['Field energy (uWh)'], 3)} & {format_float(row['Field time (ms)'], 2)} \\\\")
        if idx != len(rows) - 1:
            lines.append(r"\midrule")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{minipage}", r"\end{table}"])
    return "\n".join(lines) + "\n"

def render_memory_tex(rows: list[dict[str, Any]]) -> str:
    """Render the deployment footprint table using the manuscript tabular style."""
    lines = [
        r"\begin{table}[!t]",
        r"\centering",
        r"\caption{Deployment footprint and heap headroom.}",
        r"\label{tab:memory}",
        r"\scriptsize",
        r"\setlength{\tabcolsep}{3.0pt}",
        r"\renewcommand{\arraystretch}{1.05}",
        r"\begin{tabular} {",
        r"l",
        r"S[table-format=1.2]",
        r"S[table-format=1.2]",
        r"S[table-format=3.1]",
        r"S[table-format=2.1]",
        r"}",
        r"\toprule",
        r"\textbf{Model} &",
        r"\textbf{Flash (kB)} &",
        r"\textbf{Arena (kB)} &",
        r"\multicolumn{2}{c}{\textbf{Min Heap Free (kB)}} \\",
        r"\cmidrule(lr){4-5}",
        r"&",
        r"&",
        r"&",
        r"\textbf{Replay} &",
        r"\textbf{Field} \\",
        r"\midrule",
    ]
    for idx, row in enumerate(rows):
        lines.append(f"{latex_escape(row['Model'])} & {format_float(row['Model flash (kB)'], 2)} & {format_float(row['Tensor arena (kB)'], 2)} & {format_float(row['Min free heap Replay (kB)'], 1)} & {format_float(row['Min free heap Field (kB)'], 1)} \\\\")
        if idx != len(rows) - 1:
            lines.append(r"\midrule")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}"])
    return "\n".join(lines) + "\n"

def render_idle_tex(rows: list[dict[str, Any]]) -> str:
    """Render the IDLE baseline table using the manuscript tabular style."""
    lines = [
        r"\begin{table}[!t]",
        r"\centering",
        r"\caption{Baseline IDLE current and power.}",
        r"\label{tab:idle}",
        r"\scriptsize",
        r"\setlength{\tabcolsep}{3.2pt}",
        r"\renewcommand{\arraystretch}{1.05}",
        r"\begin{tabular}{ll",
        r"S[table-format=2.2]",
        r"S[table-format=3.2]}",
        r"\toprule",
        r"\textbf{Model} & \textbf{Mode} &",
        r"{$\mathbf{I_{\text{mean}}}$ \textbf{(mA)}} &",
        r"{$\mathbf{P_{\text{mean}}}$ \textbf{(mW)}} \\",
        r"\midrule",
    ]
    for idx, row in enumerate(rows):
        lines.append(rf"\multirow{{2}}{{*}}{{{latex_escape(row['Model'])}}}")
        lines.append(f"& Replay & {format_float(row['Replay idle current (mA)'], 2)} & {format_float(row['Replay idle power (mW)'], 2)} \\\\")
        lines.append(f"& Field  & {format_float(row['Field idle current (mA)'], 2)} & {format_float(row['Field idle power (mW)'], 2)} \\\\")
        if idx != len(rows) - 1:
            lines.append(r"\midrule")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}"])
    return "\n".join(lines) + "\n"

def assemble_model_record(bundle: BundleRecord) -> dict[str, Any]:
    """Combine workbook-derived agreement fields with Replay and Field log aggregates for one model."""
    require(bundle.prediction_workbook is not None, f"Prediction workbook missing for {bundle.model}.")
    require(bundle.model_io_workbook is not None, f"Model I/O workbook missing for {bundle.model}.")

    prediction = parse_prediction_artifacts(bundle.prediction_workbook)
    model_io = parse_model_io_artifacts(bundle.model_io_workbook)
    selected_logs = choose_latest_log_by_mode(bundle.log_files)
    require("Replay" in selected_logs, f"Replay log missing for {bundle.model}.")
    require("Field" in selected_logs, f"Field log missing for {bundle.model}.")

    replay_metrics = parse_log_metrics(selected_logs["Replay"])
    field_metrics = parse_log_metrics(selected_logs["Field"])

    blocks = model_io["by_block"]
    critical = blocks.get("Input / critical tensor columns", {})
    raw_tensor = blocks.get("Postprocess / raw tensor dump columns", {})
    raw_output = blocks.get("Postprocess / raw output columns", {})
    final_prediction = blocks.get("Postprocess / final prediction columns", {})

    prediction_summary = prediction["summary"]
    return {
        "Model": bundle.model,
        "Workbook_Root": str(bundle.workbook_root) if bundle.workbook_root else "",
        "Log_Root": str(bundle.log_root) if bundle.log_root else "",
        "Prediction_Workbook": str(bundle.prediction_workbook),
        "Model_IO_Workbook": str(bundle.model_io_workbook),
        "Replay_Log_File": selected_logs["Replay"].name,
        "Field_Log_File": selected_logs["Field"].name,
        "Critical_Input_Status": critical.get("Status", "UNKNOWN"),
        "Raw_Tensor_Status": raw_tensor.get("Status", "UNKNOWN"),
        "Decoded_Raw_Status": raw_output.get("Status", "UNKNOWN"),
        "Postprocessed_Status": final_prediction.get("Status", "UNKNOWN"),
        "Predictions_2dp": normalize_prediction_agreement(prediction.get("prediction_line") or prediction_summary.get("Prediction status", "UNKNOWN")),
        "Latest_Metrics_4dp": prediction.get("latest_match") or prediction_summary.get("Latest metrics rounded matches", "UNKNOWN"),
        "Prediction_T_Matches": prediction_summary.get("Prediction T matches"),
        "Prediction_H_Matches": prediction_summary.get("Prediction H matches"),
        "Replay_Invoke_Energy_uWh_Mean": replay_metrics["invoke_energy_uwh_mean"],
        "Replay_Event_Energy_uWh_Mean": replay_metrics["pipeline_energy_uwh_mean"],
        "Replay_Invoke_Time_ms_Mean": replay_metrics["invoke_time_ms_mean"],
        "Replay_Event_Time_ms_Mean": replay_metrics["pipeline_time_ms_mean"],
        "Replay_Minimum_Free_Heap_kB": replay_metrics["min_free_heap_kb"],
        "Replay_Idle_Voltage_V_Mean": replay_metrics["idle_voltage_v_mean"],
        "Replay_Idle_Current_mA_Mean": replay_metrics["idle_current_ma_mean"],
        "Replay_Idle_Power_mW_Mean": replay_metrics["idle_power_mw_mean"],
        "Field_Invoke_Energy_uWh_Mean": field_metrics["invoke_energy_uwh_mean"],
        "Field_Event_Energy_uWh_Mean": field_metrics["pipeline_energy_uwh_mean"],
        "Field_Invoke_Time_ms_Mean": field_metrics["invoke_time_ms_mean"],
        "Field_Event_Time_ms_Mean": field_metrics["pipeline_time_ms_mean"],
        "Field_Minimum_Free_Heap_kB": field_metrics["min_free_heap_kb"],
        "Field_Idle_Voltage_V_Mean": field_metrics["idle_voltage_v_mean"],
        "Field_Idle_Current_mA_Mean": field_metrics["idle_current_ma_mean"],
        "Field_Idle_Power_mW_Mean": field_metrics["idle_power_mw_mean"],
        "Model_Flash_kB": field_metrics["model_flash_kb"],
        "Tensor_Arena_kB": field_metrics["arena_kb"],
        "Total_Footprint_kB": field_metrics["total_footprint_kb"],
    }


def main() -> int:
    """Resolve artifact locations, parse the inputs, and write all output tables."""
    inferred_repo_root = script_repo_root()

    parser = argparse.ArgumentParser(
        description="Generate LiteML-Edge paper tables from default workbook and log directories."
    )
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=inferred_repo_root,
        help="LiteML repository root. When omitted, the script infers it from its own location.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "out_tables",
        help="Directory where output files will be written.",
    )
    args = parser.parse_args()

    repo_root = args.repo_root.resolve()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    default_bundles = discover_default_bundles(repo_root)
    fallback_bundles = discover_fallback_bundles(repo_root)
    bundles = merge_bundle_maps(default_bundles, fallback_bundles)

    available_models = [
        model for model in MODEL_ORDER
        if bundles.get(model) and bundles[model].prediction_workbook and bundles[model].model_io_workbook and bundles[model].log_files
    ]
    require(available_models, f"No valid LiteML-Edge workbook/log sets were found under {repo_root}.")

    model_rows = [assemble_model_record(bundles[model]) for model in available_models]
    replay_rows = build_replay_conformance_rows(model_rows)
    energy_rows = build_energy_rows(model_rows)
    memory_rows = build_memory_rows(model_rows)
    idle_rows = build_idle_rows(model_rows)

    write_csv(output_dir / "paper_model_sources.csv", model_rows)
    write_csv(output_dir / "table_replay_conformance.csv", replay_rows)
    write_csv(output_dir / "table_infer_energy.csv", energy_rows)
    write_csv(output_dir / "table_memory.csv", memory_rows)
    write_csv(output_dir / "table_idle.csv", idle_rows)

    write_workbook(
        output_dir / "liteml_edge_paper_tables.xlsx",
        {
            "model_sources": model_rows,
            "replay_conformance": replay_rows,
            "infer_energy": energy_rows,
            "memory": memory_rows,
            "idle": idle_rows,
        },
    )

    (output_dir / "table_replay_conformance.tex").write_text(render_replay_conformance_tex(replay_rows), encoding="utf-8")
    (output_dir / "table_infer_energy.tex").write_text(render_energy_tex(energy_rows), encoding="utf-8")
    (output_dir / "table_memory.tex").write_text(render_memory_tex(memory_rows), encoding="utf-8")
    (output_dir / "table_idle.tex").write_text(render_idle_tex(idle_rows), encoding="utf-8")

    manifest = {
        "repo_root": str(repo_root),
        "models_found": [row["Model"] for row in model_rows],
        "default_workbook_dirs": {k: str(v) for k, v in default_workbook_dirs(repo_root).items()},
        "default_log_dirs": {k: [str(p) for p in v] for k, v in default_log_dirs(repo_root).items()},
        "outputs": [
            "paper_model_sources.csv",
            "table_replay_conformance.csv",
            "table_infer_energy.csv",
            "table_memory.csv",
            "table_idle.csv",
            "liteml_edge_paper_tables.xlsx",
            "table_replay_conformance.tex",
            "table_infer_energy.tex",
            "table_memory.tex",
            "table_idle.tex",
        ],
    }
    (output_dir / "run_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print(f"LiteML-Edge paper tables generated in: {output_dir}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ExtractionError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)

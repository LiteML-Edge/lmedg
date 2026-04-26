"""
plot_rolling24_scatter_offline_ondevice.py

Generate scatter plots (ground truth vs prediction) for Rolling(24),
comparing OFFLINE (Python post-quantization) vs ON-DEVICE (Firmware Replay).

- Treat output as ONE Figure with TWO panels labeled (a) and (b).
- Avoid per-axis titles that make it look like two separate figures.
- Put Rolling(24) + variable in the LaTeX caption (recommended).

Inputs (Excel):
  A) ONE sheet containing TWO table blocks (recommended):
       - "Predictions_rolling24_Conv1d_Tiny_Python"
       - "Predictions_rolling24_Firmware_Conv1d_Tiny_Replay"
     Each block: a title row, then a header row, then data rows, then a blank row.
  B) TWO sheets (offline and on-device) with the same column schema.

Outputs:
  - PNG @ 600 dpi 
  - PDF (vector) for camera-ready inclusion
  - command: python plot_rolling24_scatter_offline_ondevice.py --excel data.xlsx --variable T --no_grid
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle


# ----------------------------
# Formatting (IEEE 1-column)
# ----------------------------
def _set_ieee_style():
    """
    IEEE LATAM (IEEEtran) 1-column friendly settings.
    Key idea: match figure physical width (~3.5in) so text is not scaled up later in LaTeX.
    """
    plt.rcParams.update({
        "font.family": "serif",
        "font.size": 7,            # base text
        "axes.titlesize": 6,       # panel label size
        "axes.labelsize": 7,       # axis labels (keep compact for 1-col)
        "legend.fontsize": 6,
        "xtick.labelsize": 6,
        "ytick.labelsize": 6,

        "axes.linewidth": 0.6,
        "xtick.major.width": 0.6,
        "ytick.major.width": 0.6,
        "xtick.major.size": 3,
        "ytick.major.size": 3,

        "mathtext.fontset": "dejavuserif",

        "figure.dpi": 100,         # runtime; export DPI controlled in savefig
        "savefig.dpi": 600,
    })


def _coerce_float(x):
    """Convert strings with comma decimal to float; keep NaN if empty."""
    if x is None:
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return np.nan
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan


# ----------------------------
# Data loading
# ----------------------------
def _extract_block_from_sheet(raw: pd.DataFrame, title: str) -> pd.DataFrame:
    """
    Extract a table block that starts after a row containing `title`.
    The next non-empty row is treated as header; rows continue until a blank row.
    """
    title_rows = raw.index[
        raw.apply(
            lambda r: r.astype(str).str.contains(re.escape(title), case=False, na=False).any(),
            axis=1
        )
    ]
    if len(title_rows) == 0:
        raise ValueError(f"Block title '{title}' not found in sheet.")
    start = int(title_rows[0])

    hdr = None
    for i in range(start + 1, len(raw)):
        row = raw.iloc[i]
        non_empty = sum([str(v).strip().lower() not in ("", "nan") for v in row.values])
        if non_empty >= 2:
            hdr = i
            break
    if hdr is None:
        raise ValueError(f"Could not find header row after '{title}'.")

    headers = [str(x).strip() for x in raw.iloc[hdr].values]
    keep_idx = [j for j, h in enumerate(headers) if h and h.lower() != "nan"]
    headers = [headers[j] for j in keep_idx]

    data_rows = []
    for i in range(hdr + 1, len(raw)):
        row = raw.iloc[i].values
        row = [row[j] for j in keep_idx]
        if all([str(v).strip().lower() in ("", "nan") for v in row]):
            break
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=headers)
    df.columns = [c.strip() for c in df.columns]

    for c in df.columns:
        df[c] = df[c].map(_coerce_float)

    return df


def _load_from_excel_one_sheet(path: Path, sheet: str,
                              offline_title: str, ondevice_title: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
    offline_df = _extract_block_from_sheet(raw, offline_title)
    ondev_df = _extract_block_from_sheet(raw, ondevice_title)
    return offline_df, ondev_df


def _load_from_excel_two_sheets(path: Path, offline_sheet: str, ondevice_sheet: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    off = pd.read_excel(path, sheet_name=offline_sheet, engine="openpyxl")
    dev = pd.read_excel(path, sheet_name=ondevice_sheet, engine="openpyxl")
    for df in (off, dev):
        df.columns = [str(c).strip() for c in df.columns]
        for c in df.columns:
            df[c] = df[c].map(_coerce_float)
    return off, dev


def _list_sheets(path: Path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames)


# ----------------------------
# Metrics and plotting
# ----------------------------
def _mae_rmse(y_true: np.ndarray, y_pred: np.ndarray) -> tuple[float, float]:
    diff = y_true - y_pred
    mae = float(np.mean(np.abs(diff)))
    rmse = float(np.sqrt(np.mean(diff * diff)))
    return mae, rmse


def _r2_score(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    """Coefficient of determination R^2 (returns NaN if undefined)."""
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    if y_true.size == 0:
        return float("nan")
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    y_mean = float(np.mean(y_true))
    ss_tot = float(np.sum((y_true - y_mean) ** 2))
    if ss_tot == 0.0:
        return float("nan")
    return 1.0 - (ss_res / ss_tot)


def _nice_limits(gt: np.ndarray, pred: np.ndarray) -> tuple[float, float]:
    lo = float(min(gt.min(), pred.min()))
    hi = float(max(gt.max(), pred.max()))
    rng = hi - lo
    pad = 0.04 * rng if rng > 0 else 0.5
    return lo - pad, hi + pad


def _panel_label(ax, text: str):
    """
    IEEE-friendly panel label (no box): e.g., '(a) Offline (Python)'.
    Placed slightly above the axes.
    """
    ax.text(
        0.02, 1.01, text,
        transform=ax.transAxes,
        va="bottom", ha="left",
        fontsize=7, fontweight="normal",
        clip_on=False
    )


def _scatter(ax, gt, pred, unit: str, *, show_grid: bool, panel: str | None):
    gt = np.asarray(gt, dtype=float)
    pred = np.asarray(pred, dtype=float)

    mask = np.isfinite(gt) & np.isfinite(pred)
    gt = gt[mask]
    pred = pred[mask]

    n = int(gt.size)
    mae, rmse = _mae_rmse(gt, pred)
    r2 = _r2_score(gt, pred)

    # Points: smaller for 1-col figures
    ax.scatter(gt, pred, s=10, alpha=0.9, linewidths=0.0)

    # Identity line: slightly thinner
    if n > 0:
        lo, hi = _nice_limits(gt, pred)
        ax.plot([lo, hi], [lo, hi], linestyle="--", linewidth=1.0, color="black")
        ax.set_xlim(lo, hi)
        ax.set_ylim(lo, hi)

    if show_grid:
        ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.30)
    else:
        ax.grid(False)

    ax.set_xlabel(f"Ground truth ({unit})")
    ax.set_ylabel(f"Prediction ({unit})")
    ax.set_aspect("equal", adjustable="box")

    # Metrics box: compact but readable in 1-col
    txt = (
        f"N = {n}\n"
        f"MAE = {mae:.3f} {unit}\n"
        f"RMSE = {rmse:.3f} {unit}"
        # f"$R^2$ = {r2:.3f}"
    )
    ax.text(
        0.03, 0.97, txt,
        transform=ax.transAxes,
        va="top", ha="left",
        fontsize=6,
        bbox=dict(
            boxstyle="round,pad=0.20",
            facecolor="white",
            edgecolor="black",
            linewidth=0.6
        )
    )

    if panel:
        _panel_label(ax, panel)


def _get_col(df: pd.DataFrame, names: list[str]) -> pd.Series:
    for n in names:
        if n in df.columns:
            return df[n]
    raise KeyError(f"Missing column. Tried: {names}. Available: {list(df.columns)}")


def make_figure(
    offline_df: pd.DataFrame,
    ondev_df: pd.DataFrame,
    variable: str,
    out_prefix: Path,
    *,
    dpi: int = 600,
    panel_headers: bool = True,
    outer_box: bool = False,
    no_grid: bool = False,
):
    """
    variable:
      - 'T' -> uses T_in (gt) and Tp (pred)
      - 'H' -> uses H_in (gt) and H_p (pred)
    Layout is fixed vertical (2x1) and sized for IEEE 1-column width.
    """
    var = variable.upper().strip()
    if var == "T":
        gt_off = _get_col(offline_df, ["T_in", "Tin"])
        pr_off = _get_col(offline_df, ["Tp", "T_p", "Tpred"])
        gt_dev = _get_col(ondev_df, ["Tin", "T_in"])
        pr_dev = _get_col(ondev_df, ["Tp", "T_p", "Tpred"])
        unit = "°C"
        var_label = "T_in"
    elif var == "H":
        gt_off = _get_col(offline_df, ["H_in", "Hin"])
        pr_off = _get_col(offline_df, ["H_p", "Hp", "Hpred"])
        gt_dev = _get_col(ondev_df, ["Hin", "H_in"])
        pr_dev = _get_col(ondev_df, ["Hp", "H_p", "Hpred"])
        unit = "%"
        var_label = "H_in"
    else:
        raise ValueError("variable must be 'T' or 'H'.")

    _set_ieee_style()

    # 1-column IEEE: width ~3.5in. Height adjusted for two stacked panels.
    fig, axes = plt.subplots(
        2, 1,
        figsize=(3.5, 5.8),
        constrained_layout=True
    )

    panels = (
        "(a) Offline (Python)",
        "(b) On-device (Firmware Replay)",
    ) if panel_headers else (None, None)

    _scatter(axes[0], gt_off, pr_off, unit, show_grid=(not no_grid), panel=panels[0])
    _scatter(axes[1], gt_dev, pr_dev, unit, show_grid=(not no_grid), panel=panels[1])

    if outer_box:
        rect = Rectangle(
            (0.01, 0.01), 0.98, 0.98,
            transform=fig.transFigure, fill=False,
            edgecolor="black", linewidth=0.8
        )
        fig.add_artist(rect)

    png_path = out_prefix.with_suffix(".png")
    pdf_path = out_prefix.with_suffix(".pdf")

    # pad_inches prevents clipping of panel labels
    fig.savefig(png_path, dpi=600, bbox_inches="tight", pad_inches=0.04)
    fig.savefig(pdf_path, bbox_inches="tight", pad_inches=0.04)
    plt.close(fig)

    return png_path, pdf_path


# ----------------------------
# CLI (IEEE-clean)
# ----------------------------
def build_argparser():
    p = argparse.ArgumentParser(
        description="Generate Rolling(24) scatter plots: OFFLINE vs ON-DEVICE (IEEE-clean, 1-column, vertical)."
    )

    p.add_argument("--excel", type=str, required=True, help="Path to Excel file (.xlsx).")

    p.add_argument("--sheet", type=str, default=None, help="Sheet name with BOTH blocks (default: first sheet).")
    p.add_argument("--offline_title", type=str, default="Predictions_rolling24_Conv1D_Tiny_Python",
                   help="Block title for OFFLINE section inside the sheet.")
    p.add_argument("--ondevice_title", type=str, default="Predictions_rolling24_Conv1D_Tiny_Firmware_Replay",
                   help="Block title for ON-DEVICE section inside the sheet.")

    p.add_argument("--offline_sheet", type=str, default=None, help="Offline sheet name (two-sheet mode).")
    p.add_argument("--ondevice_sheet", type=str, default=None, help="On-device sheet name (two-sheet mode).")

    p.add_argument("--variable", type=str, choices=["T", "H"], default="T",
                   help="T (temperature) or H (humidity).")

    p.add_argument("--outdir", type=str, default=".", help="Output directory.")
    p.add_argument("--basename", type=str, default=None, help="Base output filename (no extension).")
    p.add_argument("--dpi", type=int, default=600, help="PNG DPI (default 600).")

    p.add_argument("--no_panel_labels", action="store_true",
                   help="Disable (a)/(b) panel labels (not recommended).")
    p.add_argument("--outer_box", action="store_true",
                   help="Add one bounding box around the whole figure (optional).")
    p.add_argument("--no_grid", action="store_true",
                   help="Disable grid lines (recommended if it looks busy).")

    return p


def main():
    args = build_argparser().parse_args()

    excel = Path(args.excel)
    if not excel.exists():
        raise FileNotFoundError(excel)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    sheetnames = _list_sheets(excel)
    sheet = args.sheet if args.sheet else sheetnames[0]

    if args.offline_sheet and args.ondevice_sheet:
        offline_df, ondev_df = _load_from_excel_two_sheets(excel, args.offline_sheet, args.ondevice_sheet)
    else:
        try:
            offline_df, ondev_df = _load_from_excel_one_sheet(excel, sheet, args.offline_title, args.ondevice_title)
        except Exception as e:
            msg = (
                f"{e}\n\n"
                f"Available sheets in '{excel.name}': {sheetnames}\n"
                f"Tip: use --sheet \"{sheetnames[0]}\" (or your sheet name)."
            )
            raise SystemExit(msg)

    if args.basename:
        base = args.basename
    else:
        base = f"rolling24_scatter_offline_ondevice_{'T_in' if args.variable=='T' else 'H_in'}_ab"

    out_prefix = outdir / base

    png_path, pdf_path = make_figure(
        offline_df, ondev_df, args.variable, out_prefix,
        dpi=args.dpi,
        panel_headers=(not args.no_panel_labels),
        outer_box=args.outer_box,
        no_grid=args.no_grid,
    )

    print(f"Saved:\n  {png_path}\n  {pdf_path}")


if __name__ == "__main__":
    main()
"""
Script: compare_model_io_to_log_mlp.py
Module role:
    Compare Python reference spreadsheets for preprocessing and immediate model
    input/output stages against firmware logs and generate an audit workbook.

Technical summary:
    This script resolves the latest quantized-model metrics directory, locates
    the reference spreadsheets, parses firmware debug blocks, aligns Python and
    firmware records, evaluates exact and tolerance-based agreement, and exports
    an Excel workbook summarizing structural availability and stage-wise
    correspondence.

Inputs:
    - Quantized-model reference spreadsheets exported by the Python pipeline
    - Firmware TXT, LOG, or ZIP log containing debug blocks

Outputs:
    - Excel comparison workbook for preprocessing and model I/O validation

Notes:
    This script assumes the repository project structure and the referenced
    input artifacts and firmware logs. The computational logic, comparison
    workflow, and workbook generation procedure are preserved.
"""
from __future__ import annotations

import csv
import os
import re
import sys
import struct
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
MAX_EXCEL_CELL_CHARS = 32767

# ========================= Bootstrap local utils =========================
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base)
            break

if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from utils.global_utils.paths_mlp import (  # type: ignore
    PROJECT_ROOT,
    QUANTIZED_MODEL_METRICS,
)

# ========================= Dynamic path resolution =========================
RUN_DIR_REGEX = re.compile(r"^run\.v(\d+)$", re.IGNORECASE)


def extract_run_version(path: Path) -> int:
    match = RUN_DIR_REGEX.match(path.name)
    if not match:
        return -1
    return int(match.group(1))


def is_plain_run_dir(path: Path) -> bool:
    return path.name.lower() == "run"


def candidate_has_results_dir(run_dir: Path) -> bool:
    return (run_dir / "quantization_metrics_results").exists() and (
        run_dir / "quantization_metrics_results"
    ).is_dir()


def find_latest_run_dir(metrics_root: Path) -> Path:
    if not metrics_root.exists():
        raise FileNotFoundError(f"Metrics root directory not found: {metrics_root}")

    candidates: List[Path] = []

    for p in metrics_root.iterdir():
        if not p.is_dir():
            continue

        if is_plain_run_dir(p) and candidate_has_results_dir(p):
            candidates.append(p)
            continue

        if RUN_DIR_REGEX.match(p.name) and candidate_has_results_dir(p):
            candidates.append(p)
            continue

    if not candidates:
        available_children = []
        try:
            available_children = [x.name for x in metrics_root.iterdir()]
        except Exception:
            pass

        raise FileNotFoundError(
            f"No valid run directory found under metrics root: {metrics_root}\n"
            f"Expected one of these direct child layouts:\n"
            f"  - {metrics_root / 'run' / 'quantization_metrics_results'}\n"
            f"  - {metrics_root / 'run.vNNN' / 'quantization_metrics_results'}\n"
            f"Available direct children: {available_children}"
        )

    versioned_candidates = [p for p in candidates if RUN_DIR_REGEX.match(p.name)]
    plain_candidates = [p for p in candidates if is_plain_run_dir(p)]

    print(f"[INFO] configured_metrics_root: {metrics_root}")
    print("[INFO] detected run directories:")
    for p in sorted(candidates, key=lambda x: str(x).lower()):
        print(f"       - {p}")

    if versioned_candidates:
        latest = max(
            versioned_candidates,
            key=lambda p: (extract_run_version(p), p.stat().st_mtime),
        )
        return latest

    return max(plain_candidates, key=lambda p: p.stat().st_mtime)


def find_reference_xlsx(
    results_dir: Path,
    exact_name: str,
    fallback_patterns: List[str],
) -> Path:
    exact_path = results_dir / exact_name
    if exact_path.exists():
        return exact_path

    candidates: List[Path] = []
    seen = set()

    for pattern in fallback_patterns:
        for p in results_dir.glob(pattern):
            if p.is_file() and p.suffix.lower() == ".xlsx":
                key = str(p.resolve())
                if key not in seen:
                    seen.add(key)
                    candidates.append(p)

    if not candidates:
        raise FileNotFoundError(
            f"Could not find reference spreadsheet.\n"
            f"results_dir={results_dir}\n"
            f"exact_name={exact_name}\n"
            f"fallback_patterns={fallback_patterns}"
        )

    return max(candidates, key=lambda p: p.stat().st_mtime)


LOG_NAME_REGEX = re.compile(r"^device-monitor-(\d{6})-(\d{6})\.log$", re.IGNORECASE)
REPLAY_TOKEN = "[DBG_REPLAY_CSV]"


def extract_log_datetime_from_name(path: Path) -> Optional[datetime]:
    match = LOG_NAME_REGEX.match(path.name)
    if not match:
        return None
    date_part, time_part = match.groups()
    try:
        return datetime.strptime(date_part + time_part, "%y%m%d%H%M%S")
    except ValueError:
        return None


def _log_sort_key(path: Path) -> Tuple[datetime, float]:
    dt = extract_log_datetime_from_name(path)
    if dt is None:
        dt = datetime.min
    return dt, path.stat().st_mtime


def _log_contains_token(path: Path, token: str) -> bool:
    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line in handle:
                if token in line:
                    return True
    except Exception:
        return False
    return False


def find_latest_log_file(log_dir: Path) -> Path:
    if not log_dir.exists():
        raise FileNotFoundError(f"Log directory not found: {log_dir}")

    candidates = [p for p in log_dir.glob("device-monitor-*.log") if p.is_file()]
    if not candidates:
        raise FileNotFoundError(f"No device-monitor-*.log found in: {log_dir}")

    ordered_candidates = sorted(candidates, key=_log_sort_key, reverse=True)

    for candidate in ordered_candidates:
        if _log_contains_token(candidate, REPLAY_TOKEN):
            print(f"[INFO] Selected latest Replay log: {candidate}")
            return candidate

    raise FileNotFoundError(
        f"No Replay log found in: {log_dir} "
        f"(expected token {REPLAY_TOKEN})"
    )


def choose_writable_output_path(preferred_path: Path) -> Path:
    preferred_path.parent.mkdir(parents=True, exist_ok=True)

    if not preferred_path.exists():
        return preferred_path

    try:
        with open(preferred_path, "a+b"):
            pass
        return preferred_path
    except PermissionError:
        stamp = datetime.now().strftime("%y%m%d-%H%M%S")
        alt_path = preferred_path.with_name(
            f"{preferred_path.stem}_{stamp}{preferred_path.suffix}"
        )
        print(
            f"[WARN] Output workbook is locked: {preferred_path}\n"
            f"[WARN] Falling back to: {alt_path}"
        )
        return alt_path



def resolve_runtime_paths() -> Tuple[Path, Path, Path, Optional[Path], Optional[Path], Path, Path]:
    metrics_root = Path(QUANTIZED_MODEL_METRICS)
    latest_run_dir = find_latest_run_dir(metrics_root)

    results_dir = latest_run_dir / "quantization_metrics_results"
    if not results_dir.exists():
        raise FileNotFoundError(
            f"quantization_metrics_results not found in selected run directory: {results_dir}"
        )

    input_ref_path = find_reference_xlsx(
        results_dir=results_dir,
        exact_name="environment_quantized_dbg_model_input_reference_mlp.xlsx",
        fallback_patterns=[
            "*dbg_model_input_reference*mlp*.xlsx",
            "*input_reference*mlp*.xlsx",
            "*input*reference*mlp*.xlsx",
            "*dbg*input*mlp*.xlsx",
        ],
    )

    output_ref_path = find_reference_xlsx(
        results_dir=results_dir,
        exact_name="environment_quantized_dbg_model_output_reference_mlp.xlsx",
        fallback_patterns=[
            "*dbg_model_output_reference*mlp*.xlsx",
            "*output_reference*mlp*.xlsx",
            "*output*reference*mlp*.xlsx",
            "*dbg*output*mlp*.xlsx",
        ],
    )

    try:
        output_raw_ref_path: Optional[Path] = find_reference_xlsx(
            results_dir=results_dir,
            exact_name="environment_quantized_dbg_model_output_raw_reference_mlp.xlsx",
            fallback_patterns=[
                "*dbg_model_output_raw_reference*mlp*.xlsx",
                "*output_raw_reference*mlp*.xlsx",
                "*output*raw*reference*mlp*.xlsx",
                "*dbg*output*raw*mlp*.xlsx",
            ],
        )
    except FileNotFoundError:
        output_raw_ref_path = None

    try:
        preprocess_ref_path: Optional[Path] = find_reference_xlsx(
            results_dir=results_dir,
            exact_name="environment_quantized_preprocess_debug_reference_mlp.xlsx",
            fallback_patterns=[
                "*preprocess_debug_reference*mlp*.xlsx",
                "*preprocess*debug*reference*mlp*.xlsx",
                "*preprocess*reference*mlp*.xlsx",
            ],
        )
    except FileNotFoundError:
        preprocess_ref_path = None

    workbook_dir = (
        Path(PROJECT_ROOT)
        / "utils"
        / "workbook_mlp"
    )
    
    log_dir = (
        Path(PROJECT_ROOT)
        / "firmwares"
        / "environment_mlp"
        / "PlatfIO_ESP32_Wemos_mlp"
        / "logs"
    )
    log_txt_path = find_latest_log_file(log_dir)

    preferred_out_xlsx_path = workbook_dir / "model_io_comparison_mlp.xlsx"
    out_xlsx_path = choose_writable_output_path(preferred_out_xlsx_path)

    return latest_run_dir, input_ref_path, output_ref_path, output_raw_ref_path, preprocess_ref_path, log_txt_path, out_xlsx_path


ABS_TOLERANCE = 2e-6
INPUT_TENSOR_TOLERANCE = ABS_TOLERANCE
RAW_OUTPUT_TOLERANCE = ABS_TOLERANCE
SEMANTIC_OUTPUT_TOLERANCE = ABS_TOLERANCE
FINAL_OUTPUT_TOLERANCE = 2e-5
SEMANTIC_DEBUG_TOLERANCE = 2e-5

# ========================= Tags / schema =========================
TAG_IN = "DBG_MODEL_IN_CSV"
TAG_PRE_RAW = "DBG_PRE_RAW_CSV"
TAG_PRE_SMOOTH = "DBG_PRE_SMOOTH_CSV"
TAG_OUT = "DBG_MODEL_OUT_CSV"
TAG_OUT_BITS = "DBG_MODEL_OUT_BITS_CSV"
TAG_OUT_RAW = "DBG_MODEL_OUT_RAW_CSV"
TAG_OUT_STABILITY = "DBG_MODEL_OUT_STABILITY_CSV"

LEGACY_INPUT_KEY = ["idx", "epoch", "step"]
LEGACY_OUTPUT_KEY = ["idx", "epoch"]
LEGACY_INPUT_COMPARE_COLS = [f"p{i}" for i in range(12)] + [f"x{i}" for i in range(12)]
LEGACY_OUTPUT_COMPARE_COLS = ["o0_raw", "o1_raw", "y0", "y1"]

NEW_INPUT_KEY = ["idx", "epoch"]
NEW_INPUT_STEP_KEY = ["idx", "epoch", "step"]
NEW_OUTPUT_KEY = ["idx", "epoch"]
RAW_TENSOR_OUTPUT_KEY = ["idx", "epoch", "out_idx"]
RAW_TENSOR_BASE_COMPARE_PREFERRED = [
    "tensor_index",
    "type_code",
    "type_name",
    "bytes_total",
    "bytes_dumped",
    "dims_size",
]
NEW_INPUT_STATE_COMPARE_PREFERRED = [
    "gt_Tin_true",
    "gt_Hin_true",
    "state_Tout_phys_raw",
    "state_Hout_phys_raw",
    "state_Tin_lag1_phys_raw",
    "state_Hin_lag1_phys_raw",
    "state_Tout_lag1_phys_raw",
    "state_Hout_lag1_phys_raw",
    "state_Tin_lag2_phys_raw",
    "state_Hin_lag2_phys_raw",
    "state_sin_hour",
    "state_cos_hour",
    "state_weekday",
    "state_month",
]
NEW_OUTPUT_COMPARE_PREFERRED = [
    "out_o0_float",
    "out_o1_float",
    "out_o0_bits_hex",
    "out_o1_bits_hex",
    "y_T_scaled",
    "y_H_scaled",
    "d_T_pred",
    "d_H_pred",
    "p_Tprev_phys",
    "p_Hprev_phys",
    "p_T_pred",
    "p_H_pred",
]
OPTIONAL_OUTPUT_COMPARE_FIELDS = {
    "out_o0_bits_hex",
    "out_o1_bits_hex",
}
BITWISE_OUTPUT_FIELDS = ["out_o0_bits_hex", "out_o1_bits_hex"]

RAW_OUTPUT_FIELDS = {
    "out_o0_float",
    "out_o1_float",
    "out_o0_bits_hex",
    "out_o1_bits_hex",
    # Legacy compatibility
    "o0_raw",
    "o1_raw",
}
SEMANTIC_OUTPUT_FIELDS = {
    "y_T_scaled",
    "y_H_scaled",
    # Legacy compatibility
    "y0",
    "y1",
}
FINAL_OUTPUT_FIELDS = {
    "d_T_pred",
    "d_H_pred",
    "p_T_pred",
    "p_H_pred",
}


def _is_exact_compare_field(col: str) -> bool:
    return (
        col.endswith("_bits_hex")
        or col == "type_name"
        or bool(re.match(r"^b\d+_hex$", col))
    )


def split_input_compare_columns(compare_cols: Sequence[str]) -> Tuple[List[str], List[str]]:
    critical = [
        c
        for c in compare_cols
        if re.match(r"^in_x\d+_float$", c) or re.match(r"^x\d+$", c)
    ]
    semantic = [c for c in compare_cols if c not in critical]
    return critical, semantic


def split_output_compare_columns(compare_cols: Sequence[str]) -> Tuple[List[str], List[str], List[str], List[str]]:
    raw = [c for c in compare_cols if c in RAW_OUTPUT_FIELDS]
    semantic = [c for c in compare_cols if c in SEMANTIC_OUTPUT_FIELDS]
    final = [c for c in compare_cols if c in FINAL_OUTPUT_FIELDS]
    aux = [c for c in compare_cols if c not in set(raw) | set(semantic) | set(final)]
    if not (raw or semantic or final):
        final = list(compare_cols)
        aux = []
    return raw, semantic, final, aux


def _normalize_tolerance_value(tolerance: object) -> object:
    if isinstance(tolerance, dict):
        return "stage-aware / per-column"
    try:
        return float(tolerance)  # type: ignore[arg-type]
    except Exception:
        return tolerance


def make_empty_summary(ref_rows: int, log_rows: int, tolerance: object, merge_keys: Sequence[str]) -> Dict:
    normalized_tolerance = _normalize_tolerance_value(tolerance)
    return {
        "rows_reference": int(ref_rows),
        "rows_log": int(log_rows),
        "rows_both": min(int(ref_rows), int(log_rows)),
        "rows_only_reference": max(int(ref_rows) - int(log_rows), 0),
        "rows_only_log": max(int(log_rows) - int(ref_rows), 0),
        "rows_with_diff": 0,
        "comparisons_with_diff": 0,
        "total_comparisons": 0,
        "max_abs_diff": 0.0,
        "exact_counts": {},
        "tolerance": normalized_tolerance,
        "tolerance_label": str(normalized_tolerance),
        "total_compare_cols": 0,
        "merge_keys": list(merge_keys),
        "duplicate_key_alignment_applied": False,
        "not_applicable": True,
    }


def classify_result(summary: Dict) -> str:
    if summary.get("not_applicable"):
        return "N/A"

    no_missing = summary["rows_only_reference"] == 0 and summary["rows_only_log"] == 0
    no_diff = summary["rows_with_diff"] == 0

    if no_missing and no_diff:
        return "MATCH"
    if no_missing and not no_diff:
        return "MISMATCH"
    return "INCOMPLETE"


def compare_subset(
    ref_df: pd.DataFrame,
    log_df: pd.DataFrame,
    key_cols: Sequence[str],
    compare_cols: Sequence[str],
    tolerance: object,
    applicability_mode: Optional[str] = None,
) -> Tuple[pd.DataFrame, Dict]:
    if compare_cols:
        return merge_and_compare(
            ref_df,
            log_df,
            key_cols,
            compare_cols,
            tolerance,
            applicability_mode=applicability_mode,
        )
    return pd.DataFrame(), make_empty_summary(len(ref_df), len(log_df), tolerance, key_cols)

# ========================= Excel style =========================
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
BAD_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
INFO_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")

THIN_GRAY = Side(style="thin", color="C9C9C9")
BORDER_THIN = Border(
    left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
)

TITLE_FONT = Font(bold=True, size=15)
SUBTITLE_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(bold=True)

# ========================= Basic IO =========================
def read_first_sheet_xlsx(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Reference spreadsheet not found: {path}")
    return pd.read_excel(path, sheet_name=0)


def read_sheet_xlsx(path: Path, sheet_name: str) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Reference spreadsheet not found: {path}")
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def read_text_from_log_txt(path: Path) -> str:
    if path.suffix.lower() not in {".txt", ".log"}:
        raise ValueError(f"Expected a .txt or .log firmware log, got: {path}")
    if not path.exists():
        raise FileNotFoundError(f"Firmware log not found: {path}")
    return path.read_text(encoding="utf-8", errors="replace")


# ========================= Parsing =========================
def parse_tagged_csv(log_text: str, tag: str, log_path: Path) -> pd.DataFrame:
    rows: List[List[str]] = []
    header: List[str] | None = None
    pattern = re.compile(rf".*?\[{re.escape(tag)}\]\s*(.*)$")

    for line in log_text.splitlines():
        match = pattern.match(line)
        if not match:
            continue

        payload = match.group(1).strip()
        if not payload:
            continue

        parsed = next(csv.reader([payload]))
        parsed = [x.strip() for x in parsed]

        if header is None:
            header = parsed
            continue

        if parsed == header:
            continue

        if len(parsed) != len(header):
            continue

        rows.append(parsed)

    if header is None:
        raise ValueError(f"Tag [{tag}] not found in log: {log_path}")

    return pd.DataFrame(rows, columns=header)




def parse_optional_tagged_csv(log_text: str, tag: str) -> Optional[pd.DataFrame]:
    # Important: scan line-by-line, because a regex search over the full log text
    # without multiline mode can miss valid tagged lines that are not at end-of-string.
    token = f"[{tag}]"
    if not any(token in line for line in log_text.splitlines()):
        return None
    return parse_tagged_csv(log_text, tag, Path(f"<{tag}>"))


def merge_output_log_blocks(output_log_raw: pd.DataFrame, output_bits_log_raw: Optional[pd.DataFrame]) -> pd.DataFrame:
    if output_bits_log_raw is None or output_bits_log_raw.empty:
        return output_log_raw

    key_cols = [c for c in ["idx", "epoch"] if c in output_log_raw.columns and c in output_bits_log_raw.columns]
    if len(key_cols) != 2:
        return output_log_raw

    bit_cols = [c for c in ["out_o0_bits_hex", "out_o1_bits_hex"] if c in output_bits_log_raw.columns]
    if not bit_cols:
        return output_log_raw

    bits_only = output_bits_log_raw[key_cols + bit_cols].copy()
    bits_only = bits_only.drop_duplicates(subset=key_cols, keep="last")

    merged = output_log_raw.merge(bits_only, on=key_cols, how="left", sort=True)
    return merged
def _float_to_hex_bits(value: object) -> Optional[str]:
    try:
        f32 = float(value)
    except (TypeError, ValueError):
        return None
    return f"0x{struct.unpack('<I', struct.pack('<f', float(f32)))[0]:08X}"


def ensure_output_bit_columns(df: pd.DataFrame, *, allow_synthesize_from_float: bool) -> Tuple[pd.DataFrame, List[str]]:
    out = df.copy()
    synthesized: List[str] = []
    pairs = [("out_o0_float", "out_o0_bits_hex"), ("out_o1_float", "out_o1_bits_hex")]
    for float_col, bits_col in pairs:
        if bits_col in out.columns:
            continue
        if not allow_synthesize_from_float or float_col not in out.columns:
            continue
        series = out[float_col].map(_float_to_hex_bits)
        if series.notna().any():
            out[bits_col] = series
            synthesized.append(bits_col)
    return out, synthesized


def merge_output_log_blocks_robust(output_log_raw: pd.DataFrame, output_bits_log_raw: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Merge the optional bitwise output block into the main output block robustly.

    Primary path: merge by idx/epoch.
    Fallback path: if bits are still absent but both blocks have the same row count,
    align row-by-row after preserving original order.
    """
    merged = merge_output_log_blocks(output_log_raw, output_bits_log_raw)
    shared_bit_cols = [c for c in BITWISE_OUTPUT_FIELDS if c in merged.columns]
    if shared_bit_cols:
        return merged
    if output_bits_log_raw is None or output_bits_log_raw.empty:
        return merged

    bit_cols = [c for c in BITWISE_OUTPUT_FIELDS if c in output_bits_log_raw.columns]
    if not bit_cols:
        return merged

    # Fallback: preserve row order if counts match exactly.
    if len(output_log_raw) != len(output_bits_log_raw):
        return merged

    left = output_log_raw.reset_index(drop=True).copy()
    right = output_bits_log_raw.reset_index(drop=True).copy()
    for col in bit_cols:
        left[col] = right[col].values
    return left


def force_shared_output_bit_compare_columns(
    output_ref_df: pd.DataFrame,
    output_log_df: pd.DataFrame,
    compare_cols: Sequence[str],
) -> Tuple[List[str], List[str]]:
    """Force shared bitwise fields into semantic output comparison when available."""
    compare_set = set(compare_cols)
    forced: List[str] = []
    for col in BITWISE_OUTPUT_FIELDS:
        if col in output_ref_df.columns and col in output_log_df.columns and col not in compare_set:
            compare_set.add(col)
            forced.append(col)
    ordered = [c for c in NEW_OUTPUT_COMPARE_PREFERRED if c in compare_set]
    return ordered, forced


def normalize_numeric(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        converted = pd.to_numeric(out[col], errors="coerce")
        if converted.notna().all():
            out[col] = converted
    return out


def _extract_numeric_suffix(name: str, prefix: str, suffix: str) -> int:
    match = re.match(rf"^{re.escape(prefix)}(\d+){re.escape(suffix)}$", name)
    if not match:
        return 10**9
    return int(match.group(1))


def _sorted_matching_columns(columns: Iterable[str], pattern: str, *, prefix: str, suffix: str) -> List[str]:
    regex = re.compile(pattern)
    matched = [c for c in columns if regex.match(c)]
    return sorted(matched, key=lambda c: _extract_numeric_suffix(c, prefix, suffix))


def _natural_sort_key(value: str):
    parts = re.split(r"(\d+)", str(value))
    key = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part)
    return tuple(key)


def _build_semantic_input_compare_cols(columns: Sequence[str]) -> List[str]:
    compare_cols: List[str] = [c for c in NEW_INPUT_STATE_COMPARE_PREFERRED if c in columns]

    for suffix in ("_phys_raw", "_phys_clip", "_scaled"):
        compare_cols.extend(
            _sorted_matching_columns(
                columns,
                rf"^in_f\d+{re.escape(suffix)}$",
                prefix="in_f",
                suffix=suffix,
            )
        )

    compare_cols.extend(
        _sorted_matching_columns(
            columns,
            r"^in_x\d+_float$",
            prefix="in_x",
            suffix="_float",
        )
    )
    return compare_cols


def _align_duplicate_keys(
    ref_df: pd.DataFrame,
    log_df: pd.DataFrame,
    key_cols: Sequence[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    merge_keys = list(key_cols)
    if not key_cols:
        return ref_df.copy(), log_df.copy(), merge_keys

    ref_dup = ref_df.duplicated(subset=list(key_cols), keep=False).any()
    log_dup = log_df.duplicated(subset=list(key_cols), keep=False).any()
    if not (ref_dup or log_dup):
        return ref_df.copy(), log_df.copy(), merge_keys

    aligned_ref = ref_df.copy()
    aligned_log = log_df.copy()
    aligned_ref["_dup_ordinal"] = aligned_ref.groupby(list(key_cols), dropna=False).cumcount()
    aligned_log["_dup_ordinal"] = aligned_log.groupby(list(key_cols), dropna=False).cumcount()
    merge_keys = list(key_cols) + ["_dup_ordinal"]
    return aligned_ref, aligned_log, merge_keys


def detect_input_schema(df: pd.DataFrame) -> Tuple[List[str], List[str], str]:
    cols = list(df.columns)
    if all(c in df.columns for c in LEGACY_INPUT_KEY) and all(c in df.columns for c in LEGACY_INPUT_COMPARE_COLS):
        return LEGACY_INPUT_KEY, LEGACY_INPUT_COMPARE_COLS, "legacy_step_payload"

    semantic_step_compare_cols = _build_semantic_input_compare_cols(cols)
    if all(c in df.columns for c in NEW_INPUT_STEP_KEY) and semantic_step_compare_cols:
        return NEW_INPUT_STEP_KEY, semantic_step_compare_cols, "semantic_step_payload"

    tensor_float_cols = _sorted_matching_columns(
        cols,
        r"^in_x\d+_float$",
        prefix="in_x",
        suffix="_float",
    )
    if all(c in df.columns for c in NEW_INPUT_KEY) and tensor_float_cols:
        compare_cols = [c for c in NEW_INPUT_STATE_COMPARE_PREFERRED if c in cols] + tensor_float_cols
        return NEW_INPUT_KEY, compare_cols, "semantic_tensor_float"

    raise KeyError(
        "Could not detect input debug schema. Expected either the legacy schema "
        "(idx, epoch, step, p*, x*) or a semantic schema with step-based input fields "
        "(idx, epoch, step, in_f*_..., in_x*_float) or flattened tensor fields (idx, epoch, in_x*_float)."
    )




def detect_output_raw_tensor_schema(df: pd.DataFrame) -> Tuple[List[str], List[str], str]:
    if not all(c in df.columns for c in RAW_TENSOR_OUTPUT_KEY):
        raise KeyError(
            "Could not detect raw output tensor schema. Missing required key columns idx/epoch/out_idx."
        )

    compare_cols: List[str] = [
        c for c in RAW_TENSOR_BASE_COMPARE_PREFERRED if c in df.columns
    ]
    compare_cols.extend(
        _sorted_matching_columns(
            df.columns,
            r"^dim\d+$",
            prefix="dim",
            suffix="",
        )
    )
    compare_cols.extend(
        _sorted_matching_columns(
            df.columns,
            r"^b\d+_hex$",
            prefix="b",
            suffix="_hex",
        )
    )

    if not compare_cols:
        raise KeyError(
            "Could not detect raw output tensor schema. Expected raw tensor metadata and/or dumped byte columns."
        )

    return RAW_TENSOR_OUTPUT_KEY, compare_cols, "raw_output_tensor"

def detect_output_schema(df: pd.DataFrame) -> Tuple[List[str], List[str], str]:
    if all(c in df.columns for c in LEGACY_OUTPUT_KEY) and all(c in df.columns for c in LEGACY_OUTPUT_COMPARE_COLS):
        return LEGACY_OUTPUT_KEY, LEGACY_OUTPUT_COMPARE_COLS, "legacy_output"

    if not all(c in df.columns for c in NEW_OUTPUT_KEY):
        raise KeyError(
            "Could not detect output debug schema. Missing required key columns idx/epoch."
        )

    compare_cols = [c for c in NEW_OUTPUT_COMPARE_PREFERRED if c in df.columns]
    if compare_cols:
        return NEW_OUTPUT_KEY, compare_cols, "semantic_output"

    raise KeyError(
        "Could not detect output debug schema. Expected legacy columns "
        "(o0_raw, o1_raw, y0, y1) or new semantic columns such as out_o*_float / y_* / p_*"
    )


def reconcile_semantic_output_columns(
    python_compare_cols: Sequence[str],
    firmware_compare_cols: Sequence[str],
) -> Tuple[List[str], List[str], List[str]]:
    py_cols = list(python_compare_cols)
    fw_cols = list(firmware_compare_cols)

    py_set = set(py_cols)
    fw_set = set(fw_cols)
    symmetric_missing = py_set.symmetric_difference(fw_set)
    non_optional_missing = sorted(c for c in symmetric_missing if c not in OPTIONAL_OUTPUT_COMPARE_FIELDS)
    if non_optional_missing:
        raise KeyError(
            "Output schema mismatch between Python reference and firmware log: "
            f"non-optional columns differ: {non_optional_missing} | "
            f"python={py_cols} | firmware={fw_cols}"
        )

    shared_cols = [c for c in NEW_OUTPUT_COMPARE_PREFERRED if c in py_set and c in fw_set]
    if not shared_cols:
        raise KeyError(
            "Output schema mismatch between Python reference and firmware log: "
            "no shared semantic output columns were found."
        )

    python_only_optional = [c for c in py_cols if c in OPTIONAL_OUTPUT_COMPARE_FIELDS and c not in fw_set]
    firmware_only_optional = [c for c in fw_cols if c in OPTIONAL_OUTPUT_COMPARE_FIELDS and c not in py_set]
    return shared_cols, python_only_optional, firmware_only_optional


def reconcile_raw_output_tensor_columns(
    python_compare_cols: Sequence[str],
    firmware_compare_cols: Sequence[str],
) -> Tuple[List[str], List[str], List[str]]:
    py_cols = list(python_compare_cols)
    fw_cols = list(firmware_compare_cols)

    py_set = set(py_cols)
    fw_set = set(fw_cols)
    shared_set = py_set & fw_set

    shared_cols: List[str] = []
    for col in RAW_TENSOR_BASE_COMPARE_PREFERRED:
        if col in shared_set:
            shared_cols.append(col)
    shared_cols.extend(
        sorted((c for c in shared_set if re.fullmatch(r"dim\d+", c)), key=_natural_sort_key)
    )
    shared_cols.extend(
        sorted((c for c in shared_set if re.fullmatch(r"b\d+_hex", c)), key=_natural_sort_key)
    )

    if not shared_cols:
        raise KeyError(
            "Raw output tensor schema mismatch between Python reference and firmware log: "
            "no shared raw tensor columns were found."
        )

    python_only_optional = [c for c in py_cols if c not in fw_set]
    firmware_only_optional = [c for c in fw_cols if c not in py_set]
    return shared_cols, python_only_optional, firmware_only_optional


def promote_shared_output_bit_fields(
    output_ref_df: pd.DataFrame,
    output_log_df: pd.DataFrame,
    compare_cols: Sequence[str],
) -> Tuple[List[str], List[str]]:
    compare_set = set(compare_cols)
    shared_bits = [
        col for col in BITWISE_OUTPUT_FIELDS
        if col in output_ref_df.columns and col in output_log_df.columns
    ]
    promoted: List[str] = []
    for col in shared_bits:
        if col not in compare_set:
            compare_set.add(col)
            promoted.append(col)
    ordered = [c for c in NEW_OUTPUT_COMPARE_PREFERRED if c in compare_set]
    return ordered, promoted


def prepare_dataframe(
    df: pd.DataFrame,
    key_cols: Sequence[str],
    compare_cols: Sequence[str],
) -> pd.DataFrame:
    expected = list(key_cols) + list(compare_cols)
    missing = [c for c in expected if c not in df.columns]
    if missing:
        raise KeyError(f"Missing expected columns: {missing}")
    return normalize_numeric(df[expected].copy())


def _raw_tensor_applicable_mask(
    merged: pd.DataFrame,
    col: str,
    both_mask: pd.Series,
) -> pd.Series:
    dim_match = re.fullmatch(r"dim(\d+)", col)
    if dim_match:
        dim_idx = int(dim_match.group(1))
        dims_py = pd.to_numeric(merged.get("dims_size_py"), errors="coerce")
        dims_fw = pd.to_numeric(merged.get("dims_size_fw"), errors="coerce")
        return both_mask & dims_py.notna() & dims_fw.notna() & (dims_py > dim_idx) & (dims_fw > dim_idx)

    byte_match = re.fullmatch(r"b(\d+)_hex", col)
    if byte_match:
        byte_idx = int(byte_match.group(1))
        dumped_py = pd.to_numeric(merged.get("bytes_dumped_py"), errors="coerce")
        dumped_fw = pd.to_numeric(merged.get("bytes_dumped_fw"), errors="coerce")
        return both_mask & dumped_py.notna() & dumped_fw.notna() & (dumped_py > byte_idx) & (dumped_fw > byte_idx)

    return both_mask


def detect_output_stability_schema(df: pd.DataFrame) -> Tuple[List[str], List[str], str]:
    key_cols = ["idx", "epoch", "out_idx"]
    if not all(c in df.columns for c in key_cols):
        raise KeyError(
            "Could not detect output stability schema. Missing required key columns idx/epoch/out_idx."
        )

    preferred = [
        "type_a_code", "type_b_code",
        "type_a_name", "type_b_name",
        "bytes_a_total", "bytes_b_total",
        "bytes_a_dumped", "bytes_b_dumped",
        "dims_a_size", "dims_b_size",
        "ptr_a_hex", "ptr_b_hex",
        "ptr_equal", "raw_equal",
    ]
    compare_cols: List[str] = [c for c in preferred if c in df.columns]
    compare_cols.extend(sorted((c for c in df.columns if re.fullmatch(r"dim\d+_a", c)), key=_natural_sort_key))
    compare_cols.extend(sorted((c for c in df.columns if re.fullmatch(r"dim\d+_b", c)), key=_natural_sort_key))
    compare_cols.extend(sorted((c for c in df.columns if re.fullmatch(r"b\d+_a_hex", c)), key=_natural_sort_key))
    compare_cols.extend(sorted((c for c in df.columns if re.fullmatch(r"b\d+_b_hex", c)), key=_natural_sort_key))

    if not compare_cols:
        raise KeyError(
            "Could not detect output stability schema. Expected stability metadata and/or snapshot columns."
        )

    return key_cols, compare_cols, "output_stability"


def _normalized_text_series(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper()


def analyze_output_stability(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    if df is None or df.empty:
        return pd.DataFrame([{"Message": "No DBG_MODEL_OUT_STABILITY_CSV block found in firmware log."}]), make_empty_summary(0, 0, 0.0, ["idx", "epoch", "out_idx"])

    work = df.copy()
    key_cols, compare_cols, _ = detect_output_stability_schema(work)
    work = work.sort_values(by=[c for c in key_cols if c in work.columns]).reset_index(drop=True)

    numeric_cols = [
        "type_a_code", "type_b_code",
        "bytes_a_total", "bytes_b_total",
        "bytes_a_dumped", "bytes_b_dumped",
        "dims_a_size", "dims_b_size",
        "ptr_equal", "raw_equal",
    ]
    numeric_cols.extend([c for c in work.columns if re.fullmatch(r"dim\d+_[ab]", c)])
    for col in numeric_cols:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce")

    diff_fields_by_row: List[List[str]] = []
    exact_counts: Dict[str, int] = {c: 0 for c in compare_cols}
    comparisons_with_diff = 0
    total_comparisons = 0

    def _eq_text(col_a: str, col_b: str) -> pd.Series:
        a = _normalized_text_series(work[col_a]) if col_a in work.columns else pd.Series([""] * len(work))
        b = _normalized_text_series(work[col_b]) if col_b in work.columns else pd.Series([""] * len(work))
        return a.eq(b)

    comparisons: Dict[str, pd.Series] = {}
    for a, b, name in [
        ("type_a_code", "type_b_code", "type_code_equal"),
        ("bytes_a_total", "bytes_b_total", "bytes_total_equal"),
        ("bytes_a_dumped", "bytes_b_dumped", "bytes_dumped_equal"),
        ("dims_a_size", "dims_b_size", "dims_size_equal"),
        ("ptr_equal", None, "ptr_equal"),
        ("raw_equal", None, "raw_equal"),
    ]:
        if a in work.columns and b and b in work.columns:
            comparisons[name] = work[a].eq(work[b])
        elif a in work.columns and b is None:
            comparisons[name] = work[a].eq(1)

    if "type_a_name" in work.columns and "type_b_name" in work.columns:
        comparisons["type_name_equal"] = _eq_text("type_a_name", "type_b_name")

    # pairwise dim/byte comparisons with applicability by row
    dim_pairs = []
    dim_indices = sorted({int(m.group(1)) for c in work.columns for m in [re.fullmatch(r"dim(\d+)_a", c)] if m} |
                         {int(m.group(1)) for c in work.columns for m in [re.fullmatch(r"dim(\d+)_b", c)] if m})
    for di in dim_indices:
        ca, cb = f"dim{di:02d}_a", f"dim{di:02d}_b"
        if ca in work.columns and cb in work.columns:
            applicable = pd.Series(True, index=work.index)
            if "dims_a_size" in work.columns:
                applicable &= work["dims_a_size"].fillna(-1).gt(di)
            if "dims_b_size" in work.columns:
                applicable &= work["dims_b_size"].fillna(-1).gt(di)
            comp = pd.Series(True, index=work.index)
            comp.loc[applicable] = work.loc[applicable, ca].eq(work.loc[applicable, cb])
            comparisons[f"dim{di:02d}_equal"] = comp
            dim_pairs.append((ca, cb, applicable, comp))

    byte_pairs = []
    byte_indices = sorted({int(m.group(1)) for c in work.columns for m in [re.fullmatch(r"b(\d+)_a_hex", c)] if m} |
                          {int(m.group(1)) for c in work.columns for m in [re.fullmatch(r"b(\d+)_b_hex", c)] if m})
    for bi in byte_indices:
        ca, cb = f"b{bi:02d}_a_hex", f"b{bi:02d}_b_hex"
        if ca in work.columns and cb in work.columns:
            applicable = pd.Series(True, index=work.index)
            if "bytes_a_dumped" in work.columns:
                applicable &= work["bytes_a_dumped"].fillna(-1).gt(bi)
            if "bytes_b_dumped" in work.columns:
                applicable &= work["bytes_b_dumped"].fillna(-1).gt(bi)
            comp = pd.Series(True, index=work.index)
            if applicable.any():
                a = _normalized_text_series(work.loc[applicable, ca])
                b = _normalized_text_series(work.loc[applicable, cb])
                comp.loc[applicable] = a.eq(b)
            comparisons[f"b{bi:02d}_equal"] = comp
            byte_pairs.append((ca, cb, applicable, comp))

    # exact counts for sheet summary; count only applicable rows for dim/byte pairs
    for col in compare_cols:
        if col in work.columns:
            if col in {"type_a_name", "type_b_name", "ptr_a_hex", "ptr_b_hex"}:
                exact_counts[col] = len(work)
            else:
                exact_counts[col] = int(work[col].notna().sum())

    # comparison-level accounting
    for name, comp in comparisons.items():
        dim_match = re.fullmatch(r"dim(\d+)_equal", name)
        byte_match = re.fullmatch(r"b(\d+)_equal", name)
        if dim_match:
            di = int(dim_match.group(1))
            applicable = pd.Series(True, index=work.index)
            if "dims_a_size" in work.columns:
                applicable &= work["dims_a_size"].fillna(-1).gt(di)
            if "dims_b_size" in work.columns:
                applicable &= work["dims_b_size"].fillna(-1).gt(di)
        elif byte_match:
            bi = int(byte_match.group(1))
            applicable = pd.Series(True, index=work.index)
            if "bytes_a_dumped" in work.columns:
                applicable &= work["bytes_a_dumped"].fillna(-1).gt(bi)
            if "bytes_b_dumped" in work.columns:
                applicable &= work["bytes_b_dumped"].fillna(-1).gt(bi)
        else:
            applicable = pd.Series(True, index=work.index)
        total_comparisons += int(applicable.sum())
        comparisons_with_diff += int((~comp.loc[applicable].fillna(True)).sum())

    row_has_diff = pd.Series(False, index=work.index)
    for comp in comparisons.values():
        row_has_diff |= ~comp.fillna(True)

    # Most affected fields
    for i in work.index:
        diffs: List[str] = []
        for name, comp in comparisons.items():
            try:
                ok = bool(comp.loc[i])
            except Exception:
                ok = True
            if not ok:
                diffs.append(name)
        diff_fields_by_row.append(diffs)

    work["row_has_diff"] = row_has_diff
    work["Most_Affected_Fields"] = [", ".join(items[:8]) if items else "" for items in diff_fields_by_row]
    work["Row_Status"] = work["row_has_diff"].map(lambda x: "MISMATCH" if x else "MATCH")
    work["Comment"] = work["row_has_diff"].map(lambda x: "Output tensor changed between two consecutive reads without a new Invoke()." if x else "Output tensor remained stable between two consecutive reads without a new Invoke().")

    mismatch = work.loc[work["row_has_diff"]].copy()

    summary = {
        "rows_reference": int(len(work)),
        "rows_log": int(len(work)),
        "rows_both": int(len(work)),
        "rows_only_reference": 0,
        "rows_only_log": 0,
        "rows_with_diff": int(row_has_diff.sum()),
        "comparisons_with_diff": int(comparisons_with_diff),
        "total_comparisons": int(total_comparisons),
        "max_abs_diff": 0.0,
        "exact_counts": exact_counts,
        "tolerance": 0.0,
        "total_compare_cols": len(compare_cols),
        "merge_keys": list(key_cols),
        "duplicate_key_alignment_applied": False,
        "not_applicable": False,
    }
    return mismatch, summary


def humanize_output_stability(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame([{"Message": "No instability was found in DBG_MODEL_OUT_STABILITY_CSV."}])

    keep_front = [c for c in [
        "idx", "epoch", "out_idx", "Row_Status", "Most_Affected_Fields", "Comment",
        "type_a_code", "type_b_code", "type_a_name", "type_b_name",
        "bytes_a_total", "bytes_b_total", "bytes_a_dumped", "bytes_b_dumped",
        "dims_a_size", "dims_b_size", "ptr_a_hex", "ptr_b_hex", "ptr_equal", "raw_equal",
    ] if c in df.columns]
    other_cols = [c for c in df.columns if c not in keep_front and c not in {"row_has_diff"}]
    ordered = df[keep_front + other_cols].copy()
    return ordered


# ========================= Compare =========================
def _resolve_tolerance_for_column(
    tolerance: Union[float, Dict[str, float]],
    col: str,
) -> float:
    if isinstance(tolerance, dict):
        if col in tolerance:
            return float(tolerance[col])
        raise KeyError(f"Missing per-column tolerance for compared column: {col}")
    return float(tolerance)


def _tolerance_label(tolerance: Union[float, Dict[str, float]]) -> str:
    if isinstance(tolerance, dict):
        return "stage-aware / per-column"
    return str(float(tolerance))


def merge_and_compare(
    ref_df: pd.DataFrame,
    log_df: pd.DataFrame,
    key_cols: Sequence[str],
    compare_cols: Sequence[str],
    tolerance: Union[float, Dict[str, float]],
    applicability_mode: Optional[str] = None,
) -> Tuple[pd.DataFrame, Dict]:
    ref_aligned, log_aligned, merge_keys = _align_duplicate_keys(ref_df, log_df, key_cols)
    merged = ref_aligned.merge(
        log_aligned,
        on=merge_keys,
        how="outer",
        suffixes=("_py", "_fw"),
        indicator=True,
        sort=True,
    )

    numeric_diff_cols: List[str] = []
    numeric_exceed_cols: List[str] = []
    exact_mismatch_cols: List[str] = []
    exact_counts: Dict[str, int] = {}
    max_abs_diff = 0.0
    comparisons_with_diff = 0
    total_comparisons = 0

    both_mask = merged["_merge"].eq("both")

    for col in compare_cols:
        py_col = f"{col}_py"
        fw_col = f"{col}_fw"
        diff_col = f"diff_{col}"
        exact_col = f"exact_{col}"
        exceed_col = f"exceeds_tol_{col}"

        applicable_mask = both_mask
        if applicability_mode == "raw_output_tensor":
            applicable_mask = _raw_tensor_applicable_mask(merged, col, both_mask)

        value_mask = applicable_mask & merged[py_col].notna() & merged[fw_col].notna()

        if _is_exact_compare_field(col):
            merged[diff_col] = pd.NA
            merged[exact_col] = pd.NA
            merged[exceed_col] = False
            if value_mask.any():
                py_norm = merged.loc[value_mask, py_col].astype(str).str.strip().str.upper()
                fw_norm = merged.loc[value_mask, fw_col].astype(str).str.strip().str.upper()
                merged.loc[value_mask, exact_col] = py_norm.eq(fw_norm)
                merged.loc[value_mask, exceed_col] = ~merged.loc[value_mask, exact_col].fillna(True)
            exact_mismatch_cols.append(exact_col)
            exact_counts[col] = int(merged.loc[value_mask, exact_col].fillna(False).sum())
            total_comparisons += int(value_mask.sum())
            comparisons_with_diff += int(merged.loc[value_mask, exceed_col].fillna(False).sum())
            continue

        py_num = pd.to_numeric(merged[py_col], errors="coerce")
        fw_num = pd.to_numeric(merged[fw_col], errors="coerce")
        numeric_mask = value_mask & py_num.notna() & fw_num.notna()

        if value_mask.any() and not numeric_mask.loc[value_mask].all():
            merged[diff_col] = pd.NA
            merged[exact_col] = pd.NA
            merged[exceed_col] = False
            if value_mask.any():
                py_norm = merged.loc[value_mask, py_col].astype(str).str.strip().str.upper()
                fw_norm = merged.loc[value_mask, fw_col].astype(str).str.strip().str.upper()
                merged.loc[value_mask, exact_col] = py_norm.eq(fw_norm)
                merged.loc[value_mask, exceed_col] = ~merged.loc[value_mask, exact_col].fillna(True)
            exact_mismatch_cols.append(exact_col)
            exact_counts[col] = int(merged.loc[value_mask, exact_col].fillna(False).sum())
            total_comparisons += int(value_mask.sum())
            comparisons_with_diff += int(merged.loc[value_mask, exceed_col].fillna(False).sum())
            continue

        col_tolerance = _resolve_tolerance_for_column(tolerance, col)
        merged[diff_col] = py_num - fw_num
        merged[exact_col] = pd.NA
        merged[exceed_col] = False
        if numeric_mask.any():
            abs_diff = merged.loc[numeric_mask, diff_col].abs()
            within_tol = abs_diff <= col_tolerance
            exact_series = merged[exact_col].copy()
            exact_series.loc[numeric_mask] = within_tol
            merged[exact_col] = exact_series
            exceed_series = merged[exceed_col].copy()
            exceed_series.loc[numeric_mask] = ~within_tol
            merged[exceed_col] = exceed_series
        numeric_diff_cols.append(diff_col)
        numeric_exceed_cols.append(exceed_col)

        exact_counts[col] = int(merged.loc[value_mask, exact_col].fillna(False).sum())
        total_comparisons += int(numeric_mask.sum())
        comparisons_with_diff += int(merged.loc[numeric_mask, exceed_col].fillna(False).sum())

        if numeric_mask.any():
            col_max = float(merged.loc[numeric_mask, diff_col].abs().max())
            max_abs_diff = max(max_abs_diff, col_max)

    if numeric_diff_cols:
        merged["row_max_abs_diff"] = merged[numeric_diff_cols].abs().max(axis=1, skipna=True)
    else:
        merged["row_max_abs_diff"] = pd.NA

    if numeric_exceed_cols:
        numeric_exceed_eval = merged[numeric_exceed_cols].copy().fillna(False)
        merged["row_has_numeric_diff"] = both_mask & numeric_exceed_eval.any(axis=1)
    else:
        merged["row_has_numeric_diff"] = False

    if exact_mismatch_cols:
        exact_eval = merged[exact_mismatch_cols].copy().fillna(True)
        merged["row_has_exact_diff"] = both_mask & (~exact_eval.all(axis=1))
    else:
        merged["row_has_exact_diff"] = False

    merged["row_has_diff"] = merged["row_has_numeric_diff"] | merged["row_has_exact_diff"]

    mismatch = merged[(merged["_merge"] != "both") | (merged["row_has_diff"])].copy()

    summary = {
        "rows_reference": int(len(ref_df)),
        "rows_log": int(len(log_df)),
        "rows_both": int((merged["_merge"] == "both").sum()),
        "rows_only_reference": int((merged["_merge"] == "left_only").sum()),
        "rows_only_log": int((merged["_merge"] == "right_only").sum()),
        "rows_with_diff": int(merged["row_has_diff"].sum()),
        "comparisons_with_diff": int(comparisons_with_diff),
        "total_comparisons": int(total_comparisons),
        "max_abs_diff": float(max_abs_diff),
        "exact_counts": exact_counts,
        "tolerance": _normalize_tolerance_value(tolerance),
        "tolerance_label": _tolerance_label(tolerance),
        "total_compare_cols": len(compare_cols),
        "merge_keys": list(merge_keys),
        "duplicate_key_alignment_applied": bool(list(merge_keys) != list(key_cols)),
    }
    return mismatch, summary




PREPROCESS_STAGE_TOLERANCE = SEMANTIC_DEBUG_TOLERANCE

PRE_LAGS_FIELDS = [
    "state_Tin_lag1_phys_raw",
    "state_Hin_lag1_phys_raw",
    "state_Tout_lag1_phys_raw",
    "state_Hout_lag1_phys_raw",
    "state_Tin_lag2_phys_raw",
    "state_Hin_lag2_phys_raw",
]
PRE_TIME_FIELDS = [
    "state_sin_hour",
    "state_cos_hour",
    "state_weekday",
    "state_month",
]
PRE_PHYS_FIELDS_REGEX = r"^in_f\d+_phys_raw$"
PRE_CLIP_FIELDS_REGEX = r"^in_f\d+_phys_clip$"
PRE_SCALED_FIELDS_REGEX = r"^in_f\d+_scaled$"
PRE_MODEL_IN_FIELDS_REGEX = r"^in_x\d+_float$"

def _available_cols(df: pd.DataFrame, preferred: Sequence[str]) -> List[str]:
    return [c for c in preferred if c in df.columns]

def _regex_cols(df: pd.DataFrame, pattern: str) -> List[str]:
    rx = re.compile(pattern)
    return sorted([c for c in df.columns if rx.match(c)], key=_natural_sort_key)

def _prepare_optional_preprocess_df(df: Optional[pd.DataFrame], key_cols: Sequence[str], compare_cols: Sequence[str]) -> pd.DataFrame:
    if df is None or df.empty or not compare_cols:
        return pd.DataFrame(columns=list(key_cols))
    expected = [c for c in list(key_cols) + list(compare_cols) if c in df.columns]
    if not all(k in df.columns for k in key_cols):
        return pd.DataFrame(columns=list(key_cols))
    return normalize_numeric(df[expected].copy())

def _build_preprocess_stage(
    ref_df: pd.DataFrame,
    log_df: pd.DataFrame,
    key_cols: Sequence[str],
    compare_cols: Sequence[str],
    title: str,
):
    if not compare_cols or ref_df.empty or log_df.empty:
        summary = make_empty_summary(len(ref_df), len(log_df), PREPROCESS_STAGE_TOLERANCE, key_cols)
        return make_summary_df(title, summary), pd.DataFrame([{"Message": f"No comparable data available for {title}."}]), summary
    mismatch, summary = compare_subset(ref_df, log_df, key_cols, compare_cols, PREPROCESS_STAGE_TOLERANCE)
    human_df = humanize_input_mismatch(mismatch, compare_cols)
    return make_summary_df(title, summary), human_df, summary

# ========================= Human interpretation =========================
def pct(num: int, den: int) -> float:
    if den <= 0:
        return 0.0
    return 100.0 * float(num) / float(den)


def comparison_pct(summary: Dict) -> float:
    return pct(int(summary.get("comparisons_with_diff", 0)), int(summary.get("total_comparisons", 0)))


def _stage_match_fraction(summary: Dict) -> str:
    total = int(summary.get("rows_both", 0))
    diff = int(summary.get("rows_with_diff", 0))
    ok = max(total - diff, 0)
    return f"{ok}/{total}" if total > 0 else "0/0"


def _top_exact_count_fields(summary: Dict, top_n: int = 3) -> str:
    counts = summary.get("exact_counts", {}) or {}
    if not counts:
        return "N/A"
    ordered = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]))
    return ", ".join(f"{name}={count}" for name, count in ordered[:top_n])


def _stage_interpretation(stage_name: str, summary: Dict, purpose: str, failure_hint: str) -> str:
    status = classify_result(summary)
    matched = _stage_match_fraction(summary)
    rate = pct(summary.get("rows_with_diff", 0), summary.get("rows_both", 0))
    tol = summary.get("tolerance_label", summary.get("tolerance", ""))
    if status == "MATCH":
        return (
            f"{stage_name}: MATCH. {matched} matched rows stayed within tolerance "
            f"({tol}). This stage supports {purpose}."
        )
    if status == "MISMATCH":
        return (
            f"{stage_name}: MISMATCH. {matched} matched rows stayed within tolerance and "
            f"{rate:.2f}% exceeded it. This suggests {failure_hint}."
        )
    if status == "INCOMPLETE":
        return (
            f"{stage_name}: INCOMPLETE. Some rows exist only on one side, so the comparison is "
            f"not fully 1:1 yet. Resolve row coverage before using this stage as final evidence."
        )
    return f"{stage_name}: N/A for the detected schema."


def _recommended_next_action(
    input_critical_summary: Dict,
    output_tensor_raw_summary: Dict,
    output_raw_summary: Dict,
    output_semantic_summary: Dict,
    output_final_summary: Dict,
) -> str:
    input_status = classify_result(input_critical_summary)
    tensor_raw_status = classify_result(output_tensor_raw_summary)
    raw_status = classify_result(output_raw_summary)
    semantic_status = classify_result(output_semantic_summary)
    final_status = classify_result(output_final_summary)
    if input_status != "MATCH":
        return (
            "Prioritize the input tensor first: review merge keys, row alignment, flatten order, and the exact values of in_x*_float before interpreting any output mismatch."
        )
    if tensor_raw_status not in {"MATCH", "N/A"}:
        return (
            "Input is already closed. Focus on the raw output tensor dump immediately after Invoke(): inspect only the failing rows in type/shape/bytes and b*_hex to confirm the mismatch is already present in the tensor buffer."
        )
    if raw_status != "MATCH":
        return (
            "The raw tensor dump is closed. Focus on immediate output decode/copy: inspect only the rows that fail in out_o*_float/out_o*_bits_hex and confirm how the tensor buffer is being interpreted as floats."
        )
    if semantic_status != "MATCH":
        return (
            "Tensor dump and decoded raw output are closed, so the next step is semantic mapping: verify output ordering and the assignment from positional outputs to temperature/humidity targets."
        )
    if final_status != "MATCH":
        return (
            "Input, tensor dump, decoded raw output, and semantic output are closed. Focus on inverse-transform and absolute reconstruction, especially d_*_pred and p_*_pred tolerance handling."
        )
    return (
        "The critical path is closed. Any remaining work is optional cleanup in semantic/state debug fields or presentation-level rounding checks."
    )


def build_human_summary(
    input_critical_summary: Dict,
    output_tensor_raw_summary: Dict,
    output_raw_summary: Dict,
    output_semantic_summary: Dict,
    output_final_summary: Dict,
    input_semantic_summary: Optional[Dict] = None,
    output_aux_summary: Optional[Dict] = None,
) -> Dict[str, str]:
    input_status = classify_result(input_critical_summary)
    tensor_raw_status = classify_result(output_tensor_raw_summary)
    raw_status = classify_result(output_raw_summary)
    semantic_status = classify_result(output_semantic_summary)
    final_status = classify_result(output_final_summary)

    input_rows = input_critical_summary["rows_both"]
    tensor_raw_rows = output_tensor_raw_summary["rows_both"]
    raw_rows = output_raw_summary["rows_both"]
    semantic_rows = output_semantic_summary["rows_both"]
    final_rows = output_final_summary["rows_both"]

    input_row_diff_pct = pct(input_critical_summary["rows_with_diff"], input_rows)
    tensor_raw_row_diff_pct = pct(output_tensor_raw_summary["rows_with_diff"], tensor_raw_rows)
    raw_row_diff_pct = pct(output_raw_summary["rows_with_diff"], raw_rows)
    semantic_row_diff_pct = pct(output_semantic_summary["rows_with_diff"], semantic_rows)
    final_row_diff_pct = pct(output_final_summary["rows_with_diff"], final_rows)

    input_diff_pct = comparison_pct(input_critical_summary)
    tensor_raw_diff_pct = comparison_pct(output_tensor_raw_summary)
    raw_diff_pct = comparison_pct(output_raw_summary)
    semantic_diff_pct = comparison_pct(output_semantic_summary)
    final_diff_pct = comparison_pct(output_final_summary)

    input_semantic_status = classify_result(input_semantic_summary or make_empty_summary(0, 0, SEMANTIC_DEBUG_TOLERANCE, []))
    output_aux_status = classify_result(output_aux_summary or make_empty_summary(0, 0, SEMANTIC_DEBUG_TOLERANCE, []))

    if input_status != "MATCH":
        conclusion = "The critical model input tensor still diverges; fix input packaging before interpreting any LSTM output mismatch."
        interpretation = (
            "A discrepancy or row-alignment issue exists before inference. The output comparison cannot be considered primary until the input tensor matches."
        )
    elif tensor_raw_status not in {"MATCH", "N/A"}:
        conclusion = "The critical model input tensor agrees, but the immediate raw output tensor dump diverges."
        interpretation = (
            "The primary discrepancy starts immediately after inference in the output tensor buffer/runtime path, not in input packaging."
        )
    elif raw_status != "MATCH":
        conclusion = "The raw output tensor dump agrees, but the decoded raw model output diverges."
        interpretation = (
            "Investigate immediate output-tensor decoding/copy, because the tensor bytes already match while the interpreted floats do not."
        )
    elif semantic_status != "MATCH":
        conclusion = "Decoded raw model output agrees, but semantic output mapping diverges."
        interpretation = (
            "Investigate the mapping from positional output tensors to semantic targets (temperature/humidity)."
        )
    elif final_status != "MATCH":
        conclusion = "Tensor dump, decoded raw output, and semantic output agree, but final prediction reconstruction diverges."
        interpretation = (
            "Investigate inverse-transform and absolute reconstruction steps (d_* and p_*), including tolerance choice for reconstructed values."
        )
    else:
        conclusion = "Critical input, raw tensor dump, decoded raw output, semantic output, and final prediction all agree within their stage tolerances."
        interpretation = (
            "No evidence of a discrepancy was found in the critical model I/O path. Any remaining differences are restricted to auxiliary semantic/state debug fields."
        )

    return {
        "overall_conclusion": conclusion,
        "technical_interpretation": interpretation,
        "critical_input_status": input_status,
        "tensor_raw_status": tensor_raw_status,
        "raw_output_status": raw_status,
        "semantic_output_status": semantic_status,
        "final_prediction_status": final_status,
        "critical_input_row_mismatch_rate": f"{input_row_diff_pct:.2f}%",
        "tensor_raw_row_mismatch_rate": f"{tensor_raw_row_diff_pct:.2f}%",
        "raw_output_row_mismatch_rate": f"{raw_row_diff_pct:.2f}%",
        "semantic_output_row_mismatch_rate": f"{semantic_row_diff_pct:.2f}%",
        "final_prediction_row_mismatch_rate": f"{final_row_diff_pct:.2f}%",
        "critical_input_mismatch_rate": f"{input_diff_pct:.2f}%",
        "tensor_raw_mismatch_rate": f"{tensor_raw_diff_pct:.2f}%",
        "raw_output_mismatch_rate": f"{raw_diff_pct:.2f}%",
        "semantic_output_mismatch_rate": f"{semantic_diff_pct:.2f}%",
        "final_prediction_mismatch_rate": f"{final_diff_pct:.2f}%",
        "critical_input_match_fraction": _stage_match_fraction(input_critical_summary),
        "tensor_raw_match_fraction": _stage_match_fraction(output_tensor_raw_summary),
        "raw_output_match_fraction": _stage_match_fraction(output_raw_summary),
        "semantic_output_match_fraction": _stage_match_fraction(output_semantic_summary),
        "final_prediction_match_fraction": _stage_match_fraction(output_final_summary),
        "critical_input_interpretation": _stage_interpretation(
            "Critical input",
            input_critical_summary,
            purpose="the claim that the model is receiving the same tensor on both sides",
            failure_hint="a pre-inference mismatch in packing, flattening, or row alignment",
        ),
        "tensor_raw_interpretation": _stage_interpretation(
            "Raw tensor dump",
            output_tensor_raw_summary,
            purpose="the claim that the immediate output tensor buffer agrees byte-for-byte and metadata-for-metadata",
            failure_hint="the mismatch is already present in the output tensor buffer immediately after inference",
        ),
        "raw_output_interpretation": _stage_interpretation(
            "Decoded raw output",
            output_raw_summary,
            purpose="the claim that runtime inference and immediate tensor decode agree",
            failure_hint="the mismatch appears during immediate output-tensor decoding/copy after the tensor buffer itself",
        ),
        "semantic_output_interpretation": _stage_interpretation(
            "Semantic output",
            output_semantic_summary,
            purpose="the claim that output ordering/mapping is consistent",
            failure_hint="the mismatch is present at the semantic stage; review semantic mapping only if the decoded raw output is already clean, otherwise treat this as downstream propagation",
        ),
        "final_prediction_interpretation": _stage_interpretation(
            "Final prediction",
            output_final_summary,
            purpose="the claim that post-processing and reconstruction agree",
            failure_hint="the mismatch persists through inverse-transform and final reconstruction; if tensor/raw/semantic output already fails, treat this mostly as downstream propagation rather than the root cause",
        ),
        "critical_input_top_non_exact": _top_exact_count_fields(input_critical_summary),
        "tensor_raw_top_non_exact": _top_exact_count_fields(output_tensor_raw_summary),
        "raw_output_top_non_exact": _top_exact_count_fields(output_raw_summary),
        "semantic_output_top_non_exact": _top_exact_count_fields(output_semantic_summary),
        "final_prediction_top_non_exact": _top_exact_count_fields(output_final_summary),
        "recommended_next_action": _recommended_next_action(
            input_critical_summary,
            output_tensor_raw_summary,
            output_raw_summary,
            output_semantic_summary,
            output_final_summary,
        ),
        "semantic_input_status": input_semantic_status,
        "semantic_aux_output_status": output_aux_status,
        "semantic_input_mismatch_rate": (
            "N/A"
            if input_semantic_summary is None or input_semantic_summary.get("not_applicable")
            else f"{comparison_pct(input_semantic_summary):.2f}%"
        ),
        "semantic_aux_output_mismatch_rate": (
            "N/A"
            if output_aux_summary is None or output_aux_summary.get("not_applicable")
            else f"{comparison_pct(output_aux_summary):.2f}%"
        ),
        "semantic_input_interpretation": (
            "Auxiliary state/debug fields are consistent enough for semantic inspection."
            if input_semantic_status == "MATCH"
            else "Auxiliary state/debug fields still diverge. This does not invalidate the critical tensor result, but it means the semantic layer is not yet 1:1."
        ),
        "semantic_aux_output_interpretation": (
            "Auxiliary output/debug fields are consistent enough for semantic inspection."
            if output_aux_status == "MATCH"
            else "Auxiliary output/debug fields still diverge. Treat them as secondary evidence only."
        ),
    }


def build_general_overview(
    input_full_summary: Dict,
    output_full_summary: Dict,
    input_critical_summary: Dict,
    output_tensor_raw_summary: Dict,
    output_raw_summary: Dict,
    output_semantic_summary: Dict,
    output_final_summary: Dict,
    input_semantic_summary: Optional[Dict] = None,
    output_aux_summary: Optional[Dict] = None,
    output_stability_summary: Optional[Dict] = None,
) -> pd.DataFrame:
    entries = [
        ("Input / full compared columns (stage-aware)", input_full_summary),
        ("Input / critical tensor columns", input_critical_summary),
        ("Postprocess / full compared columns (stage-aware)", output_full_summary),
        ("Postprocess / raw tensor dump columns", output_tensor_raw_summary),
        ("Postprocess / raw output columns", output_raw_summary),
        ("Postprocess / semantic output columns", output_semantic_summary),
        ("Postprocess / final prediction columns", output_final_summary),
    ]
    if input_semantic_summary is not None:
        entries.append(("Input / semantic-state columns", input_semantic_summary))
    if output_aux_summary is not None:
        entries.append(("Postprocess / semantic-aux columns", output_aux_summary))
    if output_stability_summary is not None:
        entries.append(("Postprocess / stability columns (firmware-only)", output_stability_summary))

    rows = []
    for name, summary in entries:
        rows.append(
            {
                "Block": name,
                "Python_Reference_Rows": summary["rows_reference"],
                "Firmware_Log_Rows": summary["rows_log"],
                "Matched_Rows": summary["rows_both"],
                "Python_Only_Rows": summary["rows_only_reference"],
                "Firmware_Only_Rows": summary["rows_only_log"],
                "Rows_Above_Tolerance": summary["rows_with_diff"],
                "Row_Mismatch_Rate_Percent": round(
                    pct(summary["rows_with_diff"], summary["rows_both"]), 2
                ),
                "Comparisons_Above_Tolerance": int(summary.get("comparisons_with_diff", 0)),
                "Total_Comparisons": int(summary.get("total_comparisons", 0)),
                "Scalar_Mismatch_Rate_Percent": round(comparison_pct(summary), 2),
                "Maximum_Absolute_Difference": summary["max_abs_diff"],
                "Compared_Columns": summary.get("total_compare_cols", 0),
                "Status": classify_result(summary),
            }
        )
    return pd.DataFrame(rows)



TOLERANCE_CALIBRATION_FACTORS: Tuple[float, float, float] = (1.0, 1.5, 2.0)


def _summary_total_comparisons(summary: Dict) -> int:
    return int(summary.get("total_comparisons", 0))


def _summary_rows_both(summary: Dict) -> int:
    return int(summary.get("rows_both", 0))


def _coerce_float(value: object) -> Optional[float]:
    try:
        return float(value)  # type: ignore[arg-type]
    except Exception:
        return None


def _tolerance_calibration_status(summary: Dict, adopted_tolerance: object) -> str:
    if summary.get("not_applicable"):
        return "NOT_APPLICABLE"
    if _summary_rows_both(summary) <= 0 or _summary_total_comparisons(summary) <= 0:
        return "NO_DATA"
    status = classify_result(summary)
    if status != "MATCH":
        return "NOT_ELIGIBLE_MISMATCH"
    adopted = _coerce_float(adopted_tolerance)
    if adopted is not None and adopted == 0.0:
        return "EXACT_COMPARE"
    return "ELIGIBLE_MATCH"


def _tolerance_protocol_note(stage_name: str, summary: Dict, adopted_tolerance: object) -> str:
    calibration_status = _tolerance_calibration_status(summary, adopted_tolerance)
    if calibration_status == "EXACT_COMPARE":
        return (
            "This stage is audited as exact equality. Scientific justification is structural: raw bytes/metadata must match exactly, so no empirical tolerance is used."
        )
    if calibration_status == "NOT_APPLICABLE":
        return "This stage is not applicable for the detected schema, so no calibration evidence is available."
    if calibration_status == "NO_DATA":
        return "This stage had no aligned comparisons, so no empirical tolerance evidence can be extracted from this workbook."
    if calibration_status == "NOT_ELIGIBLE_MISMATCH":
        return (
            "This stage is not eligible for tolerance calibration from this workbook because it is not MATCH. Using a failing run to set tolerance would risk normalizing a real discrepancy."
        )
    observed_max = float(summary.get("max_abs_diff", 0.0) or 0.0)
    return (
        "This stage is eligible for empirical tolerance justification because the comparison closed as MATCH. "
        f"Observed max abs diff in this workbook = {observed_max:.10g}. Use multiple validated MATCH runs to justify a publication-grade bound; do not rely on a single workbook alone."
    )


def build_tolerance_protocol_df(stage_specs: Sequence[Tuple[str, Dict, object, str]]) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    for stage_name, summary, adopted_tolerance, rationale in stage_specs:
        observed_max = float(summary.get("max_abs_diff", 0.0) or 0.0)
        calibration_status = _tolerance_calibration_status(summary, adopted_tolerance)
        row = {
            "Stage": stage_name,
            "Current_Status": classify_result(summary),
            "Calibration_Status": calibration_status,
            "Rows_Matched": _summary_rows_both(summary),
            "Rows_Above_Tolerance": int(summary.get("rows_with_diff", 0)),
            "Total_Comparisons": _summary_total_comparisons(summary),
            "Current_Adopted_Tolerance": adopted_tolerance,
            "Observed_Max_Absolute_Difference": observed_max,
            "Observed_Scalar_Mismatch_Rate_Percent": round(comparison_pct(summary), 6),
            "Observed_Row_Mismatch_Rate_Percent": round(pct(summary.get("rows_with_diff", 0), summary.get("rows_both", 0)), 6),
            "Suggested_Tolerance_1.0x_Max": None,
            "Suggested_Tolerance_1.5x_Max": None,
            "Suggested_Tolerance_2.0x_Max": None,
            "Scientific_Rationale": rationale,
            "Protocol_Note": _tolerance_protocol_note(stage_name, summary, adopted_tolerance),
        }

        if calibration_status == "EXACT_COMPARE":
            row["Suggested_Tolerance_1.0x_Max"] = 0.0
            row["Suggested_Tolerance_1.5x_Max"] = 0.0
            row["Suggested_Tolerance_2.0x_Max"] = 0.0
        elif calibration_status == "ELIGIBLE_MATCH":
            row["Suggested_Tolerance_1.0x_Max"] = observed_max * TOLERANCE_CALIBRATION_FACTORS[0]
            row["Suggested_Tolerance_1.5x_Max"] = observed_max * TOLERANCE_CALIBRATION_FACTORS[1]
            row["Suggested_Tolerance_2.0x_Max"] = observed_max * TOLERANCE_CALIBRATION_FACTORS[2]

        rows.append(row)

    protocol_rows = [
        {
            "Stage": "PROTOCOL",
            "Current_Status": "INFO",
            "Calibration_Status": "FIXED_TOLERANCES_ACTIVE",
            "Rows_Matched": None,
            "Rows_Above_Tolerance": None,
            "Total_Comparisons": None,
            "Current_Adopted_Tolerance": None,
            "Observed_Max_Absolute_Difference": None,
            "Observed_Scalar_Mismatch_Rate_Percent": None,
            "Observed_Row_Mismatch_Rate_Percent": None,
            "Suggested_Tolerance_1.0x_Max": None,
            "Suggested_Tolerance_1.5x_Max": None,
            "Suggested_Tolerance_2.0x_Max": None,
            "Scientific_Rationale": "The workbook still uses the fixed tolerances defined in the script for pass/fail decisions.",
            "Protocol_Note": "The three Suggested_Tolerance_* columns are empirical support values derived from observed max abs diff in this workbook only when the stage is MATCH. They do not overwrite the adopted tolerances.",
        },
        {
            "Stage": "PROTOCOL",
            "Current_Status": "INFO",
            "Calibration_Status": "PUBLICATION_GUIDANCE",
            "Rows_Matched": None,
            "Rows_Above_Tolerance": None,
            "Total_Comparisons": None,
            "Current_Adopted_Tolerance": None,
            "Observed_Max_Absolute_Difference": None,
            "Observed_Scalar_Mismatch_Rate_Percent": None,
            "Observed_Row_Mismatch_Rate_Percent": None,
            "Suggested_Tolerance_1.0x_Max": None,
            "Suggested_Tolerance_1.5x_Max": None,
            "Suggested_Tolerance_2.0x_Max": None,
            "Scientific_Rationale": "A defensible scientific tolerance should be justified from multiple validated MATCH runs, not from a single failing workbook.",
            "Protocol_Note": "Recommended practice: collect observed max abs diff across trusted runs, then adopt a bound with an explicit safety factor and document both the dataset and factor in the report.",
        },
    ]
    return pd.DataFrame(protocol_rows + rows)

def make_summary_df(title: str, summary: Dict) -> pd.DataFrame:
    rows = [
        ("Analyzed block", title),
        ("Python reference rows", summary["rows_reference"]),
        ("Firmware log rows", summary["rows_log"]),
        ("Matched rows", summary["rows_both"]),
        ("Python-only rows", summary["rows_only_reference"]),
        ("Firmware-only rows", summary["rows_only_log"]),
        ("Rows above tolerance", summary["rows_with_diff"]),
        ("Row mismatch rate", f"{pct(summary['rows_with_diff'], summary['rows_both']):.2f}%"),
        ("Comparisons above tolerance", int(summary.get("comparisons_with_diff", 0))),
        ("Total comparisons", int(summary.get("total_comparisons", 0))),
        ("Scalar mismatch rate", f"{comparison_pct(summary):.2f}%"),
        ("Maximum absolute difference", summary["max_abs_diff"]),
        ("Tolerance", summary.get("tolerance_label", summary["tolerance"])),
        ("Merge keys used", ", ".join(summary.get("merge_keys", []))),
        ("Duplicate-key alignment applied", summary.get("duplicate_key_alignment_applied", False)),
        ("Final status", classify_result(summary)),
    ]

    for k, v in summary["exact_counts"].items():
        rows.append((f"Exact-match rows in column {k}", v))

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _availability_value_present(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, pd.DataFrame):
        return not value.empty
    if isinstance(value, Path):
        return value.exists()
    return bool(value)


def build_data_availability_df(rows: Sequence[Tuple[str, object, str, str]]) -> pd.DataFrame:
    records = []
    for block, value, missing_policy, note in rows:
        present = _availability_value_present(value)
        records.append(
            {
                "Block": block,
                "Present": "YES" if present else "NO",
                "Availability_Status": "AVAILABLE" if present else "MISSING_DATA",
                "Missing_Data_Policy": missing_policy,
                "Notes": note if note else ("" if present else "Required data is absent."),
            }
        )
    return pd.DataFrame(records)


def summarize_missing_blocks(availability_df: pd.DataFrame) -> str:
    if availability_df.empty:
        return "No availability audit rows were generated."
    missing = availability_df.loc[availability_df["Present"].eq("NO"), "Block"].tolist()
    if not missing:
        return "No required or monitored data blocks are missing."
    return "Missing blocks: " + ", ".join(missing)


def top_diff_fields(row: pd.Series, compare_cols: Sequence[str], top_n: int = 4) -> str:
    pairs = []
    exact_mismatches: List[str] = []

    for col in compare_cols:
        diff_col = f"diff_{col}"
        exact_col = f"exact_{col}"

        val = row.get(diff_col)
        if pd.notna(val):
            pairs.append((col, abs(float(val))))
            continue

        exact_val = row.get(exact_col)
        if exact_val is False:
            exact_mismatches.append(col)

    pairs.sort(key=lambda x: x[1], reverse=True)

    items: List[str] = [f"{name}={value:.8f}" for name, value in pairs[:top_n]]
    remaining = max(top_n - len(items), 0)
    items.extend([f"{name}=EXACT_MISMATCH" for name in exact_mismatches[:remaining]])

    return ", ".join(items)


def _row_sort_columns(df: pd.DataFrame) -> List[str]:
    return [c for c in ["row_max_abs_diff", "idx", "epoch", "step"] if c in df.columns]


def humanize_input_mismatch(df: pd.DataFrame, compare_cols: Sequence[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            [{"Message": "No discrepancies were found in the immediate model input block."}]
        )

    rows = []
    order_cols = _row_sort_columns(df)
    ordered = df.sort_values(by=order_cols, ascending=[False] + [True] * (len(order_cols) - 1))

    for _, row in ordered.iterrows():
        out_row = {
            "idx": row.get("idx"),
            "epoch": row.get("epoch"),
            "Row_Status": row.get("_merge"),
            "Maximum_Row_Absolute_Difference": row.get("row_max_abs_diff"),
            "Most_Affected_Fields": top_diff_fields(row, compare_cols, top_n=min(6, len(compare_cols))),
            "Comment": (
                "Difference detected in the immediate model input block."
                if row.get("_merge") == "both"
                else "Unmatched row between Python reference and firmware log."
            ),
        }
        if "step" in df.columns:
            out_row["step"] = row.get("step")
        rows.append(out_row)
    return pd.DataFrame(rows)


def humanize_output_mismatch(df: pd.DataFrame, compare_cols: Sequence[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            [{"Message": "No discrepancies were found in the immediate model output block."}]
        )

    rows = []
    order_cols = _row_sort_columns(df)
    ordered = df.sort_values(by=order_cols, ascending=[False] + [True] * (len(order_cols) - 1))

    for _, row in ordered.iterrows():
        out_row = {
            "idx": row.get("idx"),
            "epoch": row.get("epoch"),
            "Row_Status": row.get("_merge"),
            "Maximum_Row_Absolute_Difference": row.get("row_max_abs_diff"),
            "Most_Affected_Fields": top_diff_fields(row, compare_cols, top_n=min(6, len(compare_cols))),
            "Comment": (
                "Difference detected in the immediate model output block."
                if row.get("_merge") == "both"
                else "Unmatched row between Python reference and firmware log."
            ),
        }
        for col in compare_cols:
            out_row[f"{col}_Python"] = row.get(f"{col}_py")
            out_row[f"{col}_Firmware"] = row.get(f"{col}_fw")
            out_row[f"diff_{col}"] = row.get(f"diff_{col}")
            exact_col = f"exact_{col}"
            if exact_col in row.index:
                out_row[exact_col] = row.get(exact_col)
        rows.append(out_row)
    return pd.DataFrame(rows)


# ========================= Excel helpers =========================
def autosize(ws, min_width: int = 10, max_width: int = 46) -> None:
    for col_cells in ws.columns:
        idx = col_cells[0].column
        values = ["" if c.value is None else str(c.value) for c in col_cells]
        width = max(len(v) for v in values) + 2
        width = max(min_width, min(width, max_width))
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def apply_status_fill(cell) -> None:
    if cell.value in {"MATCH", "VALID", "AVAILABLE"}:
        cell.fill = GOOD_FILL
    elif cell.value in {"MISMATCH", "MISSING_DATA"}:
        cell.fill = BAD_FILL
    elif cell.value == "INCOMPLETE":
        cell.fill = WARN_FILL
    elif cell.value == "N/A":
        cell.fill = INFO_FILL


def _excel_safe_value(value):
    if value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, str):
        value = ILLEGAL_XML_CHARS_RE.sub("", value)
        if len(value) > MAX_EXCEL_CELL_CHARS:
            value = value[:MAX_EXCEL_CELL_CHARS]
        return value

    return value


def write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(start_row, j, col)

    style_header(ws[start_row])

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            safe_value = _excel_safe_value(value)
            cell = ws.cell(i, j, safe_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER_THIN

            header_name = str(ws.cell(start_row, j).value)
            if header_name == "Status":
                apply_status_fill(cell)

    ws.freeze_panes = ws.cell(start_row + 1, start_col)
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def fill_row(ws, row_idx: int, start_col: int, end_col: int, fill: PatternFill) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row_idx, col_idx).fill = fill


def add_executive_summary_sheet(
    wb: Workbook,
    input_full_summary: Dict,
    output_full_summary: Dict,
    input_critical_summary: Dict,
    output_tensor_raw_summary: Dict,
    output_raw_summary: Dict,
    output_semantic_summary: Dict,
    output_final_summary: Dict,
    input_semantic_summary: Optional[Dict],
    output_aux_summary: Optional[Dict],
    output_stability_summary: Optional[Dict],
    latest_run_dir: Path,
    input_ref_path: Path,
    output_ref_path: Path,
    output_raw_ref_path: Optional[Path],
    preprocess_ref_path: Optional[Path],
    log_txt_path: Path,
    out_xlsx_path: Path,
    input_schema_name: str,
    output_schema_name: str,
    output_tensor_raw_schema_name: str,
    input_compare_cols: Sequence[str],
    output_compare_cols: Sequence[str],
    input_critical_cols: Sequence[str],
    output_tensor_raw_cols: Sequence[str],
    output_raw_cols: Sequence[str],
    output_semantic_cols: Sequence[str],
    output_final_cols: Sequence[str],
    input_semantic_cols: Sequence[str],
    output_aux_cols: Sequence[str],
    availability_df: pd.DataFrame,
) -> None:
    ws = wb.active
    ws.title = "Executive_Summary"

    human = build_human_summary(
        input_critical_summary,
        output_tensor_raw_summary,
        output_raw_summary,
        output_semantic_summary,
        output_final_summary,
        input_semantic_summary,
        output_aux_summary,
    )

    ws["A1"] = "LiteML-Edge Preprocess, Immediate Model I/O, and Postprocess Validation Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "MODEL: MLP -> Python vs Firmware Comparison"
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = "Scope"
    ws["B3"] = (
        "Comparison between Python reference spreadsheets and firmware log CSV traces "
        "for immediate model input and immediate model output."
    )

    ws["A5"] = "Overall conclusion"
    ws["B5"] = human["overall_conclusion"]

    ws["A6"] = "Technical interpretation"
    ws["B6"] = human["technical_interpretation"]

    ws["A8"] = "Stage-by-stage decision"
    ws["A8"].font = SUBTITLE_FONT

    stage_rows = [
        (9,  "Critical input", human["critical_input_status"], human["critical_input_match_fraction"], human["critical_input_row_mismatch_rate"], human["critical_input_mismatch_rate"], human["critical_input_interpretation"], human["critical_input_top_non_exact"]),
        (10, "Raw tensor dump", human["tensor_raw_status"], human["tensor_raw_match_fraction"], human["tensor_raw_row_mismatch_rate"], human["tensor_raw_mismatch_rate"], human["tensor_raw_interpretation"], human["tensor_raw_top_non_exact"]),
        (11, "Decoded raw output", human["raw_output_status"], human["raw_output_match_fraction"], human["raw_output_row_mismatch_rate"], human["raw_output_mismatch_rate"], human["raw_output_interpretation"], human["raw_output_top_non_exact"]),
        (12, "Semantic output", human["semantic_output_status"], human["semantic_output_match_fraction"], human["semantic_output_row_mismatch_rate"], human["semantic_output_mismatch_rate"], human["semantic_output_interpretation"], human["semantic_output_top_non_exact"]),
        (13, "Final prediction", human["final_prediction_status"], human["final_prediction_match_fraction"], human["final_prediction_row_mismatch_rate"], human["final_prediction_mismatch_rate"], human["final_prediction_interpretation"], human["final_prediction_top_non_exact"]),
    ]
    ws["A9"] = "Stage"
    ws["B9"] = "Status"
    ws["C9"] = "Rows within tolerance"
    ws["D9"] = "Row mismatch rate"
    ws["E9"] = "Scalar mismatch rate"
    ws["F9"] = "What this means"
    ws["G9"] = "Least exact columns"
    style_header(ws[9])
    for r, stage_name, status, fraction, row_rate, scalar_rate, explanation, weakest in stage_rows:
        ws[f"A{r+1}"] = stage_name
        ws[f"B{r+1}"] = status
        ws[f"C{r+1}"] = fraction
        ws[f"D{r+1}"] = row_rate
        ws[f"E{r+1}"] = scalar_rate
        ws[f"F{r+1}"] = explanation
        ws[f"G{r+1}"] = weakest
        for c in range(1, 8):
            ws.cell(r+1, c).border = BORDER_THIN
            ws.cell(r+1, c).alignment = Alignment(vertical="top", wrap_text=True)
        apply_status_fill(ws[f"B{r+1}"])

    ws["A16"] = "Recommended next action"
    ws["B16"] = human["recommended_next_action"]

    ws["A18"] = "Semantic/state-only status"
    ws["A18"].font = SUBTITLE_FONT
    ws["A19"] = "Semantic input status"
    ws["B19"] = human["semantic_input_status"]
    ws["C19"] = human["semantic_input_mismatch_rate"]
    ws["D19"] = human["semantic_input_interpretation"]
    ws["A20"] = "Semantic auxiliary output status"
    ws["B20"] = human["semantic_aux_output_status"]
    ws["C20"] = human["semantic_aux_output_mismatch_rate"]
    ws["D20"] = human["semantic_aux_output_interpretation"]

    ws["A22"] = "Detected schemas"
    ws["A22"].font = SUBTITLE_FONT
    ws["A23"] = "Input schema"
    ws["B23"] = f"{input_schema_name} | full={len(input_compare_cols)} | critical={len(input_critical_cols)} | semantic={len(input_semantic_cols)}"
    ws["A24"] = "Output schema"
    ws["B24"] = (
        f"{output_schema_name} | full={len(output_compare_cols)} | tensor_raw={len(output_tensor_raw_cols)} | raw={len(output_raw_cols)} | "
        f"semantic={len(output_semantic_cols)} | final={len(output_final_cols)} | aux={len(output_aux_cols)}"
    )
    ws["A25"] = "Raw tensor schema"
    ws["B25"] = f"{output_tensor_raw_schema_name} | compared={len(output_tensor_raw_cols)}"

    ws["A27"] = "Stage tolerances"
    ws["A27"].font = SUBTITLE_FONT
    ws["A28"] = "Input tensor tolerance"
    ws["B28"] = INPUT_TENSOR_TOLERANCE
    ws["A29"] = "Raw tensor dump tolerance"
    ws["B29"] = 0.0
    ws["A30"] = "Decoded raw output tolerance"
    ws["B30"] = RAW_OUTPUT_TOLERANCE
    ws["A31"] = "Semantic output tolerance"
    ws["B31"] = SEMANTIC_OUTPUT_TOLERANCE
    ws["A32"] = "Final prediction tolerance"
    ws["B32"] = FINAL_OUTPUT_TOLERANCE
    ws["A33"] = "Semantic/debug tolerance"
    ws["B33"] = SEMANTIC_DEBUG_TOLERANCE

    ws["A35"] = "Data sources"
    ws["A35"].font = SUBTITLE_FONT

    ws["A36"] = "Selected run directory"
    ws["B36"] = str(latest_run_dir)

    ws["A37"] = "Input reference spreadsheet"
    ws["B37"] = str(input_ref_path)

    ws["A38"] = "Output reference spreadsheet"
    ws["B38"] = str(output_ref_path)

    ws["A39"] = "Output raw-tensor reference spreadsheet"
    ws["B39"] = str(output_raw_ref_path) if output_raw_ref_path is not None else "N/A"

    ws["A40"] = "Preprocess reference spreadsheet"
    ws["B40"] = str(preprocess_ref_path) if preprocess_ref_path is not None else "N/A"

    ws["A41"] = "Firmware log file"
    ws["B41"] = str(log_txt_path)

    ws["A42"] = "Workbook file"
    ws["B42"] = str(out_xlsx_path)

    ws["A43"] = "Duplicate-key alignment"
    ws["B43"] = (
        f"input_full={input_full_summary.get('duplicate_key_alignment_applied', False)} | "
        f"output_full={output_full_summary.get('duplicate_key_alignment_applied', False)} | "
        f"input_critical={input_critical_summary.get('duplicate_key_alignment_applied', False)} | "
        f"output_tensor_raw={output_tensor_raw_summary.get('duplicate_key_alignment_applied', False)} | "
        f"output_raw={output_raw_summary.get('duplicate_key_alignment_applied', False)} | "
        f"output_semantic={output_semantic_summary.get('duplicate_key_alignment_applied', False)} | "
        f"output_final={output_final_summary.get('duplicate_key_alignment_applied', False)}"
    )

    ws["A45"] = "Comparison validation policy"
    ws["A45"].font = SUBTITLE_FONT
    ws["A46"] = "Validation basis"
    ws["B46"] = "STRUCTURAL_AVAILABILITY"
    ws["C46"] = "Comparison validity depends on required blocks/fields being present in the reference workbooks and firmware log."
    ws["A47"] = "Availability sheet"
    ws["B47"] = "Data_Availability"
    ws["C47"] = "Use this sheet to verify whether any required input/output/preprocess block is absent."
    ws["A48"] = "Preprocess reference workbook"
    ws["B48"] = str(preprocess_ref_path) if preprocess_ref_path is not None else "N/A"
    ws["A49"] = "Validation note"
    ws["B49"] = "Missing required data is surfaced explicitly and should be treated as the validity gate for this comparison."

    ws["A51"] = "Data availability"
    ws["A51"].font = SUBTITLE_FONT
    ws["A52"] = "Availability verdict"
    missing_blocks_text = summarize_missing_blocks(availability_df)
    ws["B52"] = "MISSING_DATA" if "Missing blocks:" in missing_blocks_text else "AVAILABLE"
    ws["C52"] = missing_blocks_text
    apply_status_fill(ws["B52"])

    ws["A54"] = "How to read this sheet"
    ws["A54"].font = SUBTITLE_FONT

    ws["A55"] = "Critical input"
    ws["B55"] = "Decides whether the model is receiving the same tensor on both sides. If this is not MATCH, stop here and fix input packaging first."
    ws["A56"] = "Raw tensor dump"
    ws["B56"] = "Decides whether the mismatch already exists in the output tensor buffer immediately after Invoke(). This stage compares tensor metadata and dumped bytes directly."
    ws["A57"] = "Decoded raw output"
    ws["B57"] = "Shows whether the mismatch appears while interpreting the tensor buffer as immediate output floats. Review this stage only after Raw tensor dump is MATCH or N/A."
    ws["A58"] = "Semantic output"
    ws["B58"] = "Shows whether the mismatch persists after semantic mapping. Review output ordering/mapping only if decoded raw output is already MATCH; otherwise this stage is usually downstream propagation."
    ws["A59"] = "Final prediction"
    ws["B59"] = "Shows whether the mismatch persists through inverse-transform and absolute reconstruction. If tensor/raw/semantic output already fails, this stage is usually downstream propagation, not the root cause."
    ws["A60"] = "Semantic/state-only"
    ws["B60"] = "Auxiliary evidence only. Differences here do not invalidate a MATCH in the critical input tensor."

    ws["A62"] = "Firmware-only output stability"
    ws["A62"].font = SUBTITLE_FONT
    stability_summary_local = output_stability_summary or make_empty_summary(0, 0, 0.0, ["idx", "epoch", "out_idx"])
    stability_status = classify_result(stability_summary_local)
    stability_rate = (
        "N/A"
        if stability_summary_local.get("not_applicable")
        else f"{comparison_pct(stability_summary_local):.2f}%"
    )
    stability_fraction = _stage_match_fraction(stability_summary_local)
    if stability_summary_local.get("not_applicable"):
        stability_interpretation = "No DBG_MODEL_OUT_STABILITY_CSV block was found in the firmware log."
    elif stability_status == "MATCH":
        stability_interpretation = "The output tensor remained stable across two consecutive reads without a new Invoke()."
    elif stability_status == "MISMATCH":
        stability_interpretation = "The output tensor changed across two consecutive reads without a new Invoke(). Investigate buffer stability/materialization around the output tensor."
    else:
        stability_interpretation = "Unexpected firmware-only stability state."
    ws["A63"] = "Output stability status"
    ws["B63"] = stability_status
    ws["C63"] = stability_fraction
    ws["D63"] = stability_rate
    ws["E63"] = stability_interpretation
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}63"].border = BORDER_THIN
        ws[f"{col}63"].alignment = Alignment(vertical="top", wrap_text=True)
    apply_status_fill(ws["B63"])

    ws["A66"] = "Status guide"
    ws["A66"].font = SUBTITLE_FONT

    ws["A67"] = "MATCH"
    ws["B67"] = "All matched rows agree within tolerance and no unmatched rows exist."

    ws["A68"] = "MISMATCH"
    ws["B68"] = "Rows are aligned, but at least one matched row exceeds the tolerance."

    ws["A69"] = "INCOMPLETE"
    ws["B69"] = "One or more rows exist only in the Python reference or only in the firmware log."

    ws["A70"] = "N/A"
    ws["B70"] = "That auxiliary subgroup does not exist in the detected schema."

    for row in range(3, 71):
        if row == 9:
            continue
        ws[f"A{row}"].font = LABEL_FONT if row not in {8, 18, 22, 27, 35, 45, 51, 54, 62, 66} else SUBTITLE_FONT
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = BORDER_THIN
            ws[f"{col}{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    for marker in ["A5", "A18", "A22", "A27", "A35", "A45", "A51", "A54", "A62", "A66"]:
        ws[marker].fill = SECTION_FILL
    fill_row(ws, 5, 1, 2, INFO_FILL)
    fill_row(ws, 6, 1, 2, INFO_FILL)
    fill_row(ws, 16, 1, 2, INFO_FILL)
    apply_status_fill(ws["B19"])
    apply_status_fill(ws["B20"])

    ws.freeze_panes = "A8"
    autosize(ws, min_width=18, max_width=95)


def add_field_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Field_Guide")
    ws["A1"] = "Workbook Field Guide"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Section", "Sheet / Field", "Meaning", "How to interpret"),
        ("Workbook tabs", "Executive_Summary", "Decision-oriented overview of the comparison by stage.", "Read this first. It tells whether input tensor, raw tensor dump, decoded raw output, semantic output, and final prediction agree."),
        ("Workbook tabs", "Overview", "Compact block-level counts for rows, mismatches, and maximum absolute difference.", "Use it to see coverage and how many rows exceeded tolerance in each block."),
        ("Workbook tabs", "Data_Availability", "Presence/absence audit for reference workbooks and firmware-log blocks.", "Read this when any stage looks unexpectedly clean or N/A. Missing data is made explicit here instead of being treated as silent success."),
        ("Workbook tabs", "Tolerance_Protocol", "Empirical support sheet for scientific tolerance justification.", "This sheet does not change pass/fail. It reports observed max deviations and calibration eligibility so the fixed tolerances can be justified from validated MATCH runs."),
        ("Workbook tabs", "Input_Summary / Post_Summary", "Stage-aware full summaries using all compared columns in each block.", "These sheets preserve the same per-stage tolerance policy used in the detailed blocks. Postprocess_* is the standardized naming for the post-inference audit blocks."),
        ("Workbook tabs", "Input_Critical_Summary", "Summary using only the real model input tensor columns.", "This decides whether pre-inference packaging is closed."),
        ("Workbook tabs", "Preprocess_*_Summary", "Optional stage-by-stage preprocessing audit blocks.", "These tabs localize divergence before Invoke() across raw acquisition, smoothing, lags, time, physical vector, clip, scaling, and final model-input payload."),
        ("Workbook tabs", "Post_TensorRaw_Sum", "Summary using only immediate raw output tensor metadata/bytes dumped right after Invoke().", "If this fails while input is clean, the mismatch already exists in the output tensor buffer immediately after inference."),
        ("Workbook tabs", "Post_Raw_Sum", "Summary using only decoded immediate raw output values.", "If tensor dump is already MATCH but this fails, investigate how the output tensor is being decoded/copied into floats."),
        ("Workbook tabs", "Post_Semantic_Sum", "Summary after raw outputs are mapped to semantic targets.", "Use it to verify whether mismatch is only propagation or a semantic-mapping issue."),
        ("Workbook tabs", "Post_Final_Sum", "Summary after inverse-transform and absolute reconstruction.", "If raw and semantic stages are clean but this fails, investigate post-processing/reconstruction."),
        ("Workbook tabs", "Post_Stability_Sum", "Firmware-only summary of two consecutive reads of the same output tensor without a new Invoke().", "If this fails, investigate output-buffer stability/materialization around the inference result on the firmware side."),
        ("Workbook tabs", "Input_Semantic_Summary / Post_Aux_Sum", "Summaries for auxiliary state/debug fields.", "These sheets help semantic auditing, but they do not decide whether the critical model path matches."),
        ("Workbook tabs", "Preprocess_*_Differences", "Row-level mismatch details for each preprocessing checkpoint.", "Use these tabs to identify whether divergence starts in raw acquisition, smoothing, lag state, time derivation, physical features, clip, scaling, or final model-input payload."),
        ("Workbook tabs", "*_Differences sheets / Postprocess_*_Differences", "Row-level details for rows with mismatch or missing coverage.", "Start from the summary sheet, then open the corresponding Differences sheet for the failing stage. Postprocess_*_Differences is the naming used for the post-inference blocks."),
        ("Workbook tabs", "Preprocess_*_Window_Ref", "Python-only detailed 24-step preprocess traces exported from the new workbook.", "These tabs are reference/context only. They are not audited 1:1 against firmware because the firmware log does not expose equivalent window-detail blocks."),

        ("Common fields", "Status", "Block result after row coverage and tolerance checks.", "MATCH = all matched rows are within tolerance and coverage is complete. MISMATCH = same rows exist on both sides but at least one row exceeded tolerance. INCOMPLETE = row coverage is not fully 1:1. N/A = stage not applicable for detected schema."),
        ("Common fields", "Scalar mismatch rate", "Percent of scalar comparisons that exceeded the applicable tolerance.", "This is comparison-level, not row-level. Use it together with Row mismatch rate."),
        ("Common fields", "Row mismatch rate", "Percent of matched rows with at least one compared field above tolerance.", "This is row-level. It explains counts like 21/24 more directly."),
        ("Common fields", "Rows within tolerance", "Matched rows that stayed within tolerance for the stage.", "This is the clearest measure of stage agreement at row level."),
        ("Common fields", "_merge / Row_Status", "Join status between Python reference and firmware log.", "both = row exists on both sides. left_only = only Python. right_only = only firmware."),
        ("Common fields", "row_max_abs_diff / Maximum_Row_Absolute_Difference", "Largest absolute difference among compared fields in that row.", "Use it to rank the most important mismatching rows first."),
        ("Common fields", "Most_Affected_Fields", "Top fields with the largest absolute difference in the row.", "This points to the columns that explain why the row failed."),
        ("Common fields", "diff_<field>", "Python value minus firmware value for a single field.", "Check the sign and magnitude to see bias direction and severity."),
        ("Common fields", "<field>_Python / <field>_Firmware", "Side-by-side values for the same compared field.", "Use them to inspect the row visually before deciding where the mismatch starts."),

        ("Input critical", "in_x*_float / x*", "Actual model input tensor values converted to float for comparison.", "These columns decide whether the model received the same tensor on both sides."),
        ("Input semantic", "state_*", "Semantic/logical state used to assemble the input features.", "Helpful for debugging the semantic layer, but not the primary evidence for input-tensor agreement."),
        ("Input semantic", "in_f*_phys_raw", "Physical feature value before clipping.", "Use it to inspect whether raw physical values are being assembled consistently."),
        ("Input semantic", "in_f*_phys_clip", "Physical feature value after clipping.", "Use it to verify clipping limits and per-feature clipping consistency."),
        ("Input semantic", "in_f*_scaled", "Feature value after normalization.", "Useful when the semantic layer matters, but still not the final tensor evidence if in_x*_float already matches."),
        ("Legacy input", "p* / x*", "Legacy input schema columns retained for backward compatibility.", "Interpret p* and x* only in the context of the legacy schema detected by the script."),

        ("Output tensor raw", "out_idx / tensor_index / type_code / type_name / bytes_total / dims_size / dim* / b*_hex", "Immediate output tensor metadata and dumped bytes right after Invoke().", "This is the lowest-level proof. If this fails while input is MATCH, the mismatch is already present in the runtime/output buffer path."),
        ("Output raw", "out_o*_float / o*_raw / out_o*_bits_hex", "Decoded immediate raw output values after interpreting the tensor buffer, plus optional bitwise float32 hex patterns.", "If tensor raw is already MATCH but this fails, the discrepancy is in immediate decode/copy rather than in the raw tensor buffer itself. When bits_hex is present, it is compared by exact string equality, not by tolerance."),
        ("Output semantic", "y_T_scaled / y_H_scaled / y0 / y1", "Semantic target values in the normalized target domain.", "If raw output is already failing, this stage usually shows downstream propagation. Review mapping only if raw output is clean."),
        ("Output final", "d_T_pred / d_H_pred", "Physical residual prediction after inverse-transform.", "Use this stage to inspect whether inverse-transform behaves consistently."),
        ("Output final", "p_T_pred / p_H_pred", "Final absolute prediction after lag-1 reconstruction.", "This is the final prediction level. If only this stage fails, investigate reconstruction/post-processing."),
        ("Postprocess auxiliary", "p_Tprev_phys / p_Hprev_phys", "Lag-1 physical baselines used in absolute reconstruction.", "These are support fields for reconstruction analysis, not the main proof of inference mismatch. This block belongs to Postprocess_Aux."),

        ("Keys and ordering", "idx", "Logical sample index used by the debug export.", "Use it to locate the same sample across stage-specific sheets."),
        ("Keys and ordering", "epoch", "Epoch timestamp associated with the exported sample.", "Together with idx, it helps confirm row identity and temporal order."),
        ("Keys and ordering", "step", "Per-window or per-feature step identifier in step-based input exports.", "Only present in schemas that compare input row-by-row before flattening."),
    ]

    for r, row in enumerate(rows, start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)

    style_header(ws[3])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, min_width=16, max_width=80)


# ========================= Main =========================
def main() -> None:

    latest_run_dir, input_ref_path, output_ref_path, output_raw_ref_path, preprocess_ref_path, log_txt_path, out_xlsx_path = resolve_runtime_paths()
    print("[INFO] Comparing immediate model I/O against firmware log...")
    print(f"[INFO] latest_run: {latest_run_dir}")
    print(f"[INFO] input_ref : {input_ref_path}")
    print(f"[INFO] output_ref: {output_ref_path}")
    print(f"[INFO] output_raw_ref: {output_raw_ref_path}")
    print(f"[INFO] preprocess_ref: {preprocess_ref_path}")
    print(f"[INFO] log_file  : {log_txt_path}")
    print(f"[INFO] out_xlsx  : {out_xlsx_path}")
    input_ref_raw = read_first_sheet_xlsx(input_ref_path)
    output_ref_raw = read_first_sheet_xlsx(output_ref_path)
    output_tensor_raw_ref_raw = read_first_sheet_xlsx(output_raw_ref_path) if output_raw_ref_path is not None else None
    preprocess_raw_ref_raw = read_sheet_xlsx(preprocess_ref_path, "PRE_RAW_CSV") if preprocess_ref_path is not None else pd.DataFrame()
    preprocess_smooth_ref_raw = read_sheet_xlsx(preprocess_ref_path, "PRE_SMOOTH_CSV") if preprocess_ref_path is not None else pd.DataFrame()
    preprocess_raw_window_ref_raw = read_sheet_xlsx(preprocess_ref_path, "PRE_RAW_WINDOW_CSV") if preprocess_ref_path is not None else pd.DataFrame()
    preprocess_smooth_window_ref_raw = read_sheet_xlsx(preprocess_ref_path, "PRE_SMOOTH_WINDOW_CSV") if preprocess_ref_path is not None else pd.DataFrame()
    log_text = read_text_from_log_txt(log_txt_path)
    input_log_raw = parse_tagged_csv(log_text, TAG_IN, log_txt_path)
    preprocess_raw_log_raw = parse_optional_tagged_csv(log_text, TAG_PRE_RAW)
    preprocess_smooth_log_raw = parse_optional_tagged_csv(log_text, TAG_PRE_SMOOTH)
    output_log_raw = parse_tagged_csv(log_text, TAG_OUT, log_txt_path)
    output_bits_log_raw = parse_optional_tagged_csv(log_text, TAG_OUT_BITS)
    output_tensor_raw_log_raw = parse_optional_tagged_csv(log_text, TAG_OUT_RAW)
    output_stability_log_raw = parse_optional_tagged_csv(log_text, TAG_OUT_STABILITY)
    output_log_raw = merge_output_log_blocks_robust(output_log_raw, output_bits_log_raw)
    output_ref_raw, synthesized_python_output_bits = ensure_output_bit_columns(
        output_ref_raw,
        allow_synthesize_from_float=True,
    )
    output_log_raw, synthesized_firmware_output_bits = ensure_output_bit_columns(
        output_log_raw,
        allow_synthesize_from_float=False,
    )

    input_key_cols, input_compare_cols, input_schema_name = detect_input_schema(input_ref_raw)
    input_log_key_cols, input_log_compare_cols, input_log_schema_name = detect_input_schema(input_log_raw)
    if input_key_cols != input_log_key_cols or list(input_compare_cols) != list(input_log_compare_cols):
        raise KeyError(
            "Input schema mismatch between Python reference and firmware log: "
            f"python={input_schema_name} keys={input_key_cols} cols={list(input_compare_cols)[:6]}... | "
            f"firmware={input_log_schema_name} keys={input_log_key_cols} cols={list(input_log_compare_cols)[:6]}..."
        )

    output_key_cols, output_compare_cols, output_schema_name = detect_output_schema(output_ref_raw)
    output_log_key_cols, output_log_compare_cols, output_log_schema_name = detect_output_schema(output_log_raw)
    output_optional_missing_python: List[str] = []
    output_optional_missing_firmware: List[str] = []
    if output_key_cols != output_log_key_cols:
        raise KeyError(
            "Output schema mismatch between Python reference and firmware log: "
            f"python={output_schema_name} keys={output_key_cols} | "
            f"firmware={output_log_schema_name} keys={output_log_key_cols}"
        )
    promoted_shared_output_bits: List[str] = []
    forced_shared_output_bits: List[str] = []
    if output_schema_name == output_log_schema_name == "semantic_output":
        output_compare_cols, output_optional_missing_python, output_optional_missing_firmware = reconcile_semantic_output_columns(
            output_compare_cols,
            output_log_compare_cols,
        )
        output_compare_cols, promoted_shared_output_bits = promote_shared_output_bit_fields(
            output_ref_raw,
            output_log_raw,
            output_compare_cols,
        )
        output_compare_cols, forced_shared_output_bits = force_shared_output_bit_compare_columns(
            output_ref_raw,
            output_log_raw,
            output_compare_cols,
        )
    elif list(output_compare_cols) != list(output_log_compare_cols):
        raise KeyError(
            "Output schema mismatch between Python reference and firmware log: "
            f"python={output_schema_name} keys={output_key_cols} cols={list(output_compare_cols)} | "
            f"firmware={output_log_schema_name} keys={output_log_key_cols} cols={list(output_log_compare_cols)}"
        )

    print(f"[INFO] input schema : {input_schema_name} | compared columns={len(input_compare_cols)}")

    output_compare_cols, forced_shared_output_bits_after_prepare = force_shared_output_bit_compare_columns(
        output_ref_raw,
        output_log_raw,
        output_compare_cols,
    )

    input_ref = prepare_dataframe(input_ref_raw, input_key_cols, input_compare_cols)
    output_ref = prepare_dataframe(output_ref_raw, output_key_cols, output_compare_cols)
    input_log = prepare_dataframe(input_log_raw, input_key_cols, input_compare_cols)
    output_log = prepare_dataframe(output_log_raw, output_key_cols, output_compare_cols)

    output_tensor_raw_schema_name = "raw_output_tensor"
    output_tensor_raw_key_cols: List[str] = RAW_TENSOR_OUTPUT_KEY
    output_tensor_raw_cols: List[str] = []
    output_tensor_raw_optional_missing_python: List[str] = []
    output_tensor_raw_optional_missing_firmware: List[str] = []
    output_tensor_raw_ref = pd.DataFrame(columns=RAW_TENSOR_OUTPUT_KEY)
    output_tensor_raw_log = pd.DataFrame(columns=RAW_TENSOR_OUTPUT_KEY)
    output_tensor_raw_available = (
        output_tensor_raw_ref_raw is not None and
        output_tensor_raw_log_raw is not None and
        not output_tensor_raw_ref_raw.empty and
        not output_tensor_raw_log_raw.empty
    )
    if output_tensor_raw_available:
        output_tensor_raw_key_cols, output_tensor_raw_cols, output_tensor_raw_schema_name = detect_output_raw_tensor_schema(output_tensor_raw_ref_raw)
        output_tensor_raw_log_key_cols, output_tensor_raw_log_cols, output_tensor_raw_log_schema_name = detect_output_raw_tensor_schema(output_tensor_raw_log_raw)
        if output_tensor_raw_key_cols != output_tensor_raw_log_key_cols:
            raise KeyError(
                "Raw output tensor key mismatch between Python reference and firmware log: "
                f"python={output_tensor_raw_schema_name} keys={output_tensor_raw_key_cols} | "
                f"firmware={output_tensor_raw_log_schema_name} keys={output_tensor_raw_log_key_cols}"
            )
        output_tensor_raw_cols, output_tensor_raw_optional_missing_python, output_tensor_raw_optional_missing_firmware = reconcile_raw_output_tensor_columns(
            output_tensor_raw_cols,
            output_tensor_raw_log_cols,
        )
        output_tensor_raw_ref = prepare_dataframe(output_tensor_raw_ref_raw, output_tensor_raw_key_cols, output_tensor_raw_cols)
        output_tensor_raw_log = prepare_dataframe(output_tensor_raw_log_raw, output_tensor_raw_key_cols, output_tensor_raw_cols)

    output_stability_schema_name = "output_stability"
    output_stability_cols: List[str] = []
    output_stability_mismatch = pd.DataFrame()
    output_stability_summary = make_empty_summary(0, 0, 0.0, ["idx", "epoch", "out_idx"])
    output_stability_human_df = pd.DataFrame([{"Message": "No DBG_MODEL_OUT_STABILITY_CSV block found in firmware log."}])
    if output_stability_log_raw is not None and not output_stability_log_raw.empty:
        _, output_stability_cols, output_stability_schema_name = detect_output_stability_schema(output_stability_log_raw)
        output_stability_mismatch, output_stability_summary = analyze_output_stability(output_stability_log_raw)
        output_stability_human_df = humanize_output_stability(output_stability_mismatch)

    preprocess_key_cols = ["idx", "epoch"]
    preprocess_raw_cols = [c for c in preprocess_raw_ref_raw.columns if c not in preprocess_key_cols and c in preprocess_raw_log_raw.columns] if preprocess_raw_log_raw is not None else []
    preprocess_smooth_cols = [c for c in preprocess_smooth_ref_raw.columns if c not in preprocess_key_cols and c in preprocess_smooth_log_raw.columns] if preprocess_smooth_log_raw is not None else []
    preprocess_raw_ref = _prepare_optional_preprocess_df(preprocess_raw_ref_raw, preprocess_key_cols, preprocess_raw_cols)
    preprocess_raw_log = _prepare_optional_preprocess_df(preprocess_raw_log_raw, preprocess_key_cols, preprocess_raw_cols)
    preprocess_smooth_ref = _prepare_optional_preprocess_df(preprocess_smooth_ref_raw, preprocess_key_cols, preprocess_smooth_cols)
    preprocess_smooth_log = _prepare_optional_preprocess_df(preprocess_smooth_log_raw, preprocess_key_cols, preprocess_smooth_cols)

    preprocess_raw_window_ref_df = (
        preprocess_raw_window_ref_raw.copy()
        if preprocess_raw_window_ref_raw is not None and not preprocess_raw_window_ref_raw.empty
        else pd.DataFrame([
            {"Message": "PRE_RAW_WINDOW_CSV not found in preprocess workbook."}
        ])
    )
    preprocess_smooth_window_ref_df = (
        preprocess_smooth_window_ref_raw.copy()
        if preprocess_smooth_window_ref_raw is not None and not preprocess_smooth_window_ref_raw.empty
        else pd.DataFrame([
            {"Message": "PRE_SMOOTH_WINDOW_CSV not found in preprocess workbook."}
        ])
    )

    availability_df = build_data_availability_df([
        ("Input reference workbook", input_ref_raw, "required", "Primary Python input reference used for the comparison."),
        ("Output reference workbook", output_ref_raw, "required", "Primary Python output reference used for the comparison."),
        ("Output raw-tensor reference workbook", output_tensor_raw_ref_raw, "optional but recommended", "Needed to prove whether divergence starts immediately after Invoke()."),
        ("Preprocess reference workbook", preprocess_ref_path, "optional but recommended for the new workbook", "Needed to audit PRE_RAW/PRE_SMOOTH and the new preprocess workbook tabs."),
        ("PRE_RAW_CSV reference sheet", preprocess_raw_ref_raw, "required when PRE_RAW is claimed in firmware", "Python-side logical-sample raw preprocess reference."),
        ("PRE_SMOOTH_CSV reference sheet", preprocess_smooth_ref_raw, "required when PRE_SMOOTH is claimed in firmware", "Python-side logical-sample smoothed preprocess reference."),
        ("PRE_RAW_WINDOW_CSV reference sheet", preprocess_raw_window_ref_raw, "reference/context only", "Detailed Python-only window trace; not audited 1:1 against firmware."),
        ("PRE_SMOOTH_WINDOW_CSV reference sheet", preprocess_smooth_window_ref_raw, "reference/context only", "Detailed Python-only smoothed window trace; not audited 1:1 against firmware."),
        ("DBG_MODEL_IN_CSV in firmware log", input_log_raw, "required", "Critical immediate input block."),
        ("DBG_PRE_RAW_CSV in firmware log", preprocess_raw_log_raw, "required when PRE_RAW_CSV reference sheet is present", "Missing here means raw preprocess cannot be audited from the log."),
        ("DBG_PRE_SMOOTH_CSV in firmware log", preprocess_smooth_log_raw, "required when PRE_SMOOTH_CSV reference sheet is present", "Missing here means smoothed preprocess cannot be audited from the log."),
        ("DBG_MODEL_OUT_CSV in firmware log", output_log_raw, "required", "Primary immediate output block."),
        ("DBG_MODEL_OUT_BITS_CSV in firmware log", output_bits_log_raw, "optional", "Supplemental bitwise output block when emitted separately."),
        ("DBG_MODEL_OUT_RAW_CSV in firmware log", output_tensor_raw_log_raw, "optional but recommended", "Raw output tensor dump used to prove whether divergence starts immediately after Invoke()."),
        ("DBG_MODEL_OUT_STABILITY_CSV in firmware log", output_stability_log_raw, "optional", "Firmware-only output stability audit block."),
    ])

    input_critical_cols, input_semantic_cols = split_input_compare_columns(input_compare_cols)
    output_raw_cols, output_semantic_cols, output_final_cols, output_aux_cols = split_output_compare_columns(output_compare_cols)
    preprocess_lags_cols = _available_cols(input_ref, PRE_LAGS_FIELDS)
    preprocess_time_cols = _available_cols(input_ref, PRE_TIME_FIELDS)
    preprocess_phys_cols = _regex_cols(input_ref, PRE_PHYS_FIELDS_REGEX)
    preprocess_clip_cols = _regex_cols(input_ref, PRE_CLIP_FIELDS_REGEX)
    preprocess_scaled_cols = _regex_cols(input_ref, PRE_SCALED_FIELDS_REGEX)
    preprocess_model_in_cols = _regex_cols(input_ref, PRE_MODEL_IN_FIELDS_REGEX)
    print(f"[INFO] output schema: {output_schema_name} | compared columns={len(output_compare_cols)} | raw={len(output_raw_cols)} semantic={len(output_semantic_cols)} final={len(output_final_cols)} aux={len(output_aux_cols)}")
    print(f"[INFO] output bits block in log: {'present' if output_bits_log_raw is not None else 'absent'}")
    print(f"[INFO] raw tensor block in log: {'present' if output_tensor_raw_log_raw is not None else 'absent'}")
    print(f"[INFO] stability block in log: {'present' if output_stability_log_raw is not None else 'absent'}")
    print(f"[INFO] raw tensor reference: {'present' if output_raw_ref_path is not None else 'absent'}")
    promoted_all_output_bits = list(dict.fromkeys(promoted_shared_output_bits + forced_shared_output_bits + forced_shared_output_bits_after_prepare))
    if promoted_all_output_bits:
        print(f"[INFO] promoted shared output bit fields into raw output comparison: {promoted_all_output_bits}")
    if synthesized_python_output_bits:
        print(f"[INFO] synthesized output bit columns in Python reference from out_o*_float: {synthesized_python_output_bits}")
    if synthesized_firmware_output_bits:
        print(f"[INFO] synthesized output bit columns in firmware log from out_o*_float: {synthesized_firmware_output_bits}")
    if output_optional_missing_firmware:
        print(f"[INFO] output optional columns missing in firmware log (comparison skipped for these fields): {output_optional_missing_firmware}")
    if output_optional_missing_python:
        print(f"[INFO] output optional columns missing in Python reference (comparison skipped for these fields): {output_optional_missing_python}")
    if output_tensor_raw_optional_missing_firmware:
        print(f"[INFO] raw tensor optional columns missing in firmware log (comparison skipped for these fields): {output_tensor_raw_optional_missing_firmware}")
    if output_tensor_raw_optional_missing_python:
        print(f"[INFO] raw tensor optional columns missing in Python reference (comparison skipped for these fields): {output_tensor_raw_optional_missing_python}")

    input_full_tolerance_map = {
        **{col: INPUT_TENSOR_TOLERANCE for col in input_critical_cols},
        **{col: SEMANTIC_DEBUG_TOLERANCE for col in input_semantic_cols},
    }
    output_full_tolerance_map = {
        **{col: RAW_OUTPUT_TOLERANCE for col in output_raw_cols},
        **{col: SEMANTIC_OUTPUT_TOLERANCE for col in output_semantic_cols},
        **{col: FINAL_OUTPUT_TOLERANCE for col in output_final_cols},
        **{col: SEMANTIC_DEBUG_TOLERANCE for col in output_aux_cols},
    }

    input_full_mismatch, input_full_summary = merge_and_compare(
        input_ref,
        input_log,
        input_key_cols,
        input_compare_cols,
        input_full_tolerance_map,
    )
    output_full_mismatch, output_full_summary = merge_and_compare(
        output_ref,
        output_log,
        output_key_cols,
        output_compare_cols,
        output_full_tolerance_map,
    )

    input_critical_mismatch, input_critical_summary = compare_subset(
        input_ref,
        input_log,
        input_key_cols,
        input_critical_cols,
        INPUT_TENSOR_TOLERANCE,
    )
    output_tensor_raw_mismatch, output_tensor_raw_summary = compare_subset(
        output_tensor_raw_ref,
        output_tensor_raw_log,
        output_tensor_raw_key_cols,
        output_tensor_raw_cols,
        0.0,
        applicability_mode="raw_output_tensor",
    )
    output_raw_mismatch, output_raw_summary = compare_subset(
        output_ref,
        output_log,
        output_key_cols,
        output_raw_cols,
        RAW_OUTPUT_TOLERANCE,
    )
    output_semantic_mismatch, output_semantic_summary = compare_subset(
        output_ref,
        output_log,
        output_key_cols,
        output_semantic_cols,
        SEMANTIC_OUTPUT_TOLERANCE,
    )
    output_final_mismatch, output_final_summary = compare_subset(
        output_ref,
        output_log,
        output_key_cols,
        output_final_cols,
        FINAL_OUTPUT_TOLERANCE,
    )
    input_semantic_mismatch, input_semantic_summary = compare_subset(
        input_ref,
        input_log,
        input_key_cols,
        input_semantic_cols,
        SEMANTIC_DEBUG_TOLERANCE,
    )
    output_aux_mismatch, output_aux_summary = compare_subset(
        output_ref,
        output_log,
        output_key_cols,
        output_aux_cols,
        SEMANTIC_DEBUG_TOLERANCE,
    )

    preprocess_raw_summary_df, preprocess_raw_human_df, preprocess_raw_summary = _build_preprocess_stage(
        preprocess_raw_ref,
        preprocess_raw_log,
        preprocess_key_cols,
        preprocess_raw_cols,
        "Preprocess / raw acquisition fields",
    )
    preprocess_smooth_summary_df, preprocess_smooth_human_df, preprocess_smooth_summary = _build_preprocess_stage(
        preprocess_smooth_ref,
        preprocess_smooth_log,
        preprocess_key_cols,
        preprocess_smooth_cols,
        "Preprocess / smoothed-condition fields",
    )
    preprocess_lags_summary_df, preprocess_lags_human_df, preprocess_lags_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_lags_cols,
        "Preprocess / lag-state fields",
    )
    preprocess_time_summary_df, preprocess_time_human_df, preprocess_time_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_time_cols,
        "Preprocess / time-derived fields",
    )
    preprocess_phys_summary_df, preprocess_phys_human_df, preprocess_phys_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_phys_cols,
        "Preprocess / physical feature vector (pre-clip)",
    )
    preprocess_clip_summary_df, preprocess_clip_human_df, preprocess_clip_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_clip_cols,
        "Preprocess / clipped physical feature vector",
    )
    preprocess_scaled_summary_df, preprocess_scaled_human_df, preprocess_scaled_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_scaled_cols,
        "Preprocess / scaled feature vector",
    )
    preprocess_model_in_summary_df, preprocess_model_in_human_df, preprocess_model_in_summary = _build_preprocess_stage(
        input_ref,
        input_log,
        input_key_cols,
        preprocess_model_in_cols,
        "Preprocess / final model-input payload fields",
    )

    tolerance_protocol_df = build_tolerance_protocol_df([
        (
            "Immediate input tensor",
            input_critical_summary,
            INPUT_TENSOR_TOLERANCE,
            "Immediate input tensor should remain under a strict float-domain bound because it is the last pre-inference checkpoint and directly determines whether both sides invoked the same model input.",
        ),
        (
            "Immediate raw output tensor dump",
            output_tensor_raw_summary,
            0.0,
            "Raw tensor metadata and dumped bytes are structural evidence from the output buffer itself, so this stage is justified as exact equality.",
        ),
        (
            "Immediate decoded raw output",
            output_raw_summary,
            RAW_OUTPUT_TOLERANCE,
            "Decoded raw output remains close to the tensor buffer and therefore should use a tight numeric tolerance anchored to validated Python-vs-firmware float agreement.",
        ),
        (
            "Immediate semantic output",
            output_semantic_summary,
            SEMANTIC_OUTPUT_TOLERANCE,
            "Semantic output is downstream from raw decode but still pre-reconstruction, so it should retain a tight tolerance while allowing minimal float propagation.",
        ),
        (
            "Immediate final prediction",
            output_final_summary,
            FINAL_OUTPUT_TOLERANCE,
            "Final prediction includes inverse-transform and absolute reconstruction, so its tolerance may legitimately be wider than the raw/semantic stages.",
        ),
        (
            "Semantic/state debug",
            input_semantic_summary,
            SEMANTIC_DEBUG_TOLERANCE,
            "Auxiliary semantic/state fields are secondary evidence and may accumulate additional float handling, so they are tracked with the semantic/debug tolerance rather than the critical tensor bound.",
        ),
        (
            "Semantic/aux output debug",
            output_aux_summary,
            SEMANTIC_DEBUG_TOLERANCE,
            "Auxiliary postprocess fields are not the primary validity gate, but their observed deviations still support the scientific discussion of tolerance choices.",
        ),
        (
            "Preprocess raw",
            preprocess_raw_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Raw preprocess checkpoints are debug-support evidence; when aligned and MATCH, they can support empirical justification for the semantic/debug tolerance family.",
        ),
        (
            "Preprocess smooth",
            preprocess_smooth_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Smoothed preprocess checkpoints are debug-support evidence; they should not be calibrated from failing runs.",
        ),
        (
            "Preprocess lags",
            preprocess_lags_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Lag-state checkpoints support the scientific rationale for semantic/debug tolerance but should remain secondary to the critical input tensor.",
        ),
        (
            "Preprocess time",
            preprocess_time_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Time-derived checkpoints are deterministic fields whose tolerated deviation should be justified only from validated MATCH runs.",
        ),
        (
            "Preprocess physical vector",
            preprocess_phys_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Physical feature-vector checkpoints support semantic/debug calibration and should be documented separately from the critical tensor tolerance.",
        ),
        (
            "Preprocess clipped vector",
            preprocess_clip_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Clipped physical-vector checkpoints support semantic/debug calibration and should be justified from validated MATCH runs.",
        ),
        (
            "Preprocess scaled vector",
            preprocess_scaled_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Scaled feature-vector checkpoints help justify semantic/debug tolerance because they are numerically close to the model payload but still pre-inference support fields.",
        ),
        (
            "Preprocess model-input payload",
            preprocess_model_in_summary,
            PREPROCESS_STAGE_TOLERANCE,
            "Final preprocess payload fields support the bridge between semantic/debug checkpoints and the critical input tensor tolerance.",
        ),
    ])

    overview_df = build_general_overview(
        input_full_summary,
        output_full_summary,
        input_critical_summary,
        output_tensor_raw_summary,
        output_raw_summary,
        output_semantic_summary,
        output_final_summary,
        input_semantic_summary,
        output_aux_summary,
        output_stability_summary,
    )
    input_summary_df = make_summary_df("Immediate model input / full compared columns (stage-aware aggregate)", input_full_summary)
    output_summary_df = make_summary_df("Immediate postprocess / full compared columns (stage-aware aggregate)", output_full_summary)
    input_critical_summary_df = make_summary_df("Immediate model input / critical tensor columns", input_critical_summary)
    output_tensor_raw_summary_df = make_summary_df("Immediate postprocess / raw tensor dump columns", output_tensor_raw_summary)
    output_raw_summary_df = make_summary_df("Immediate postprocess / decoded raw output columns", output_raw_summary)
    output_semantic_summary_df = make_summary_df("Immediate postprocess / semantic output columns", output_semantic_summary)
    output_final_summary_df = make_summary_df("Immediate postprocess / final prediction columns", output_final_summary)
    input_semantic_summary_df = make_summary_df("Immediate model input / semantic-state columns", input_semantic_summary)
    output_aux_summary_df = make_summary_df("Immediate postprocess / semantic-aux columns", output_aux_summary)
    output_stability_summary_df = make_summary_df("Immediate postprocess / stability columns (firmware-only)", output_stability_summary)
    input_human_df = humanize_input_mismatch(input_critical_mismatch, input_critical_cols)
    output_tensor_raw_human_df = humanize_output_mismatch(output_tensor_raw_mismatch, output_tensor_raw_cols)
    output_raw_human_df = humanize_output_mismatch(output_raw_mismatch, output_raw_cols)
    output_semantic_human_df = humanize_output_mismatch(output_semantic_mismatch, output_semantic_cols)
    output_final_human_df = humanize_output_mismatch(output_final_mismatch, output_final_cols)
    input_semantic_human_df = humanize_input_mismatch(input_semantic_mismatch, input_semantic_cols)
    output_aux_human_df = humanize_output_mismatch(output_aux_mismatch, output_aux_cols)

    wb = Workbook()
    add_executive_summary_sheet(
        wb,
        input_full_summary,
        output_full_summary,
        input_critical_summary,
        output_tensor_raw_summary,
        output_raw_summary,
        output_semantic_summary,
        output_final_summary,
        input_semantic_summary,
        output_aux_summary,
        output_stability_summary,
        latest_run_dir,
        input_ref_path,
        output_ref_path,
        output_raw_ref_path,
        preprocess_ref_path,
        log_txt_path,
        out_xlsx_path,
        input_schema_name,
        output_schema_name,
        output_tensor_raw_schema_name,
        input_compare_cols,
        output_compare_cols,
        input_critical_cols,
        output_tensor_raw_cols,
        output_raw_cols,
        output_semantic_cols,
        output_final_cols,
        input_semantic_cols,
        output_aux_cols,
        availability_df,
    )

    ws = wb.create_sheet("Overview")
    write_dataframe(ws, overview_df)

    ws = wb.create_sheet("Data_Availability")
    write_dataframe(ws, availability_df)

    ws = wb.create_sheet("Tolerance_Protocol")
    write_dataframe(ws, tolerance_protocol_df)

    ws = wb.create_sheet("Preprocess_Raw_Summary")
    write_dataframe(ws, preprocess_raw_summary_df)

    ws = wb.create_sheet("Preprocess_Raw_Window_Ref")
    write_dataframe(ws, preprocess_raw_window_ref_df)

    ws = wb.create_sheet("Preprocess_Smooth_Summary")
    write_dataframe(ws, preprocess_smooth_summary_df)

    ws = wb.create_sheet("Preprocess_Smooth_Window_Ref")
    write_dataframe(ws, preprocess_smooth_window_ref_df)

    ws = wb.create_sheet("Preprocess_Lags_Summary")
    write_dataframe(ws, preprocess_lags_summary_df)

    ws = wb.create_sheet("Preprocess_Time_Summary")
    write_dataframe(ws, preprocess_time_summary_df)

    ws = wb.create_sheet("Preprocess_Phys_Summary")
    write_dataframe(ws, preprocess_phys_summary_df)

    ws = wb.create_sheet("Preprocess_Clip_Summary")
    write_dataframe(ws, preprocess_clip_summary_df)

    ws = wb.create_sheet("Preprocess_Scaled_Summary")
    write_dataframe(ws, preprocess_scaled_summary_df)

    ws = wb.create_sheet("Preprocess_ModelIn_Summary")
    write_dataframe(ws, preprocess_model_in_summary_df)

    ws = wb.create_sheet("Input_Summary")
    write_dataframe(ws, input_summary_df)

    ws = wb.create_sheet("Input_Critical_Summary")
    write_dataframe(ws, input_critical_summary_df)

    ws = wb.create_sheet("Input_Semantic_Summary")
    write_dataframe(ws, input_semantic_summary_df)

    ws = wb.create_sheet("Post_Summary")
    write_dataframe(ws, output_summary_df)

    ws = wb.create_sheet("Post_TensorRaw_Sum")
    write_dataframe(ws, output_tensor_raw_summary_df)

    ws = wb.create_sheet("Post_Raw_Sum")
    write_dataframe(ws, output_raw_summary_df)

    ws = wb.create_sheet("Post_Semantic_Sum")
    write_dataframe(ws, output_semantic_summary_df)

    ws = wb.create_sheet("Post_Final_Sum")
    write_dataframe(ws, output_final_summary_df)

    ws = wb.create_sheet("Post_Aux_Sum")
    write_dataframe(ws, output_aux_summary_df)

    ws = wb.create_sheet("Post_Stability_Sum")
    write_dataframe(ws, output_stability_summary_df)

    ws = wb.create_sheet("Preprocess_Raw_Differences")
    write_dataframe(ws, preprocess_raw_human_df)

    ws = wb.create_sheet("Preprocess_Smooth_Differences")
    write_dataframe(ws, preprocess_smooth_human_df)

    ws = wb.create_sheet("Preprocess_Lags_Differences")
    write_dataframe(ws, preprocess_lags_human_df)

    ws = wb.create_sheet("Preprocess_Time_Differences")
    write_dataframe(ws, preprocess_time_human_df)

    ws = wb.create_sheet("Preprocess_Phys_Differences")
    write_dataframe(ws, preprocess_phys_human_df)

    ws = wb.create_sheet("Preprocess_Clip_Differences")
    write_dataframe(ws, preprocess_clip_human_df)

    ws = wb.create_sheet("Preprocess_Scaled_Differences")
    write_dataframe(ws, preprocess_scaled_human_df)

    ws = wb.create_sheet("Preprocess_ModelIn_Differences")
    write_dataframe(ws, preprocess_model_in_human_df)

    ws = wb.create_sheet("Input_Critical_Differences")
    write_dataframe(ws, input_human_df)

    ws = wb.create_sheet("Input_Semantic_Differences")
    write_dataframe(ws, input_semantic_human_df)

    ws = wb.create_sheet("Post_TensorRaw_Diff")
    write_dataframe(ws, output_tensor_raw_human_df)

    ws = wb.create_sheet("Post_Raw_Diff")
    write_dataframe(ws, output_raw_human_df)

    ws = wb.create_sheet("Post_Semantic_Diff")
    write_dataframe(ws, output_semantic_human_df)

    ws = wb.create_sheet("Post_Final_Diff")
    write_dataframe(ws, output_final_human_df)

    ws = wb.create_sheet("Post_Aux_Diff")
    write_dataframe(ws, output_aux_human_df)

    ws = wb.create_sheet("Post_Stability_Diff")
    write_dataframe(ws, output_stability_human_df)

    add_field_guide_sheet(wb)

    desired_order = [
        "Executive_Summary",
        "Overview",
        "Data_Availability",
        "Tolerance_Protocol",
        "Preprocess_Raw_Summary",
        "Preprocess_Raw_Window_Ref",
        "Preprocess_Smooth_Summary",
        "Preprocess_Smooth_Window_Ref",
        "Preprocess_Lags_Summary",
        "Preprocess_Time_Summary",
        "Preprocess_Phys_Summary",
        "Preprocess_Clip_Summary",
        "Preprocess_Scaled_Summary",
        "Preprocess_ModelIn_Summary",
        "Input_Summary",
        "Input_Critical_Summary",
        "Input_Semantic_Summary",
        "Post_Summary",
        "Post_TensorRaw_Sum",
        "Post_Raw_Sum",
        "Post_Semantic_Sum",
        "Post_Final_Sum",
        "Post_Aux_Sum",
        "Post_Stability_Sum",
        "Preprocess_Raw_Differences",
        "Preprocess_Smooth_Differences",
        "Preprocess_Lags_Differences",
        "Preprocess_Time_Differences",
        "Preprocess_Phys_Differences",
        "Preprocess_Clip_Differences",
        "Preprocess_Scaled_Differences",
        "Preprocess_ModelIn_Differences",
        "Input_Critical_Differences",
        "Input_Semantic_Differences",
        "Post_TensorRaw_Diff",
        "Post_Raw_Diff",
        "Post_Semantic_Diff",
        "Post_Final_Diff",
        "Post_Aux_Diff",
        "Post_Stability_Diff",
        "Field_Guide",
    ]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx_path)
    print(f"[OK] Saved human-readable comparison workbook: {out_xlsx_path}")


if __name__ == "__main__":
    main()

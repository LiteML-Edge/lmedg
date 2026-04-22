"""
Script: compare_predictions_metrics_to_log_Conv1D_Tiny.py
Module role:
    Compare Python rolling24 prediction and metric references against firmware
    replay logs and generate an audit workbook.

Technical summary:
    This script resolves the latest quantized-model metrics directory, locates
    the reference spreadsheets, parses replay and metric blocks from the
    firmware log, aligns Python and firmware records, evaluates agreement for
    predictions and aggregated metrics, and exports an Excel workbook for the
    final validation stage.

Inputs:
    - Rolling24 prediction and metric spreadsheets exported by the Python
      pipeline
    - Firmware TXT, LOG, or ZIP log containing replay and metric blocks

Outputs:
    - Excel comparison workbook for final predictions and rolling24 metrics

Notes:
    This script assumes the repository project structure and the referenced
    input artifacts and firmware logs. The computational logic, comparison
    workflow, and workbook generation procedure are preserved.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ========================= Bootstrap local utils =========================
ROOT = os.environ.get("RUNNER_PROJECT_ROOT")
if not ROOT:
    HERE = Path(__file__).resolve()
    for base in [HERE, *HERE.parents, Path.cwd(), *Path.cwd().parents]:
        if (base / "utils").exists():
            ROOT = str(base)
            break

if ROOT and ROOT not in sys.path:
    sys.path.insert(0, ROOT)

PROJECT_ROOT = None
QUANTIZED_MODEL_METRICS = None
try:
    from utils.global_utils.paths_Conv1D_Tiny import PROJECT_ROOT, QUANTIZED_MODEL_METRICS  # type: ignore
except Exception:
    pass

# ========================= Dynamic path resolution =========================
RUN_DIR_REGEX = re.compile(r"^run\.v(\d+)$", re.IGNORECASE)
TAG_REPLAY = "DBG_REPLAY_CSV"
TAG_HOUR_METRICS = "HOUR_METRICS"
DEFAULT_EXCEL_ROW_START = 25
DEFAULT_EXCEL_ROW_END = 48
DEFAULT_ROUND_DECIMALS = 2
DEFAULT_METRICS_ROUND_DECIMALS = 4

LOG_NAME_REGEX = re.compile(r"^device-monitor-(\d{6})-(\d{6})\.(log|txt|zip)$", re.IGNORECASE)


def extract_run_version(path: Path) -> int:
    match = RUN_DIR_REGEX.match(path.name)
    return int(match.group(1)) if match else -1


def is_plain_run_dir(path: Path) -> bool:
    return path.name.lower() == "run"


def candidate_has_results_dir(run_dir: Path) -> bool:
    return (run_dir / "quantization_metrics_results").is_dir()


def normalize_metrics_root(metrics_path: Path) -> Path:
    """Return the quantized_model root even if the configured path points to run/, run.vXXX/ or quantization_metrics_results/."""
    p = metrics_path
    if p.name.lower() == "quantization_metrics_results":
        return p.parent.parent if p.parent.name.lower().startswith("run") else p.parent
    if is_plain_run_dir(p) or RUN_DIR_REGEX.match(p.name):
        return p.parent
    if candidate_has_results_dir(p):
        # already a run dir-like path containing quantization_metrics_results
        return p.parent if (is_plain_run_dir(p) or RUN_DIR_REGEX.match(p.name)) else p
    return p


def find_latest_run_dir(metrics_root: Path) -> Path:
    if not metrics_root.exists():
        raise FileNotFoundError(f"Metrics root directory not found: {metrics_root}")

    candidates: List[Path] = []
    for p in metrics_root.iterdir():
        if not p.is_dir():
            continue
        if is_plain_run_dir(p) and candidate_has_results_dir(p):
            candidates.append(p)
            continue
        if RUN_DIR_REGEX.match(p.name) and candidate_has_results_dir(p):
            candidates.append(p)

    if not candidates:
        raise FileNotFoundError(f"No valid run directory found under metrics root: {metrics_root}")

    versioned = [p for p in candidates if RUN_DIR_REGEX.match(p.name)]
    if versioned:
        return max(versioned, key=lambda p: (extract_run_version(p), p.stat().st_mtime))
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_xlsx(results_dir: Path, exact_name: str, fallback_patterns: Iterable[str]) -> Path:
    exact = results_dir / exact_name
    if exact.exists():
        return exact

    candidates: List[Path] = []
    seen: set[str] = set()
    for pattern in fallback_patterns:
        for p in results_dir.glob(pattern):
            if p.is_file() and p.suffix.lower() == ".xlsx":
                key = str(p.resolve())
                if key not in seen:
                    seen.add(key)
                    candidates.append(p)

    if not candidates:
        raise FileNotFoundError(
            f"Could not find spreadsheet in {results_dir}. exact_name={exact_name} patterns={list(fallback_patterns)}"
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def extract_log_datetime_from_name(path: Path) -> datetime | None:
    match = LOG_NAME_REGEX.match(path.name)
    if not match:
        return None
    date_part, time_part, _suffix = match.groups()
    try:
        return datetime.strptime(date_part + time_part, "%y%m%d%H%M%S")
    except ValueError:
        return None


def _log_sort_key(path: Path) -> tuple[datetime, float]:
    dt = extract_log_datetime_from_name(path)
    if dt is None:
        dt = datetime.min
    return dt, path.stat().st_mtime


def _log_or_zip_contains_all_tokens(path: Path, tokens: List[str]) -> bool:
    try:
        text = read_text_from_log_or_zip(path)
    except Exception:
        return False
    return all(token in text for token in tokens)


def find_latest_log_or_zip(log_dir: Path) -> Path:
    if not log_dir.exists():
        raise FileNotFoundError(f"Log directory not found: {log_dir}")

    candidates = [
        p for p in log_dir.glob("device-monitor-*.*")
        if p.is_file() and p.suffix.lower() in {".log", ".txt", ".zip"}
    ]
    if not candidates:
        raise FileNotFoundError(f"No device-monitor-*.log/.txt/.zip found in: {log_dir}")

    required_tokens = [f"[{TAG_REPLAY}]", "[HOUR] METRICS"]
    ordered_candidates = sorted(candidates, key=_log_sort_key, reverse=True)

    for candidate in ordered_candidates:
        if _log_or_zip_contains_all_tokens(candidate, required_tokens):
            print(f"[INFO] Selected latest Replay log/zip: {candidate}")
            return candidate

    raise FileNotFoundError(
        f"No Replay log/zip with required tokens found in: {log_dir}. "
        f"Required tokens: {required_tokens}"
    )


def choose_writable_output_path(preferred_path: Path) -> Path:
    preferred_path.parent.mkdir(parents=True, exist_ok=True)
    if not preferred_path.exists():
        return preferred_path
    try:
        with open(preferred_path, "a+b"):
            pass
        return preferred_path
    except PermissionError:
        stamp = datetime.now().strftime("%y%m%d-%H%M%S")
        return preferred_path.with_name(f"{preferred_path.stem}_{stamp}{preferred_path.suffix}")


def resolve_runtime_paths(
    cli_pred: str | None,
    cli_log: str | None,
    cli_out: str | None,
    cli_metrics: str | None,
) -> Tuple[Path, Path, Path, Path | None, Path]:
    latest_run_dir: Path | None = None
    results_dir: Path | None = None

    if cli_pred:
        pred_path = Path(cli_pred)
    else:
        if not QUANTIZED_MODEL_METRICS:
            raise FileNotFoundError("Prediction spreadsheet path not provided and utils.paths_Conv1D_Tiny is unavailable.")
        metrics_root = normalize_metrics_root(Path(QUANTIZED_MODEL_METRICS))
        latest_run_dir = find_latest_run_dir(metrics_root)
        results_dir = latest_run_dir / "quantization_metrics_results"
        pred_path = find_xlsx(
            results_dir,
            "environment_quantized_predictions_rolling24_Conv1D_Tiny.xlsx",
            [
                "*predictions_rolling24*Conv1D_Tiny*.xlsx",
                "*predictions*Conv1D_Tiny*.xlsx",
                "*rolling24*Conv1D_Tiny*.xlsx",
            ],
        )

    if cli_metrics:
        metrics_xlsx_path = Path(cli_metrics)
    else:
        if results_dir is None:
            if not QUANTIZED_MODEL_METRICS:
                raise FileNotFoundError("Metrics spreadsheet path not provided and utils.paths_Conv1D_Tiny is unavailable.")
            metrics_root = normalize_metrics_root(Path(QUANTIZED_MODEL_METRICS))
            latest_run_dir = find_latest_run_dir(metrics_root)
            results_dir = latest_run_dir / "quantization_metrics_results"
        metrics_xlsx_path = find_xlsx(
            results_dir,
            "environment_quantized_metrics_rolling24_Conv1D_Tiny.xlsx",
            [
                "*metrics_rolling24*Conv1D_Tiny*.xlsx",
                "*metrics*Conv1D_Tiny*.xlsx",
                "*rolling24*Conv1D_Tiny*.xlsx",
            ],
        )

    if cli_log:
        requested_log = Path(cli_log)
        required_tokens = [f"[{TAG_REPLAY}]", "[HOUR] METRICS"]
        if _log_or_zip_contains_all_tokens(requested_log, required_tokens):
            log_path = requested_log
        else:
            print(
                f"[WARN] Provided log is not Replay-compatible: {requested_log}\n"
                f"[WARN] Falling back to the latest Replay log/zip in the same folder."
            )
            log_path = find_latest_log_or_zip(requested_log.parent)
    else:
        if not PROJECT_ROOT:
            raise FileNotFoundError("Log path not provided and utils.paths_Conv1D_Tiny is unavailable.")
        
    workbook_dir = (Path(PROJECT_ROOT) / "utils" / "workbook_Conv1D_Tiny")

    log_dir = Path(PROJECT_ROOT) / "firmwares" / "environment_Conv1D_Tiny" / "PlatfIO_ESP32_Wemos_Conv1D_Tiny" / "logs"
    log_path = find_latest_log_or_zip(log_dir)

    if cli_out:
        out_path = choose_writable_output_path(Path(cli_out))
    else:
        preferred = workbook_dir / "predictions_metrics_vs_log_comparison_Conv1D_Tiny.xlsx"
    out_path = choose_writable_output_path(preferred)

    return pred_path, log_path, out_path, latest_run_dir, metrics_xlsx_path


# ========================= Column helpers =========================
def normalize_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).lower())


def resolve_columns(df: pd.DataFrame, alias_map: Dict[str, List[str]]) -> Dict[str, str]:
    normalized = {normalize_name(c): c for c in df.columns}
    resolved: Dict[str, str] = {}

    for canonical, aliases in alias_map.items():
        for alias in aliases:
            hit = normalized.get(normalize_name(alias))
            if hit is not None:
                resolved[canonical] = hit
                break
        else:
            raise KeyError(
                f"Could not resolve required column '{canonical}'. Available columns: {list(df.columns)}"
            )
    return resolved


# ========================= Input readers =========================
def read_prediction_block(path: Path, excel_row_start: int, excel_row_end: int) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Prediction spreadsheet not found: {path}")

    full_df = pd.read_excel(path, sheet_name=0)
    aliases = {
        "datetime_end": ["datetime_end", "datetime end", "datetime"],
        "Tin_true": ["T_in_ground truth", "T_in_ground_truth", "Tin_ground_truth", "Tin_true", "gt_Tin_true"],
        "Tin_pred": ["T_in_pred", "Tin_pred", "p_T_pred"],
        "Hin_true": ["H_in_ground truth", "H_in_ground_truth", "Hin_ground_truth", "Hin_true", "gt_Hin_true"],
        "Hin_pred": ["H_in_pred", "Hin_pred", "p_H_pred"],
    }
    cols = resolve_columns(full_df, aliases)

    start_idx = excel_row_start - 2
    end_idx = excel_row_end - 1
    if start_idx < 0 or end_idx > len(full_df):
        raise IndexError(
            f"Requested Excel rows {excel_row_start}:{excel_row_end}, but sheet has {len(full_df)} data rows."
        )

    block = full_df.iloc[start_idx:end_idx].copy().reset_index(drop=True)
    out = pd.DataFrame(
        {
            "sheet_order": range(len(block)),
            "datetime_end": pd.to_datetime(block[cols["datetime_end"]]),
            "Tin_true_py": pd.to_numeric(block[cols["Tin_true"]], errors="coerce"),
            "Tin_pred_py": pd.to_numeric(block[cols["Tin_pred"]], errors="coerce"),
            "Hin_true_py": pd.to_numeric(block[cols["Hin_true"]], errors="coerce"),
            "Hin_pred_py": pd.to_numeric(block[cols["Hin_pred"]], errors="coerce"),
        }
    )
    return out


def read_text_from_log_or_zip(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".log", ".txt"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".zip":
        with zipfile.ZipFile(path) as zf:
            names = [n for n in zf.namelist() if n.lower().endswith((".log", ".txt"))]
            if not names:
                raise FileNotFoundError(f"No .log/.txt file found inside zip: {path}")
            target = max(names)
            return zf.read(target).decode("utf-8", errors="replace")
    raise ValueError(f"Unsupported log file type: {path}")


def parse_tagged_csv(log_text: str, tag: str, source_path: Path) -> pd.DataFrame:
    rows: List[List[str]] = []
    header: List[str] | None = None
    pattern = re.compile(rf".*?\[{re.escape(tag)}\]\s*(.*)$")

    for line in log_text.splitlines():
        match = pattern.match(line)
        if not match:
            continue
        payload = match.group(1).strip()
        if not payload:
            continue
        parsed = next(csv.reader([payload]))
        parsed = [x.strip() for x in parsed]
        if header is None:
            header = parsed
            continue
        if parsed == header or len(parsed) != len(header):
            continue
        rows.append(parsed)

    if header is None:
        raise ValueError(f"Tag [{tag}] not found in log: {source_path}")

    df = pd.DataFrame(rows, columns=header)
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().all():
            df[col] = converted
    return df


def prepare_replay_log(log_text: str, source_path: Path) -> pd.DataFrame:
    replay = parse_tagged_csv(log_text, TAG_REPLAY, source_path)

    aliases = {
        "idx": ["idx"],
        "epoch": ["epoch"],
        "Tin_true": ["Tin_true", "gt_Tin_true"],
        "Hin_true": ["Hin_true", "gt_Hin_true"],
        "Tp": ["Tp", "p_T_pred"],
        "Hp": ["Hp", "p_H_pred"],
    }
    cols = resolve_columns(replay, aliases)

    out = pd.DataFrame(
        {
            "idx_log": pd.to_numeric(replay[cols["idx"]], errors="coerce").astype("Int64"),
            "epoch": pd.to_numeric(replay[cols["epoch"]], errors="coerce").astype("Int64"),
            "datetime_end": pd.to_datetime(pd.to_numeric(replay[cols["epoch"]], errors="coerce"), unit="s"),
            "Tin_true_log": pd.to_numeric(replay[cols["Tin_true"]], errors="coerce"),
            "Hin_true_log": pd.to_numeric(replay[cols["Hin_true"]], errors="coerce"),
            "Tp_log": pd.to_numeric(replay[cols["Tp"]], errors="coerce"),
            "Hp_log": pd.to_numeric(replay[cols["Hp"]], errors="coerce"),
        }
    )
    return out


def read_metrics_latest(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Metrics spreadsheet not found: {path}")

    full_df = pd.read_excel(path, sheet_name=0)
    aliases = {
        "datetime_end": ["datetime_end", "datetime end", "datetime"],
        "N": ["N"],
        "MAE": ["MAE"],
        "RMSE": ["RMSE"],
        "R2": ["R2"],
        "MAE_T": ["MAE_T", "MAE Tin", "MAE_Tin"],
        "RMSE_T": ["RMSE_T", "RMSE Tin", "RMSE_Tin"],
        "R2_T": ["R2_T", "R2 Tin", "R2_Tin"],
        "MAE_H": ["MAE_H", "MAE Hin", "MAE_Hin"],
        "RMSE_H": ["RMSE_H", "RMSE Hin", "RMSE_Hin"],
        "R2_H": ["R2_H", "R2 Hin", "R2_Hin"],
    }
    cols = resolve_columns(full_df, aliases)
    latest = full_df.iloc[-1:].copy().reset_index(drop=True)

    out = pd.DataFrame(
        {
            "datetime_end": pd.to_datetime(latest[cols["datetime_end"]]),
            "N_sheet": pd.to_numeric(latest[cols["N"]], errors="coerce"),
            "MAE_sheet": pd.to_numeric(latest[cols["MAE"]], errors="coerce"),
            "RMSE_sheet": pd.to_numeric(latest[cols["RMSE"]], errors="coerce"),
            "R2_sheet": pd.to_numeric(latest[cols["R2"]], errors="coerce"),
            "MAE_T_sheet": pd.to_numeric(latest[cols["MAE_T"]], errors="coerce"),
            "RMSE_T_sheet": pd.to_numeric(latest[cols["RMSE_T"]], errors="coerce"),
            "R2_T_sheet": pd.to_numeric(latest[cols["R2_T"]], errors="coerce"),
            "MAE_H_sheet": pd.to_numeric(latest[cols["MAE_H"]], errors="coerce"),
            "RMSE_H_sheet": pd.to_numeric(latest[cols["RMSE_H"]], errors="coerce"),
            "R2_H_sheet": pd.to_numeric(latest[cols["R2_H"]], errors="coerce"),
        }
    )
    return out


def parse_float_or_nan(text_value: str) -> float:
    s = str(text_value).strip()
    if s.lower() in {"na", "nan", ""}:
        return float("nan")
    return float(s)


def prepare_hour_metrics_latest(log_text: str, source_path: Path) -> pd.DataFrame:
    pattern = re.compile(
        r".*?\[HOUR\] METRICS \| "
        r"MAE=(\S+) RMSE=(\S+) R2=(\S+) \| "
        r"MAE_T=(\S+) RMSE_T=(\S+) R2_T=(\S+) \| "
        r"MAE_H=(\S+) RMSE_H=(\S+) R2_H=(\S+)"
    )

    parsed_rows: List[Dict[str, object]] = []
    for line_no, line in enumerate(log_text.splitlines(), start=1):
        m = pattern.match(line)
        if not m:
            continue
        vals = list(m.groups())
        parsed_rows.append(
            {
                "line_no": line_no,
                "MAE_log": parse_float_or_nan(vals[0]),
                "RMSE_log": parse_float_or_nan(vals[1]),
                "R2_log": parse_float_or_nan(vals[2]),
                "MAE_T_log": parse_float_or_nan(vals[3]),
                "RMSE_T_log": parse_float_or_nan(vals[4]),
                "R2_T_log": parse_float_or_nan(vals[5]),
                "MAE_H_log": parse_float_or_nan(vals[6]),
                "RMSE_H_log": parse_float_or_nan(vals[7]),
                "R2_H_log": parse_float_or_nan(vals[8]),
            }
        )

    if not parsed_rows:
        raise ValueError(f"No [HOUR] METRICS lines found in log: {source_path}")

    metrics_df = pd.DataFrame(parsed_rows)
    metric_cols = [c for c in metrics_df.columns if c.endswith("_log")]
    valid_mask = metrics_df[metric_cols].notna().any(axis=1)
    valid_rows = metrics_df[valid_mask].copy()
    if valid_rows.empty:
        raise ValueError(f"No valid numeric [HOUR] METRICS line found in log: {source_path}")

    latest = valid_rows.iloc[-1:].copy().reset_index(drop=True)
    latest["metrics_line_no"] = latest["line_no"]
    return latest.drop(columns=["line_no"])


# ========================= Compare helpers =========================
def classify_rows(merged: pd.DataFrame) -> str:
    if (merged["_merge"] == "both").all():
        return "MATCH"
    return "INCOMPLETE"


def _align_duplicate_datetimes(
    py_df: pd.DataFrame,
    log_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], bool]:
    merge_keys = ["datetime_end"]
    py_dup = py_df.duplicated(subset=merge_keys, keep=False).any()
    log_dup = log_df.duplicated(subset=merge_keys, keep=False).any()
    if not (py_dup or log_dup):
        return py_df.copy(), log_df.copy(), merge_keys, False

    py_aligned = py_df.copy()
    log_aligned = log_df.copy()
    py_aligned["_dup_ordinal"] = py_aligned.groupby(merge_keys, dropna=False).cumcount()
    log_aligned["_dup_ordinal"] = log_aligned.groupby(merge_keys, dropna=False).cumcount()
    return py_aligned, log_aligned, ["datetime_end", "_dup_ordinal"], True


def compare_predictions(py_df: pd.DataFrame, log_df: pd.DataFrame, round_decimals: int) -> Tuple[pd.DataFrame, Dict[str, object]]:
    py_aligned, log_aligned, merge_keys, used_duplicate_alignment = _align_duplicate_datetimes(py_df, log_df)
    merged = py_aligned.merge(log_aligned, on=merge_keys, how="outer", indicator=True, sort=True)

    merged["Tin_true_diff"] = merged["Tin_true_py"] - merged["Tin_true_log"]
    merged["Hin_true_diff"] = merged["Hin_true_py"] - merged["Hin_true_log"]
    merged["Tin_pred_diff"] = merged["Tin_pred_py"] - merged["Tp_log"]
    merged["Hin_pred_diff"] = merged["Hin_pred_py"] - merged["Hp_log"]

    merged["Tin_true_log_round"] = merged["Tin_true_log"].round(round_decimals)
    merged["Hin_true_log_round"] = merged["Hin_true_log"].round(round_decimals)
    merged["Tp_log_round"] = merged["Tp_log"].round(round_decimals)
    merged["Hp_log_round"] = merged["Hp_log"].round(round_decimals)

    merged["match_Tin_true_rounded"] = merged["Tin_true_py"].round(round_decimals) == merged["Tin_true_log_round"]
    merged["match_Hin_true_rounded"] = merged["Hin_true_py"].round(round_decimals) == merged["Hin_true_log_round"]
    merged["match_Tin_pred_rounded"] = merged["Tin_pred_py"].round(round_decimals) == merged["Tp_log_round"]
    merged["match_Hin_pred_rounded"] = merged["Hin_pred_py"].round(round_decimals) == merged["Hp_log_round"]

    both = merged[merged["_merge"] == "both"].copy()

    def count_matches(col: str) -> int:
        return int(both[col].fillna(False).sum())

    summary: Dict[str, object] = {
        "spreadsheet_rows": int(len(py_df)),
        "log_rows": int(len(log_df)),
        "matched_rows": int(len(both)),
        "spreadsheet_only_rows": int((merged["_merge"] == "left_only").sum()),
        "log_only_rows": int((merged["_merge"] == "right_only").sum()),
        "row_alignment_status": classify_rows(merged),
        "round_decimals": int(round_decimals),
        "gt_T_matches": count_matches("match_Tin_true_rounded"),
        "gt_H_matches": count_matches("match_Hin_true_rounded"),
        "pred_T_matches": count_matches("match_Tin_pred_rounded"),
        "pred_H_matches": count_matches("match_Hin_pred_rounded"),
        "gt_T_max_abs_diff": float(both["Tin_true_diff"].abs().max()) if len(both) else 0.0,
        "gt_H_max_abs_diff": float(both["Hin_true_diff"].abs().max()) if len(both) else 0.0,
        "pred_T_max_abs_diff": float(both["Tin_pred_diff"].abs().max()) if len(both) else 0.0,
        "pred_H_max_abs_diff": float(both["Hin_pred_diff"].abs().max()) if len(both) else 0.0,
        "merge_keys": merge_keys,
        "duplicate_key_alignment_applied": bool(used_duplicate_alignment),
    }

    summary["ground_truth_status"] = (
        "MATCH"
        if summary["gt_T_matches"] == summary["matched_rows"] and summary["gt_H_matches"] == summary["matched_rows"] and summary["row_alignment_status"] == "MATCH"
        else ("INCOMPLETE" if summary["row_alignment_status"] != "MATCH" else "MISMATCH")
    )
    summary["prediction_status"] = (
        "MATCH"
        if summary["pred_T_matches"] == summary["matched_rows"] and summary["pred_H_matches"] == summary["matched_rows"] and summary["row_alignment_status"] == "MATCH"
        else ("INCOMPLETE" if summary["row_alignment_status"] != "MATCH" else "MISMATCH")
    )

    merged["row_has_any_difference_rounded"] = ~(
        merged["match_Tin_true_rounded"].fillna(False)
        & merged["match_Hin_true_rounded"].fillna(False)
        & merged["match_Tin_pred_rounded"].fillna(False)
        & merged["match_Hin_pred_rounded"].fillna(False)
    )
    return merged, summary


def pct(num: int, den: int) -> float:
    return 0.0 if den <= 0 else 100.0 * float(num) / float(den)


def build_overview_df(summary: Dict[str, object]) -> pd.DataFrame:
    matched = int(summary["matched_rows"])
    return pd.DataFrame(
        [
            {
                "Block": "Row alignment by datetime_end",
                "Spreadsheet_Rows": summary["spreadsheet_rows"],
                "Log_Rows": summary["log_rows"],
                "Matched_Rows": summary["matched_rows"],
                "Spreadsheet_Only_Rows": summary["spreadsheet_only_rows"],
                "Log_Only_Rows": summary["log_only_rows"],
                "Status": summary["row_alignment_status"],
            },
            {
                "Block": f"Ground truth agreement at {summary['round_decimals']} decimals",
                "Spreadsheet_Rows": summary["spreadsheet_rows"],
                "Log_Rows": summary["log_rows"],
                "Matched_Rows": f"T={summary['gt_T_matches']}/{matched} | H={summary['gt_H_matches']}/{matched}",
                "Spreadsheet_Only_Rows": "-",
                "Log_Only_Rows": "-",
                "Status": summary["ground_truth_status"],
            },
            {
                "Block": f"Prediction agreement at {summary['round_decimals']} decimals",
                "Spreadsheet_Rows": summary["spreadsheet_rows"],
                "Log_Rows": summary["log_rows"],
                "Matched_Rows": f"T={summary['pred_T_matches']}/{matched} | H={summary['pred_H_matches']}/{matched}",
                "Spreadsheet_Only_Rows": "-",
                "Log_Only_Rows": "-",
                "Status": summary["prediction_status"],
            },
        ]
    )


def build_summary_df(
    summary: Dict[str, object],
    metrics_summary: Dict[str, object] | None = None,
    metrics_round_decimals: int = DEFAULT_METRICS_ROUND_DECIMALS,
) -> pd.DataFrame:
    matched = int(summary["matched_rows"])
    rows = [
        ("Spreadsheet rows", summary["spreadsheet_rows"]),
        ("Log rows", summary["log_rows"]),
        ("Matched rows", summary["matched_rows"]),
        ("Spreadsheet-only rows", summary["spreadsheet_only_rows"]),
        ("Log-only rows", summary["log_only_rows"]),
        ("Row alignment status", summary["row_alignment_status"]),
        ("Rounded comparison decimals", summary["round_decimals"]),
        ("Ground truth status", summary["ground_truth_status"]),
        ("Prediction status", summary["prediction_status"]),
        ("Ground truth T matches", f"{summary['gt_T_matches']}/{matched} ({pct(int(summary['gt_T_matches']), matched):.2f}%)"),
        ("Ground truth H matches", f"{summary['gt_H_matches']}/{matched} ({pct(int(summary['gt_H_matches']), matched):.2f}%)"),
        ("Prediction T matches", f"{summary['pred_T_matches']}/{matched} ({pct(int(summary['pred_T_matches']), matched):.2f}%)"),
        ("Prediction H matches", f"{summary['pred_H_matches']}/{matched} ({pct(int(summary['pred_H_matches']), matched):.2f}%)"),
        ("Max |Tin_true_py - Tin_true_log|", summary["gt_T_max_abs_diff"]),
        ("Max |Hin_true_py - Hin_true_log|", summary["gt_H_max_abs_diff"]),
        ("Max |Tin_pred_py - Tp_log|", summary["pred_T_max_abs_diff"]),
        ("Max |Hin_pred_py - Hp_log|", summary["pred_H_max_abs_diff"]),
    ]
    if metrics_summary is not None:
        rows.extend(
            [
                ("Latest metrics status", metrics_summary["metrics_status"]),
                ("Latest metrics rounded matches", f"{metrics_summary['metrics_matches']}/{metrics_summary['metrics_compared']}"),
                ("Latest metrics max abs diff", metrics_summary["metrics_max_abs_diff"]),
                ("Latest valid [HOUR] METRICS line", metrics_summary.get("metrics_line_no")),
            ]
        )
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def build_human_differences(merged: pd.DataFrame, round_decimals: int) -> pd.DataFrame:
    mism = merged[(merged["_merge"] != "both") | (merged["row_has_any_difference_rounded"])].copy()
    if mism.empty:
        return pd.DataFrame([
            {"Message": f"No differences were found after alignment and rounding to {round_decimals} decimals."}
        ])

    order_cols = ["datetime_end"]
    if "idx_log" in mism.columns:
        order_cols = ["idx_log", "datetime_end"]
    mism = mism.sort_values(order_cols)

    rows = []
    for _, row in mism.iterrows():
        rows.append(
            {
                "datetime_end": row.get("datetime_end"),
                "idx_log": row.get("idx_log"),
                "epoch": row.get("epoch"),
                "Row_Status": row.get("_merge"),
                "Tin_true_Python": row.get("Tin_true_py"),
                "Tin_true_Log": row.get("Tin_true_log"),
                "Tin_true_Log_Rounded": row.get("Tin_true_log_round"),
                "Tin_true_Diff": row.get("Tin_true_diff"),
                "Tin_true_Match_Rounded": row.get("match_Tin_true_rounded"),
                "Hin_true_Python": row.get("Hin_true_py"),
                "Hin_true_Log": row.get("Hin_true_log"),
                "Hin_true_Log_Rounded": row.get("Hin_true_log_round"),
                "Hin_true_Diff": row.get("Hin_true_diff"),
                "Hin_true_Match_Rounded": row.get("match_Hin_true_rounded"),
                "Tin_pred_Python": row.get("Tin_pred_py"),
                "Tp_Log": row.get("Tp_log"),
                "Tp_Log_Rounded": row.get("Tp_log_round"),
                "Tin_pred_Diff": row.get("Tin_pred_diff"),
                "Tin_pred_Match_Rounded": row.get("match_Tin_pred_rounded"),
                "Hin_pred_Python": row.get("Hin_pred_py"),
                "Hp_Log": row.get("Hp_log"),
                "Hp_Log_Rounded": row.get("Hp_log_round"),
                "Hin_pred_Diff": row.get("Hin_pred_diff"),
                "Hin_pred_Match_Rounded": row.get("match_Hin_pred_rounded"),
            }
        )
    return pd.DataFrame(rows)




def compare_latest_metrics(
    sheet_metrics_df: pd.DataFrame,
    log_metrics_df: pd.DataFrame,
    round_decimals: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, object]]:
    merged = pd.concat([sheet_metrics_df.reset_index(drop=True), log_metrics_df.reset_index(drop=True)], axis=1)

    metric_pairs = [
        ("MAE_sheet", "MAE_log", "MAE"),
        ("RMSE_sheet", "RMSE_log", "RMSE"),
        ("R2_sheet", "R2_log", "R2"),
        ("MAE_T_sheet", "MAE_T_log", "MAE_T"),
        ("RMSE_T_sheet", "RMSE_T_log", "RMSE_T"),
        ("R2_T_sheet", "R2_T_log", "R2_T"),
        ("MAE_H_sheet", "MAE_H_log", "MAE_H"),
        ("RMSE_H_sheet", "RMSE_H_log", "RMSE_H"),
        ("R2_H_sheet", "R2_H_log", "R2_H"),
    ]

    rows = []
    rounded_matches = 0
    max_abs_diff = 0.0
    for sheet_col, log_col, label in metric_pairs:
        sheet_val = float(merged.at[0, sheet_col]) if pd.notna(merged.at[0, sheet_col]) else float("nan")
        log_val = float(merged.at[0, log_col]) if pd.notna(merged.at[0, log_col]) else float("nan")
        diff = sheet_val - log_val if pd.notna(sheet_val) and pd.notna(log_val) else float("nan")
        rounded_match = (
            pd.notna(sheet_val)
            and pd.notna(log_val)
            and round(sheet_val, round_decimals) == round(log_val, round_decimals)
        )
        rounded_matches += int(bool(rounded_match))
        if pd.notna(diff):
            max_abs_diff = max(max_abs_diff, abs(float(diff)))
        rows.append(
            {
                "Metric": label,
                "Spreadsheet_Latest": sheet_val,
                "Log_Latest": log_val,
                "Diff": diff,
                "Spreadsheet_Rounded": round(sheet_val, round_decimals) if pd.notna(sheet_val) else sheet_val,
                "Log_Rounded": round(log_val, round_decimals) if pd.notna(log_val) else log_val,
                "Match_Rounded": rounded_match,
            }
        )

    detail_df = pd.DataFrame(rows)
    summary = {
        "metrics_compared": len(metric_pairs),
        "metrics_matches": rounded_matches,
        "metrics_status": "MATCH" if rounded_matches == len(metric_pairs) else "MISMATCH",
        "metrics_max_abs_diff": max_abs_diff,
        "metrics_line_no": int(log_metrics_df.at[0, "metrics_line_no"]) if "metrics_line_no" in log_metrics_df.columns and pd.notna(log_metrics_df.at[0, "metrics_line_no"]) else None,
        "N_sheet": float(sheet_metrics_df.at[0, "N_sheet"]) if pd.notna(sheet_metrics_df.at[0, "N_sheet"]) else float("nan"),
    }
    overview_df = pd.DataFrame(
        [
            {
                "Block": f"Latest valid rolling24 metrics at {round_decimals} decimals",
                "Spreadsheet_Rows": 1,
                "Log_Rows": 1,
                "Matched_Rows": f"{rounded_matches}/{len(metric_pairs)}",
                "Spreadsheet_Only_Rows": "-",
                "Log_Only_Rows": "-",
                "Status": summary["metrics_status"],
            }
        ]
    )
    return detail_df, overview_df, summary

# ========================= Excel style helpers =========================
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
BAD_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
INFO_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")

THIN_GRAY = Side(style="thin", color="C9C9C9")
BORDER_THIN = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
TITLE_FONT = Font(bold=True, size=15)
SUBTITLE_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(bold=True)


def autosize(ws, min_width: int = 10, max_width: int = 46) -> None:
    for col_cells in ws.columns:
        idx = col_cells[0].column
        values = ["" if c.value is None else str(c.value) for c in col_cells]
        width = min(max(len(v) for v in values) + 2, max_width)
        width = max(min_width, width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def apply_status_fill(cell) -> None:
    if cell.value == "MATCH":
        cell.fill = GOOD_FILL
    elif cell.value == "MISMATCH":
        cell.fill = BAD_FILL
    elif cell.value == "INCOMPLETE":
        cell.fill = WARN_FILL


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(start_row, j, col)
    style_header(ws[start_row])

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            cell = ws.cell(i, j, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER_THIN
            header_name = str(ws.cell(start_row, j).value)
            if header_name == "Status":
                apply_status_fill(cell)

    ws.freeze_panes = ws.cell(start_row + 1, start_col)
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def fill_row(ws, row_idx: int, start_col: int, end_col: int, fill: PatternFill) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row_idx, col_idx).fill = fill


def add_executive_summary_sheet(
    wb: Workbook,
    summary: Dict[str, object],
    metrics_summary: Dict[str, object],
    pred_path: Path,
    metrics_xlsx_path: Path,
    log_path: Path,
    out_xlsx_path: Path,
    latest_run_dir: Path | None,
    excel_row_start: int,
    excel_row_end: int,
    metrics_round_decimals: int,
) -> None:
    ws = wb.active
    ws.title = "Executive_Summary"

    ws["A1"] = "Rolling24 Predictions vs Firmware Replay Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "MODEL: CONV1D_TINY -> Python vs Firmware Comparison"
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = "Scope"
    ws["B3"] = (
        "Compares the final prediction spreadsheet block against DBG_REPLAY_CSV from the firmware log. "
        "Spreadsheet rows are selected by Excel row number and aligned to the log by datetime_end."
    )

    ws["A5"] = "Prediction block rows"
    ws["B5"] = f"Excel rows {excel_row_start}:{excel_row_end}"

    ws["A6"] = "Row alignment status"
    ws["B6"] = summary["row_alignment_status"]

    ws["A7"] = "Ground truth status"
    ws["B7"] = summary["ground_truth_status"]

    ws["A8"] = "Prediction status"
    ws["B8"] = summary["prediction_status"]

    ws["A9"] = "Latest metrics status"
    ws["B9"] = metrics_summary["metrics_status"]

    ws["A10"] = "Ground truth rounded matches"
    ws["B10"] = (
        f"T={summary['gt_T_matches']}/{summary['matched_rows']} | "
        f"H={summary['gt_H_matches']}/{summary['matched_rows']}"
    )

    ws["A11"] = "Prediction rounded matches"
    ws["B11"] = (
        f"T={summary['pred_T_matches']}/{summary['matched_rows']} | "
        f"H={summary['pred_H_matches']}/{summary['matched_rows']}"
    )

    ws["A12"] = "Prediction rounded comparison precision"
    ws["B12"] = f"{summary['round_decimals']} decimals"

    ws["A13"] = "Latest metrics rounded comparison precision"
    ws["B13"] = f"{metrics_round_decimals} decimals"

    ws["A14"] = "Latest metrics rounded matches"
    ws["B14"] = f"{metrics_summary['metrics_matches']}/{metrics_summary['metrics_compared']}"

    ws["A15"] = "Largest raw |T_in_pred - Tp|"
    ws["B15"] = summary["pred_T_max_abs_diff"]

    ws["A16"] = "Largest raw |H_in_pred - Hp|"
    ws["B16"] = summary["pred_H_max_abs_diff"]

    ws["A17"] = "Duplicate-key alignment"
    ws["B17"] = summary.get("duplicate_key_alignment_applied", False)

    ws["A18"] = "Data sources"
    ws["A18"].font = SUBTITLE_FONT

    ws["A19"] = "Prediction spreadsheet"
    ws["B19"] = str(pred_path)

    ws["A20"] = "Metrics spreadsheet"
    ws["B20"] = str(metrics_xlsx_path)

    ws["A21"] = "Log / zip"
    ws["B21"] = str(log_path)

    ws["A22"] = "Comparison workbook"
    ws["B22"] = str(out_xlsx_path)

    ws["A23"] = "Selected run directory"
    ws["B23"] = str(latest_run_dir) if latest_run_dir else "(not used; explicit file path mode)"

    ws["A24"] = "Interpretation"
    ws["A24"].font = SUBTITLE_FONT
    ws["A25"] = "Rounded comparison note"
    ws["B25"] = (
        "Prediction values are compared using the configured prediction rounding precision, while rolling24 metrics are compared using the configured metrics rounding precision."
    )

    for row in range(3, 27):
        ws[f"A{row}"].font = LABEL_FONT if row not in {18, 24} else SUBTITLE_FONT
        ws[f"A{row}"].border = BORDER_THIN
        ws[f"B{row}"].border = BORDER_THIN
        ws[f"A{row}"].alignment = Alignment(vertical="top", wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    ws["A3"].fill = SECTION_FILL
    ws["A18"].fill = SECTION_FILL
    ws["A25"].fill = SECTION_FILL
    fill_row(ws, 26, 1, 2, INFO_FILL)
    apply_status_fill(ws["B6"])
    apply_status_fill(ws["B7"])
    apply_status_fill(ws["B8"])
    apply_status_fill(ws["B9"])
    ws.freeze_panes = "A4"
    autosize(ws, min_width=18, max_width=95)


def add_field_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Field_Guide")
    ws["A1"] = "Workbook Field Guide"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Section", "Sheet / Field", "Meaning", "How to interpret"),
        ("Workbook tabs", "Executive_Summary", "High-level decision sheet for final prediction agreement.", "Read this first to understand row alignment, ground-truth agreement, prediction agreement, and data sources."),
        ("Workbook tabs", "Overview", "Compact view of row alignment, ground-truth agreement, and prediction agreement.", "Use it to confirm coverage and the global status before reading row-level differences."),
        ("Workbook tabs", "Summary", "Detailed metric list for match counts and largest raw absolute differences.", "Use it when you need the exact counts and maxima reported by the comparison."),
        ("Workbook tabs", "Aligned_Data", "Python prediction block and firmware replay rows aligned by datetime_end.", "This is the main evidence table for row-by-row inspection after alignment."),
        ("Workbook tabs", "Differences", "Filtered view of only rows with mismatch or incomplete alignment.", "Open this when Overview or Summary reports a mismatch."),
        ("Workbook tabs", "Latest_Metrics", "Compares the latest valid rolling24 metrics from the spreadsheet against the latest valid [HOUR] METRICS line from the firmware log.", "Use it to confirm final MAE/RMSE/R2 agreement between the latest spreadsheet metrics and the latest valid firmware metrics."),

        ("Common fields", "Row alignment status", "Whether spreadsheet rows and firmware replay rows aligned 1:1 by datetime_end.", "MATCH means coverage is complete. INCOMPLETE means some rows exist only on one side."),
        ("Common fields", "Ground truth status", "Agreement status for ground-truth values after the configured rounding precision.", "This confirms whether the spreadsheet and the replay log describe the same reference target values."),
        ("Common fields", "Prediction status", "Agreement status for final predictions after the configured rounding precision.", "This is the final display-level agreement check between spreadsheet and firmware replay."),
        ("Common fields", "Prediction rounded comparison precision", "Number of decimals used for prediction display-level agreement.", "The workbook still shows raw differences separately; rounded match is for user-facing prediction agreement."),
        ("Common fields", "Latest metrics rounded comparison precision", "Number of decimals used for the latest metrics comparison.", "This is separate from prediction rounding and defaults to 4 decimals."),
        ("Common fields", "_merge / Row_Status", "Join status between spreadsheet rows and firmware replay rows.", "both = aligned row exists on both sides. left_only = only spreadsheet. right_only = only log."),
        ("Common fields", "<field>_Diff", "Raw spreadsheet value minus raw firmware replay value.", "Magnitude shows severity; sign shows direction of bias."),
        ("Common fields", "<field>_Match_Rounded", "Boolean flag for equality after rounding to the configured precision.", "True means display-level agreement, even if a tiny raw difference still exists."),

        ("Alignment keys", "datetime_end", "Main alignment key used to join spreadsheet predictions and firmware replay rows.", "This should identify the same logical final prediction timestamp on both sides."),
        ("Alignment keys", "sheet_order", "Original row order inside the selected spreadsheet block.", "Use it to recover the original spreadsheet sequence after alignment."),
        ("Alignment keys", "idx_log", "Firmware replay index from DBG_REPLAY_CSV.", "Useful for tracing the same final prediction back to the firmware log."),
        ("Alignment keys", "epoch", "Firmware replay epoch timestamp in seconds.", "This is the raw timestamp underlying datetime_end in the replay log."),

        ("Ground truth fields", "Tin_true_py / Hin_true_py", "Ground-truth values read from the spreadsheet block.", "These come from the Python rolling24 prediction sheet."),
        ("Ground truth fields", "Tin_true_log / Hin_true_log", "Ground-truth values read from DBG_REPLAY_CSV.", "Use them to verify that the spreadsheet and firmware replay refer to the same target values."),
        ("Ground truth fields", "Tin_true_log_round / Hin_true_log_round", "Rounded firmware ground-truth values used for display-level match.", "Compare them against the spreadsheet values rounded the same way."),

        ("Prediction fields", "Tin_pred_py / Hin_pred_py", "Final predictions read from the spreadsheet block.", "These are the Python-side final values shown to the user."),
        ("Prediction fields", "Tp_log / Hp_log", "Final predictions read from DBG_REPLAY_CSV in the firmware log.", "These are the firmware-side final absolute predictions."),
        ("Prediction fields", "Tp_log_round / Hp_log_round", "Rounded firmware final predictions used for display-level match.", "Compare them against the displayed spreadsheet values."),

        ("Schema compatibility", "Tin_true / Hin_true / Tp / Hp", "Legacy replay field names accepted by the script.", "The script also accepts the newer semantic names gt_Tin_true / gt_Hin_true / p_T_pred / p_H_pred."),
    ]

    for r, row in enumerate(rows, start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)

    style_header(ws[3])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, min_width=16, max_width=80)


# ========================= Main =========================
def main() -> None:
    parser = argparse.ArgumentParser(description="Compare final rolling24 predictions against DBG_REPLAY_CSV from firmware log.")
    parser.add_argument("--pred-xlsx", dest="pred_xlsx", help="Path to environment_quantized_predictions_rolling24_Conv1D_Tiny.xlsx")
    parser.add_argument("--log", dest="log_path", help="Path to device-monitor .log/.txt or .zip")
    parser.add_argument("--metrics-xlsx", dest="metrics_xlsx", help="Path to environment_quantized_metrics_rolling24_Conv1D_Tiny.xlsx")
    parser.add_argument("--out", dest="out_path", help="Path to output .xlsx workbook")
    parser.add_argument("--excel-row-start", type=int, default=DEFAULT_EXCEL_ROW_START, help="First Excel row (inclusive) to read from spreadsheet")
    parser.add_argument("--excel-row-end", type=int, default=DEFAULT_EXCEL_ROW_END, help="Last Excel row (inclusive) to read from spreadsheet")
    parser.add_argument("--round-decimals", type=int, default=DEFAULT_ROUND_DECIMALS, help="Decimals used for spreadsheet/display-level prediction match")
    parser.add_argument("--metrics-round-decimals", type=int, default=DEFAULT_METRICS_ROUND_DECIMALS, help="Decimals used for latest metrics comparison")
    args = parser.parse_args()

    pred_path, log_path, out_xlsx_path, latest_run_dir, metrics_xlsx_path = resolve_runtime_paths(
        args.pred_xlsx,
        args.log_path,
        args.out_path,
        args.metrics_xlsx,
    )

    print("[INFO] Comparing final rolling24 predictions against firmware replay log...")
    print(f"[INFO] pred_xlsx    : {pred_path}")
    print(f"[INFO] metrics_xlsx : {metrics_xlsx_path}")
    print(f"[INFO] log_path     : {log_path}")
    print(f"[INFO] out_xlsx     : {out_xlsx_path}")
    print(f"[INFO] excel rows: {args.excel_row_start}:{args.excel_row_end}")
    print(f"[INFO] pred round dp    : {args.round_decimals}")
    print(f"[INFO] metrics round dp : {args.metrics_round_decimals}")

    py_df = read_prediction_block(pred_path, args.excel_row_start, args.excel_row_end)
    metrics_sheet_df = read_metrics_latest(metrics_xlsx_path)
    log_text = read_text_from_log_or_zip(log_path)
    log_df = prepare_replay_log(log_text, log_path)
    log_metrics_df = prepare_hour_metrics_latest(log_text, log_path)

    merged, summary = compare_predictions(py_df, log_df, args.round_decimals)
    overview_df = build_overview_df(summary)
    detailed_df = build_human_differences(merged, args.round_decimals)
    metrics_detail_df, metrics_overview_df, metrics_summary = compare_latest_metrics(
        metrics_sheet_df,
        log_metrics_df,
        args.metrics_round_decimals,
    )
    summary_df = build_summary_df(summary, metrics_summary, args.metrics_round_decimals)

    wb = Workbook()
    add_executive_summary_sheet(
        wb,
        summary,
        metrics_summary,
        pred_path,
        metrics_xlsx_path,
        log_path,
        out_xlsx_path,
        latest_run_dir,
        args.excel_row_start,
        args.excel_row_end,
        args.metrics_round_decimals,
    )

    ws = wb.create_sheet("Overview")
    write_dataframe(ws, pd.concat([overview_df, metrics_overview_df], ignore_index=True))

    ws = wb.create_sheet("Summary")
    write_dataframe(ws, summary_df)

    ws = wb.create_sheet("Aligned_Data")
    aligned_cols = [
        "datetime_end", "sheet_order", "idx_log", "epoch", "_merge",
        "Tin_true_py", "Tin_true_log", "Tin_true_log_round", "Tin_true_diff", "match_Tin_true_rounded",
        "Hin_true_py", "Hin_true_log", "Hin_true_log_round", "Hin_true_diff", "match_Hin_true_rounded",
        "Tin_pred_py", "Tp_log", "Tp_log_round", "Tin_pred_diff", "match_Tin_pred_rounded",
        "Hin_pred_py", "Hp_log", "Hp_log_round", "Hin_pred_diff", "match_Hin_pred_rounded",
    ]
    aligned_present = [c for c in aligned_cols if c in merged.columns]
    write_dataframe(ws, merged[aligned_present].copy())

    ws = wb.create_sheet("Differences")
    write_dataframe(ws, detailed_df)

    ws = wb.create_sheet("Latest_Metrics")
    write_dataframe(ws, metrics_detail_df)

    add_field_guide_sheet(wb)

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx_path)
    print(f"[OK] Saved workbook: {out_xlsx_path}")


if __name__ == "__main__":
    main()
